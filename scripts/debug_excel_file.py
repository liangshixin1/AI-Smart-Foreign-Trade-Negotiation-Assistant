#!/usr/bin/env python3
"""调试Excel文件内容"""

import sys
import openpyxl

def debug_excel(file_path):
    """调试Excel文件的详细内容"""
    print("=" * 80)
    print(f"调试Excel文件: {file_path}")
    print("=" * 80)
    print()

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        print(f"Sheet列表: {wb.sheetnames}")
        print()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print("=" * 80)
            print(f"Sheet: {sheet_name}")
            print("=" * 80)
            print(f"最大行: {ws.max_row}")
            print(f"最大列: {ws.max_column}")
            print()

            # 读取所有行
            row_count = 0
            empty_count = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                if any(row):
                    row_count += 1
                    # 显示前5列
                    display_row = [str(cell)[:30] if cell is not None else "" for cell in row[:5]]
                    print(f"  行{row_idx}: {display_row}")
                else:
                    empty_count += 1
                    if empty_count <= 3:
                        print(f"  行{row_idx}: (空行)")

            print()
            print(f"总结: {row_count} 行有数据, {empty_count} 行为空")
            print()

        wb.close()

    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python debug_excel_file.py <excel文件路径>")
        sys.exit(1)

    debug_excel(sys.argv[1])
