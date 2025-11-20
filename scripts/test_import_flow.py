#!/usr/bin/env python3
"""测试三表导入流程 - 诊断工具"""

import os
import sys

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

def test_import_flow():
    """测试三表导入的完整流程"""
    print("=" * 80)
    print("三表导入诊断工具")
    print("=" * 80)
    print()

    # 1. 生成模板
    print("步骤 1: 生成Excel模板...")
    try:
        from services.knowledge_graph_batch_importer import generate_smart_templates
        excel_bytes = generate_smart_templates()
        print(f"✅ 模板生成成功，大小: {len(excel_bytes)} 字节")
    except Exception as e:
        print(f"❌ 模板生成失败: {e}")
        return

    # 2. 解析模板（测试示例数据）
    print("\n步骤 2: 解析模板中的示例数据...")
    try:
        from io import BytesIO
        from services.knowledge_graph_batch_importer import KnowledgeGraphBatchImporter
        from services.graph_service import GraphService

        importer = KnowledgeGraphBatchImporter(GraphService())

        # 解析流程表
        flow_data, flow_errors = importer._parse_flow_table(BytesIO(excel_bytes))
        print(f"   流程数据解析结果:")
        print(f"   - 解析到 {len(flow_data)} 个阶段")
        print(f"   - 错误数: {len(flow_errors)}")

        if flow_data:
            print(f"   - 阶段列表:")
            for stage in flow_data:
                print(f"     • {stage.get('name')} (顺序: {stage.get('_order')})")
        else:
            print(f"   ⚠️  警告：未解析到任何阶段！")

        if flow_errors:
            print(f"   - 错误详情:")
            for err in flow_errors:
                print(f"     • [{err.severity}] {err.message}")

        # 解析知识点表
        points_data, points_errors = importer._parse_points_table_from_workbook(
            BytesIO(excel_bytes), sheet_name="知识点主表"
        )
        print(f"\n   知识点数据解析结果:")
        print(f"   - 解析到 {len(points_data)} 个知识点")
        print(f"   - 错误数: {len(points_errors)}")

        # 解析案例表
        examples_data, examples_errors = importer._parse_examples_table_from_workbook(
            BytesIO(excel_bytes), sheet_name="案例库表"
        )
        print(f"\n   案例数据解析结果:")
        print(f"   - 解析到 {len(examples_data)} 个案例")
        print(f"   - 错误数: {len(examples_errors)}")

    except Exception as e:
        print(f"❌ 解析失败: {e}")
        import traceback
        traceback.print_exc()
        return

    # 3. 检查Excel文件结构
    print("\n步骤 3: 检查Excel文件结构...")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True)
        print(f"   Sheet列表: {wb.sheetnames}")

        # 检查谈判流程表
        if "谈判流程" in wb.sheetnames:
            ws = wb["谈判流程"]
            print(f"\n   '谈判流程' Sheet:")
            print(f"   - 最大行数: {ws.max_row}")
            print(f"   - 最大列数: {ws.max_column}")

            # 读取前5行
            print(f"   - 前5行数据预览:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                print(f"     行{row_idx}: {row}")

        wb.close()
    except Exception as e:
        print(f"❌ 检查文件结构失败: {e}")
        import traceback
        traceback.print_exc()

    # 4. 检查Neo4j连接
    print("\n步骤 4: 检查Neo4j连接...")
    try:
        from services import graph_service

        if graph_service.is_configured():
            print("   ✅ Neo4j已配置")

            # 尝试连接
            driver = graph_service._get_driver()
            driver.verify_connectivity()
            print("   ✅ Neo4j连接正常")

            # 查询现有Stage节点
            with driver.session() as session:
                result = session.run("MATCH (s:Stage) RETURN count(s) as count")
                stage_count = result.single()["count"]
                print(f"   - 现有Stage节点数: {stage_count}")

                if stage_count > 0:
                    result = session.run("MATCH (s:Stage) RETURN s.name as name LIMIT 10")
                    print(f"   - Stage节点列表:")
                    for record in result:
                        print(f"     • {record['name']}")
        else:
            print("   ❌ Neo4j未配置")

    except Exception as e:
        print(f"   ⚠️  Neo4j连接失败: {e}")

    print("\n" + "=" * 80)
    print("诊断完成")
    print("=" * 80)


if __name__ == "__main__":
    test_import_flow()
