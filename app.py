import copy
import json
import os
import uuid
from itertools import chain
from typing import Dict, Iterable, List, Optional, Tuple

from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    jsonify,
    request,
    send_from_directory,
    stream_with_context,
)
from openai import OpenAI
from openpyxl import load_workbook

import database
from levels import CHAPTERS, flatten_scenario_for_template

load_dotenv()
database.init_database()
database.seed_default_levels(CHAPTERS)

DEEPSEEK_BASE = "https://api.deepseek.com"
MODEL = "deepseek-chat"

app = Flask(__name__, static_folder="static")

DEFAULT_DIFFICULTY = "balanced"
DIFFICULTY_PROFILES: Dict[str, Dict[str, str]] = {
    "friendly": {
        "label": "友好型 · 引导与鼓励",
        "description": "语气温暖，适度让步以支持学生梳理思路并建立自信。",
        "prompt_suffix": (
            "你是一位友好型的谈判对手，会主动给予积极反馈、概括要点，"
            "并在学生出现疏漏时给出提醒或示范句型。适度分享行业见解，鼓励学生提出更多澄清问题，"
            "在价格或条款上可在原底线上最多让步约 5% 以换取长期合作。"
        ),
        "tone_hint": "语气更温暖、积极，主动给予肯定与指导。",
        "bottom_line_hint": "可在原有底线上额外让步约 5%，强调合作诚意。",
    },
    "balanced": {
        "label": "默认 · 平衡博弈",
        "description": "保持专业礼貌，兼顾自身立场与合作机会。",
        "prompt_suffix": "",
        "tone_hint": "",
        "bottom_line_hint": "",
    },
    "tough": {
        "label": "强硬型 · 严守底线",
        "description": "语气坚定谨慎，强调风险控制与公司底线。",
        "prompt_suffix": (
            "你是一位强硬型的谈判对手，强调风控与底线意识。"
            "当学生尝试议价时，请要求其提供充分理由或额外让步，"
            "并重申关键条款的重要性。只有在获得确凿价值回报时才考虑微量让步。"
        ),
        "tone_hint": "语气更为坚定，明确指出风险与不可退让的条件。",
        "bottom_line_hint": "底线不可轻易突破，除非学生提供充分价值交换。",
    },
    "shrewd": {
        "label": "精明型 · 灵活试探",
        "description": "善于条件交换与试探，关注整体收益最大化。",
        "prompt_suffix": (
            "你是一位精明型的谈判对手，善于抛出条件交换并观察学生反应。"
            "请通过试探问题与设定多种方案，引导学生思考让步条件，"
            "在关键价格或条款上保持敏锐并要求对等回报。"
        ),
        "tone_hint": "语气务实敏锐，喜欢提出条件交换与方案比较。",
        "bottom_line_hint": "可根据学生的回报方案灵活调整底线范围。",
    },
}

SCENARIO_DIVERSITY_HINT = (
    "请在设计谈判情境时兼顾制造业、服务业、数字贸易、农业、电子产品业、汽车业、文化创意产业等多元行业，"
    "避免始终聚焦于一个产品、一个案例，使学生能够接触不同的外贸品类。"
)

ROLE_ENFORCEMENT_HINT = (
    "学生在本场景中必须明确扮演来自中国的买家或卖家，可根据任务设置选择进口商或出口商。"
    "请确保 student_role 字段中包含‘中国’字样，并给出行业、职位描述。"
)

CONVERSATION_DIVERSITY_HINT = (
    "在与学生的每轮对话中，选择贴合场景的行业背景示例，可结合原材料、工业品、"
    "生活消费品、服务解决方案等不同类型，刻意避免反复引用电子产品为例。"
)

ENGLISH_ENFORCEMENT_HINT = (
    "All assistant-facing outputs, including scenario briefings and conversation replies, must be written entirely in English."
    " Avoid inserting Chinese characters unless the student explicitly provides them or requests bilingual content."
)


def _get_difficulty_profile(key: Optional[str]) -> Dict[str, str]:
    normalized = str(key or DEFAULT_DIFFICULTY)
    return DIFFICULTY_PROFILES.get(normalized, DIFFICULTY_PROFILES[DEFAULT_DIFFICULTY])


def _apply_difficulty_profile(
    scenario: Dict[str, object], difficulty_key: str
) -> Tuple[Dict[str, object], Dict[str, str]]:
    profile = _get_difficulty_profile(difficulty_key)
    scenario_copy: Dict[str, object] = copy.deepcopy(scenario)
    scenario_copy["difficulty_key"] = difficulty_key
    scenario_copy["difficulty_label"] = profile["label"]
    scenario_copy["difficulty_description"] = profile["description"]

    tone_hint = profile.get("tone_hint")
    if tone_hint:
        base_tone = scenario_copy.get("communication_tone", "") or ""
        if tone_hint not in base_tone:
            scenario_copy["communication_tone"] = (
                f"{base_tone}（{tone_hint}）" if base_tone else tone_hint
            )

    product = scenario_copy.get("product")
    if isinstance(product, dict):
        price_expectation = product.get("price_expectation")
        if isinstance(price_expectation, dict):
            hint = profile.get("bottom_line_hint")
            if hint:
                ai_bottom_line = price_expectation.get("ai_bottom_line")
                if isinstance(ai_bottom_line, str) and ai_bottom_line.strip():
                    if hint not in ai_bottom_line:
                        price_expectation["ai_bottom_line"] = f"{ai_bottom_line}（{hint}）"
                else:
                    price_expectation["ai_bottom_line"] = hint

    return scenario_copy, profile


def _inject_difficulty_metadata(item: Dict[str, object]) -> None:
    difficulty_key = str(item.get("difficulty") or DEFAULT_DIFFICULTY)
    profile = _get_difficulty_profile(difficulty_key)
    item["difficulty"] = difficulty_key
    item["difficultyLabel"] = profile["label"]
    if "difficultyDescription" not in item or not item["difficultyDescription"]:
        item["difficultyDescription"] = profile["description"]


def _ensure_level_hierarchy(include_prompts: bool = False) -> List[Dict[str, object]]:
    """Load level hierarchy, reseeding defaults if database is empty."""

    chapters = database.list_level_hierarchy(include_prompts=include_prompts)
    if chapters:
        return chapters

    database.seed_default_levels(CHAPTERS)
    return database.list_level_hierarchy(include_prompts=include_prompts)


class MissingKeyError(RuntimeError):
    pass


def _require_key(env_name: str) -> str:
    key = os.getenv(env_name)
    if not key:
        raise MissingKeyError(f"Missing environment variable: {env_name}")
    return key


def _create_client(api_key: str) -> OpenAI:
    return OpenAI(api_key=api_key, base_url=DEEPSEEK_BASE)


def _complete_chat(api_key: str, messages: List[Dict[str, str]], temperature: float = 0.7) -> str:
    client = _create_client(api_key)
    response = client.chat.completions.create(
        model=MODEL,
        messages=messages,
        temperature=temperature,
    )
    if not response.choices:
        raise RuntimeError("Empty response from chat completion API")
    return response.choices[0].message.content or ""


def _stream_chat(api_key: str, messages: List[Dict[str, str]], temperature: float = 0.7):
    client = _create_client(api_key)
    stream = client.chat.completions.create(
        model=MODEL,
        messages=messages,
        temperature=temperature,
        stream=True,
    )
    for chunk in stream:
        for choice in chunk.choices or []:
            delta = getattr(choice, "delta", None)
            if delta and getattr(delta, "content", None):
                yield delta.content


def _extract_json_block(text: str) -> Dict[str, object]:
    cleaned = text.strip().strip("`")
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("JSON block not found in response")
    json_string = cleaned[start : end + 1]
    return json.loads(json_string)


def _prepare_scenario_payload(raw: Dict[str, object]) -> Dict[str, object]:
    difficulty_key = (raw.get("difficulty_key") or DEFAULT_DIFFICULTY)
    profile = _get_difficulty_profile(difficulty_key)
    return {
        "title": raw.get("scenario_title", ""),
        "summary": raw.get("scenario_summary", ""),
        "studentRole": raw.get("student_role", ""),
        "studentCompany": raw.get("student_company", {}) or {},
        "aiRole": raw.get("ai_role", ""),
        "aiCompany": raw.get("ai_company", {}) or {},
        "aiRules": raw.get("ai_rules", []) or [],
        "product": raw.get("product", {}) or {},
        "marketLandscape": raw.get("market_landscape", ""),
        "timeline": raw.get("timeline", ""),
        "logistics": raw.get("logistics", ""),
        "risks": raw.get("risks", []) or [],
        "negotiationTargets": raw.get("negotiation_targets", []) or [],
        "communicationTone": raw.get("communication_tone", ""),
        "checklist": raw.get("checklist", []) or [],
        "knowledgePoints": raw.get("knowledge_points", []) or [],
        "difficulty": difficulty_key,
        "difficultyLabel": raw.get("difficulty_label") or profile["label"],
        "difficultyDescription": raw.get("difficulty_description")
        or profile["description"],
    }


def _infer_student_trade_role(section: Dict[str, object]) -> str:
    text_segments: List[str] = []
    for key in ("description", "environment_user_message", "title"):
        value = section.get(key)
        if isinstance(value, str):
            text_segments.append(value)
    text_blob = " ".join(text_segments)
    lowered = text_blob.lower()
    if any(keyword in text_blob for keyword in ("卖家", "出口", "供货", "供應")):
        return "seller"
    if any(keyword in lowered for keyword in ("sell", "export")):
        return "seller"
    return "buyer"


def _ensure_chinese_student_role(scenario: Dict[str, object], trade_role: str) -> None:
    student_role = scenario.get("student_role")
    normalized = student_role.strip() if isinstance(student_role, str) else ""
    if "中国" not in normalized:
        normalized = f"中国{normalized}" if normalized else "中国外贸业务代表"

    if trade_role == "seller":
        if not any(keyword in normalized for keyword in ("卖", "出口", "供货", "供应")):
            normalized = f"中国卖家代表（{normalized}）"
    else:
        if not any(keyword in normalized for keyword in ("买", "采购", "进口")):
            normalized = f"中国买家代表（{normalized}）"

    scenario["student_role"] = normalized


def _render_prompts_from_section(
    section: Dict[str, object],
    scenario: Dict[str, object],
    difficulty_key: str,
    difficulty_profile: Dict[str, str],
) -> Tuple[str, str]:
    flat_context = flatten_scenario_for_template(scenario)
    if not flat_context.get("knowledge_points_hint"):
        flat_context["knowledge_points_hint"] = (
            "報盤結構, 議價策略, 跨文化溝通"
            if section.get("expects_bargaining")
            else "英文商務函電寫作, 信息提取, 跨文化表達"
        )

    conversation_prompt = section["conversation_prompt_template"].format_map(
        flat_context
    )
    prompt_suffix = difficulty_profile.get("prompt_suffix")
    if prompt_suffix:
        conversation_prompt = f"{conversation_prompt}\n\n[難度設定]\n{prompt_suffix}"
    if CONVERSATION_DIVERSITY_HINT not in conversation_prompt:
        conversation_prompt = (
            f"{conversation_prompt}\n\n[案例多样性提醒]\n{CONVERSATION_DIVERSITY_HINT}"
        )
    if ROLE_ENFORCEMENT_HINT not in conversation_prompt:
        conversation_prompt = (
            f"{conversation_prompt}\n\n[角色约束]\n请始终以学生为中国买家或中国卖家来组织对话，"
            "在回应中适时引用中国市场或供应链视角。"
        )
    if ENGLISH_ENFORCEMENT_HINT not in conversation_prompt:
        conversation_prompt = (
            f"{conversation_prompt}\n\n[Language Requirement]\n{ENGLISH_ENFORCEMENT_HINT}"
        )

    evaluation_prompt = section["evaluation_prompt_template"].format_map(flat_context)
    return conversation_prompt, evaluation_prompt


def _normalize_text(value: object) -> str:
    if isinstance(value, str):
        return value.strip()
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()
    return str(value).strip() if value is not None else ""


def _normalize_text_list(value: object) -> List[str]:
    if isinstance(value, list):
        result: List[str] = []
        for item in value:
            text = _normalize_text(item)
            if text:
                result.append(text)
        return result
    if isinstance(value, str):
        items = [segment.strip() for segment in value.splitlines()]
        return [item for item in items if item]
    return []


def _normalize_company(data: object) -> Dict[str, str]:
    if not isinstance(data, dict):
        return {}
    return {
        "name": _normalize_text(
            data.get("name")
            or data.get("company")
            or data.get("companyName")
            or data.get("display")
        ),
        "profile": _normalize_text(
            data.get("profile")
            or data.get("description")
            or data.get("summary")
        ),
    }


def _normalize_price_expectation(data: object) -> Dict[str, str]:
    if not isinstance(data, dict):
        return {}
    return {
        "student_target": _normalize_text(
            data.get("student_target")
            or data.get("studentTarget")
            or data.get("target")
            or data.get("student")
        ),
        "ai_bottom_line": _normalize_text(
            data.get("ai_bottom_line")
            or data.get("aiBottomLine")
            or data.get("bottomLine")
            or data.get("ai")
        ),
    }


def _normalize_product(data: object) -> Dict[str, object]:
    if not isinstance(data, dict):
        return {}
    result: Dict[str, object] = {
        "name": _normalize_text(data.get("name")),
        "specifications": _normalize_text(
            data.get("specifications")
            or data.get("specs")
            or data.get("features")
        ),
        "quantity_requirement": _normalize_text(
            data.get("quantity_requirement")
            or data.get("quantityRequirement")
            or data.get("quantity")
        ),
    }
    price_data = data.get("price_expectation") or data.get("priceExpectation")
    price = _normalize_price_expectation(price_data)
    if price:
        result["price_expectation"] = price
    if data.get("highlights"):
        result["highlights"] = _normalize_text_list(data.get("highlights"))
    return result


def _first_non_empty(mapping: Dict[str, object], keys: Iterable[str]) -> str:
    for key in keys:
        value = mapping.get(key)
        text = _normalize_text(value)
        if text:
            return text
    return ""


def _assemble_scenario_from_blueprint(
    blueprint: Dict[str, object], difficulty_key: str
) -> Tuple[Dict[str, object], Dict[str, str]]:
    scenario = {
        "scenario_title": _first_non_empty(
            blueprint, ["scenarioTitle", "title", "name"]
        ),
        "scenario_summary": _first_non_empty(
            blueprint, ["scenarioSummary", "summary", "description"]
        ),
        "student_role": _first_non_empty(
            blueprint, ["studentRole", "student_role", "student"]
        ),
        "student_company": _normalize_company(
            blueprint.get("studentCompany") or blueprint.get("student_company")
        ),
        "ai_role": _first_non_empty(
            blueprint, ["aiRole", "assistantRole", "ai_role"]
        ),
        "ai_company": _normalize_company(
            blueprint.get("aiCompany") or blueprint.get("ai_company")
        ),
        "ai_rules": _normalize_text_list(
            blueprint.get("aiRules") or blueprint.get("ai_rules")
        ),
        "product": _normalize_product(
            blueprint.get("product") or blueprint.get("productInfo")
        ),
        "market_landscape": _first_non_empty(
            blueprint, ["marketLandscape", "market_landscape", "market"]
        ),
        "timeline": _first_non_empty(
            blueprint, ["timeline", "delivery", "schedule"]
        ),
        "logistics": _first_non_empty(
            blueprint, ["logistics", "tradeTerms", "shipping"]
        ),
        "risks": _normalize_text_list(blueprint.get("risks")),
        "negotiation_targets": _normalize_text_list(
            blueprint.get("negotiationTargets") or blueprint.get("negotiation_targets")
        ),
        "communication_tone": _first_non_empty(
            blueprint, ["communicationTone", "communication_tone", "tone"]
        ),
        "checklist": _normalize_text_list(
            blueprint.get("checklist") or blueprint.get("taskChecklist")
        ),
        "knowledge_points": _normalize_text_list(
            blueprint.get("knowledgePoints") or blueprint.get("knowledge_points")
        ),
        "opening_message": _first_non_empty(
            blueprint, ["openingMessage", "opening_message", "opening"]
        ),
    }
    scenario, profile = _apply_difficulty_profile(scenario, difficulty_key)
    return scenario, profile


def _build_custom_assignment_prompts(
    scenario: Dict[str, object], difficulty_profile: Dict[str, str]
) -> Tuple[str, str]:
    flat = flatten_scenario_for_template(scenario)
    ai_rules = scenario.get("ai_rules", []) or []
    rules_block = "\n".join(f"- {rule}" for rule in ai_rules if isinstance(rule, str) and rule.strip())
    if not rules_block:
        rules_block = "- Maintain consistency with the scenario details and protect your company's interests."
    tone = flat.get("communication_tone") or "Professional and courteous business English"
    conversation_prompt = f"""
You are {flat.get('ai_role') or 'the supplier representative'} from {flat.get('ai_company_name') or 'the partner company'}.
The student is {flat.get('student_role') or 'a Chinese trade professional'} representing {flat.get('student_company_name') or 'their company'}.

Scenario briefing:
- Product focus: {flat.get('product_name') or 'N/A'} ({flat.get('product_specs') or 'specifications TBD'}).
- Quantity / capacity: {flat.get('product_quantity') or 'Discuss with the student'}.
- Market situation: {flat.get('market_landscape') or 'Use industry-relevant details'}.
- Logistics & timeline: {flat.get('logistics') or 'Negotiate feasible terms'}.
- Negotiation targets: {scenario.get('negotiation_targets') or []}.

Ground rules:
{rules_block}

Conduct the negotiation entirely in English. Adopt a tone that is {tone}. Guide the student to articulate clear proposals, ask clarifying questions, and explore win-win trade-offs. Reference real-world trade considerations whenever helpful.
""".strip()
    suffix = difficulty_profile.get("prompt_suffix")
    if suffix:
        conversation_prompt = f"{conversation_prompt}\n\n[Difficulty]\n{suffix}"
    if ENGLISH_ENFORCEMENT_HINT not in conversation_prompt:
        conversation_prompt = f"{conversation_prompt}\n\n[Language Requirement]\n{ENGLISH_ENFORCEMENT_HINT}"
    if ROLE_ENFORCEMENT_HINT not in conversation_prompt:
        conversation_prompt = f"{conversation_prompt}\n\n[Role Reminder]\n{ROLE_ENFORCEMENT_HINT}"

    knowledge_hint = "、".join(scenario.get("knowledge_points", []) or []) or "Negotiation strategy, Cross-cultural communication"
    evaluation_prompt = f"""
You are an experienced trade negotiation coach. Review the scenario summary and the dialogue transcript to evaluate the student's performance. Respond in JSON with:
{{
  "score": integer 0-100,
  "score_label": "Short label summarizing performance",
  "commentary": "Detailed Chinese feedback highlighting strengths and improvements",
  "action_items": ["3 concrete next steps"],
  "knowledge_points": ["Key knowledge points, prefer: {knowledge_hint}"],
  "bargaining_win_rate": "0-100 if bargaining outcome is discussed, else null"
}}

Focus on language quality, clarity of negotiation strategy, data support for proposals, and etiquette.
""".strip()
    return conversation_prompt, evaluation_prompt


def _serialize_blueprint(record: Dict[str, object]) -> Dict[str, object]:
    payload = {
        "id": record["id"],
        "title": record.get("title", ""),
        "description": record.get("description", ""),
        "difficulty": record.get("difficulty", DEFAULT_DIFFICULTY),
        "blueprint": record.get("blueprint", {}),
        "createdAt": record.get("createdAt"),
        "updatedAt": record.get("updatedAt"),
        "difficultyDescription": record.get("difficultyDescription", ""),
    }
    _inject_difficulty_metadata(payload)
    blueprint_data = payload.get("blueprint", {})
    if isinstance(blueprint_data, dict):
        payload["scenarioPreview"] = _prepare_scenario_payload(blueprint_data)
    else:
        payload["scenarioPreview"] = {}
    return payload


def _serialize_assignment(record: Dict[str, object]) -> Dict[str, object]:
    scenario_data = record.get("scenario", {}) or {}
    payload = {
        "id": record["id"],
        "title": record.get("title", ""),
        "description": record.get("description", ""),
        "difficulty": record.get("difficulty", DEFAULT_DIFFICULTY),
        "chapterId": record.get("chapterId"),
        "sectionId": record.get("sectionId"),
        "blueprintId": record.get("blueprintId"),
        "scenario": _prepare_scenario_payload(scenario_data),
        "createdAt": record.get("createdAt"),
        "updatedAt": record.get("updatedAt"),
        "dueAt": record.get("dueAt"),
    }
    if "assignedCount" in record:
        payload["assignedCount"] = record.get("assignedCount", 0)
    if "completedCount" in record:
        payload["completedCount"] = record.get("completedCount", 0)
    if "inProgressCount" in record:
        payload["inProgressCount"] = record.get("inProgressCount", 0)
    if "studentIds" in record and isinstance(record.get("studentIds"), list):
        payload["studentIds"] = record.get("studentIds")
    if "status" in record:
        payload["status"] = record.get("status")
    if "sessionId" in record:
        payload["sessionId"] = record.get("sessionId")
    if "submittedAt" in record:
        payload["submittedAt"] = record.get("submittedAt")
    _inject_difficulty_metadata(payload)
    return payload


def _normalize_student_header(value: object) -> str:
    text = _normalize_text(value).lower()
    if text in {"id", "账号", "學號", "学号", "user", "userid"}:
        return "id"
    if text in {"姓名", "name", "display", "nickname"}:
        return "name"
    if text in {"password", "密码", "pass", "pwd"}:
        return "password"
    return ""


def _parse_student_records(file_storage) -> List[Dict[str, str]]:
    workbook = load_workbook(file_storage, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        first_row = next(rows)
    except StopIteration:
        return []

    headers = [_normalize_student_header(cell) for cell in (first_row or [])]
    if not any(headers):
        headers = ["id", "name", "password"]
        rows = chain([first_row], rows)

    records: List[Dict[str, str]] = []
    for row in rows:
        if not row:
            continue
        entry: Dict[str, str] = {"id": "", "name": "", "password": ""}
        for index, cell in enumerate(row):
            if index >= len(headers):
                continue
            key = headers[index]
            if not key:
                continue
            entry[key] = _normalize_text(cell)
        entry["id"] = _normalize_text(entry["id"])
        entry["name"] = _normalize_text(entry["name"]) or entry["id"]
        entry["password"] = _normalize_text(entry["password"]) or entry["id"]
        if entry["id"] and entry["password"]:
            records.append(entry)
    return records


def _generate_scenario_for_section(
    section: Dict[str, object], difficulty_key: str
) -> Tuple[Dict[str, object], Dict[str, str]]:
    try:
        generator_key = _require_key("DEEPSEEK_GENERATOR_KEY")
    except MissingKeyError as exc:
        raise RuntimeError(str(exc)) from exc

    messages = [
        {"role": "system", "content": section["environment_prompt_template"]},
        {"role": "system", "content": SCENARIO_DIVERSITY_HINT},
        {"role": "system", "content": ROLE_ENFORCEMENT_HINT},
        {"role": "system", "content": ENGLISH_ENFORCEMENT_HINT},
        {"role": "user", "content": section["environment_user_message"]},
    ]

    raw_response = _complete_chat(generator_key, messages, temperature=0.85)
    scenario = _extract_json_block(raw_response)
    trade_role = _infer_student_trade_role(section)
    _ensure_chinese_student_role(scenario, trade_role)
    scenario, profile = _apply_difficulty_profile(scenario, difficulty_key)
    return scenario, profile


def _prepare_opening_message(
    conversation_prompt: str, scenario: Dict[str, object]
) -> str:
    """Derive the assistant's opening message based on the system prompt."""

    opening_message = ""
    try:
        collab_key = _require_key("DEEPSEEK_COLLAB_KEY")
    except MissingKeyError:
        collab_key = ""

    if collab_key and conversation_prompt.strip():
        try:
            opening_message = _complete_chat(
                collab_key,
                [{"role": "system", "content": conversation_prompt}],
                temperature=0.7,
            ).strip()
        except Exception:  # pragma: no cover - tolerate upstream variance
            opening_message = ""

    if not opening_message and isinstance(scenario, dict):
        raw_opening = scenario.get("opening_message")
        if isinstance(raw_opening, str):
            opening_message = raw_opening.strip()

    return opening_message


def _build_transcript(history: List[Dict[str, str]], scenario: Dict[str, object]) -> str:
    lines: List[str] = []
    lines.append(f"場景標題: {scenario.get('scenario_title', '')}")
    lines.append(f"場景摘要: {scenario.get('scenario_summary', '')}")
    lines.append(f"學生角色: {scenario.get('student_role', '')}")
    lines.append(f"AI 角色: {scenario.get('ai_role', '')}")
    lines.append(f"產品資訊: {json.dumps(scenario.get('product', {}), ensure_ascii=False)}")
    lines.append(f"市場與物流: {scenario.get('market_landscape', '')}；{scenario.get('logistics', '')}")
    lines.append("對話逐字稿：")

    ai_name = "AI"
    ai_company = scenario.get("ai_company", {}) or {}
    if isinstance(ai_company, dict):
        ai_company_name = ai_company.get("name")
        if isinstance(ai_company_name, str) and ai_company_name:
            ai_name = ai_company_name

    for message in history:
        role = message.get("role")
        content = message.get("content", "")
        if role == "user":
            speaker = "學生"
        elif role == "assistant":
            speaker = ai_name
        else:
            speaker = role or "系統"
        lines.append(f"{speaker}: {content}")
    return "\n".join(lines)


def _get_section(chapter_id: str, section_id: str) -> Dict[str, object]:
    section = database.get_section_template(chapter_id, section_id)
    if not section:
        raise KeyError("Invalid section id")
    return section


def _extract_token() -> Optional[str]:
    auth_header = request.headers.get("Authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header[7:].strip()
    token = request.headers.get("X-Auth-Token")
    if token:
        return token.strip()
    return None


def _as_bool(value: object, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        normalized = value.strip().lower()
        return normalized in {"1", "true", "yes", "on", "y"}
    return default


def _require_user(required_role: Optional[str] = None) -> Tuple[Optional[Dict[str, object]], Optional[Tuple[Dict[str, str], int]]]:
    token = _extract_token()
    if not token:
        return None, ({"error": "Authentication required"}, 401)
    user = database.get_user_by_token(token)
    if not user:
        return None, ({"error": "Invalid or expired token"}, 401)
    if required_role and user.get("role") != required_role:
        return None, ({"error": "Forbidden"}, 403)
    return user, None


@app.route("/")
def index() -> str:
    return send_from_directory(app.static_folder, "index.html")


@app.post("/api/login")
def login():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username or not password:
        return jsonify({"error": "username and password are required"}), 400

    user = database.authenticate_user(username, password)
    if not user:
        return jsonify({"error": "Invalid credentials"}), 401

    token = database.issue_auth_token(int(user["id"]))
    payload = {"token": token, "user": user}
    return jsonify(payload)


@app.get("/api/levels")
def list_levels():
    chapters = _ensure_level_hierarchy(include_prompts=False)
    return jsonify({"chapters": chapters})


@app.get("/api/blueprints")
def list_blueprints_endpoint():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    records = database.list_blueprints(int(user["id"]))
    payload = [_serialize_blueprint(record) for record in records]
    return jsonify({"blueprints": payload})


@app.post("/api/blueprints")
def create_blueprint_endpoint():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    blueprint_raw = data.get("blueprint") or {}
    if not isinstance(blueprint_raw, dict):
        return jsonify({"error": "blueprint must be an object"}), 400

    difficulty_key = str(
        data.get("difficulty") or blueprint_raw.get("difficulty") or DEFAULT_DIFFICULTY
    ).lower()
    if difficulty_key not in DIFFICULTY_PROFILES:
        difficulty_key = DEFAULT_DIFFICULTY

    scenario, profile = _assemble_scenario_from_blueprint(blueprint_raw, difficulty_key)
    title = _normalize_text(data.get("title")) or scenario.get("scenario_title") or "未命名关卡"
    description = _normalize_text(data.get("description")) or scenario.get("scenario_summary", "")

    record = database.create_blueprint(
        owner_id=int(user["id"]),
        title=title,
        description=description,
        difficulty=difficulty_key,
        blueprint=scenario,
    )
    record["difficultyDescription"] = profile.get("description")
    return jsonify({"blueprint": _serialize_blueprint(record)}), 201


@app.put("/api/blueprints/<blueprint_id>")
def update_blueprint_endpoint(blueprint_id: str):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    existing = database.get_blueprint(blueprint_id)
    if not existing or int(existing.get("ownerId")) != int(user["id"]):
        return jsonify({"error": "Blueprint not found"}), 404

    data = request.get_json(force=True)
    blueprint_raw = data.get("blueprint")
    difficulty_key = str(data.get("difficulty") or existing.get("difficulty") or DEFAULT_DIFFICULTY).lower()
    if difficulty_key not in DIFFICULTY_PROFILES:
        difficulty_key = DEFAULT_DIFFICULTY

    updates: Dict[str, object] = {
        "title": data.get("title"),
        "description": data.get("description"),
        "difficulty": difficulty_key,
    }
    scenario = None
    profile = _get_difficulty_profile(difficulty_key)
    if isinstance(blueprint_raw, dict):
        scenario, profile = _assemble_scenario_from_blueprint(blueprint_raw, difficulty_key)
        updates["blueprint"] = scenario

    updated = database.update_blueprint(
        blueprint_id,
        title=_normalize_text(updates.get("title")) if updates.get("title") is not None else None,
        description=_normalize_text(updates.get("description"))
        if updates.get("description") is not None
        else None,
        difficulty=difficulty_key,
        blueprint=scenario if scenario is not None else None,
    )
    if not updated:
        return jsonify({"error": "Blueprint not found"}), 404
    updated["difficultyDescription"] = profile.get("description")
    return jsonify({"blueprint": _serialize_blueprint(updated)})


@app.delete("/api/blueprints/<blueprint_id>")
def delete_blueprint_endpoint(blueprint_id: str):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    existing = database.get_blueprint(blueprint_id)
    if not existing or int(existing.get("ownerId")) != int(user["id"]):
        return jsonify({"error": "Blueprint not found"}), 404

    database.delete_blueprint(blueprint_id)
    return jsonify({"status": "deleted"})


@app.post("/api/generator/scenario")
def generator_scenario():
    user, error = _require_user()
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    chapter_id = data.get("chapterId")
    section_id = data.get("sectionId")
    difficulty_key = str(data.get("difficulty") or DEFAULT_DIFFICULTY).lower()
    if difficulty_key not in DIFFICULTY_PROFILES:
        difficulty_key = DEFAULT_DIFFICULTY

    if not chapter_id or not section_id:
        return jsonify({"error": "chapterId and sectionId are required"}), 400

    try:
        section = _get_section(chapter_id, section_id)
    except KeyError as exc:
        return jsonify({"error": str(exc)}), 404

    try:
        scenario, profile = _generate_scenario_for_section(section, difficulty_key)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500
    except Exception as exc:  # pragma: no cover - tolerate upstream variance
        return jsonify({"error": f"Failed to generate scenario: {exc}"}), 500

    payload = {
        "scenario": scenario,
        "difficulty": difficulty_key,
        "difficultyLabel": profile.get("label"),
        "difficultyDescription": profile.get("description"),
        "chapterId": chapter_id,
        "sectionId": section_id,
    }
    return jsonify(payload)


@app.post("/api/start_level")
def start_level():
    user, error = _require_user(required_role="student")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    chapter_id = data.get("chapterId")
    section_id = data.get("sectionId")
    difficulty_key = str(data.get("difficulty") or DEFAULT_DIFFICULTY).lower()
    if difficulty_key not in DIFFICULTY_PROFILES:
        difficulty_key = DEFAULT_DIFFICULTY

    if not chapter_id or not section_id:
        return jsonify({"error": "chapterId and sectionId are required"}), 400

    try:
        section = _get_section(chapter_id, section_id)
    except KeyError as exc:
        return jsonify({"error": str(exc)}), 404

    try:
        scenario, difficulty_profile = _generate_scenario_for_section(
            section, difficulty_key
        )
    except Exception as exc:  # pragma: no cover - protects against LLM format variance
        return (
            jsonify(
                {
                    "error": "Failed to generate scenario",
                    "details": str(exc),
                }
            ),
            500,
        )

    flat_context = flatten_scenario_for_template(scenario)
    if not flat_context.get("knowledge_points_hint"):
        flat_context["knowledge_points_hint"] = (
            "報盤結構, 議價策略, 跨文化溝通"
            if section.get("expects_bargaining")
            else "英文商務函電寫作, 信息提取, 跨文化表達"
        )
    conversation_prompt = section["conversation_prompt_template"].format_map(flat_context)
    prompt_suffix = difficulty_profile.get("prompt_suffix")
    if prompt_suffix:
        conversation_prompt = f"{conversation_prompt}\n\n[難度設定]\n{prompt_suffix}"
    if CONVERSATION_DIVERSITY_HINT not in conversation_prompt:
        conversation_prompt = (
            f"{conversation_prompt}\n\n[案例多样性提醒]\n{CONVERSATION_DIVERSITY_HINT}"
        )
    if ROLE_ENFORCEMENT_HINT not in conversation_prompt:
        conversation_prompt = (
            f"{conversation_prompt}\n\n[角色约束]\n请始终以学生为中国买家或中国卖家来组织对话，"
            "在回应中适时引用中国市场或供应链视角。"
        )
    if ENGLISH_ENFORCEMENT_HINT not in conversation_prompt:
        conversation_prompt = (
            f"{conversation_prompt}\n\n[Language Requirement]\n{ENGLISH_ENFORCEMENT_HINT}"
        )
    evaluation_prompt = section["evaluation_prompt_template"].format_map(flat_context)

    session_id = uuid.uuid4().hex

    database.create_session(
        session_id=session_id,
        user_id=int(user["id"]),
        chapter_id=chapter_id,
        section_id=section_id,
        system_prompt=conversation_prompt,
        evaluation_prompt=evaluation_prompt,
        scenario=scenario,
        expects_bargaining=bool(section.get("expects_bargaining")),
        difficulty=difficulty_key,
    )

    opening_message = _prepare_opening_message(conversation_prompt, scenario)
    if opening_message:
        database.add_message(session_id, "assistant", opening_message)

    payload = {
        "sessionId": session_id,
        "scenario": _prepare_scenario_payload(scenario),
        "openingMessage": opening_message,
        "knowledgePoints": scenario.get("knowledge_points", []) or [],
        "chapterId": chapter_id,
        "sectionId": section_id,
        "difficulty": difficulty_key,
    }
    return jsonify(payload)


@app.post("/api/assignments")
def create_assignment_endpoint():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    blueprint_id = data.get("blueprintId")
    raw_scenario = data.get("scenario")
    blueprint_raw = data.get("blueprint")

    difficulty_key = str(
        data.get("difficulty")
        or (raw_scenario or {}).get("difficulty")
        or (blueprint_raw or {}).get("difficulty")
        or DEFAULT_DIFFICULTY
    ).lower()
    if difficulty_key not in DIFFICULTY_PROFILES:
        difficulty_key = DEFAULT_DIFFICULTY

    scenario: Optional[Dict[str, object]] = None
    profile: Dict[str, str] = _get_difficulty_profile(difficulty_key)

    if blueprint_id:
        blueprint = database.get_blueprint(blueprint_id)
        if not blueprint or int(blueprint.get("ownerId")) != int(user["id"]):
            return jsonify({"error": "Blueprint not found"}), 404
        scenario = copy.deepcopy(blueprint.get("blueprint") or {})
        scenario, profile = _apply_difficulty_profile(scenario, difficulty_key)
    elif isinstance(raw_scenario, dict) and "scenario_title" in raw_scenario:
        scenario = copy.deepcopy(raw_scenario)
        scenario, profile = _apply_difficulty_profile(scenario, difficulty_key)
    elif isinstance(blueprint_raw, dict):
        scenario, profile = _assemble_scenario_from_blueprint(blueprint_raw, difficulty_key)
    else:
        return jsonify({"error": "scenario or blueprint data is required"}), 400

    chapter_id = data.get("chapterId")
    section_id = data.get("sectionId")
    description = _normalize_text(data.get("description")) or scenario.get("scenario_summary", "")
    title = _normalize_text(data.get("title")) or scenario.get("scenario_title") or "统一作业"

    conversation_prompt: str
    evaluation_prompt: str
    if chapter_id and section_id:
        try:
            section = _get_section(chapter_id, section_id)
        except KeyError:
            return jsonify({"error": "Invalid chapterId or sectionId"}), 404
        conversation_prompt, evaluation_prompt = _render_prompts_from_section(
            section, scenario, difficulty_key, profile
        )
    else:
        conversation_prompt, evaluation_prompt = _build_custom_assignment_prompts(
            scenario, profile
        )

    student_ids = []
    raw_students = data.get("studentIds") or []
    if isinstance(raw_students, list):
        for value in raw_students:
            try:
                student_ids.append(int(value))
            except (TypeError, ValueError):
                continue

    assignment_id = f"assignment-{uuid.uuid4().hex[:12]}"
    record = database.create_assignment(
        assignment_id=assignment_id,
        owner_id=int(user["id"]),
        title=title,
        description=description,
        chapter_id=chapter_id,
        section_id=section_id,
        scenario=scenario,
        conversation_prompt=conversation_prompt,
        evaluation_prompt=evaluation_prompt,
        difficulty=difficulty_key,
        blueprint_id=blueprint_id,
        due_at=data.get("dueAt"),
        student_ids=student_ids,
    )

    response_payload = _serialize_assignment(record)
    response_payload["difficultyDescription"] = profile.get("description")
    response_payload["studentIds"] = student_ids
    return jsonify({"assignment": response_payload}), 201


@app.get("/api/assignments")
def list_assignments_endpoint():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    records = database.list_assignments_by_teacher(int(user["id"]))
    payload = [_serialize_assignment(record) for record in records]
    return jsonify({"assignments": payload})


@app.get("/api/student/assignments")
def list_student_assignments():
    user, error = _require_user(required_role="student")
    if error:
        body, status = error
        return jsonify(body), status

    records = database.list_assignments_for_student(int(user["id"]))
    payload = [_serialize_assignment(record) for record in records]
    return jsonify({"assignments": payload})


@app.post("/api/assignments/<assignment_id>/start")
def start_assignment(assignment_id: str):
    user, error = _require_user(required_role="student")
    if error:
        body, status = error
        return jsonify(body), status

    record = database.get_assignment_for_student(assignment_id, int(user["id"]))
    if not record:
        return jsonify({"error": "Assignment not found"}), 404

    scenario = record.get("scenario") or {}
    difficulty_key = record.get("difficulty") or DEFAULT_DIFFICULTY
    if record.get("sessionId"):
        session = database.get_session(record["sessionId"])
        if session and int(session["user_id"]) == int(user["id"]):
            evaluation = database.get_latest_evaluation(session["id"])
            if evaluation:
                _inject_difficulty_metadata(evaluation)
            payload = {
                "sessionId": session["id"],
                "scenario": _prepare_scenario_payload(scenario),
                "assignmentId": assignment_id,
                "knowledgePoints": scenario.get("knowledge_points", []) or [],
                "difficulty": difficulty_key,
            }
            _inject_difficulty_metadata(payload)
            payload["evaluation"] = evaluation
            return jsonify(payload)

    session_id = uuid.uuid4().hex
    expects_bargaining = False
    product = scenario.get("product") if isinstance(scenario, dict) else {}
    if isinstance(product, dict):
        price_expectation = product.get("price_expectation") or {}
        expects_bargaining = bool(
            isinstance(price_expectation, dict)
            and (
                _normalize_text(price_expectation.get("student_target"))
                or _normalize_text(price_expectation.get("ai_bottom_line"))
            )
        )

    database.create_session(
        session_id=session_id,
        user_id=int(user["id"]),
        chapter_id=record.get("chapterId"),
        section_id=record.get("sectionId"),
        system_prompt=record.get("conversationPrompt"),
        evaluation_prompt=record.get("evaluationPrompt"),
        scenario=scenario,
        expects_bargaining=expects_bargaining,
        difficulty=difficulty_key,
        assignment_id=assignment_id,
    )
    database.link_assignment_session(assignment_id, int(user["id"]), session_id)

    conversation_prompt = _normalize_text(record.get("conversationPrompt"))
    opening_message = _prepare_opening_message(conversation_prompt, scenario)
    if opening_message:
        database.add_message(session_id, "assistant", opening_message)

    payload = {
        "sessionId": session_id,
        "scenario": _prepare_scenario_payload(scenario),
        "assignmentId": assignment_id,
        "knowledgePoints": scenario.get("knowledge_points", []) or [],
        "openingMessage": opening_message,
        "difficulty": difficulty_key,
    }
    _inject_difficulty_metadata(payload)
    return jsonify(payload), 201


@app.post("/api/admin/students/import")
def import_students_endpoint():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "file is required"}), 400
    try:
        records = _parse_student_records(file)
    except Exception as exc:  # pragma: no cover - file parsing safety
        return jsonify({"error": f"Failed to parse file: {exc}"}), 400

    if not records:
        return jsonify({"error": "No valid student rows found"}), 400

    summary = database.bulk_import_students(records)
    summary["total"] = len(records)
    return jsonify({"result": summary})


@app.post("/api/admin/students/<int:student_id>/password")
def admin_reset_student_password(student_id: int):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    new_password = _normalize_text(data.get("newPassword"))
    if len(new_password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400

    detail = database.get_student_detail(student_id)
    if not detail:
        return jsonify({"error": "Student not found"}), 404

    database.update_user_password(student_id, new_password)
    return jsonify({"status": "updated"})


@app.post("/api/account/password")
def update_own_password():
    user, error = _require_user()
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    current_password = data.get("currentPassword")
    new_password = _normalize_text(data.get("newPassword"))
    if len(new_password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400

    if not current_password or not database.verify_user_password(int(user["id"]), current_password):
        return jsonify({"error": "Current password is incorrect"}), 400

    database.update_user_password(int(user["id"]), new_password)
    return jsonify({"status": "updated"})


@app.post("/api/account/profile")
def update_own_profile():
    user, error = _require_user()
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    display_name = _normalize_text(data.get("displayName"))
    if not display_name:
        return jsonify({"error": "Display name is required"}), 400

    database.update_user_profile(int(user["id"]), display_name)
    updated = database.get_user(int(user["id"]))
    return jsonify({"status": "updated", "user": updated})


@app.post("/api/chat")
def chat():
    user, error = _require_user(required_role="student")
    if error:
        body, status = error
        return jsonify(body), status

    try:
        collab_key = _require_key("DEEPSEEK_COLLAB_KEY")
    except MissingKeyError as exc:  # pragma: no cover
        return jsonify({"error": str(exc)}), 500

    data = request.get_json(force=True)
    session_id = data.get("sessionId")
    user_message = (data.get("message") or "").strip()

    if not session_id or not user_message:
        return jsonify({"error": "sessionId and message are required"}), 400

    session = database.get_session(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    if int(user["id"]) != int(session["user_id"]):
        return jsonify({"error": "Forbidden"}), 403

    database.add_message(session_id, "user", user_message)

    history_rows = database.get_messages(session_id)
    history: List[Dict[str, str]] = [
        {"role": row["role"], "content": row["content"]} for row in history_rows
    ]

    messages = [{"role": "system", "content": session["system_prompt"]}]
    messages.extend(history)

    stream_requested = _as_bool(request.args.get("stream"))

    if stream_requested:

        def event_stream():
            chunks: List[str] = []
            try:
                for delta in _stream_chat(collab_key, messages, temperature=0.7):
                    chunks.append(delta)
                    payload = json.dumps({"content": delta})
                    yield f"event: chunk\ndata: {payload}\n\n"
            except Exception as exc:  # pragma: no cover - tolerate upstream errors
                database.remove_last_message(session_id)
                error_payload = json.dumps({"error": str(exc)})
                yield f"event: error\ndata: {error_payload}\n\n"
                return

            ai_reply = "".join(chunks).strip()
            if not ai_reply:
                ai_reply = "(no valid reply received)"

            database.add_message(session_id, "assistant", ai_reply)

            evaluation = _evaluate_session(session_id, session)
            latest_evaluation = database.get_latest_evaluation(session_id)
            if latest_evaluation:
                evaluation = latest_evaluation

            reply_payload = json.dumps({"reply": ai_reply})
            yield f"event: summary\ndata: {reply_payload}\n\n"

            evaluation_payload = json.dumps({"evaluation": evaluation})
            yield f"event: evaluation\ndata: {evaluation_payload}\n\n"

            yield "event: done\ndata: {}\n\n"

        response = Response(stream_with_context(event_stream()), mimetype="text/event-stream")
        response.headers["Cache-Control"] = "no-cache"
        return response

    try:
        ai_reply = _complete_chat(collab_key, messages, temperature=0.7).strip()
    except Exception as exc:  # pragma: no cover
        database.remove_last_message(session_id)
        return jsonify({"error": f"Failed to fetch assistant reply: {exc}"}), 500

    database.add_message(session_id, "assistant", ai_reply)

    evaluation = _evaluate_session(session_id, session)

    latest_evaluation = database.get_latest_evaluation(session_id)
    if latest_evaluation:
        evaluation = latest_evaluation

    return jsonify({"reply": ai_reply, "evaluation": evaluation})


def _evaluate_session(session_id: str, session: Dict[str, object]) -> Dict[str, object]:
    try:
        critic_key = _require_key("DEEPSEEK_CRITIC_KEY")
    except MissingKeyError:
        scenario = session.get("scenario", {}) if session else {}
        return {
            "score": None,
            "scoreLabel": None,
            "commentary": "未配置批判評估 API Key。",
            "actionItems": [],
            "knowledgePoints": scenario.get("knowledge_points", []) or [],
            "bargainingWinRate": None,
        }

    scenario = session.get("scenario", {})
    scenario_knowledge = scenario.get("knowledge_points", []) or []
    history_rows = database.get_messages(session_id)
    transcript_history = [{"role": row["role"], "content": row["content"]} for row in history_rows]
    transcript = _build_transcript(transcript_history, scenario)
    evaluation_prompt = session.get("evaluation_prompt", "")

    messages = [
        {"role": "system", "content": str(evaluation_prompt)},
        {
            "role": "user",
            "content": transcript,
        },
    ]

    try:
        raw = _complete_chat(critic_key, messages, temperature=0.2)
        data = _extract_json_block(raw)
    except Exception:  # pragma: no cover - tolerate evaluation failures
        return {
            "score": None,
            "scoreLabel": None,
            "commentary": "評估暫時無法提供，請稍後再試。",
            "actionItems": [],
            "knowledgePoints": scenario_knowledge,
            "bargainingWinRate": None,
        }

    score = data.get("score")
    score_label = data.get("score_label")
    action_items = data.get("action_items", []) or []
    knowledge_points = data.get("knowledge_points", []) or scenario_knowledge
    if not isinstance(action_items, list):
        action_items = [action_items]
    if not isinstance(knowledge_points, list):
        knowledge_points = [knowledge_points] if knowledge_points else []
    bargaining_win_rate = data.get("bargaining_win_rate") if session.get("expects_bargaining") else None

    result = {
        "score": score,
        "scoreLabel": score_label,
        "commentary": data.get("commentary", ""),
        "actionItems": action_items,
        "knowledgePoints": knowledge_points,
        "bargainingWinRate": bargaining_win_rate,
    }

    database.save_evaluation(session_id, result)
    if session.get("assignment_id"):
        database.mark_assignment_completed_by_session(session_id)
    return result


@app.get("/api/sessions")
def list_sessions_for_user_endpoint():
    user, error = _require_user()
    if error:
        body, status = error
        return jsonify(body), status

    target_user_id = int(user["id"])
    if user["role"] == "teacher":
        query_param = request.args.get("userId")
        if not query_param:
            return jsonify({"error": "userId is required for teacher queries"}), 400
        target_user_id = int(query_param)

    sessions = database.list_sessions_for_user(target_user_id)
    for session in sessions:
        _inject_difficulty_metadata(session)
    return jsonify({"sessions": sessions})


@app.get("/api/student/dashboard")
def get_student_dashboard_endpoint():
    user, error = _require_user(required_role="student")
    if error:
        body, status = error
        return jsonify(body), status

    dashboard = database.get_student_dashboard(int(user["id"]))
    for entry in dashboard.get("timeline", []):
        _inject_difficulty_metadata(entry)
    return jsonify(dashboard)


@app.get("/api/sessions/<session_id>")
def get_session_detail(session_id: str):
    user, error = _require_user()
    if error:
        body, status = error
        return jsonify(body), status

    session = database.get_session(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    if user["role"] == "student" and int(session["user_id"]) != int(user["id"]):
        return jsonify({"error": "Forbidden"}), 403

    history = database.get_messages(session_id)
    evaluation = database.get_latest_evaluation(session_id)

    payload = {
        "session": {
            "id": session["id"],
            "chapterId": session["chapter_id"],
            "sectionId": session["section_id"],
            "scenario": _prepare_scenario_payload(session["scenario"]),
            "expectsBargaining": session["expects_bargaining"],
            "difficulty": session.get("difficulty"),
        },
        "messages": history,
        "evaluation": evaluation,
    }
    _inject_difficulty_metadata(payload["session"])
    return jsonify(payload)


@app.post("/api/sessions/<session_id>/reset")
def reset_session_endpoint(session_id: str):
    user, error = _require_user(required_role="student")
    if error:
        body, status = error
        return jsonify(body), status

    session = database.get_session(session_id)
    if not session or int(session["user_id"]) != int(user["id"]):
        return jsonify({"error": "Session not found"}), 404

    database.reset_session(session_id)
    scenario = session["scenario"]
    conversation_prompt = _normalize_text(session.get("system_prompt"))
    opening_message = _prepare_opening_message(conversation_prompt, scenario)
    if opening_message:
        database.add_message(session_id, "assistant", opening_message)

    payload = {
        "sessionId": session_id,
        "scenario": _prepare_scenario_payload(scenario),
        "openingMessage": opening_message,
        "knowledgePoints": scenario.get("knowledge_points", []) or [],
        "chapterId": session["chapter_id"],
        "sectionId": session["section_id"],
        "difficulty": session.get("difficulty") or DEFAULT_DIFFICULTY,
    }
    return jsonify(payload)


@app.get("/api/admin/students")
def list_students_progress_endpoint():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    students = database.list_students_progress()
    return jsonify({"students": students})


@app.get("/api/admin/students/<int:student_id>")
def get_student_detail_endpoint(student_id: int):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    detail = database.get_student_detail(student_id)
    if not detail:
        return jsonify({"error": "Student not found"}), 404
    for session in detail.get("sessions", []):
        _inject_difficulty_metadata(session)
    return jsonify(detail)


@app.get("/api/admin/analytics")
def get_admin_analytics_endpoint():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    analytics = database.get_class_analytics()
    return jsonify(analytics)


@app.get("/api/admin/levels")
def get_admin_levels():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    chapters = _ensure_level_hierarchy(include_prompts=True)
    return jsonify({"chapters": chapters})


@app.post("/api/admin/chapters")
def create_admin_chapter():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"error": "title is required"}), 400

    description = (data.get("description") or "").strip()
    order_index = data.get("orderIndex")
    try:
        order_value = int(order_index) if order_index is not None else None
    except (TypeError, ValueError):
        return jsonify({"error": "orderIndex must be an integer"}), 400

    chapter_id = (data.get("id") or "").strip() or None
    chapter = database.create_chapter(
        title=title,
        description=description,
        order_index=order_value,
        chapter_id=chapter_id,
    )
    return jsonify({"chapter": chapter}), 201


@app.put("/api/admin/chapters/<chapter_id>")
def update_admin_chapter(chapter_id: str):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    kwargs: Dict[str, object] = {}
    if "title" in data:
        kwargs["title"] = (data.get("title") or "").strip()
    if "description" in data:
        kwargs["description"] = (data.get("description") or "").strip()
    if "orderIndex" in data:
        try:
            kwargs["order_index"] = int(data.get("orderIndex"))
        except (TypeError, ValueError):
            return jsonify({"error": "orderIndex must be an integer"}), 400

    chapter = database.update_chapter(chapter_id, **kwargs)
    if not chapter:
        return jsonify({"error": "Chapter not found"}), 404
    return jsonify({"chapter": chapter})


@app.delete("/api/admin/chapters/<chapter_id>")
def delete_admin_chapter(chapter_id: str):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    existing = database.get_chapter(chapter_id)
    if not existing:
        return jsonify({"error": "Chapter not found"}), 404
    database.delete_chapter(chapter_id)
    return ("", 204)


@app.post("/api/admin/sections")
def create_admin_section():
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    chapter_id = (data.get("chapterId") or "").strip()
    if not chapter_id:
        return jsonify({"error": "chapterId is required"}), 400

    title = (data.get("title") or "").strip()
    description = (data.get("description") or "").strip()
    env_prompt = (data.get("environmentPromptTemplate") or "").strip()
    env_user = (data.get("environmentUserMessage") or "").strip()
    convo_prompt = (data.get("conversationPromptTemplate") or "").strip()
    eval_prompt = (data.get("evaluationPromptTemplate") or "").strip()

    if not all([title, description, env_prompt, env_user, convo_prompt, eval_prompt]):
        return jsonify({"error": "title, description and all prompt templates are required"}), 400

    expects_bargaining = _as_bool(data.get("expectsBargaining"), False)
    order_index = data.get("orderIndex")
    try:
        order_value = int(order_index) if order_index is not None else None
    except (TypeError, ValueError):
        return jsonify({"error": "orderIndex must be an integer"}), 400

    section_id = (data.get("id") or "").strip() or None
    section = database.create_section(
        chapter_id=chapter_id,
        title=title,
        description=description,
        environment_prompt_template=env_prompt,
        environment_user_message=env_user,
        conversation_prompt_template=convo_prompt,
        evaluation_prompt_template=eval_prompt,
        expects_bargaining=expects_bargaining,
        order_index=order_value,
        section_id=section_id,
    )
    if not section:
        return jsonify({"error": "Chapter not found"}), 404
    return jsonify({"section": section}), 201


@app.put("/api/admin/sections/<section_id>")
def update_admin_section(section_id: str):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    data = request.get_json(force=True)
    kwargs: Dict[str, object] = {}
    if "chapterId" in data:
        kwargs["chapter_id"] = (data.get("chapterId") or "").strip()
    if "title" in data:
        kwargs["title"] = (data.get("title") or "").strip()
    if "description" in data:
        kwargs["description"] = (data.get("description") or "").strip()
    if "environmentPromptTemplate" in data:
        kwargs["environment_prompt_template"] = (
            data.get("environmentPromptTemplate") or ""
        ).strip()
    if "environmentUserMessage" in data:
        kwargs["environment_user_message"] = (
            data.get("environmentUserMessage") or ""
        ).strip()
    if "conversationPromptTemplate" in data:
        kwargs["conversation_prompt_template"] = (
            data.get("conversationPromptTemplate") or ""
        ).strip()
    if "evaluationPromptTemplate" in data:
        kwargs["evaluation_prompt_template"] = (
            data.get("evaluationPromptTemplate") or ""
        ).strip()
    if "expectsBargaining" in data:
        kwargs["expects_bargaining"] = _as_bool(data.get("expectsBargaining"))
    if "orderIndex" in data:
        try:
            kwargs["order_index"] = int(data.get("orderIndex"))
        except (TypeError, ValueError):
            return jsonify({"error": "orderIndex must be an integer"}), 400

    section = database.update_section(section_id, **kwargs)
    if not section:
        return jsonify({"error": "Section not found"}), 404
    return jsonify({"section": section})


@app.delete("/api/admin/sections/<section_id>")
def delete_admin_section(section_id: str):
    user, error = _require_user(required_role="teacher")
    if error:
        body, status = error
        return jsonify(body), status

    section = database.get_section(section_id)
    if not section:
        return jsonify({"error": "Section not found"}), 404
    database.delete_section(section_id)
    return ("", 204)


if __name__ == "__main__":  # pragma: no cover
    app.run(host="0.0.0.0", debug=True)
