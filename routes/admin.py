"""教师端管理接口集合。"""

from __future__ import annotations

import re
import threading
import uuid
from concurrent.futures import TimeoutError
from itertools import chain
from typing import Dict, List, Optional, Sequence

from flask import Blueprint, jsonify, request, current_app
from openpyxl import load_workbook

import database
from services import ai_matching
from services import docx_importer
from services import draft_knowledge_service
from services import embedding_service
from services import graph_service
from services import knowledge_job_service
from services import rag_matcher
from services.auth_service import current_user, require_role
from services.scenario_generator import ensure_level_hierarchy, inject_difficulty_metadata
from utils.normalizers import normalize_text
from utils.validators import as_bool

bp = Blueprint("admin", __name__)


def _normalize_student_header(value: object) -> str:
    text = normalize_text(value).lower()
    if text in {"id", "账号", "學號", "学号", "user", "userid"}:
        return "id"
    if text in {"姓名", "name", "display", "nickname"}:
        return "name"
    if text in {"password", "密码", "pass", "pwd"}:
        return "password"
    return ""


def _parse_student_records(file_storage) -> List[Dict[str, str]]:
    workbook = load_workbook(file_storage, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        first_row = next(rows)
    except StopIteration:
        return []

    headers = [_normalize_student_header(cell) for cell in (first_row or [])]
    if not any(headers):
        headers = ["id", "name", "password"]
        rows = chain([first_row], rows)

    records: List[Dict[str, str]] = []
    for row in rows:
        if not row:
            continue
        entry: Dict[str, str] = {"id": "", "name": "", "password": ""}
        for index, cell in enumerate(row):
            if index >= len(headers):
                continue
            key = headers[index]
            if not key:
                continue
            entry[key] = normalize_text(cell)
        entry["id"] = normalize_text(entry["id"])
        entry["name"] = normalize_text(entry["name"]) or entry["id"]
        entry["password"] = normalize_text(entry["password"]) or entry["id"]
        if entry["id"] and entry["password"]:
            records.append(entry)
    return records


def _sync_graph_background() -> None:
    """Synchronize content to Neo4j knowledge graph in background.

    This function is fail-safe and will not raise exceptions that could
    break the main business logic.
    """
    if not graph_service.is_configured():
        current_app.logger.debug("Knowledge graph not configured, skipping sync")
        return

    try:
        graph_service.sync_static_content()
        current_app.logger.debug("Knowledge graph sync completed successfully")
    except graph_service.GraphUnavailableError as exc:
        current_app.logger.warning("Skipping graph sync (service unavailable): %s", exc)
    except Exception as exc:  # pragma: no cover - logging safeguard
        current_app.logger.exception("Failed to sync knowledge graph: %s", exc)


def _set_lesson_knowledge_points(
    lesson_id: str, points: Sequence[object], *, allow_empty: bool = False
) -> None:
    """Attach knowledge points to a lesson in the graph, if configured."""

    if not points and not allow_empty:
        return
    if not graph_service.is_configured():
        current_app.logger.debug(
            "Knowledge graph not configured, skipping lesson knowledge sync"
        )
        return

    try:
        graph_service.set_lesson_knowledge_points(lesson_id, points)
        current_app.logger.debug(
            "Synchronized %s knowledge points for lesson %s",
            len(points),
            lesson_id,
        )
    except graph_service.GraphUnavailableError as exc:
        current_app.logger.warning(
            "Skipping lesson knowledge sync (service unavailable): %s", exc
        )
    except Exception as exc:  # pragma: no cover - logging safeguard
        current_app.logger.exception(
            "Failed to synchronize lesson knowledge points: %s", exc
        )


def _precompute_lesson_graph_async(lesson_id: str, *, allow_unpublished: bool = False) -> None:
    """Precompute KP detection + lesson graph cache in background."""

    if not graph_service.is_configured():
        current_app.logger.debug("Knowledge graph not configured, skip lesson precompute")
        return

    def _worker():
        try:
            graph_service.compute_lesson_knowledge_and_graph(
                lesson_id,
                include_unpublished=allow_unpublished,
            )
            current_app.logger.info("Lesson graph precomputed for %s", lesson_id)
        except graph_service.GraphUnavailableError as exc:
            current_app.logger.warning("Graph unavailable during lesson precompute: %s", exc)
        except Exception as exc:  # pragma: no cover - defensive log
            current_app.logger.exception("Failed to precompute lesson graph: %s", exc)

    threading.Thread(target=_worker, daemon=True).start()


def _invalidate_lesson_graph_cache(lesson_id: str) -> None:
    try:
        graph_service.invalidate_lesson_graph_cache(lesson_id)
    except Exception:
        current_app.logger.debug("Lesson cache invalidation skipped for %s", lesson_id)


@bp.post("/api/admin/students/import")
@require_role("teacher")
def import_students():
    """批量导入学生账号，支持 Excel 格式。"""
    current_user()
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "file is required"}), 400
    try:
        records = _parse_student_records(file)
    except Exception as exc:
        return jsonify({"error": f"Failed to parse file: {exc}"}), 400

    if not records:
        return jsonify({"error": "No valid student rows found"}), 400

    summary = database.bulk_import_students(records)
    summary["total"] = len(records)
    return jsonify({"result": summary})


@bp.post("/api/admin/theory/import-docx")
@require_role("teacher")
def import_theory_from_docx():
    """Parse a Word document and return a draft theory outline for preview."""

    current_user()
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "file is required"}), 400
    if not file.filename.lower().endswith(".docx"):
        return jsonify({"error": "仅支持 .docx 文档"}), 400

    size_bytes = getattr(file, "content_length", None)
    if size_bytes and size_bytes > 10 * 1024 * 1024:
        return jsonify({"error": "文档体积过大，请控制在 10MB 以内后重试"}), 400

    current_app.logger.info(
        "Word import requested: name=%s size=%s", file.filename, size_bytes or "unknown"
    )

    try:
        outline = docx_importer.parse_docx_outline_with_timeout(file, timeout_seconds=25)
    except TimeoutError:
        current_app.logger.error("Word import timed out for file %s", file.filename)
        return jsonify({"error": "解析超时，请简化文档或拆分后重试"}), 504
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover - defensive logging
        current_app.logger.exception("Failed to parse Word document: %s", exc)
        return jsonify({"error": "无法解析该 Word 文档，请检查格式后重试"}), 500

    return jsonify({"import": outline})


@bp.post("/api/admin/theory/import-docx/drafts")
@require_role("teacher")
def import_theory_docx_generate_drafts():
    """Parse a Word doc and generate draft knowledge points (MVP, persistent)."""

    current_user()
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "file is required"}), 400
    if not file.filename.lower().endswith(".docx"):
        return jsonify({"error": "仅支持 .docx 文档"}), 400
    job_id = knowledge_job_service.create_job()
    try:
        outline = docx_importer.parse_docx_outline(file)
        drafts: List[Dict[str, object]] = []
        chapters = outline.get("chapters") or []
        for chapter in chapters:
            topics = chapter.get("topics") or []
            for topic in topics:
                lessons = topic.get("lessons") or []
                for lesson in lessons:
                    lesson_title = lesson.get("title") or "未命名知识点"
                    body_html = lesson.get("contentHtml") or ""
                    summary = lesson.get("summary") or (body_html[:180] if body_html else "")
                    drafts.append(
                        {
                            "id": str(uuid.uuid4()),
                            "name": lesson_title,
                            "summary": summary,
                            "bodyHtml": body_html,
                            "content": body_html,
                            "tags": [],
                            "status": "draft",
                        }
                    )
        knowledge_job_service.insert_drafts(job_id, drafts)
        knowledge_job_service.update_job(job_id, status="completed", total=len(drafts), processed=len(drafts))
        return jsonify({"jobId": job_id, "drafts": drafts, "status": "completed"}), 200
    except Exception as exc:  # pragma: no cover - defensive logging
        current_app.logger.exception("Failed to parse Word document: %s", exc)
        knowledge_job_service.update_job(job_id, status="failed")
        return jsonify({"error": "无法解析该 Word 文档，请检查格式后重试"}), 500


@bp.get("/api/admin/theory/drafts/<batch_id>")
@require_role("teacher")
def list_theory_drafts(batch_id: str):
    """Fetch generated knowledge point drafts by batch."""

    current_user()
    drafts = draft_knowledge_service.get_batch(batch_id)
    if not drafts:
        drafts = knowledge_job_service.list_drafts(batch_id)
    return jsonify({"batchId": batch_id, "drafts": drafts}), 200


@bp.post("/api/admin/theory/drafts/<batch_id>/approve")
@require_role("teacher")
def approve_theory_drafts(batch_id: str):
    """Approve drafts -> create/merge knowledge points in Neo4j (no relations)."""

    current_user()
    body = request.get_json(force=True, silent=True) or {}
    ids = body.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "ids is required"}), 400
    approved = draft_knowledge_service.approve(batch_id, ids)
    created: List[Dict[str, object]] = []
    for draft in approved:
        try:
            payload = {
                "name": draft.get("name", ""),
                "summary": draft.get("summary", ""),
                "content": draft.get("content", ""),
                "bodyHtml": draft.get("bodyHtml", ""),
                "tags": draft.get("tags") or [],
            }
            graph_service.create_knowledge_point(payload)
            created.append(payload)
        except Exception as exc:  # pragma: no cover - continue best-effort
            current_app.logger.warning("Failed to create draft knowledge point: %s", exc)
            continue
    knowledge_job_service.mark_drafts(batch_id, ids, "approved")
    return jsonify({"created": created, "count": len(created)}), 200


@bp.post("/api/ai/knowledge-points/match")
@require_role("teacher")
def ai_match_knowledge_point():
    """Auto-match selected text to an existing knowledge point (lightweight heuristic)."""

    current_user()
    data = request.get_json(force=True, silent=True) or {}
    selection_text = normalize_text(data.get("selectionText") or "")
    selection_html = data.get("selectionHtml") or ""
    candidate_names = data.get("candidateNames") or []
    lesson_id = data.get("lessonId") or ""
    lesson_context = data.get("lessonContext") or {}
    if not selection_text:
        return jsonify({"error": "selectionText is required"}), 400

    try:
        overview = graph_service.get_knowledge_management_overview()
    except Exception as exc:  # pragma: no cover - defensive log
        current_app.logger.exception("Knowledge overview failed: %s", exc)
        return jsonify({"error": "知识点索引获取失败"}), 500

    records = overview.get("knowledge_cards") or []
    if candidate_names:
        candidate_set = {normalize_text(n) for n in candidate_names if n}
        records = [r for r in records if normalize_text(r.get("name")) in candidate_set]

    def _cosine(a: Optional[List[float]], b: Optional[List[float]]) -> float:
        if not a or not b or len(a) != len(b):
            return 0.0
        dot = sum(x * y for x, y in zip(a, b))
        norm_a = sum(x * x for x in a) ** 0.5
        norm_b = sum(y * y for y in b) ** 0.5
        return dot / (norm_a * norm_b) if norm_a and norm_b else 0.0

    def _extract_tokens(value: str) -> List[str]:
        """适配中英文的轻量分词，中文拆成2-3字 ngram，英文按单词。"""
        normalized = normalize_text(value)
        raw_parts = re.findall(r"[\w\d]+|[\u4e00-\u9fa5]+", normalized)
        tokens: List[str] = []
        for part in raw_parts:
            if not part or len(part) <= 1:
                continue
            if re.search(r"[\u4e00-\u9fa5]", part):
                ngrams = set()
                for n in (2, 3):
                    ngrams.update(part[i : i + n] for i in range(0, max(len(part) - n + 1, 0)))
                tokens.extend([ng for ng in ngrams if len(ng) >= 2])
            else:
                tokens.append(part)
        return tokens

    selection_tokens = _extract_tokens(selection_text)

    # 语义向量相似度（sentence-transformers，若模型不可用则为空）
    embed_scores: Dict[int, float] = {}
    selection_vec: Optional[List[float]] = None
    try:
        selection_vecs = embedding_service.embed_texts([selection_text])
        selection_vec = selection_vecs[0] if selection_vecs else None
        if selection_vec:
            candidate_texts = [
                " ".join(
                    filter(
                        None,
                        [
                            rec.get("name"),
                            rec.get("summary"),
                            rec.get("bodyHtml"),
                            rec.get("content"),
                            rec.get("description"),
                        ],
                    )
                )
                for rec in records
            ]
            candidate_vecs = embedding_service.embed_texts(candidate_texts)
            if candidate_vecs:
                for rec, vec in zip(records, candidate_vecs):
                    embed_scores[id(rec)] = _cosine(selection_vec, vec)
    except Exception as exc:  # pragma: no cover - safety
        current_app.logger.warning("Embedding similarity failed: %s", exc)

    def _score(record: dict) -> float:
        name = normalize_text(record.get("name") or "")
        summary = normalize_text(record.get("summary") or "")
        body = normalize_text(record.get("bodyHtml") or record.get("content") or "")
        score = 0.0
        # 适度提高中文匹配的局部敏感度
        if selection_text and selection_text in name:
            score += 3
        if selection_text and selection_text in summary:
            score += 2
        if selection_text and selection_text in body:
            score += 1.5
        for tok in selection_tokens:
            if tok in name:
                score += 0.6
            if tok in summary:
                score += 0.45
            if tok in body:
                score += 0.35
        lesson_ids = record.get("lessons") or []
        if lesson_id and lesson_id in lesson_ids:
            score += 0.5
        if embed_scores:
            score += 3.2 * embed_scores.get(id(record), 0.0)
        return score

    ranked = sorted(records, key=_score, reverse=True)
    top_candidates = ranked[:12]

    source = "heuristic"
    confidence = 0.0
    best = top_candidates[0] if top_candidates else None
    reason = ""

    try:
        if top_candidates:
            best, confidence, reason = ai_matching.match_knowledge_point(
                selection_text, top_candidates, lesson_context=lesson_context
            )
            source = "deepseek"
    except Exception as exc:
        current_app.logger.warning("Deepseek match failed, fallback to heuristic: %s", exc)
        best = top_candidates[0] if top_candidates else None
        confidence = _score(best) if best else 0.0
        source = "fallback"
        reason = str(exc)

    payload = {
        "match": best or {},
        "confidence": confidence,
        "selection": {"text": selection_text, "html": selection_html},
        "source": source,
        "reason": reason,
        "embed_score": embed_scores.get(id(best)) if best and embed_scores else None,
        "lesson_context": lesson_context,
    }
    return jsonify(payload), 200


@bp.post("/api/ai/knowledge-points/match-rag")
@require_role("teacher")
def ai_match_knowledge_point_rag():
    """RAG风格匹配：chunk + embedding-like 排序（Beta）。"""

    current_user()
    data = request.get_json(force=True, silent=True) or {}
    selection_text = normalize_text(data.get("selectionText") or "")
    selection_html = data.get("selectionHtml") or ""
    candidate_names = data.get("candidateNames") or []
    if not selection_text:
        return jsonify({"error": "selectionText is required"}), 400

    try:
        overview = graph_service.get_knowledge_management_overview()
    except Exception as exc:  # pragma: no cover
        current_app.logger.exception("Knowledge overview failed: %s", exc)
        return jsonify({"error": "知识点索引获取失败"}), 500

    records = overview.get("knowledge_cards") or []
    if candidate_names:
        candidate_set = {normalize_text(n) for n in candidate_names if n}
        records = [r for r in records if normalize_text(r.get("name")) in candidate_set]

    match, confidence, context = rag_matcher.match(selection_text, records)
    payload = {
        "match": match,
        "confidence": confidence,
        "selection": {"text": selection_text, "html": selection_html},
        "context": context,
        "source": "rag-beta",
    }
    return jsonify(payload), 200


@bp.post("/api/admin/students/<int:student_id>/password")
@require_role("teacher")
def reset_student_password(student_id: int):
    """教师重置学生密码，方便线下教学支援。"""
    data = request.get_json(force=True)
    new_password = normalize_text(data.get("newPassword"))
    if len(new_password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400

    detail = database.get_student_detail(student_id)
    if not detail:
        return jsonify({"error": "Student not found"}), 404

    database.update_user_password(student_id, new_password)
    return jsonify({"status": "updated"})


@bp.get("/api/admin/students")
@require_role("teacher")
def list_students_progress():
    students = database.list_students_progress()
    return jsonify({"students": students})


@bp.get("/api/admin/students/<int:student_id>")
@require_role("teacher")
def get_student_detail(student_id: int):
    detail = database.get_student_detail(student_id)
    if not detail:
        return jsonify({"error": "Student not found"}), 404
    for session in detail.get("sessions", []):
        inject_difficulty_metadata(session)
    return jsonify(detail)


@bp.get("/api/admin/analytics")
@require_role("teacher")
def get_admin_analytics():
    analytics = database.get_class_analytics()
    return jsonify(analytics)


@bp.get("/api/admin/levels")
@require_role("teacher")
def get_admin_levels():
    chapters = ensure_level_hierarchy(include_prompts=True)
    return jsonify({"chapters": chapters})


@bp.post("/api/admin/chapters")
@require_role("teacher")
def create_admin_chapter():
    data = request.get_json(force=True)
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"error": "title is required"}), 400

    description = (data.get("description") or "").strip()
    order_index = data.get("orderIndex")
    try:
        order_value = int(order_index) if order_index is not None else None
    except (TypeError, ValueError):
        return jsonify({"error": "orderIndex must be an integer"}), 400

    chapter_id = (data.get("id") or "").strip() or None
    chapter = database.create_chapter(
        title=title,
        description=description,
        order_index=order_value,
        chapter_id=chapter_id,
    )
    _sync_graph_background()
    return jsonify({"chapter": chapter}), 201


@bp.put("/api/admin/chapters/<chapter_id>")
@require_role("teacher")
def update_admin_chapter(chapter_id: str):
    data = request.get_json(force=True)
    kwargs: Dict[str, object] = {}
    if "title" in data:
        kwargs["title"] = (data.get("title") or "").strip()
    if "description" in data:
        kwargs["description"] = (data.get("description") or "").strip()
    if "orderIndex" in data:
        try:
            kwargs["order_index"] = int(data.get("orderIndex"))
        except (TypeError, ValueError):
            return jsonify({"error": "orderIndex must be an integer"}), 400

    chapter = database.update_chapter(chapter_id, **kwargs)
    if not chapter:
        return jsonify({"error": "Chapter not found"}), 404
    _sync_graph_background()
    return jsonify({"chapter": chapter})


@bp.delete("/api/admin/chapters/<chapter_id>")
@require_role("teacher")
def delete_admin_chapter(chapter_id: str):
    existing = database.get_chapter(chapter_id)
    if not existing:
        return jsonify({"error": "Chapter not found"}), 404
    database.delete_chapter(chapter_id)
    _sync_graph_background()
    return ("", 204)


@bp.post("/api/admin/sections")
@require_role("teacher")
def create_admin_section():
    data = request.get_json(force=True)
    chapter_id = (data.get("chapterId") or "").strip()
    if not chapter_id:
        return jsonify({"error": "chapterId is required"}), 400

    title = (data.get("title") or "").strip()
    description = (data.get("description") or "").strip()
    env_prompt = (data.get("environmentPromptTemplate") or "").strip()
    env_user = (data.get("environmentUserMessage") or "").strip()
    convo_prompt = (data.get("conversationPromptTemplate") or "").strip()
    eval_prompt = (data.get("evaluationPromptTemplate") or "").strip()

    if not all([title, description, env_prompt, env_user, convo_prompt, eval_prompt]):
        return jsonify({"error": "title, description and all prompt templates are required"}), 400

    expects_bargaining = as_bool(data.get("expectsBargaining"), False)
    order_index = data.get("orderIndex")
    try:
        order_value = int(order_index) if order_index is not None else None
    except (TypeError, ValueError):
        return jsonify({"error": "orderIndex must be an integer"}), 400

    section_id = (data.get("id") or "").strip() or None
    section = database.create_section(
        chapter_id=chapter_id,
        title=title,
        description=description,
        environment_prompt_template=env_prompt,
        environment_user_message=env_user,
        conversation_prompt_template=convo_prompt,
        evaluation_prompt_template=eval_prompt,
        expects_bargaining=expects_bargaining,
        order_index=order_value,
        section_id=section_id,
    )
    if not section:
        return jsonify({"error": "Chapter not found"}), 404
    _sync_graph_background()
    return jsonify({"section": section}), 201


@bp.put("/api/admin/sections/<section_id>")
@require_role("teacher")
def update_admin_section(section_id: str):
    data = request.get_json(force=True)
    kwargs: Dict[str, object] = {}
    if "chapterId" in data:
        kwargs["chapter_id"] = (data.get("chapterId") or "").strip()
    if "title" in data:
        kwargs["title"] = (data.get("title") or "").strip()
    if "description" in data:
        kwargs["description"] = (data.get("description") or "").strip()
    if "environmentPromptTemplate" in data:
        kwargs["environment_prompt_template"] = (
            data.get("environmentPromptTemplate") or ""
        ).strip()
    if "environmentUserMessage" in data:
        kwargs["environment_user_message"] = (
            data.get("environmentUserMessage") or ""
        ).strip()
    if "conversationPromptTemplate" in data:
        kwargs["conversation_prompt_template"] = (
            data.get("conversationPromptTemplate") or ""
        ).strip()
    if "evaluationPromptTemplate" in data:
        kwargs["evaluation_prompt_template"] = (
            data.get("evaluationPromptTemplate") or ""
        ).strip()
    if "expectsBargaining" in data:
        kwargs["expects_bargaining"] = as_bool(data.get("expectsBargaining"))
    if "orderIndex" in data:
        try:
            kwargs["order_index"] = int(data.get("orderIndex"))
        except (TypeError, ValueError):
            return jsonify({"error": "orderIndex must be an integer"}), 400

    section = database.update_section(section_id, **kwargs)
    if not section:
        return jsonify({"error": "Section not found"}), 404
    _sync_graph_background()
    return jsonify({"section": section})


@bp.delete("/api/admin/sections/<section_id>")
@require_role("teacher")
def delete_admin_section(section_id: str):
    section = database.get_section(section_id)
    if not section:
        return jsonify({"error": "Section not found"}), 404
    database.delete_section(section_id)
    _sync_graph_background()
    return ("", 204)


@bp.get("/api/admin/theory")
@require_role("teacher")
def list_admin_theory():
    include_content = as_bool(request.args.get("includeContent"), default=True)
    records = database.list_theory_hierarchy(
        include_content=include_content, published_only=False
    )
    return jsonify({"theory": records})


@bp.post("/api/admin/theory/topics")
@require_role("teacher")
def create_admin_theory_topic():
    """创建理论主题（二级目录）。"""
    try:
        data = request.get_json(force=True) or {}
        chapter_id = normalize_text(data.get("chapterId"))
        if not chapter_id:
            current_app.logger.warning("Create theory topic failed: chapterId is required")
            return jsonify({"error": "chapterId is required"}), 400

        title = normalize_text(data.get("title")) or "未命名理论单元"
        code = normalize_text(data.get("code")) or ""
        summary = normalize_text(data.get("summary"))
        order_index_raw = data.get("orderIndex")
        order_index: Optional[int] = None
        if order_index_raw not in (None, ""):
            try:
                order_index = int(order_index_raw)
            except (TypeError, ValueError):
                current_app.logger.warning("Create theory topic failed: invalid orderIndex")
                return jsonify({"error": "orderIndex must be an integer"}), 400

        record = database.create_theory_topic(
            chapter_id=chapter_id,
            title=title,
            code=code,
            summary=summary,
            order_index=order_index,
        )
        if not record:
            current_app.logger.warning(f"Create theory topic failed: chapter {chapter_id} not found")
            return jsonify({"error": "Chapter not found"}), 404

        current_app.logger.info(f"Theory topic created successfully: {record.get('id')} - {title}")
        _sync_graph_background()
        return jsonify({"topic": record}), 201
    except Exception as exc:
        current_app.logger.exception(f"Unexpected error creating theory topic: {exc}")
        return jsonify({"error": "Internal server error", "detail": str(exc)}), 500


@bp.put("/api/admin/theory/topics/<topic_id>")
@require_role("teacher")
def update_admin_theory_topic(topic_id: str):
    data = request.get_json(force=True) or {}
    updates: Dict[str, object] = {}

    if "chapterId" in data:
        updates["chapter_id"] = normalize_text(data.get("chapterId"))
    if "title" in data:
        title = normalize_text(data.get("title"))
        if not title:
            return jsonify({"error": "title is required"}), 400
        updates["title"] = title
    if "code" in data:
        updates["code"] = normalize_text(data.get("code"))
    if "summary" in data:
        updates["summary"] = normalize_text(data.get("summary"))
    if "orderIndex" in data:
        order_value = data.get("orderIndex")
        if order_value in (None, ""):
            updates["order_index"] = None
        else:
            try:
                updates["order_index"] = int(order_value)
            except (TypeError, ValueError):
                return jsonify({"error": "orderIndex must be an integer"}), 400

    topic = database.update_theory_topic(topic_id, **updates)
    if not topic:
        return jsonify({"error": "Topic not found"}), 404
    _sync_graph_background()
    return jsonify({"topic": topic})


@bp.delete("/api/admin/theory/topics/<topic_id>")
@require_role("teacher")
def delete_admin_theory_topic(topic_id: str):
    topic = database.get_theory_topic(topic_id)
    if not topic:
        return jsonify({"error": "Topic not found"}), 404
    database.delete_theory_topic(topic_id)
    _sync_graph_background()
    return ("", 204)


@bp.post("/api/admin/theory/lessons")
@require_role("teacher")
def create_admin_theory_lesson():
    """创建理论课时（三级内容）。"""
    try:
        data = request.get_json(force=True) or {}
        topic_id = normalize_text(data.get("topicId"))
        if not topic_id:
            current_app.logger.warning("Create theory lesson failed: topicId is required")
            return jsonify({"error": "topicId is required"}), 400

        title = normalize_text(data.get("title")) or "未命名知识点"
        code = normalize_text(data.get("code")) or ""
        content_html = data.get("contentHtml") or ""
        order_index_raw = data.get("orderIndex")
        order_index: Optional[int] = None
        if order_index_raw not in (None, ""):
            try:
                order_index = int(order_index_raw)
            except (TypeError, ValueError):
                current_app.logger.warning("Create theory lesson failed: invalid orderIndex")
                return jsonify({"error": "orderIndex must be an integer"}), 400

        section_id = normalize_text(data.get("sectionId")) or None
        is_published = as_bool(data.get("isPublished"), default=False)
        knowledge_points = data.get("knowledgePoints") or []
        if knowledge_points and not isinstance(knowledge_points, list):
            current_app.logger.warning(
                "Create theory lesson failed: knowledgePoints must be a list"
            )
            return jsonify({"error": "knowledgePoints must be a list"}), 400

        lesson = database.create_theory_lesson(
            topic_id=topic_id,
            title=title,
            code=code,
            content_html=content_html,
            order_index=order_index,
            section_id=section_id,
            is_published=is_published,
        )
        if not lesson:
            current_app.logger.warning(f"Create theory lesson failed: topic {topic_id} not found or creation failed")
            return jsonify({"error": "Unable to create lesson"}), 400

        current_app.logger.info(f"Theory lesson created successfully: {lesson.get('id')} - {title}")
        if knowledge_points:
            _set_lesson_knowledge_points(lesson.get("id"), knowledge_points)
        _sync_graph_background()
        precompute_triggered = False
        if lesson.get("isPublished"):
            precompute_triggered = True
            _precompute_lesson_graph_async(lesson.get("id"), allow_unpublished=True)
        else:
            _invalidate_lesson_graph_cache(lesson.get("id"))
        return jsonify({"lesson": lesson, "precomputeTriggered": precompute_triggered}), 201
    except Exception as exc:
        current_app.logger.exception(f"Unexpected error creating theory lesson: {exc}")
        return jsonify({"error": "Internal server error", "detail": str(exc)}), 500


@bp.put("/api/admin/theory/lessons/<lesson_id>")
@require_role("teacher")
def update_admin_theory_lesson(lesson_id: str):
    data = request.get_json(force=True) or {}
    updates: Dict[str, object] = {}
    knowledge_points_payload: Sequence[object] = []
    knowledge_points_provided = False

    if "topicId" in data:
        updates["topic_id"] = normalize_text(data.get("topicId"))
    if "title" in data:
        title = normalize_text(data.get("title"))
        if not title:
            return jsonify({"error": "title is required"}), 400
        updates["title"] = title
    if "code" in data:
        updates["code"] = normalize_text(data.get("code"))
    if "contentHtml" in data:
        updates["content_html"] = data.get("contentHtml") or ""
    if "orderIndex" in data:
        order_value = data.get("orderIndex")
        if order_value in (None, ""):
            updates["order_index"] = None
        else:
            try:
                updates["order_index"] = int(order_value)
            except (TypeError, ValueError):
                return jsonify({"error": "orderIndex must be an integer"}), 400
    if "sectionId" in data:
        section_value = normalize_text(data.get("sectionId"))
        updates["section_id"] = section_value or None
    if "isPublished" in data:
        updates["is_published"] = as_bool(data.get("isPublished"))
    if "knowledgePoints" in data:
        raw_points = data.get("knowledgePoints") or []
        if raw_points and not isinstance(raw_points, list):
            return jsonify({"error": "knowledgePoints must be a list"}), 400
        knowledge_points_provided = True
        knowledge_points_payload = raw_points if isinstance(raw_points, list) else []

    lesson = database.update_theory_lesson(lesson_id, **updates)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404
    if knowledge_points_provided:
        _set_lesson_knowledge_points(
            lesson_id, knowledge_points_payload, allow_empty=True
        )
    _sync_graph_background()
    precompute_triggered = False
    if lesson.get("isPublished"):
        precompute_triggered = True
        _precompute_lesson_graph_async(lesson_id, allow_unpublished=True)
    else:
        _invalidate_lesson_graph_cache(lesson_id)
    return jsonify({"lesson": lesson, "precomputeTriggered": precompute_triggered})


@bp.delete("/api/admin/theory/lessons/<lesson_id>")
@require_role("teacher")
def delete_admin_theory_lesson(lesson_id: str):
    lesson = database.get_theory_lesson(lesson_id, include_unpublished=True)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404
    database.delete_theory_lesson(lesson_id)
    _sync_graph_background()
    _invalidate_lesson_graph_cache(lesson_id)
    return ("", 204)
