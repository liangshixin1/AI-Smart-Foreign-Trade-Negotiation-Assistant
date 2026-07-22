#!/usr/bin/env python3
"""简单的Excel模板检查工具"""

import sys

try:
    import openpyxl
    from io import BytesIO
except ImportError as e:
    print(f"缺少依赖: {e}")
    sys.exit(1)

# 模板表头定义（从源代码复制）
FLOW_TEMPLATE_HEADERS = [
    ("阶段名称", "name", True, "例如：询盘"),
    ("英文名称", "englishName", False, "例如：Inquiry"),
    ("阶段描述", "description", False, "简要说明该阶段的核心任务"),
    ("难度级别", "difficulty", False, "初级/中级/高级"),
    ("预计时长(天)", "estimatedDuration", False, "例如：7"),
    ("图标", "icon", False, "例如：🔍"),
    ("颜色", "color", False, "例如：#3B82F6"),
]

def analyze_template():
    """分析生成的模板结构"""
    print("=" * 80)
    print("Excel模板分析工具")
    print("=" * 80)
    print()

    # 创建一个简单的测试Excel
    print("创建测试Excel...")
    wb = openpyxl.Workbook()
    ws_flow = wb.active
    ws_flow.title = "谈判流程"

    # 写入表头（第1行）
    for col_idx, (header, field, required, example) in enumerate(FLOW_TEMPLATE_HEADERS, start=1):
        cell = ws_flow.cell(row=1, column=col_idx)
        cell.value = f"{header}{'*' if required else ''}"
        print(f"  表头列{col_idx}: {cell.value}")

    # 写入示例行（第2行）
    example_row = ["例如：询盘", "例如：Inquiry", "简要说明该阶段的核心任务", "初级/中级/高级", "例如：7", "例如：🔍", "例如：#3B82F6"]
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws_flow.cell(row=2, column=col_idx)
        cell.value = value

    # 写入示例数据（第3行开始）
    sample_stages = [
        ["询盘", "Inquiry", "买方向卖方询问商品信息和交易条件的阶段", "初级", "7", "🔍", "#3B82F6"],
        ["报盘", "Offer", "卖方向买方报价和交易条件的阶段", "中级", "5", "📊", "#10B981"],
        ["还盘", "Counter-Offer", "买卖双方针对报价和条件进行协商和调整", "高级", "10", "🔄", "#F59E0B"],
    ]

    print(f"\n写入 {len(sample_stages)} 个示例阶段（从第3行开始）:")
    for row_idx, stage_data in enumerate(sample_stages, start=3):
        for col_idx, value in enumerate(stage_data, start=1):
            cell = ws_flow.cell(row=row_idx, column=col_idx)
            cell.value = value
        print(f"  第{row_idx}行: {stage_data[0]}")

    # 保存到BytesIO
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    excel_bytes = excel_stream.getvalue()

    print(f"\n✅ Excel文件创建成功，大小: {len(excel_bytes)} 字节")

    # 重新读取并解析
    print("\n" + "=" * 80)
    print("解析测试")
    print("=" * 80)

    wb2 = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    ws = wb2["谈判流程"]

    print(f"\nSheet信息:")
    print(f"  - 名称: {ws.title}")
    print(f"  - 最大行: {ws.max_row}")
    print(f"  - 最大列: {ws.max_column}")

    print(f"\n读取所有行（从第1行开始）:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        if any(row):
            print(f"  第{row_idx}行: {row[:3]}...")  # 只显示前3列

    print(f"\n从第3行开始解析数据（跳过表头和示例行）:")
    parsed_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not any(row):
            print(f"  第{row_idx}行: (空行，跳过)")
            continue

        name = row[0] if len(row) > 0 else None
        if name and isinstance(name, str) and name.strip():
            parsed_count += 1
            print(f"  第{row_idx}行: ✅ 解析到阶段 '{name}'")
        else:
            print(f"  第{row_idx}行: ⚠️  没有阶段名称，跳过")

    print(f"\n总结: 成功解析 {parsed_count} 个阶段")

    wb2.close()

    if parsed_count == 0:
        print("\n⚠️  警告: 没有解析到任何阶段！")
        print("可能的原因:")
        print("  1. 数据从第3行开始，但没有数据")
        print("  2. 第一列（阶段名称）为空")
        print("  3. 所有行都是空行")
    else:
        print(f"\n✅ 解析成功！")

if __name__ == "__main__":
    analyze_template()
