"""智能知识图谱批量导入器 - 优化两表方案

特点：
1. 两张表：知识点主表（含关系）+ 案例库表
2. 不需要手动填写ID，系统自动生成
3. 关系用自然语言表达（必须先学、建议同时学、可对比学习）
4. 智能错误提示，推荐相似名称
5. 事务性导入，失败自动回滚
"""

from __future__ import annotations

import io
import logging
import re
import hashlib
from typing import Dict, List, Optional, BinaryIO, Tuple
from dataclasses import dataclass, field
from difflib import SequenceMatcher

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from services import knowledge_service
from services.graph_service import GraphService
from services import graph_service

LOGGER = logging.getLogger(__name__)


# ============================================
# 数据结构定义
# ============================================

@dataclass
class ValidationError:
    """验证错误"""
    severity: str  # ERROR / WARNING
    table: str  # knowledge_points / examples
    row: int
    field: Optional[str] = None
    value: Optional[str] = None
    message: str = ""
    suggestion: Optional[str] = None
    action_taken: Optional[str] = None


@dataclass
class ImportStatistics:
    """导入统计"""
    total: int = 0
    created: int = 0
    updated: int = 0
    failed: int = 0

    @property
    def success_rate(self) -> str:
        if self.total == 0:
            return "0%"
        return f"{(self.created + self.updated) / self.total * 100:.1f}%"


@dataclass
class ImportResult:
    """导入结果"""
    success: bool
    points_stats: ImportStatistics
    topics_stats: ImportStatistics
    relations_stats: ImportStatistics
    examples_stats: ImportStatistics
    stages_stats: ImportStatistics = field(default_factory=ImportStatistics)
    errors: List[ValidationError] = field(default_factory=list)
    warnings: List[ValidationError] = field(default_factory=list)
    execution_time: float = 0.0
    topics_by_stage: Dict[str, int] = field(default_factory=dict)

    def to_dict(self) -> Dict:
        return {
            "success": self.success,
            "statistics": {
                "stages": {
                    "total": self.stages_stats.total,
                    "created": self.stages_stats.created,
                    "updated": self.stages_stats.updated,
                    "failed": self.stages_stats.failed,
                    "success_rate": self.stages_stats.success_rate,
                },
                "topics": {
                    "total": self.topics_stats.total,
                    "created": self.topics_stats.created,
                    "updated": self.topics_stats.updated,
                    "failed": self.topics_stats.failed,
                    "success_rate": self.topics_stats.success_rate,
                },
                "points": {
                    "total": self.points_stats.total,
                    "created": self.points_stats.created,
                    "updated": self.points_stats.updated,
                    "failed": self.points_stats.failed,
                    "success_rate": self.points_stats.success_rate,
                },
                "relations": {
                    "total": self.relations_stats.total,
                    "created": self.relations_stats.created,
                    "failed": self.relations_stats.failed,
                    "success_rate": self.relations_stats.success_rate,
                },
                "examples": {
                    "total": self.examples_stats.total,
                    "created": self.examples_stats.created,
                    "failed": self.examples_stats.failed,
                    "success_rate": self.examples_stats.success_rate,
                },
            },
            "errors": [
                {
                    "severity": e.severity,
                    "table": e.table,
                    "row": e.row,
                    "field": e.field,
                    "value": e.value,
                    "message": e.message,
                    "suggestion": e.suggestion,
                    "action_taken": e.action_taken,
                }
                for e in self.errors
            ],
            "warnings": [
                {
                    "severity": w.severity,
                    "table": w.table,
                    "row": w.row,
                    "field": w.field,
                    "value": w.value,
                    "message": w.message,
                    "suggestion": w.suggestion,
                    "action_taken": w.action_taken,
                }
                for w in self.warnings
            ],
            "execution_time": f"{self.execution_time:.2f}s",
            "topicsByStage": self.topics_by_stage,
        }


# ============================================
# 枚举值和映射
# ============================================

# 知识点类型（中英文映射）
KNOWLEDGE_TYPES = {
    "terminology": "术语",
    "knowledge": "知识",
    "concept": "概念型",
    "skill": "技能型",
    "document": "文档型",
    "case": "案例型",
    "tool": "工具型",
    "theory": "理论型",
    "regulation": "法规型",
}

TYPE_CN_TO_EN = {
    "术语": "terminology",
    "术语型": "terminology",
    "term": "terminology",
    "terminology": "terminology",
    "知识": "knowledge",
    "知识型": "knowledge",
    "knowledge": "knowledge",
    "概念型": "concept",
    "概念": "concept",
    "技能型": "skill",
    "技能": "skill",
    "业务流程": "skill",
    "流程": "skill",
    "文档型": "document",
    "文档": "document",
    "案例型": "case",
    "案例": "case",
    "实际案例": "case",
    "工具型": "tool",
    "工具": "tool",
    "理论型": "theory",
    "理论": "theory",
    "法规型": "regulation",
    "法规": "regulation",
    "价格术语": "concept",  # 特殊映射
    "贸易术语": "concept",
}

# 分类名称映射（用于 KnowledgeCategory）
CATEGORY_NAME_ALIASES = {
    "skill": "技能",
    "技能": "技能",
    "技能型": "技能",
    "terminology": "术语",
    "term": "术语",
    "术语": "术语",
    "概念": "术语",
    "concept": "术语",
    "knowledge": "知识",
    "知识": "知识",
    "knowledgepoint": "知识",
    "document": "文档",
    "doc": "文档",
    "文档": "文档",
}

# 难度级别
DIFFICULTY_LEVELS = {
    "beginner": "初级",
    "intermediate": "中级",
    "advanced": "高级",
}

DIFFICULTY_CN_TO_EN = {
    "初级": "beginner",
    "初": "beginner",
    "入门": "beginner",
    "中级": "intermediate",
    "中": "intermediate",
    "进阶": "intermediate",
    "高级": "advanced",
    "高": "advanced",
    "深入": "advanced",
}

# 重要程度
IMPORTANCE_LEVELS = {
    "required": "必修",
    "recommended": "推荐",
    "optional": "选修",
}

IMPORTANCE_CN_TO_EN = {
    "必修": "required",
    "必学": "required",
    "核心": "required",
    "高": "required",
    "推荐": "recommended",
    "建议": "recommended",
    "中": "recommended",
    "选修": "optional",
    "扩展": "optional",
    "低": "optional",
}

# 关系类型（自然语言列名）
RELATION_COLUMNS = {
    "必须先学": "prerequisite",
    "必须先学这些": "prerequisite",
    "前置知识": "prerequisite",
    "依赖知识": "prerequisite",
    "建议同时学": "similar",
    "相关知识": "similar",
    "关联知识": "similar",
    "可对比学习": "contrast",
    "对比知识": "contrast",
    "文化敏感": "culture",
    "文化维度": "culture",
    "跨文化敏感": "culture",
    "应用场景": "application",
    "应用于": "application",
}

CULTURE_SENSITIVITY_OPTIONS = [
    "权力距离（PDI）",
    "个人主义/集体主义（IDV）",
    "男性气质/女性气质（MAS）",
    "不确定性规避（UAI）",
    "长期导向/短期导向（LTO）",
    "放纵/约束（IVR）",
    "高情境文化",
    "低情境文化",
    "以退为进策略（Door-in-the-face）",
    "得寸进尺法（Foot-in-the-door）",
]

# 实例类型
EXAMPLE_TYPES = {
    "actual_case": "实际案例",
    "email_template": "邮件模板",
    "document_template": "文档模板",
    "common_mistake": "常见错误",
    "dialogue_sample": "对话示例",
}

EXAMPLE_TYPE_CN_TO_EN = {
    "实际案例": "actual_case",
    "案例": "actual_case",
    "真实案例": "actual_case",
    "邮件模板": "email_template",
    "邮件": "email_template",
    "文档模板": "document_template",
    "文档": "document_template",
    "模板": "document_template",
    "常见错误": "common_mistake",
    "错误": "common_mistake",
    "典型错误": "common_mistake",
    "对话示例": "dialogue_sample",
    "对话": "dialogue_sample",
    "谈判对话": "dialogue_sample",
}

# 词汇角色映射
LEX_ROLE_ALIASES = {
    "lexeme": "lexeme",
    "词条": "lexeme",
    "术语": "lexeme",
    "term": "lexeme",
    "collocation": "collocation",
    "搭配": "collocation",
    "短语": "collocation",
    "construction": "construction",
    "句式": "construction",
    "句型": "construction",
}


# ============================================
# 模板定义
# ============================================

# Sheet 1: 谈判流程表模板（Stage节点定义）
FLOW_TEMPLATE_HEADERS = [
    ("阶段名称", "name", True, "例如：询盘"),
    ("英文名称", "englishName", False, "例如：Inquiry"),
    ("阶段描述", "description", False, "简要说明该阶段的核心任务"),
    ("难度级别", "difficulty", False, "初级/中级/高级"),
    ("预计时长(天)", "estimatedDuration", False, "例如：7"),
    ("图标", "icon", False, "例如：🔍"),
    ("颜色", "color", False, "例如：#3B82F6"),
]

# Sheet 2: 知识点主表模板（合并关系，新增"所属阶段"列）
POINTS_TEMPLATE_HEADERS = [
    ("章节", "chapter", False, "例如：第一章 询盘"),
    ("知识点名称", "name", True, "例如：询盘基本流程"),
    ("所属阶段", "stage", False, "从下拉列表选择（询盘/报盘/还盘...）"),
    ("二级主题", "topic", False, "例如：采购需求分析与供应商甄选"),
    ("知识点类型", "type", False, "术语/技能/知识（或概念型/技能型/文档型等）"),
    ("难度", "difficulty", False, "初级/中级/高级"),
    ("重要性", "importance", False, "必修/推荐/选修（或：高/中/低）"),
    ("预计学时(分钟)", "estimatedMinutes", False, "例如：30"),
    ("内容简介", "summary", False, "一句话描述"),
    ("详细描述", "description", False, "详细说明"),
    ("关键词", "keywords", False, "用逗号分隔"),
    ("必须先学", "prerequisite", False, "填写知识点名称，多个用分号分隔"),
    ("建议同时学", "similar", False, "填写知识点名称，多个用分号分隔"),
    ("可对比学习", "contrast", False, "填写知识点名称，多个用分号分隔"),
    ("文化敏感", "culture", False, "从下拉选择文化维度，多个用分号分隔"),
]

# Sheet 3: 案例库表模板
EXAMPLES_TEMPLATE_HEADERS = [
    ("关联知识点", "knowledge_point_name", True, "填写知识点名称"),
    ("案例类型", "type", True, "实际案例/邮件模板/文档模板/常见错误/对话示例"),
    ("案例标题", "title", True, "简短标题"),
    ("案例内容", "content", True, "详细内容"),
    ("关联练习关卡", "practice_id", False, "例如：6-1"),
]

# Sheet 4: 词汇网络表模板
LEXICON_TEMPLATE_HEADERS = [
    ("关联知识点名称", "anchor", True, "必须存在于知识点主表"),
    ("词汇项", "lex_item", True, "例如：offer a discount"),
    ("词汇项类型", "lex_role", True, "词条/搭配/句式"),
    ("词族", "family", False, "如: quote-* 或 ship-*"),
    ("语义类", "semantic_class", False, "如: Pricing-Concession"),
    ("语义槽位", "slots", False, "多值用分号或逗号分隔"),
    ("语气/强度", "tone", False, "softer/neutral/stronger"),
    ("思政元素", "civic_tags", False, "多选: Win-Win; Integrity; Dignity; Compliance; Zero-Sum; Dishonesty; Disrespect; Non-Compliance"),
    ("地道/标准", "idiomatic", False, "是/否，可空"),
    ("搭配组成词", "components", False, "用于搭配与词条关联"),
    ("备注", "note", False, "补充说明"),
]


# ============================================
# 通用辅助
# ============================================

def _normalize_multi_value(value: Optional[str]) -> List[str]:
    """兼容中英文分隔符, 去重并保持顺序。"""
    if value is None:
        return []
    parts = re.split(r"[，,;；、]+", str(value))
    normalized: List[str] = []
    for part in parts:
        token = part.strip()
        if token and token not in normalized:
            normalized.append(token)
    return normalized


def _is_template_hint_value(value: object) -> bool:
    """判断单元格是否是模板说明文字，而不是用户数据。"""
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    return (
        text.startswith("例如")
        or text.startswith("示例")
        or text.startswith("填写")
        or text.startswith("从下拉列表选择")
        or text.startswith("简要说明")
        or text in {
            "一句话描述",
            "详细说明",
            "用逗号分隔",
            "中文名称",
            "英文名称",
            "阶段名称",
            "阶段名称*",
            "阶段描述",
            "阶段",
            "难度级别",
            "预计时长(天)",
            "图标",
            "颜色",
            "章节",
            "知识点名称",
            "知识点名称*",
            "所属阶段",
            "二级主题",
            "知识点类型",
            "难度",
            "重要性",
            "内容简介",
            "详细描述",
            "关键词",
        }
    )


def _is_template_hint_row(row: tuple, key_indexes: List[int]) -> bool:
    """跳过模板中用于提示填写方法的行，避免导入成真实节点。"""
    values = [row[idx] for idx in key_indexes if idx < len(row)]
    visible_values = [value for value in values if value is not None and str(value).strip()]
    if not visible_values:
        return False
    return all(_is_template_hint_value(value) for value in visible_values)


def _to_bool_flag(value: Optional[str]) -> Optional[bool]:
    """将常见是/否文本转为布尔, 其他返回None。"""
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"yes", "y", "true", "1", "是", "对", "标准", "地道"}:
        return True
    if text in {"no", "n", "false", "0", "否", "非", "不标准", "不地道"}:
        return False
    return None


# ============================================
# 核心导入器类
# ============================================

class KnowledgeGraphBatchImporter:
    """智能知识图谱批量导入器（支持多节点类型）"""

    def __init__(self, graph_service: Optional[GraphService] = None):
        self.graph_service = graph_service or GraphService()
        self.known_point_names: List[str] = []  # 用于智能错误提示
        self.known_stage_names: List[str] = []  # 已知阶段名称列表

    def import_from_two_tables(
        self,
        points_file: BinaryIO,
        examples_file: Optional[BinaryIO] = None,
        mode: str = "merge",
        created_by: str = "batch-import",
    ) -> ImportResult:
        """
        两表联动导入

        Args:
            points_file: 知识点主表Excel文件
            examples_file: 案例库表Excel文件（可选）
            mode: 导入模式 (merge/replace)
            created_by: 操作用户标识

        Returns:
            ImportResult对象
        """
        import time
        start_time = time.time()

        result = ImportResult(
            success=False,
            points_stats=ImportStatistics(),
            topics_stats=ImportStatistics(),
            relations_stats=ImportStatistics(),
            examples_stats=ImportStatistics(),
        )

        try:
            # Phase 1: 解析表格数据
            LOGGER.info("Phase 1: 解析知识点表...")
            points_data, parse_errors = self._parse_points_table(points_file)
            result.errors.extend(parse_errors)

            if not points_data:
                result.errors.append(ValidationError(
                    severity="ERROR",
                    table="knowledge_points",
                    row=0,
                    message="知识点表中没有找到有效数据，请检查文件格式",
                ))
                return result

            # 建立知识点名称列表（用于后续验证）
            self.known_point_names = [p["name"] for p in points_data]

            examples_data = []
            if examples_file:
                LOGGER.info("Phase 1: 解析案例库表...")
                examples_data, example_errors = self._parse_examples_table(examples_file)
                result.errors.extend(example_errors)

            # Phase 2: 验证数据
            LOGGER.info("Phase 2: 验证数据...")
            validation_errors, validation_warnings = self._validate_all_data(
                points_data, examples_data
            )
            result.errors.extend(validation_errors)
            result.warnings.extend(validation_warnings)

            # 如果有致命错误，停止导入
            if any(e.severity == "ERROR" for e in result.errors):
                LOGGER.error(f"发现 {len([e for e in result.errors if e.severity == 'ERROR'])} 个错误，停止导入")
                result.execution_time = time.time() - start_time
                return result

            # Phase 3: 事务性导入
            LOGGER.info("Phase 3: 开始导入...")
            result.points_stats, result.relations_stats, result.examples_stats = (
                self._import_with_transaction(points_data, examples_data, created_by)
            )

            result.success = True
            result.execution_time = time.time() - start_time

            LOGGER.info(
                f"导入完成: 知识点 {result.points_stats.created}创建/"
                f"{result.points_stats.updated}更新, "
                f"关系 {result.relations_stats.created}创建, "
                f"案例 {result.examples_stats.created}创建"
            )

        except Exception as e:
            LOGGER.exception("导入失败")
            result.errors.append(ValidationError(
                severity="ERROR",
                table="system",
                row=0,
                message=f"系统错误: {str(e)}",
            ))
            result.execution_time = time.time() - start_time

        return result

    def import_from_three_sheets(
        self,
        excel_file: BinaryIO,
        mode: str = "merge",
        created_by: str = "batch-import",
    ) -> ImportResult:
        """
        三表联动导入（新版，支持多节点类型）

        Args:
            excel_file: 包含三个sheet的Excel文件
                - Sheet 1: 谈判流程（Stage节点）
                - Sheet 2: 知识点主表（支持"所属阶段"）
                - Sheet 3: 案例库（可选）
                - Sheet 4: 词汇网络表（可选，词条/搭配/句式）
            mode: 导入模式 (merge/replace)
            created_by: 操作用户标识

        Returns:
            ImportResult对象（扩展了stages_stats）
        """
        import time
        start_time = time.time()
        stages_stats = ImportStatistics()

        result = ImportResult(
            success=False,
            stages_stats=stages_stats,
            topics_stats=ImportStatistics(),
            points_stats=ImportStatistics(),
            relations_stats=ImportStatistics(),
            examples_stats=ImportStatistics(),
            topics_by_stage={},
        )

        # 添加stages统计
        stages_stats = result.stages_stats

        try:
            # Phase 1: 解析三个Sheet的数据
            LOGGER.info("Phase 1: 解析Excel文件...")

            # 读取文件内容到内存，避免多次读取同一文件流时出现问题
            if hasattr(excel_file, 'read'):
                file_content = excel_file.read()
            else:
                file_content = excel_file

            # 1.1 解析谈判流程表（Sheet 1）
            flow_data, flow_errors = self._parse_flow_table(io.BytesIO(file_content))
            result.errors.extend(flow_errors)

            # 建立阶段名称列表
            self.known_stage_names = [f["name"] for f in flow_data]
            LOGGER.info(f"解析到 {len(flow_data)} 个阶段: {self.known_stage_names}")

            # 检查：如果没有解析到阶段，添加明确的错误提示
            if len(flow_data) == 0:
                result.errors.append(ValidationError(
                    severity="ERROR",
                    table="flow",
                    row=0,
                    message="未在'谈判流程'Sheet中找到任何阶段数据",
                    suggestion="请确保在'谈判流程'Sheet的第3行及以后添加阶段数据，或保留模板中的示例数据",
                ))
                LOGGER.error("未找到任何阶段数据，停止导入")
                result.execution_time = time.time() - start_time
                return result

            # 1.2 解析知识点表（Sheet 2）
            points_data, points_errors = self._parse_points_table_from_workbook(io.BytesIO(file_content), sheet_name="知识点主表")
            result.errors.extend(points_errors)

            # 建立知识点名称列表
            self.known_point_names = [p["name"] for p in points_data]
            LOGGER.info(f"解析到 {len(points_data)} 个知识点")

            # 1.3 解析案例库表（Sheet 3，可选）
            examples_data, examples_errors = self._parse_examples_table_from_workbook(io.BytesIO(file_content), sheet_name="案例库表")
            result.errors.extend(examples_errors)
            LOGGER.info(f"解析到 {len(examples_data)} 个案例")

            # 1.4 解析词汇网络表（Sheet 4，可选）
            lexicon_data, lexicon_errors, lexicon_warnings = self._parse_lexicon_table_from_workbook(io.BytesIO(file_content), sheet_name="词汇网络表")
            result.errors.extend(lexicon_errors)
            result.warnings.extend(lexicon_warnings)
            LOGGER.info(f"解析到 {len(lexicon_data)} 条词汇记录")

            # Phase 2: 验证数据
            LOGGER.info("Phase 2: 验证数据...")
            validation_errors, validation_warnings = self._validate_three_sheets_data(
                flow_data, points_data, examples_data, lexicon_data
            )
            result.errors.extend(validation_errors)
            result.warnings.extend(validation_warnings)

            # 如果有致命错误，停止导入
            if any(e.severity == "ERROR" for e in result.errors):
                LOGGER.error(f"发现 {len([e for e in result.errors if e.severity == 'ERROR'])} 个错误，停止导入")
                result.execution_time = time.time() - start_time
                return result

            # Phase 3: 事务性导入
            LOGGER.info("Phase 3: 开始导入...")
            (
                stages_stats,
                result.topics_stats,
                result.points_stats,
                result.relations_stats,
                result.examples_stats,
                result.topics_by_stage,
                transaction_errors,
            ) = self._import_three_sheets_with_transaction(flow_data, points_data, examples_data, lexicon_data, created_by)
            result.stages_stats = stages_stats
            result.errors.extend(transaction_errors)

            result.success = True
            result.execution_time = time.time() - start_time

            LOGGER.info(
                f"导入完成: 阶段 {stages_stats.created}创建, "
                f"知识点 {result.points_stats.created}创建/{result.points_stats.updated}更新, "
                f"关系 {result.relations_stats.created}创建, "
                f"案例 {result.examples_stats.created}创建"
            )

            # 将stages_stats添加到result中（需要扩展ImportResult）
            # 这里暂时记录在日志中
            LOGGER.info(f"Stage统计: {stages_stats}")

        except Exception as e:
            LOGGER.exception("三表导入失败")
            result.errors.append(ValidationError(
                severity="ERROR",
                table="system",
                row=0,
                message=f"系统错误: {str(e)}",
            ))
            result.execution_time = time.time() - start_time

        return result

    # ========================================
    # Phase 1: 数据解析
    # ========================================

    def _parse_flow_table(self, file: BinaryIO) -> Tuple[List[Dict], List[ValidationError]]:
        """解析谈判流程表（Sheet 1）"""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl未安装，无法读取Excel文件")

        errors = []
        flow_data = []

        try:
            if hasattr(file, 'seek'):
                file.seek(0)

            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

            # 尝试找到谈判流程表
            sheet_name = None
            if "谈判流程" in wb.sheetnames:
                sheet_name = "谈判流程"
            elif len(wb.sheetnames) > 0:
                sheet_name = wb.sheetnames[0]  # 使用第一个sheet
            else:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="flow",
                    row=0,
                    message="未找到谈判流程表",
                ))
                return [], errors

            ws = wb[sheet_name]

            # 读取表头
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h.replace('*', '').strip() if h else '' for h in header_row]

            # 构建列名映射
            field_map = {}
            for idx, header in enumerate(headers):
                for template_header, field, _, _ in FLOW_TEMPLATE_HEADERS:
                    if header == template_header:
                        field_map[field] = idx
                        break

            # 检查必填字段
            if "name" not in field_map:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="flow",
                    row=1,
                    field="阶段名称",
                    message="缺少必填列：阶段名称",
                ))
                return [], errors

            hint_indexes = [field_map.get("name", 0), field_map.get("description", 2)]

            # 读取数据行：教师模板可能从第2行直接给真实阶段，也可能第2行是提示行。
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                if _is_template_hint_row(row, hint_indexes):
                    continue

                try:
                    stage = self._parse_flow_row(row, field_map, row_idx)
                    if stage:
                        flow_data.append(stage)
                except Exception as e:
                    errors.append(ValidationError(
                        severity="WARNING",
                        table="flow",
                        row=row_idx,
                        message=f"解析行数据失败: {str(e)}",
                        action_taken="SKIP_ROW",
                    ))

            wb.close()

        except Exception as e:
            errors.append(ValidationError(
                severity="ERROR",
                table="flow",
                row=0,
                message=f"读取Excel文件失败: {str(e)}",
            ))

        for order_index, stage in enumerate(flow_data, start=1):
            stage["_order"] = order_index
            stage["orderIndex"] = order_index

        return flow_data, errors

    def _parse_flow_row(self, row: tuple, field_map: Dict, row_idx: int) -> Optional[Dict]:
        """解析单行谈判流程数据"""
        stage = {}

        for field, col_idx in field_map.items():
            if col_idx >= len(row):
                continue

            value = row[col_idx]
            if value is None or (isinstance(value, str) and not value.strip()):
                continue

            # 类型转换
            if field == "estimatedDuration":
                try:
                    stage[field] = int(value)
                except (ValueError, TypeError):
                    stage[field] = 7  # 默认7天
            elif field == "difficulty":
                # 中英文转换
                stage[field] = DIFFICULTY_CN_TO_EN.get(str(value).strip(), str(value).strip())
            else:
                stage[field] = str(value).strip()

        # 必须有名称
        if not stage.get("name"):
            return None

        # 添加行号
        stage["_row"] = row_idx
        # 添加顺序（根据Excel中的行号自动生成）
        stage["_order"] = row_idx - 2  # 减去表头和示例行

        return stage

    def _parse_points_table_from_workbook(
        self, file: BinaryIO, sheet_name: str = "知识点主表"
    ) -> Tuple[List[Dict], List[ValidationError]]:
        """从工作簿中解析知识点表（支持"所属阶段"列）"""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl未安装")

        errors = []
        points_data = []

        try:
            if hasattr(file, 'seek'):
                file.seek(0)

            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

            # 查找知识点表
            if sheet_name not in wb.sheetnames:
                # 尝试其他可能的名称
                for possible_name in ["知识点主表", "知识点", "Sheet2"]:
                    if possible_name in wb.sheetnames:
                        sheet_name = possible_name
                        break
                else:
                    errors.append(ValidationError(
                        severity="WARNING",
                        table="knowledge_points",
                        row=0,
                        message=f"未找到'{sheet_name}'表，知识点导入将跳过",
                    ))
                    return [], errors

            ws = wb[sheet_name]

            # 读取表头
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h.replace('*', '').strip() if h else '' for h in header_row]

            # 构建列名映射
            field_map = {}
            for idx, header in enumerate(headers):
                for template_header, field, _, _ in POINTS_TEMPLATE_HEADERS:
                    if header == template_header:
                        field_map[field] = idx
                        break
                if header in RELATION_COLUMNS:
                    field_map[header] = idx

            # 检查必填字段
            if "name" not in field_map:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="knowledge_points",
                    row=1,
                    field="知识点名称",
                    message="缺少必填列：知识点名称",
                ))
                return [], errors

            hint_indexes = [field_map.get("name", 1), field_map.get("stage", 2), field_map.get("summary", 8)]

            # 读取数据行：跳过模板说明/示例行，只保留真实知识点。
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                if _is_template_hint_row(row, hint_indexes):
                    continue

                try:
                    point = self._parse_point_row(row, field_map, headers, row_idx)
                    if point:
                        points_data.append(point)
                except Exception as e:
                    errors.append(ValidationError(
                        severity="WARNING",
                        table="knowledge_points",
                        row=row_idx,
                        message=f"解析行数据失败: {str(e)}",
                        action_taken="SKIP_ROW",
                    ))

            wb.close()

        except Exception as e:
            errors.append(ValidationError(
                severity="ERROR",
                table="knowledge_points",
                row=0,
                message=f"读取Excel文件失败: {str(e)}",
            ))

        return points_data, errors

    def _parse_examples_table_from_workbook(
        self, file: BinaryIO, sheet_name: str = "案例库表"
    ) -> Tuple[List[Dict], List[ValidationError]]:
        """从工作簿中解析案例库表"""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl未安装")

        errors = []
        examples_data = []

        try:
            if hasattr(file, 'seek'):
                file.seek(0)

            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

            # 查找案例库表（可选）
            if sheet_name not in wb.sheetnames:
                # 尝试其他可能的名称
                for possible_name in ["案例库表", "案例库", "Sheet3"]:
                    if possible_name in wb.sheetnames:
                        sheet_name = possible_name
                        break
                else:
                    # 案例库表是可选的，没有也不报错
                    return [], errors

            ws = wb[sheet_name]

            # 读取表头
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h.replace('*', '').strip() if h else '' for h in header_row]

            # 构建列名映射
            field_map = {}
            for idx, header in enumerate(headers):
                for template_header, field, _, _ in EXAMPLES_TEMPLATE_HEADERS:
                    if header == template_header:
                        field_map[field] = idx
                        break

            # 检查必填字段
            required_fields = ["knowledge_point_name", "type", "title", "content"]
            missing_fields = [f for f in required_fields if f not in field_map]
            if missing_fields:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="examples",
                    row=1,
                    message=f"缺少必填列: {', '.join(missing_fields)}",
                ))
                return [], errors

            # 读取数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not any(row):
                    continue

                try:
                    example = {}
                    for field, col_idx in field_map.items():
                        value = row[col_idx] if col_idx < len(row) else None
                        if value is None or (isinstance(value, str) and not value.strip()):
                            continue

                        if field == "type":
                            example[field] = EXAMPLE_TYPE_CN_TO_EN.get(
                                str(value).strip(), str(value).strip()
                            )
                        else:
                            example[field] = str(value).strip()

                    if all(example.get(f) for f in required_fields):
                        example["_row"] = row_idx
                        examples_data.append(example)
                    else:
                        errors.append(ValidationError(
                            severity="WARNING",
                            table="examples",
                            row=row_idx,
                            message="缺少必填字段，已跳过",
                            action_taken="SKIP_ROW",
                        ))

                except Exception as e:
                    errors.append(ValidationError(
                        severity="WARNING",
                        table="examples",
                        row=row_idx,
                        message=f"解析失败: {str(e)}",
                        action_taken="SKIP_ROW",
                    ))

            wb.close()

        except Exception as e:
            errors.append(ValidationError(
                severity="ERROR",
                table="examples",
                row=0,
                message=f"读取Excel文件失败: {str(e)}",
            ))

        return examples_data, errors

    def _parse_lexicon_table_from_workbook(
        self, file: BinaryIO, sheet_name: str = "词汇网络表"
    ) -> Tuple[List[Dict], List[ValidationError], List[ValidationError]]:
        """解析词汇网络表（可选）"""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl未安装")

        errors: List[ValidationError] = []
        warnings: List[ValidationError] = []
        lexicon: List[Dict] = []

        try:
            if hasattr(file, "seek"):
                file.seek(0)

            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                for candidate in ["词汇网络表", "词汇网络", "Lexicon", "Sheet4"]:
                    if candidate in wb.sheetnames:
                        sheet_name = candidate
                        break
                else:
                    return [], errors, warnings

            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h.replace("*", "").strip() if h else "" for h in header_row]

            field_map: Dict[str, int] = {}
            for idx, header in enumerate(headers):
                for template_header, field, _, _ in LEXICON_TEMPLATE_HEADERS:
                    if header == template_header:
                        field_map[field] = idx
                        break

            missing_required = [f for f in ("anchor", "lex_item", "lex_role") if f not in field_map]
            if missing_required:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="lexicon",
                    row=1,
                    message=f"缺少必填列: {', '.join(missing_required)}",
                ))
                return [], errors, warnings

            seen_pairs = set()

            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not any(row):
                    continue

                anchor = str(row[field_map["anchor"]]).strip() if len(row) > field_map["anchor"] and row[field_map["anchor"]] else ""
                lex_item = str(row[field_map["lex_item"]]).strip() if len(row) > field_map["lex_item"] and row[field_map["lex_item"]] else ""
                raw_role = str(row[field_map["lex_role"]]).strip() if len(row) > field_map["lex_role"] and row[field_map["lex_role"]] else ""

                if not (anchor and lex_item and raw_role):
                    continue

                norm_role = LEX_ROLE_ALIASES.get(raw_role.lower()) or LEX_ROLE_ALIASES.get(raw_role)
                if not norm_role:
                    errors.append(ValidationError(
                        severity="ERROR",
                        table="lexicon",
                        row=row_idx,
                        field="词汇项类型",
                        value=raw_role,
                        message="词汇项类型必须是 词条/搭配/句式（或对应英文）",
                    ))
                    continue

                entry = {
                    "_row": row_idx,
                    "anchor": anchor,
                    "lex_item": lex_item,
                    "lex_role": norm_role,
                    "family": "",
                    "semantic_class": "",
                    "slots": [],
                    "tone": "",
                    "idiomatic": None,
                    "civic_tags": [],
                    "components": [],
                    "note": "",
                }

                if "family" in field_map and len(row) > field_map["family"] and row[field_map["family"]]:
                    entry["family"] = str(row[field_map["family"]]).strip()

                if "semantic_class" in field_map and len(row) > field_map["semantic_class"] and row[field_map["semantic_class"]]:
                    entry["semantic_class"] = str(row[field_map["semantic_class"]]).strip()

                if "slots" in field_map and len(row) > field_map["slots"] and row[field_map["slots"]]:
                    entry["slots"] = _normalize_multi_value(row[field_map["slots"]])

                if "tone" in field_map and len(row) > field_map["tone"] and row[field_map["tone"]]:
                    entry["tone"] = str(row[field_map["tone"]]).strip().lower()

                if "idiomatic" in field_map and len(row) > field_map["idiomatic"] and row[field_map["idiomatic"]]:
                    entry["idiomatic"] = _to_bool_flag(row[field_map["idiomatic"]])

                if "civic_tags" in field_map and len(row) > field_map["civic_tags"] and row[field_map["civic_tags"]]:
                    entry["civic_tags"] = _normalize_multi_value(row[field_map["civic_tags"]])

                if "components" in field_map and len(row) > field_map["components"] and row[field_map["components"]]:
                    entry["components"] = _normalize_multi_value(row[field_map["components"]])

                if "note" in field_map and len(row) > field_map["note"] and row[field_map["note"]]:
                    entry["note"] = str(row[field_map["note"]]).strip()

                slots_signature = ";".join(entry["slots"])
                pair_key = (anchor, lex_item, entry["semantic_class"], slots_signature)
                if pair_key in seen_pairs:
                    warnings.append(ValidationError(
                        severity="WARNING",
                        table="lexicon",
                        row=row_idx,
                        field="词汇项",
                        value=f"{lex_item} ({entry['semantic_class']})",
                        message="发现重复的词汇项，已自动去重",
                        action_taken="SKIP_DUPLICATE",
                    ))
                    continue
                seen_pairs.add(pair_key)

                lexicon.append(entry)

        except Exception as e:
            errors.append(ValidationError(
                severity="ERROR",
                table="lexicon",
                row=0,
                message=f"解析词汇网络表失败: {e}",
            ))

        return lexicon, errors, warnings

    def _parse_points_table(self, file: BinaryIO) -> Tuple[List[Dict], List[ValidationError]]:
        """解析知识点主表"""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl未安装，无法读取Excel文件")

        errors = []
        points_data = []

        try:
            if hasattr(file, 'seek'):
                file.seek(0)

            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active

            # 读取表头
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h.replace('*', '').strip() if h else '' for h in header_row]

            # 构建列名映射
            field_map = {}
            for idx, header in enumerate(headers):
                # 匹配模板列名
                for template_header, field, _, _ in POINTS_TEMPLATE_HEADERS:
                    if header == template_header:
                        field_map[field] = idx
                        break
                # 也匹配关系列名
                if header in RELATION_COLUMNS:
                    field_map[header] = idx

            # 检查必填字段
            if "name" not in field_map:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="knowledge_points",
                    row=1,
                    field="知识点名称",
                    message="缺少必填列：知识点名称",
                ))
                return [], errors

            # 读取数据行（跳过表头和示例行）
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                # 跳过空行
                if not any(row):
                    continue

                try:
                    point = self._parse_point_row(row, field_map, headers, row_idx)
                    if point:
                        points_data.append(point)
                except Exception as e:
                    errors.append(ValidationError(
                        severity="WARNING",
                        table="knowledge_points",
                        row=row_idx,
                        message=f"解析行数据失败: {str(e)}",
                        action_taken="SKIP_ROW",
                    ))

            wb.close()

        except Exception as e:
            errors.append(ValidationError(
                severity="ERROR",
                table="knowledge_points",
                row=0,
                message=f"读取Excel文件失败: {str(e)}",
            ))

        return points_data, errors

    def _parse_point_row(
        self, row: tuple, field_map: Dict, headers: List[str], row_idx: int
    ) -> Optional[Dict]:
        """解析单行知识点数据"""
        point = {}

        # 基本字段
        for field, col_idx in field_map.items():
            if field in RELATION_COLUMNS.values():
                continue  # 关系字段稍后处理

            value = row[col_idx] if col_idx < len(row) else None
            if value is None or (isinstance(value, str) and not value.strip()):
                continue

            # 类型转换
            if field == "estimatedMinutes":
                try:
                    point[field] = int(value)
                except (ValueError, TypeError):
                    point[field] = 30  # 默认值
            elif field == "keywords":
                if isinstance(value, str):
                    point[field] = [v.strip() for v in value.split(",") if v.strip()]
                else:
                    point[field] = []
            elif field == "type":
                # 中英文转换
                point[field] = TYPE_CN_TO_EN.get(str(value).strip(), str(value).strip())
            elif field == "difficulty":
                point[field] = DIFFICULTY_CN_TO_EN.get(str(value).strip(), str(value).strip())
            elif field == "importance":
                point[field] = IMPORTANCE_CN_TO_EN.get(str(value).strip(), str(value).strip())
            else:
                point[field] = str(value).strip()

        # 解析关系列
        relations = {}
        for header, col_idx in [(h, field_map.get(h)) for h in headers if h in RELATION_COLUMNS]:
            if col_idx is None or col_idx >= len(row):
                continue

            value = row[col_idx]
            if value and isinstance(value, str) and value.strip():
                relation_type = RELATION_COLUMNS[header]
                related_names = _normalize_multi_value(value)
                if related_names:
                    relations[relation_type] = related_names

        if relations:
            point["_relations"] = relations

        # 必须有名称
        if not point.get("name"):
            return None

        # 添加行号（用于错误提示）
        point["_row"] = row_idx

        return point

    def _parse_examples_table(self, file: BinaryIO) -> Tuple[List[Dict], List[ValidationError]]:
        """解析案例库表"""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl未安装")

        errors = []
        examples_data = []

        try:
            if hasattr(file, 'seek'):
                file.seek(0)

            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active

            # 读取表头
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h.replace('*', '').strip() if h else '' for h in header_row]

            # 构建列名映射
            field_map = {}
            for idx, header in enumerate(headers):
                for template_header, field, _, _ in EXAMPLES_TEMPLATE_HEADERS:
                    if header == template_header:
                        field_map[field] = idx
                        break

            # 检查必填字段
            required_fields = ["knowledge_point_name", "type", "title", "content"]
            missing_fields = [f for f in required_fields if f not in field_map]
            if missing_fields:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="examples",
                    row=1,
                    message=f"缺少必填列: {', '.join(missing_fields)}",
                ))
                return [], errors

            # 读取数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not any(row):
                    continue

                try:
                    example = {}
                    for field, col_idx in field_map.items():
                        value = row[col_idx] if col_idx < len(row) else None
                        if value is None or (isinstance(value, str) and not value.strip()):
                            continue

                        # 类型转换
                        if field == "type":
                            example[field] = EXAMPLE_TYPE_CN_TO_EN.get(
                                str(value).strip(), str(value).strip()
                            )
                        else:
                            example[field] = str(value).strip()

                    # 必填字段检查
                    if all(example.get(f) for f in required_fields):
                        example["_row"] = row_idx
                        examples_data.append(example)
                    else:
                        errors.append(ValidationError(
                            severity="WARNING",
                            table="examples",
                            row=row_idx,
                            message="缺少必填字段，已跳过",
                            action_taken="SKIP_ROW",
                        ))

                except Exception as e:
                    errors.append(ValidationError(
                        severity="WARNING",
                        table="examples",
                        row=row_idx,
                        message=f"解析失败: {str(e)}",
                        action_taken="SKIP_ROW",
                    ))

            wb.close()

        except Exception as e:
            errors.append(ValidationError(
                severity="ERROR",
                table="examples",
                row=0,
                message=f"读取Excel文件失败: {str(e)}",
            ))

        return examples_data, errors

    # ========================================
    # Phase 2: 数据验证
    # ========================================

    def _validate_all_data(
        self, points_data: List[Dict], examples_data: List[Dict]
    ) -> Tuple[List[ValidationError], List[ValidationError]]:
        """验证所有数据"""
        errors = []
        warnings = []

        # 1. 验证知识点数据
        point_errors, point_warnings = self._validate_points(points_data)
        errors.extend(point_errors)
        warnings.extend(point_warnings)

        # 2. 验证案例数据
        if examples_data:
            example_errors, example_warnings = self._validate_examples(examples_data)
            errors.extend(example_errors)
            warnings.extend(example_warnings)

        return errors, warnings

    def _validate_points(self, points: List[Dict]) -> Tuple[List[ValidationError], List[ValidationError]]:
        """验证知识点数据"""
        errors = []
        warnings = []

        # 检查名称唯一性
        name_counts = {}
        for point in points:
            name = point.get("name")
            if name:
                name_counts[name] = name_counts.get(name, 0) + 1

        duplicates = {name: count for name, count in name_counts.items() if count > 1}
        if duplicates:
            for name, count in duplicates.items():
                errors.append(ValidationError(
                    severity="ERROR",
                    table="knowledge_points",
                    row=0,
                    field="知识点名称",
                    value=name,
                    message=f"知识点名称重复{count}次: '{name}'",
                    suggestion="每个知识点名称必须唯一，请修改重复的名称",
                ))

        # 验证枚举值
        for point in points:
            row = point.get("_row", 0)

            # 类型
            if point.get("type") and point["type"] not in KNOWLEDGE_TYPES:
                warnings.append(ValidationError(
                    severity="WARNING",
                    table="knowledge_points",
                    row=row,
                    field="知识点类型",
                    value=point["type"],
                    message=f"未知的知识点类型",
                    suggestion=f"建议使用: {', '.join(KNOWLEDGE_TYPES.values())}",
                    action_taken="使用默认值 'concept'",
                ))
                point["type"] = "concept"

            # 难度
            if point.get("difficulty") and point["difficulty"] not in DIFFICULTY_LEVELS:
                warnings.append(ValidationError(
                    severity="WARNING",
                    table="knowledge_points",
                    row=row,
                    field="难度",
                    value=point["difficulty"],
                    message="未知的难度级别",
                    suggestion=f"建议使用: {', '.join(DIFFICULTY_LEVELS.values())}",
                    action_taken="使用默认值 'beginner'",
                ))
                point["difficulty"] = "beginner"

            # 重要性
            if point.get("importance") and point["importance"] not in IMPORTANCE_LEVELS:
                warnings.append(ValidationError(
                    severity="WARNING",
                    table="knowledge_points",
                    row=row,
                    field="重要性",
                    value=point["importance"],
                    message="未知的重要性级别",
                    suggestion=f"建议使用: {', '.join(IMPORTANCE_LEVELS.values())}",
                    action_taken="使用默认值 'recommended'",
                ))
                point["importance"] = "recommended"

            # 验证关系引用
            if "_relations" in point:
                for relation_type, related_names in point["_relations"].items():
                    if relation_type == "culture":
                        continue
                    for related_name in related_names:
                        if related_name not in self.known_point_names:
                            # 查找相似名称
                            suggestions = self._find_similar_names(related_name)
                            errors.append(ValidationError(
                                severity="ERROR",
                                table="knowledge_points",
                                row=row,
                                field="关系引用",
                                value=related_name,
                                message=f"引用的知识点 '{related_name}' 不存在",
                                suggestion=f"你是不是想写: {', '.join(suggestions[:3])}" if suggestions else "请检查知识点名称是否正确",
                            ))

        return errors, warnings

    def _validate_examples(self, examples: List[Dict]) -> Tuple[List[ValidationError], List[ValidationError]]:
        """验证案例数据"""
        errors = []
        warnings = []

        for example in examples:
            row = example.get("_row", 0)
            kp_name = example.get("knowledge_point_name")

            # 检查关联的知识点是否存在
            if kp_name and kp_name not in self.known_point_names:
                suggestions = self._find_similar_names(kp_name)
                errors.append(ValidationError(
                    severity="ERROR",
                    table="examples",
                    row=row,
                    field="关联知识点",
                    value=kp_name,
                    message=f"关联的知识点 '{kp_name}' 不存在",
                    suggestion=f"你是不是想写: {', '.join(suggestions[:3])}" if suggestions else "请先在知识点表中添加该知识点",
                ))

            # 验证案例类型
            if example.get("type") and example["type"] not in EXAMPLE_TYPES:
                warnings.append(ValidationError(
                    severity="WARNING",
                    table="examples",
                    row=row,
                    field="案例类型",
                    value=example["type"],
                    message="未知的案例类型",
                    suggestion=f"建议使用: {', '.join(EXAMPLE_TYPES.values())}",
                    action_taken="使用默认值 'actual_case'",
                ))
                example["type"] = "actual_case"

        return errors, warnings

    def _validate_three_sheets_data(
        self,
        flow_data: List[Dict],
        points_data: List[Dict],
        examples_data: List[Dict],
        lexicon_data: Optional[List[Dict]] = None,
    ) -> Tuple[List[ValidationError], List[ValidationError]]:
        """验证三表数据（扩展版）"""
        errors = []
        warnings = []

        # 1. 验证谈判流程数据
        flow_errors, flow_warnings = self._validate_flow_data(flow_data)
        errors.extend(flow_errors)
        warnings.extend(flow_warnings)

        # 2. 验证知识点数据（包括阶段关联）
        points_errors, points_warnings = self._validate_points_with_stages(points_data)
        errors.extend(points_errors)
        warnings.extend(points_warnings)

        # 3. 验证案例数据
        examples_errors, examples_warnings = self._validate_examples(examples_data)
        errors.extend(examples_errors)
        warnings.extend(examples_warnings)

        # 4. 验证词汇网络挂载的知识点
        lexicon_data = lexicon_data or []
        point_names = {p.get("name") for p in points_data}
        for entry in lexicon_data:
            anchor = entry.get("anchor")
            if anchor and anchor not in point_names:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="lexicon",
                    row=entry.get("_row", 0),
                    field="关联知识点名称",
                    value=anchor,
                    message="关联知识点在主表中不存在",
                    suggestion="请先在知识点主表创建该知识点",
                ))

        return errors, warnings

    def _validate_flow_data(self, flow_data: List[Dict]) -> Tuple[List[ValidationError], List[ValidationError]]:
        """验证谈判流程数据"""
        errors = []
        warnings = []

        stage_names = set()

        for stage in flow_data:
            row = stage.get("_row", 0)
            name = stage.get("name", "")

            # 检查重复
            if name in stage_names:
                errors.append(ValidationError(
                    severity="ERROR",
                    table="flow",
                    row=row,
                    field="阶段名称",
                    value=name,
                    message=f"阶段名称重复: '{name}'",
                ))
            else:
                stage_names.add(name)

            # 验证难度
            difficulty = stage.get("difficulty")
            if difficulty and difficulty not in ["beginner", "intermediate", "advanced"]:
                warnings.append(ValidationError(
                    severity="WARNING",
                    table="flow",
                    row=row,
                    field="难度级别",
                    value=difficulty,
                    message="未知的难度级别",
                    suggestion="建议使用: 初级/中级/高级",
                    action_taken="使用默认值 'intermediate'",
                ))
                stage["difficulty"] = "intermediate"

        return errors, warnings

    def _validate_points_with_stages(
        self, points_data: List[Dict]
    ) -> Tuple[List[ValidationError], List[ValidationError]]:
        """验证知识点数据（包括阶段关联）"""
        errors = []
        warnings = []

        # 先使用原有的验证逻辑
        base_errors, base_warnings = self._validate_points(points_data)
        errors.extend(base_errors)
        warnings.extend(base_warnings)

        # 额外验证: 所属阶段是否存在
        for point in points_data:
            row = point.get("_row", 0)
            stage_name = point.get("stage")

            if stage_name:
                # 检查阶段是否存在
                if stage_name not in self.known_stage_names:
                    warnings.append(ValidationError(
                        severity="WARNING",
                        table="knowledge_points",
                        row=row,
                        field="所属阶段",
                        value=stage_name,
                        message=f"阶段 '{stage_name}' 不存在",
                        suggestion=f"可用阶段: {', '.join(self.known_stage_names[:5])}",
                        action_taken="知识点将创建，但不会关联到阶段",
                    ))

        return errors, warnings

    def _find_similar_names(self, target: str, threshold: float = 0.6) -> List[str]:
        """查找相似的知识点名称（智能推荐）"""
        if not self.known_point_names:
            return []

        similarities = []
        for name in self.known_point_names:
            ratio = SequenceMatcher(None, target.lower(), name.lower()).ratio()
            if ratio >= threshold:
                similarities.append((name, ratio))

        # 按相似度排序
        similarities.sort(key=lambda x: x[1], reverse=True)
        return [name for name, _ in similarities[:5]]

    # ========================================
    # Phase 3: 事务性导入
    # ========================================

    def _import_with_transaction(
        self, points_data: List[Dict], examples_data: List[Dict], created_by: str
    ) -> Tuple[ImportStatistics, ImportStatistics, ImportStatistics]:
        """使用事务导入所有数据"""

        points_stats = ImportStatistics(total=len(points_data))
        relations_stats = ImportStatistics()
        examples_stats = ImportStatistics(total=len(examples_data))

        # 第零步：提取并创建 Stage 节点
        print("=" * 80)
        print("🔍 [STAGE DEBUG] 开始提取并创建阶段节点...")
        print(f"🔍 [STAGE DEBUG] 知识点总数: {len(points_data)}")
        LOGGER.info("从知识点中提取并创建阶段节点...")
        stages_to_create = set()
        point_stage_map = {}  # 记录每个知识点的所属阶段

        for idx, point in enumerate(points_data):
            stage_name = point.get("stage")
            if idx < 5:  # 只打印前5个作为示例
                print(f"🔍 [STAGE DEBUG] 知识点 [{idx}] '{point.get('name')}' 的所有字段: {list(point.keys())}")
                print(f"🔍 [STAGE DEBUG] 知识点 [{idx}] '{point.get('name')}' 的 stage 字段值: '{stage_name}'")
            LOGGER.debug(f"知识点 '{point.get('name')}' 的 stage 字段值: {stage_name}")
            if stage_name and isinstance(stage_name, str) and stage_name.strip():
                stages_to_create.add(stage_name.strip())
                point_stage_map[point["name"]] = stage_name.strip()

        print(f"🔍 [STAGE DEBUG] 提取到 {len(stages_to_create)} 个唯一阶段: {stages_to_create}")
        print(f"🔍 [STAGE DEBUG] point_stage_map 包含 {len(point_stage_map)} 个映射")
        print("=" * 80)
        LOGGER.info(f"提取到 {len(stages_to_create)} 个阶段: {stages_to_create}")

        # 创建 Stage 节点
        print(f"\n🔧 [STAGE CREATE] 开始创建 {len(stages_to_create)} 个 Stage 节点...")
        for stage_name in stages_to_create:
            try:
                print(f"🔧 [STAGE CREATE] 正在创建 Stage: {stage_name}")
                from services import graph_service
                driver = graph_service._get_driver()

                # 使用 MERGE 确保节点存在（如果已存在则不创建）
                merge_query = """
                MERGE (s:Stage {name: $name})
                ON CREATE SET
                    s.englishName = $englishName,
                    s.description = $description,
                    s.difficulty = $difficulty,
                    s.estimatedDuration = $estimatedDuration,
                    s.icon = $icon,
                    s.color = $color,
                    s.createdAt = datetime(),
                    s.createdBy = $createdBy,
                    s.updatedAt = datetime(),
                    s.updatedBy = $createdBy
                SET
                    s.englishName = coalesce(s.englishName, $englishName),
                    s.description = $description,
                    s.difficulty = $difficulty,
                    s.estimatedDuration = $estimatedDuration,
                    s.icon = $icon,
                    s.color = $color,
                    s.updatedAt = datetime(),
                    s.updatedBy = $createdBy
                RETURN s.name AS name
                """

                params = {
                    "name": stage_name,
                    "englishName": "",
                    "description": f"{stage_name}阶段",
                    "difficulty": "intermediate",
                    "estimatedDuration": 7,
                    "icon": "🔵",
                    "color": "#3B82F6",
                    "createdBy": created_by,
                }

                with driver.session() as session:
                    result = session.run(merge_query, params)
                    created_name = result.single()["name"]
                    print(f"✅ [STAGE CREATE] Stage 节点已确保存在: {created_name}")
                    LOGGER.info(f"✓ Stage 节点已确保存在: {created_name}")

            except Exception as e:
                print(f"❌ [STAGE CREATE] 创建 Stage 节点失败 {stage_name}: {e}")
                LOGGER.error(f"✗ 创建 Stage 节点失败 {stage_name}: {e}", exc_info=True)

        # 第一步：导入知识点节点
        LOGGER.info("导入知识点节点...")
        point_name_to_id = {}

        for point in points_data:
            # 移除内部字段
            clean_point = {k: v for k, v in point.items() if not k.startswith("_")}

            try:
                # 提取name
                point_name = clean_point.pop("name")
                node_type_label = self._determine_node_type(clean_point.get("type"))

                # 字段映射和过滤：只保留 create_knowledge_point 支持的字段
                supported_fields = {
                    "category", "type", "difficulty", "importance",
                    "summary", "description", "keywords", "tags",
                    "estimated_minutes", "image_url", "video_url",
                    "document_url", "external_url", "created_by",
                }

                # 字段名映射（camelCase -> snake_case）
                field_mapping = {
                    "estimatedMinutes": "estimated_minutes",
                    "imageUrl": "image_url",
                    "videoUrl": "video_url",
                    "documentUrl": "document_url",
                    "externalUrl": "external_url",
                    "createdBy": "created_by",
                }

                # 过滤并映射字段
                filtered_point = {}
                for key, value in clean_point.items():
                    mapped_key = field_mapping.get(key, key)
                    if mapped_key in supported_fields:
                        filtered_point[mapped_key] = value

                # 检查是否已存在
                try:
                    existing = knowledge_service.get_knowledge_point(point_name)
                except graph_service.GraphEntityNotFoundError:
                    existing = None

                if existing:
                    knowledge_service.update_knowledge_point(point_name, **filtered_point)
                    points_stats.updated += 1
                else:
                    knowledge_service.create_knowledge_point(point_name, **filtered_point)
                    points_stats.created += 1

                # 补打节点类型与标签，便于学生端按技能/术语/知识分组
                self._apply_node_type_label(point_name, node_type_label)

                # 生成ID（用name的hash作为唯一标识）
                point_id = self._generate_point_id(point_name)
                point_name_to_id[point_name] = point_id

            except Exception as e:
                LOGGER.error(f"导入知识点失败: {point.get('name', 'unknown')}: {e}")
                points_stats.failed += 1

        # 第一点五步：创建 Stage 和 KnowledgePoint 的关系
        print(f"\n🔗 [RELATION] 开始创建 Stage-KnowledgePoint 关系... (共 {len(point_stage_map)} 个)")
        LOGGER.info(f"创建 Stage 和 KnowledgePoint 的关系... (共 {len(point_stage_map)} 个)")
        stage_relation_success = 0
        stage_relation_failed = 0

        for point_name, stage_name in point_stage_map.items():
            if point_name not in point_name_to_id:
                print(f"⚠️ [RELATION] 跳过: 知识点 '{point_name}' 未成功创建")
                LOGGER.warning(f"跳过关系创建: 知识点 '{point_name}' 未成功创建")
                stage_relation_failed += 1
                continue

            try:
                from services import graph_service
                driver = graph_service._get_driver()
                rel_query = """
                MATCH (s:Stage {name: $stageName})
                MATCH (k:KnowledgePoint {name: $pointName})
                MERGE (s)-[:HAS_TOPIC]->(k)
                RETURN s.name AS stage, k.name AS point
                """
                with driver.session() as session:
                    result = session.run(rel_query, {"stageName": stage_name, "pointName": point_name})
                    record = result.single()
                    if record:
                        stage_relation_success += 1
                        if stage_relation_success <= 3:  # 只打印前3个作为示例
                            print(f"✅ [RELATION] Stage '{stage_name}' -[HAS_TOPIC]-> KnowledgePoint '{point_name}'")
                        LOGGER.debug(f"✓ Stage '{stage_name}' -[HAS_TOPIC]-> KnowledgePoint '{point_name}'")
                    else:
                        stage_relation_failed += 1
                        print(f"❌ [RELATION] 未找到节点: Stage '{stage_name}' 或 KnowledgePoint '{point_name}'")
                        LOGGER.error(f"✗ 未找到节点: Stage '{stage_name}' 或 KnowledgePoint '{point_name}'")
            except Exception as e:
                stage_relation_failed += 1
                print(f"❌ [RELATION] 创建失败 ({stage_name} -> {point_name}): {e}")
                LOGGER.error(f"✗ 创建 Stage-KnowledgePoint 关系失败 ({stage_name} -> {point_name}): {e}", exc_info=True)

        print(f"🔗 [RELATION] 完成! 成功: {stage_relation_success}, 失败: {stage_relation_failed}")
        print("=" * 80)
        LOGGER.info(f"Stage-KnowledgePoint 关系创建完成: 成功 {stage_relation_success}, 失败 {stage_relation_failed}")

        # 第二步：创建关系
        LOGGER.info("创建知识点关系...")
        for point in points_data:
            if "_relations" not in point:
                continue

            source_name = point["name"]
            if source_name not in point_name_to_id:
                continue

            for relation_type, target_names in point["_relations"].items():
                for target_name in target_names:
                    if target_name not in point_name_to_id:
                        relations_stats.failed += 1
                        continue

                    try:
                        # 创建关系
                        self._create_relation(
                            source_name, target_name, relation_type, created_by
                        )
                        relations_stats.created += 1
                        relations_stats.total += 1
                    except Exception as e:
                        LOGGER.error(f"创建关系失败: {source_name} -> {target_name}: {e}")
                        relations_stats.failed += 1
                        relations_stats.total += 1

        # 第三步：创建案例节点和关系
        if examples_data:
            LOGGER.info("创建案例数据...")
            for example in examples_data:
                kp_name = example.get("knowledge_point_name")
                if not kp_name or kp_name not in point_name_to_id:
                    examples_stats.failed += 1
                    continue

                try:
                    self._create_example(example, created_by)
                    examples_stats.created += 1
                except Exception as e:
                    LOGGER.error(f"创建案例失败: {example.get('title')}: {e}")
                    examples_stats.failed += 1

        return points_stats, relations_stats, examples_stats

    def _import_three_sheets_with_transaction(
        self,
        flow_data: List[Dict],
        points_data: List[Dict],
        examples_data: List[Dict],
        lexicon_data: List[Dict],
        created_by: str,
    ) -> Tuple[ImportStatistics, ImportStatistics, ImportStatistics, ImportStatistics, Dict[str, int], List[ValidationError]]:
        """三表联动事务性导入"""

        stages_stats = ImportStatistics(total=len(flow_data))
        topics_stats = ImportStatistics()
        points_stats = ImportStatistics(total=len(points_data) + len(lexicon_data))
        relations_stats = ImportStatistics()
        examples_stats = ImportStatistics(total=len(examples_data))
        topics_by_stage: Dict[str, int] = {}
        errors: List[ValidationError] = []

        # 导入 graph_service
        from services import graph_service

        # 第一步：导入 Stage 节点
        LOGGER.info("导入谈判流程阶段...")
        stage_name_map = {}  # 用于记录成功导入的阶段
        topic_stage_map: Dict[str, str] = {}  # Topic -> Stage
        point_topic_map: Dict[str, str] = {}  # Point -> Topic
        topic_categories: Dict[str, set] = {}  # Topic -> {category_name}

        for stage in flow_data:
            clean_stage = {k: v for k, v in stage.items() if not k.startswith("_")}
            stage_name = clean_stage.get("name")

            try:
                # 尝试获取已存在的 Stage
                try:
                    existing_stage = graph_service.get_stage(stage_name)
                    # 如果存在，可以选择更新（这里暂时跳过更新）
                    LOGGER.info(f"Stage '{stage_name}' 已存在，跳过创建")
                    stage_name_map[stage_name] = stage
                except graph_service.GraphEntityNotFoundError:
                    # 不存在，需要创建
                    # 直接调用 Neo4j 创建 Stage 节点
                    self._create_stage_node(clean_stage, created_by)
                    stages_stats.created += 1
                    stage_name_map[stage_name] = stage
                    LOGGER.info(f"Created Stage: {stage_name}")

            except Exception as e:
                LOGGER.error(f"导入 Stage 失败: {stage_name}: {e}")
                stages_stats.failed += 1
                errors.append(ValidationError(
                    severity="ERROR",
                    table="flow",
                    row=stage.get("_row", 0),
                    field="阶段名称",
                    value=stage_name,
                    message=str(e),
                ))

        # 从知识点中提取 Topic-Stage 映射
        for point in points_data:
            topic_name = point.get("topic")
            stage_name = point.get("stage")
            if topic_name and stage_name:
                topic_stage_map[str(topic_name).strip()] = str(stage_name).strip()
                point_topic_map[point.get("name")] = str(topic_name).strip()

        topics_stats.total = len(topic_stage_map)

        # 第二步：创建 PRECEDES 关系（流程先后）
        LOGGER.info("创建流程关系...")
        sorted_stages = sorted(flow_data, key=lambda s: s.get("_order", 0))

        for i in range(len(sorted_stages) - 1):
            from_stage = sorted_stages[i].get("name")
            to_stage = sorted_stages[i + 1].get("name")

            if from_stage in stage_name_map and to_stage in stage_name_map:
                try:
                    self._create_precedes_relation(from_stage, to_stage, created_by)
                    relations_stats.created += 1
                    relations_stats.total += 1
                    LOGGER.info(f"Created PRECEDES: {from_stage} -> {to_stage}")
                except Exception as e:
                    LOGGER.error(f"创建 PRECEDES 关系失败: {from_stage} -> {to_stage}: {e}")
                    relations_stats.failed += 1
                    relations_stats.total += 1

        # 第二步半：创建 Topic 节点并关联 Stage
        LOGGER.info("创建业务主题 Topic 节点...")
        for topic_name, stage_name in topic_stage_map.items():
            if stage_name not in stage_name_map:
                continue
            try:
                from services import graph_service
                driver = graph_service._get_driver()
                with driver.session() as session:
                    session.run(
                        """
                        MATCH (s:Stage {name: $stage})
                        MERGE (t:Topic {name: $topic, stage: $stage})
                        ON CREATE SET t.createdAt = datetime(), t.createdBy = $createdBy, t.updatedAt = datetime(), t.updatedBy = $createdBy
                        SET t.updatedAt = datetime(), t.updatedBy = $createdBy
                        MERGE (s)-[:CONTAIN_TOPIC]->(t)
                        """,
                        {"stage": stage_name, "topic": topic_name, "createdBy": created_by},
                    )
                    topics_stats.created += 1
                    topics_by_stage[stage_name] = topics_by_stage.get(stage_name, 0) + 1
                    relations_stats.created += 1
                    relations_stats.total += 1
            except Exception as e:
                LOGGER.error(f"创建 Topic 失败: {topic_name} -> {stage_name}: {e}")
                topics_stats.failed += 1
                relations_stats.failed += 1
                relations_stats.total += 1

        # 第三步：导入知识点节点
        LOGGER.info("导入知识点节点...")
        point_name_map = {}

        for point in points_data:
            clean_point = {k: v for k, v in point.items() if not k.startswith("_")}

            # 提取名称并移除不应直接写入的字段
            point_name = clean_point.pop("name", None)
            stage_name = clean_point.pop("stage", None)
            topic_name = clean_point.get("topic")
            raw_type = clean_point.get("type")
            category_name = self._resolve_category_name(raw_type)

            # 字段名映射（CamelCase -> snake_case，符合 knowledge_service）
            field_mapping = {
                "estimatedMinutes": "estimated_minutes",
                "imageUrl": "image_url",
                "videoUrl": "video_url",
                "documentUrl": "document_url",
                "externalUrl": "external_url",
            }

            # 仅保留 knowledge_service 支持的字段
            allowed_fields = {
                "category", "type", "difficulty", "importance",
                "summary", "description", "keywords", "tags",
                "estimated_minutes", "image_url", "video_url",
                "document_url", "external_url",
            }

            filtered_point = {}
            for key, value in clean_point.items():
                mapped_key = field_mapping.get(key, key)
                if mapped_key in allowed_fields:
                    filtered_point[mapped_key] = value
            node_type_label = self._determine_node_type(raw_type)

            if not point_name:
                LOGGER.warning("跳过缺少名称的知识点行")
                points_stats.failed += 1
                continue

            try:
                try:
                    existing = knowledge_service.get_knowledge_point(point_name)
                except graph_service.GraphEntityNotFoundError:
                    existing = None

                if existing:
                    knowledge_service.update_knowledge_point(point_name, **filtered_point)
                    points_stats.updated += 1
                else:
                    knowledge_service.create_knowledge_point(point_name, **filtered_point)
                    points_stats.created += 1

                # 标注节点类型与标签
                self._apply_node_type_label(point_name, node_type_label)

                point_name_map[point_name] = {"stage": stage_name, "topic": topic_name}
                if topic_name:
                    topic_categories.setdefault(str(topic_name).strip(), set()).add(category_name)

            except Exception as e:
                LOGGER.error(f"导入知识点失败: {point_name}: {e}")
                points_stats.failed += 1
                errors.append(ValidationError(
                    severity="ERROR",
                    table="knowledge_points",
                    row=point.get("_row", 0),
                    field="知识点名称",
                    value=point_name,
                    message=str(e),
                ))

        driver = graph_service._get_driver()

        def _merge_related_kind(source: str, target: str, kind: str):
            with driver.session() as session:
                session.run(
                    """
                    MATCH (a:KnowledgePoint {name: $source})
                    MATCH (b:KnowledgePoint {name: $target})
                    MERGE (a)-[r:RELATED_TO {kind: $kind}]->(b)
                    ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
                    """,
                    {"source": source, "target": target, "kind": kind, "createdBy": created_by},
                )

        def _link_family(name: str, family: str):
            with driver.session() as session:
                session.run(
                    """
                    MERGE (wf:WordFamily {key: $family})
                    ON CREATE SET wf.name = $family, wf.createdAt = datetime(), wf.createdBy = $createdBy
                    SET wf.updatedAt = datetime(), wf.updatedBy = $createdBy
                    WITH wf
                    MERGE (k:KnowledgePoint {name: $name})
                    MERGE (k)-[r:IN_FAMILY]->(wf)
                    ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
                    """,
                    {"family": family, "name": name, "createdBy": created_by},
                )

        def _link_class(name: str, semantic_class: str, tone: Optional[str], civic_tags: List[str], idiomatic: Optional[bool]):
            with driver.session() as session:
                session.run(
                    """
                    MERGE (sc:SemanticClass {key: $class})
                    ON CREATE SET sc.name = $class, sc.createdAt = datetime(), sc.createdBy = $createdBy
                    SET sc.updatedAt = datetime(), sc.updatedBy = $createdBy
                    WITH sc
                    MERGE (k:KnowledgePoint {name: $name})
                    MERGE (k)-[r:IN_CLASS]->(sc)
                    ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
                    SET r.tone = $tone,
                        r.civicTags = $civicTags,
                        r.idiomatic = $idiomatic
                    """,
                    {
                        "class": semantic_class,
                        "name": name,
                        "tone": tone or "",
                        "civicTags": civic_tags or [],
                        "idiomatic": idiomatic,
                        "createdBy": created_by,
                    },
                )

        def _link_slot(name: str, slot: str, tone: Optional[str], civic_tags: List[str], idiomatic: Optional[bool]):
            with driver.session() as session:
                session.run(
                    """
                    MERGE (slot:Slot {name: $slot})
                    ON CREATE SET slot.createdAt = datetime(), slot.createdBy = $createdBy
                    SET slot.updatedAt = datetime(), slot.updatedBy = $createdBy
                    WITH slot
                    MERGE (k:KnowledgePoint {name: $name})
                    MERGE (k)-[r:FITS_SLOT]->(slot)
                    ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
                    SET r.tone = $tone,
                        r.civicTags = $civicTags,
                        r.idiomatic = $idiomatic
                    """,
                    {
                        "slot": slot,
                        "name": name,
                        "tone": tone or "",
                        "civicTags": civic_tags or [],
                        "idiomatic": idiomatic,
                        "createdBy": created_by,
                    },
                )

        # 导入词汇网络子图
        if lexicon_data:
            LOGGER.info("导入词汇网络子图...")
            for entry in lexicon_data:
                lex_name = entry.get("lex_item")
                if not lex_name:
                    continue
                anchor = entry.get("anchor")
                lex_role = entry.get("lex_role", "")
                note = entry.get("note", "")
                family = entry.get("family", "")
                semantic_class = entry.get("semantic_class", "")
                slots = entry.get("slots", []) or []
                components = entry.get("components", []) or []

                try:
                    try:
                        existing_lex = knowledge_service.get_knowledge_point(lex_name)
                    except graph_service.GraphEntityNotFoundError:
                        existing_lex = None

                    summary = note or (existing_lex.get("summary") if existing_lex else "")
                    if existing_lex:
                        knowledge_service.update_knowledge_point(lex_name, lex_role=lex_role, summary=summary)
                        points_stats.updated += 1
                    else:
                        knowledge_service.create_knowledge_point(
                            lex_name,
                            lex_role=lex_role,
                            summary=summary,
                            type="concept",
                            created_by=created_by,
                        )
                        points_stats.created += 1

                    point_name_map.setdefault(lex_name, {"stage": None, "topic": None})

                    if anchor:
                        _merge_related_kind(lex_name, anchor, "lex_anchor")
                        relations_stats.created += 1
                        relations_stats.total += 1

                    if family:
                        _link_family(lex_name, family)
                        relations_stats.created += 1
                        relations_stats.total += 1

                    if semantic_class:
                        _link_class(
                            lex_name,
                            semantic_class,
                            entry.get("tone"),
                            entry.get("civic_tags", []),
                            entry.get("idiomatic"),
                        )
                        relations_stats.created += 1
                        relations_stats.total += 1

                    for slot in slots:
                        _link_slot(
                            lex_name,
                            slot,
                            entry.get("tone"),
                            entry.get("civic_tags", []),
                            entry.get("idiomatic"),
                        )
                        relations_stats.created += 1
                        relations_stats.total += 1

                    if lex_role == "collocation":
                        for component in components:
                            if not component:
                                continue
                            try:
                                comp_existing = knowledge_service.get_knowledge_point(component)
                                if comp_existing:
                                    knowledge_service.update_knowledge_point(component, lex_role="lexeme")
                                    points_stats.updated += 1
                            except graph_service.GraphEntityNotFoundError:
                                knowledge_service.create_knowledge_point(
                                    component,
                                    lex_role="lexeme",
                                    summary="",
                                    type="concept",
                                    created_by=created_by,
                                )
                                points_stats.created += 1
                            _merge_related_kind(lex_name, component, "collocation")
                            relations_stats.created += 1
                            relations_stats.total += 1

                except Exception as e:
                    LOGGER.error(f"导入词汇项失败 {lex_name}: {e}")
                    points_stats.failed += 1
                    errors.append(ValidationError(
                        severity="ERROR",
                        table="lexicon",
                        row=entry.get("_row", 0),
                        field="词汇项",
                        value=lex_name,
                        message=str(e),
                    ))
        # 第四步：为每个 Topic 创建 KnowledgeCategory 并建立 HAS_CATEGORY
        LOGGER.info("创建 Topic 下的 KnowledgeCategory ...")
        from services import graph_service
        driver = graph_service._get_driver()
        for topic_name, categories in topic_categories.items():
            stage_name = topic_stage_map.get(topic_name)
            if not stage_name:
                continue
            for category_name in categories:
                try:
                    with driver.session() as session:
                        session.run(
                            """
                            MATCH (t:Topic {name: $topic, stage: $stage})
                            MERGE (c:KnowledgeCategory {name: $category, topic: $topic, stage: $stage})
                            ON CREATE SET
                                c.type = $category,
                                c.createdAt = datetime(),
                                c.updatedAt = datetime(),
                                c.createdBy = $createdBy,
                                c.updatedBy = $createdBy
                            SET
                                c.updatedAt = datetime(),
                                c.updatedBy = $createdBy
                            MERGE (t)-[:HAS_CATEGORY]->(c)
                            """,
                            {
                                "topic": topic_name,
                                "stage": stage_name,
                                "category": category_name,
                                "createdBy": created_by,
                            },
                        )
                        relations_stats.created += 1
                        relations_stats.total += 1
                except Exception as e:
                    LOGGER.error(f"创建 KnowledgeCategory 失败: {topic_name}-{category_name}: {e}")
                    relations_stats.failed += 1
                    relations_stats.total += 1

        # 第五步：将知识点挂到对应 Category
        LOGGER.info("创建 Category-Point 关联...")
        for point_name, info in point_name_map.items():
            topic_name = info.get("topic")
            stage_name = topic_stage_map.get(topic_name) if topic_name else None
            category_name = None
            if topic_name in topic_categories:
                # 重新解析分类（确保一致）
                raw_point = next((p for p in points_data if p.get("name") == point_name), {})
                category_name = self._resolve_category_name(raw_point.get("type"))
            if topic_name and stage_name and category_name:
                try:
                    with driver.session() as session:
                        session.run(
                            """
                            MATCH (c:KnowledgeCategory {name: $category, topic: $topic, stage: $stage})
                            MATCH (k:KnowledgePoint {name: $point})
                            MERGE (c)-[:CONTAINS]->(k)
                            """,
                            {
                                "category": category_name,
                                "topic": topic_name,
                                "stage": stage_name,
                                "point": point_name,
                            },
                        )
                        relations_stats.created += 1
                        relations_stats.total += 1
                except Exception as e:
                    LOGGER.error(f"关联知识点到分类失败: {topic_name}/{category_name} -> {point_name}: {e}")
                    relations_stats.failed += 1
                    relations_stats.total += 1

        # 第六步：创建知识点之间的关系
        LOGGER.info("创建知识点关系...")
        for point in points_data:
            if "_relations" not in point:
                continue

            source_name = point["name"]
            if source_name not in point_name_map:
                continue

            for relation_type, target_names in point["_relations"].items():
                for target_name in target_names:
                    if relation_type != "culture" and target_name not in point_name_map:
                        relations_stats.failed += 1
                        relations_stats.total += 1
                        continue

                    try:
                        self._create_relation(source_name, target_name, relation_type, created_by)
                        relations_stats.created += 1
                        relations_stats.total += 1
                    except Exception as e:
                        LOGGER.error(f"创建关系失败: {source_name} -> {target_name}: {e}")
                        relations_stats.failed += 1
                        relations_stats.total += 1
                        errors.append(ValidationError(
                            severity="ERROR",
                            table="relations",
                            row=point.get("_row", 0),
                            field=relation_type,
                            value=target_name,
                            message=str(e),
                        ))

        # 第七步：创建案例节点
        if examples_data:
            LOGGER.info("创建案例数据...")
            for example in examples_data:
                kp_name = example.get("knowledge_point_name")
                if not kp_name or kp_name not in point_name_map:
                    examples_stats.failed += 1
                    continue

                try:
                    self._create_example(example, created_by)
                    examples_stats.created += 1
                except Exception as e:
                    LOGGER.error(f"创建案例失败: {example.get('title')}: {e}")
                    examples_stats.failed += 1
                    errors.append(ValidationError(
                        severity="ERROR",
                        table="examples",
                        row=example.get("_row", 0),
                        field="案例标题",
                        value=example.get("title"),
                        message=str(e),
                    ))

        return stages_stats, topics_stats, points_stats, relations_stats, examples_stats, topics_by_stage, errors

    def _create_stage_node(self, stage_data: Dict, created_by: str) -> None:
        """创建 Stage 节点（直接操作 Neo4j）"""
        from services import graph_service

        driver = graph_service._get_driver()

        query = """
        MERGE (s:Stage {name: $name})
        ON CREATE SET
            s.englishName = $englishName,
            s.description = $description,
            s.difficulty = $difficulty,
            s.estimatedDuration = $estimatedDuration,
            s.orderIndex = $orderIndex,
            s.icon = $icon,
            s.color = $color,
            s.createdAt = datetime(),
            s.createdBy = $createdBy,
            s.updatedAt = datetime(),
            s.updatedBy = $createdBy
        SET
            s.englishName = coalesce(s.englishName, $englishName),
            s.description = $description,
            s.difficulty = $difficulty,
            s.estimatedDuration = $estimatedDuration,
            s.orderIndex = $orderIndex,
            s.icon = $icon,
            s.color = $color,
            s.updatedAt = datetime(),
            s.updatedBy = $createdBy
        RETURN s.name AS name
        """

        params = {
            "name": stage_data.get("name"),
            "englishName": stage_data.get("englishName", ""),
            "description": stage_data.get("description", ""),
            "difficulty": stage_data.get("difficulty", "intermediate"),
            "estimatedDuration": stage_data.get("estimatedDuration", 7),
            "orderIndex": int(stage_data.get("orderIndex") or stage_data.get("order") or 0),
            "icon": stage_data.get("icon", "🔵"),
            "color": stage_data.get("color", "#3B82F6"),
            "createdBy": created_by,
        }

        with driver.session() as session:
            session.run(query, params)

    def _create_precedes_relation(self, from_stage: str, to_stage: str, created_by: str) -> None:
        """创建 PRECEDES 关系"""
        from services import graph_service

        driver = graph_service._get_driver()

        query = """
        MATCH (s1:Stage {name: $from_stage})
        MATCH (s2:Stage {name: $to_stage})
        MERGE (s1)-[r:PRECEDES]->(s2)
        ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
        RETURN s1.name AS from, s2.name AS to
        """

        params = {
            "from_stage": from_stage,
            "to_stage": to_stage,
            "createdBy": created_by,
        }

        with driver.session() as session:
            session.run(query, params)

    def _resolve_category_name(self, raw_type: Optional[str]) -> str:
        """根据类型字段映射 KnowledgeCategory 名称。"""
        if raw_type is not None:
            normalized = str(raw_type).strip()
            lower = normalized.lower()
            # 先尝试中文到英文的映射
            mapped = TYPE_CN_TO_EN.get(normalized) or TYPE_CN_TO_EN.get(lower)
            candidate = mapped or lower
            if candidate in CATEGORY_NAME_ALIASES:
                return CATEGORY_NAME_ALIASES[candidate]
            if normalized in CATEGORY_NAME_ALIASES:
                return CATEGORY_NAME_ALIASES[normalized]
        return "知识"

    def _determine_node_type(self, raw_type: Optional[str]) -> str:
        """根据导入列推断节点类型标签，默认为 KnowledgePoint。"""
        if raw_type is None:
            return "KnowledgePoint"
        text = str(raw_type).strip().lower()
        if text in {"skill", "技能", "技能型", "业务流程", "流程"}:
            return "Skill"
        if text in {"terminology", "term", "术语", "术语型", "概念", "concept"}:
            return "Terminology"
        if text in {"knowledge", "知识", "知识型"}:
            return "KnowledgePoint"
        return "KnowledgePoint"

    def _apply_node_type_label(self, name: str, node_type: str) -> None:
        """在 Neo4j 为知识点补打 nodeType 属性与标签（Skill/Terminology）。"""
        try:
            driver = graph_service._get_driver()
        except Exception:
            return
        label_clause = ""
        if node_type == "Skill":
            label_clause = "SET k:Skill "
        elif node_type == "Terminology":
            label_clause = "SET k:Terminology "
        with driver.session() as session:
            session.run(
                f"""
                MATCH (k:KnowledgePoint {{name: $name}})
                {label_clause}
                SET k.nodeType = $nodeType,
                    k.updatedAt = datetime()
                """,
                {"name": name, "nodeType": node_type},
            )

    def _generate_point_id(self, name: str) -> str:
        """生成知识点ID（基于名称的hash）"""
        # 使用MD5生成短ID
        return f"K{hashlib.md5(name.encode('utf-8')).hexdigest()[:8].upper()}"

    def _create_relation(
        self, source_name: str, target_name: str, relation_type: str, created_by: str
    ):
        """创建知识点关系"""
        if relation_type == "prerequisite":
            knowledge_service.add_knowledge_prerequisite(source_name, target_name)
        elif relation_type == "similar":
            self._create_direct_relation(source_name, target_name, "SUGGESTS_CO_LEARNING", created_by)
        elif relation_type == "contrast":
            knowledge_service.add_knowledge_relation(
                source_name, target_name, relation_type="CONTRASTS_WITH"
            )
        elif relation_type == "application":
            knowledge_service.add_knowledge_relation(
                source_name, target_name, relation_type="APPLIES_TO_SCENARIO"
            )
        elif relation_type == "culture":
            self._create_culture_sensitivity_relation(source_name, target_name, created_by)
        else:
            knowledge_service.add_knowledge_relation(
                source_name, target_name, relation_type=relation_type
            )

    def _create_direct_relation(
        self, source_name: str, target_name: str, rel_type: str, created_by: str
    ) -> None:
        """创建白名单外或批量导入专用的知识点关系。"""
        from services import graph_service

        driver = graph_service._get_driver()
        with driver.session() as session:
            session.run(
                f"""
                MATCH (a:KnowledgePoint {{name: $source}})
                MATCH (b:KnowledgePoint {{name: $target}})
                MERGE (a)-[r:{rel_type}]->(b)
                ON CREATE SET r.createdAt = datetime(),
                              r.createdBy = $createdBy
                SET r.relationType = $relType,
                    r.updatedAt = datetime(),
                    r.updatedBy = $createdBy
                """,
                {
                    "source": source_name,
                    "target": target_name,
                    "relType": rel_type,
                    "createdBy": created_by,
                },
            )

    def _create_culture_sensitivity_relation(
        self, source_name: str, culture_name: str, created_by: str
    ) -> None:
        """创建知识点到文化维度的敏感性关系。"""
        from services import graph_service

        clean_culture_name = str(culture_name or "").strip()
        if not clean_culture_name:
            return
        culture_id = f"culture-{hashlib.md5(clean_culture_name.encode('utf-8')).hexdigest()[:10]}"
        driver = graph_service._get_driver()
        with driver.session() as session:
            session.run(
                """
                MATCH (k:KnowledgePoint {name: $source})
                MERGE (c:CultureDimension {name: $cultureName})
                ON CREATE SET c.id = $cultureId,
                              c.theory = 'Teacher Import',
                              c.summary = $cultureName,
                              c.createdAt = datetime(),
                              c.createdBy = $createdBy
                SET c.id = coalesce(c.id, $cultureId),
                    c.updatedAt = datetime(),
                    c.updatedBy = $createdBy
                MERGE (k)-[r:HAS_CULTURAL_SENSITIVITY]->(c)
                ON CREATE SET r.createdAt = datetime(),
                              r.createdBy = $createdBy
                SET r.relationType = 'HAS_CULTURAL_SENSITIVITY',
                    r.updatedAt = datetime(),
                    r.updatedBy = $createdBy
                """,
                {
                    "source": source_name,
                    "cultureName": clean_culture_name,
                    "cultureId": culture_id,
                    "createdBy": created_by,
                },
            )

    def _create_example(self, example: Dict, created_by: str):
        """创建案例节点"""
        # 使用Neo4j创建Example节点和HAS_EXAMPLE关系
        query = """
        MATCH (kp:KnowledgePoint {name: $kp_name})
        CREATE (e:Example {
            id: $example_id,
            type: $type,
            title: $title,
            content: $content,
            createdAt: datetime(),
            createdBy: $created_by
        })
        CREATE (kp)-[:HAS_EXAMPLE]->(e)
        WITH e
        OPTIONAL MATCH (p:Practice {id: $practice_id})
        FOREACH (pr IN CASE WHEN p IS NOT NULL THEN [p] ELSE [] END |
            CREATE (e)-[:LINKED_TO_PRACTICE]->(pr)
        )
        RETURN e.id as id
        """

        example_id = f"E{hashlib.md5(example['title'].encode('utf-8')).hexdigest()[:8].upper()}"

        with self.graph_service.driver.session() as session:
            session.run(
                query,
                kp_name=example["knowledge_point_name"],
                example_id=example_id,
                type=example.get("type", "actual_case"),
                title=example["title"],
                content=example["content"],
                practice_id=example.get("practice_id"),
                created_by=created_by,
            )


# ============================================
# 智能模板生成器
# ============================================

def generate_smart_templates(
    existing_points: Optional[List[str]] = None,
    existing_stages: Optional[List[str]] = None,
    profile: str = "researcher",
) -> bytes:
    """
    生成智能Excel模板（包含数据验证和下拉菜单）- 支持多节点类型

    Args:
        existing_points: 现有知识点名称列表（用于关系列的下拉菜单）
        existing_stages: 现有阶段名称列表（用于"所属阶段"下拉菜单）
        profile: teacher=教师简版；researcher=教研员高级版

    Returns:
        Excel模板文件。教师版包含必要的谈判流程和知识点主表；教研员版保留案例库和词汇网络表。
    """
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl未安装，无法生成模板")

    normalized_profile = (profile or "researcher").strip().lower()
    is_researcher_profile = normalized_profile in {"researcher", "教研员", "advanced", "expert"}

    wb = Workbook()

    # 样式定义（所有 Sheet 共用）
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    example_font = Font(italic=True, size=9, color="666666")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # ========================================
    # Sheet 1: 谈判流程
    # ========================================
    ws_flow = wb.active
    ws_flow.title = "谈判流程"

    # 设置列宽
    flow_widths = [15, 20, 40, 12, 15, 10, 12]
    for idx, width in enumerate(flow_widths, start=1):
        ws_flow.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    # 写入表头
    for col_idx, (header, field, required, example) in enumerate(FLOW_TEMPLATE_HEADERS, start=1):
        cell = ws_flow.cell(row=1, column=col_idx)
        cell.value = f"{header}{'*' if required else ''}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        # 示例行
        example_cell = ws_flow.cell(row=2, column=col_idx)
        example_cell.value = example
        example_cell.font = example_font
        example_cell.alignment = Alignment(wrap_text=True)

    # 冻结首行
    ws_flow.freeze_panes = "A3"

    # 添加难度下拉菜单
    difficulty_validation = DataValidation(
        type="list",
        formula1='"初级,中级,高级"',
        allow_blank=True
    )
    difficulty_validation.error = "请从下拉列表中选择"
    difficulty_validation.errorTitle = "输入错误"
    ws_flow.add_data_validation(difficulty_validation)
    difficulty_validation.add("D3:D1000")  # 难度级别列

    # 添加示例数据
    sample_stages = [
        ["询盘", "Inquiry", "买方向卖方询问商品信息和交易条件的阶段", "初级", "7", "🔍", "#3B82F6"],
        ["报盘", "Offer", "卖方向买方报价和交易条件的阶段", "中级", "5", "📊", "#10B981"],
        ["还盘", "Counter-Offer", "买卖双方针对报价和条件进行协商和调整", "高级", "10", "🔄", "#F59E0B"],
    ]
    for row_idx, stage_data in enumerate(sample_stages, start=3):
        for col_idx, value in enumerate(stage_data, start=1):
            cell = ws_flow.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(italic=True, size=9, color="999999")

    # ========================================
    # Sheet 2: 知识点主表
    # ========================================
    ws_points = wb.create_sheet("知识点主表")

    # 样式定义
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    example_font = Font(italic=True, size=9, color="666666")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # 设置列宽
    column_widths = [15, 20, 15, 20, 12, 12, 12, 12, 40, 50, 25, 25, 25, 25, 25]
    for idx, width in enumerate(column_widths, start=1):
        ws_points.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    # 写入表头
    for col_idx, (header, field, required, example) in enumerate(POINTS_TEMPLATE_HEADERS, start=1):
        # 表头行
        cell = ws_points.cell(row=1, column=col_idx)
        cell.value = f"{header}{'*' if required else ''}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        # 示例行
        example_cell = ws_points.cell(row=2, column=col_idx)
        example_cell.value = example
        example_cell.font = example_font
        example_cell.alignment = Alignment(wrap_text=True)

    # 冻结首行
    ws_points.freeze_panes = "A3"

    # 添加数据验证（下拉菜单）
    _add_data_validations(ws_points, existing_points, existing_stages)

    # 添加示例数据
    _add_sample_data(ws_points)

    if is_researcher_profile:
        # ========================================
        # Sheet 3: 案例库表
        # ========================================
        ws_examples = wb.create_sheet("案例库表")

        # 设置列宽
        example_widths = [25, 20, 30, 60, 15]
        for idx, width in enumerate(example_widths, start=1):
            ws_examples.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        # 写入表头
        for col_idx, (header, field, required, example) in enumerate(EXAMPLES_TEMPLATE_HEADERS, start=1):
            cell = ws_examples.cell(row=1, column=col_idx)
            cell.value = f"{header}{'*' if required else ''}"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            example_cell = ws_examples.cell(row=2, column=col_idx)
            example_cell.value = example
            example_cell.font = example_font

        ws_examples.freeze_panes = "A3"

        # 添加案例类型下拉菜单
        example_type_validation = DataValidation(
            type="list",
            formula1='"实际案例,邮件模板,文档模板,常见错误,对话示例"',
            allow_blank=False
        )
        example_type_validation.error = "请从下拉列表中选择"
        example_type_validation.errorTitle = "输入错误"
        ws_examples.add_data_validation(example_type_validation)
        example_type_validation.add(f"B3:B1000")

        # ========================================
        # Sheet 4: 词汇网络表
        # ========================================
        ws_lex = wb.create_sheet("词汇网络表")
        lex_widths = [22, 28, 14, 16, 20, 20, 14, 28, 14, 24, 30]
        for idx, width in enumerate(lex_widths, start=1):
            ws_lex.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        for col_idx, (header, field, required, example) in enumerate(LEXICON_TEMPLATE_HEADERS, start=1):
            cell = ws_lex.cell(row=1, column=col_idx)
            cell.value = f"{header}{'*' if required else ''}"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            example_cell = ws_lex.cell(row=2, column=col_idx)
            example_cell.value = example
            example_cell.font = example_font
            example_cell.alignment = Alignment(wrap_text=True)

        ws_lex.freeze_panes = "A3"

        # 词汇项类型下拉
        lex_role_validation = DataValidation(
            type="list",
            formula1='"词条,搭配,句式"',
            allow_blank=False
        )
        lex_role_validation.error = "请从下拉列表中选择"
        lex_role_validation.errorTitle = "输入错误"
        ws_lex.add_data_validation(lex_role_validation)
        lex_role_validation.add("C3:C1000")

        # 语气下拉
        lex_tone_validation = DataValidation(
            type="list",
            formula1='"softer,neutral,stronger"',
            allow_blank=True
        )
        lex_tone_validation.error = "请从下拉列表中选择：softer/neutral/stronger"
        lex_tone_validation.errorTitle = "输入错误"
        ws_lex.add_data_validation(lex_tone_validation)
        lex_tone_validation.add("G3:G1000")

        # 思政元素下拉
        lex_civic_validation = DataValidation(
            type="list",
            formula1='"Win-Win,Integrity,Dignity,Compliance,Zero-Sum,Dishonesty,Disrespect,Non-Compliance"',
            allow_blank=True
        )
        lex_civic_validation.error = "请从下拉列表选择或使用分号分隔多选项"
        lex_civic_validation.errorTitle = "输入错误"
        ws_lex.add_data_validation(lex_civic_validation)
        lex_civic_validation.add("H3:H1000")

        # 地道性下拉
        lex_idiomatic_validation = DataValidation(
            type="list",
            formula1='"是,否"',
            allow_blank=True
        )
        lex_idiomatic_validation.error = "请选择 是/否，或留空"
        lex_idiomatic_validation.errorTitle = "输入错误"
        ws_lex.add_data_validation(lex_idiomatic_validation)
        lex_idiomatic_validation.add("I3:I1000")

    # ========================================
    # Sheet 5: 使用说明
    # ========================================
    ws_guide = wb.create_sheet("使用说明")
    ws_guide.column_dimensions['A'].width = 100

    guide_title = "📖 教研员高级知识图谱批量导入指南" if is_researcher_profile else "📖 教师版知识图谱导入指南"
    guide_content = [
        (guide_title, True, 16),
        ("", False, 10),
        ("✨ 主要特点", True, 13),
        ("1. 教师版主要填写「知识点主表」，「谈判流程」已预填常用阶段", False, 11),
        ("2. 关系用自然语言表达（必须先学、建议同时学、可对比学习、文化敏感）", False, 11),
        ("3. 下拉菜单辅助填写，避免输入错误", False, 11),
        ("4. 智能错误提示，自动推荐相似名称", False, 11),
        ("", False, 10),
        ("📝 填写步骤", True, 13),
        ("", False, 10),
        ("第一步：填写「知识点主表」", True, 12),
        ("  1. 章节：例如「第一章 询盘」", False, 11),
        ("  2. 知识点名称：必填，例如「询盘基本流程」", False, 11),
        ("  3. 知识点类型：从下拉菜单选择（术语/技能/知识，或概念型/技能型/文档型/案例型等）", False, 11),
        ("  4. 难度：从下拉菜单选择（初级/中级/高级）", False, 11),
        ("  5. 重要性：从下拉菜单选择（必修/推荐/选修）", False, 11),
        ("  6. 预计学时：填写数字，单位为分钟", False, 11),
        ("  7. 内容简介：一句话描述", False, 11),
        ("  8. 详细描述：详细说明", False, 11),
        ("  9. 关键词：用逗号分隔，例如「询盘,业务流程,初级」", False, 11),
        (" 10. 必须先学：填写知识点名称，多个用分号分隔（例如：询盘基本流程;价格术语）", False, 11),
        (" 11. 建议同时学：填写相关知识点名称", False, 11),
        (" 12. 可对比学习：填写可以对比的知识点名称", False, 11),
        (" 13. 文化敏感：从下拉选择十个固定项；多个值用分号分隔，例如「权力距离（PDI）;高情境文化」", False, 11),
        ("", False, 10),
        ("⚠️ 重要提示：", True, 12),
        ("  • 「必须先学」等关系列，请直接复制粘贴知识点名称，不要手打（避免错别字）", False, 11),
        ("  • 如果一个知识点依赖多个前置知识，用分号分隔：「A;B;C」", False, 11),
        ("  • 不需要的关系列可以留空", False, 11),
        ("", False, 10),
    ]

    if is_researcher_profile:
        guide_content.extend([
        ("第二步：填写「案例库表」（可选，教研员高级功能）", True, 12),
        ("  1. 关联知识点：填写知识点名称（必须在主表中存在）", False, 11),
        ("  2. 案例类型：从下拉菜单选择", False, 11),
        ("  3. 案例标题：简短标题", False, 11),
        ("  4. 案例内容：详细内容，可以很长", False, 11),
        ("  5. 关联练习关卡：如果有对应的练习关卡，填写关卡号（例如：6-1）", False, 11),
        ("", False, 10),
        ("第三步：填写「词汇网络表」（可选，教研员高级功能）", True, 12),
        ("  1. 关联知识点名称：必须在知识点主表存在，用于挂载词汇", False, 11),
        ("  2. 词汇项类型：词条/搭配/句式（下拉选择）", False, 11),
        ("  3. 多值字段（槽位/搭配组成词）支持逗号、分号或顿号，系统自动统一为分号并去重", False, 11),
        ("  4. 同一关联点下重复的词汇项会自动去重", False, 11),
        ("", False, 10),
        ])

    guide_content.extend([
        ("上传导入", True, 12),
        ("  1. 删除第2行的示例数据", False, 11),
        ("  2. 保存文件", False, 11),
        ("  3. 在系统中选择「批量导入」并上传", False, 11),
        ("  4. 系统会自动检查数据，如果有错误会详细提示", False, 11),
        ("", False, 10),
        ("💡 常见问题", True, 13),
        ("", False, 10),
        ("Q1: 如果我在「必须先学」列填错了知识点名称怎么办？", True, 11),
        ("A1: 系统会自动检测并推荐相似的名称，例如：", False, 11),
        ("    你写了「CIF术语」，系统会提示「你是不是想写：CIF价格术语」", False, 11),
        ("", False, 10),
        ("Q2: 我可以先导入知识点，稍后再补充关系吗？", True, 11),
        ("A2: 可以！关系列可以留空，后续再次导入时会自动更新", False, 11),
        ("", False, 10),
        ("Q3: 如果知识点名称重复了怎么办？", True, 11),
        ("A3: 系统会报错并拒绝导入，请确保每个知识点名称唯一", False, 11),
        ("", False, 10),
        ("Q4: 案例库表可以不填吗？", True, 11),
        ("A4: 可以！案例库表是可选的，如果暂时没有案例数据可以不上传", False, 11),
        ("", False, 10),
        ("📊 枚举值参考", True, 13),
        ("", False, 10),
        ("知识点类型：", True, 11),
        ("  • 概念型 - 定义、术语解释", False, 10),
        ("  • 技能型 - 操作方法、流程、技巧", False, 10),
        ("  • 文档型 - 表格、模板、格式", False, 10),
        ("  • 案例型 - 实际案例、情景分析", False, 10),
        ("  • 工具型 - 计算器、检查清单", False, 10),
        ("  • 理论型 - 理论框架、模型", False, 10),
        ("  • 法规型 - 法律、规则、标准", False, 10),
        ("", False, 10),
        ("难度级别：", True, 11),
        ("  • 初级 - 入门必学", False, 10),
        ("  • 中级 - 进阶内容", False, 10),
        ("  • 高级 - 深度专题", False, 10),
        ("", False, 10),
        ("重要性：", True, 11),
        ("  • 必修 - 核心知识点，必须掌握", False, 10),
        ("  • 推荐 - 建议学习", False, 10),
        ("  • 选修 - 扩展阅读", False, 10),
        ("", False, 10),
        ("案例类型：", True, 11),
        ("  • 实际案例 - 真实业务场景", False, 10),
        ("  • 邮件模板 - 可复用的邮件范本", False, 10),
        ("  • 文档模板 - 合同、发票等模板", False, 10),
        ("  • 常见错误 - 典型错误案例及纠正", False, 10),
        ("  • 对话示例 - 谈判对话实录", False, 10),
        ("", False, 10),
        ("🎯 最佳实践", True, 13),
        ("", False, 10),
        ("1. 先整理知识点列表，再填写关系", False, 11),
        ("2. 使用复制粘贴而不是手动输入知识点名称", False, 11),
        ("3. 预计学时设置合理（一般10-60分钟）", False, 11),
        ("4. 关键词设置3-5个为宜", False, 11),
        ("5. 教师版不需要维护词汇网络；需要表达优化、语义槽位时请使用「我是教研员」高级模板", False, 11),
        ("", False, 10),
        ("祝您使用愉快！如有问题请联系技术支持。", False, 11),
    ])

    for row_idx, (text, is_bold, font_size) in enumerate(guide_content, start=1):
        cell = ws_guide.cell(row=row_idx, column=1)
        cell.value = text
        cell.font = Font(bold=is_bold, size=font_size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _add_data_validations(ws, existing_points: Optional[List[str]] = None, existing_stages: Optional[List[str]] = None):
    """为工作表添加数据验证（下拉菜单）- 支持阶段下拉"""

    # 所属阶段下拉菜单（新增）
    if existing_stages and len(existing_stages) > 0:
        # 使用现有阶段列表
        stages_formula = f'"{",".join(existing_stages)}"'
    else:
        # 使用默认阶段列表
        stages_formula = '"询盘,报盘,还盘,接受与订货,包装与装运,付款与交货,商检,保险与仲裁,投诉,索赔与理赔"'

    stage_validation = DataValidation(
        type="list",
        formula1=stages_formula,
        allow_blank=True
    )
    stage_validation.error = "请从下拉列表中选择所属阶段"
    stage_validation.errorTitle = "输入错误"
    stage_validation.prompt = "选择该知识点所属的谈判流程阶段"
    stage_validation.promptTitle = "提示"
    ws.add_data_validation(stage_validation)
    stage_validation.add("C3:C1000")  # 所属阶段在第3列

    # 二级主题列提示（不做校验，保留自由文本）

    # 知识点类型下拉菜单
    type_validation = DataValidation(
        type="list",
        formula1='"术语,技能,知识,概念型,技能型,文档型,案例型,工具型,理论型,法规型"',
        allow_blank=True
    )
    type_validation.error = "请从下拉列表中选择知识点类型"
    type_validation.errorTitle = "输入错误"
    type_validation.prompt = "选择知识点类型"
    type_validation.promptTitle = "提示"
    ws.add_data_validation(type_validation)
    type_validation.add("E3:E1000")  # 知识点类型移到第5列

    # 难度下拉菜单
    difficulty_validation = DataValidation(
        type="list",
        formula1='"初级,中级,高级"',
        allow_blank=True
    )
    difficulty_validation.error = "请从下拉列表中选择：初级、中级、高级"
    difficulty_validation.errorTitle = "输入错误"
    ws.add_data_validation(difficulty_validation)
    difficulty_validation.add("F3:F1000")  # 难度移到第6列

    # 重要性下拉菜单
    importance_validation = DataValidation(
        type="list",
        formula1='"必修,推荐,选修"',
        allow_blank=True
    )
    importance_validation.error = "请从下拉列表中选择：必修、推荐、选修"
    importance_validation.errorTitle = "输入错误"
    ws.add_data_validation(importance_validation)
    importance_validation.add("G3:G1000")  # 重要性移到第7列

    # 预计学时数字验证（只能输入正整数）
    minutes_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=True
    )
    minutes_validation.error = "请输入大于0的整数"
    minutes_validation.errorTitle = "输入错误"
    minutes_validation.prompt = "填写预计学习时长（分钟），例如：30"
    minutes_validation.promptTitle = "提示"
    ws.add_data_validation(minutes_validation)
    minutes_validation.add("H3:H1000")  # 预计学时回到第8列

    # 如果有现有知识点，为关系列添加下拉菜单
    if existing_points and len(existing_points) > 0:
        # 创建隐藏的知识点列表sheet
        ws_hidden = ws.parent.create_sheet("_知识点列表")
        for idx, point_name in enumerate(existing_points, start=1):
            ws_hidden.cell(row=idx, column=1, value=point_name)

        # 隐藏该sheet
        ws_hidden.sheet_state = "hidden"

        # 为关系列添加下拉菜单（引用隐藏sheet）
        relation_validation = DataValidation(
            type="list",
            formula1=f"'_知识点列表'!$A$1:$A${len(existing_points)}",
            allow_blank=True
        )
        relation_validation.prompt = "从列表中选择知识点名称，多个用分号分隔"
        relation_validation.promptTitle = "提示"
        ws.add_data_validation(relation_validation)
        # 应用到知识点关系列：必须先学、建议同时学、可对比学习
        relation_validation.add("L3:N1000")

    ws_culture = ws.parent.create_sheet("_文化维度列表")
    for idx, culture_name in enumerate(CULTURE_SENSITIVITY_OPTIONS, start=1):
        ws_culture.cell(row=idx, column=1, value=culture_name)
    ws_culture.sheet_state = "hidden"

    culture_validation = DataValidation(
        type="list",
        formula1=f"'_文化维度列表'!$A$1:$A${len(CULTURE_SENSITIVITY_OPTIONS)}",
        allow_blank=True
    )
    culture_validation.prompt = "从十个文化/策略维度中选择；多个值可用分号分隔"
    culture_validation.promptTitle = "文化敏感"
    culture_validation.error = "请从文化敏感下拉列表中选择"
    culture_validation.errorTitle = "输入错误"
    ws.add_data_validation(culture_validation)
    culture_validation.add("O3:O1000")


def _add_sample_data(ws):
    """添加示例数据（第3行）- 支持多节点类型"""
    sample_data = [
        "第一章 询盘",
        "询盘基本流程",
        "询盘",  # 所属阶段
        "采购需求分析与供应商甄选",  # 二级主题
        "技能型",
        "初级",
        "必修",
        30,
        "掌握询盘的基本流程和注意事项",
        "询盘是外贸业务的第一步，包括客户背景调查、询盘接收、信息记录等环节...",
        "询盘,业务流程,沟通",
        "",  # 必须先学
        "",  # 建议同时学
        "",  # 可对比学习
        "",  # 文化敏感
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = value
        cell.font = Font(size=10, color="999999", italic=True)
        cell.alignment = Alignment(wrap_text=True)


def generate_batch_template_zip(existing_knowledge_points: Optional[List[str]] = None) -> bytes:
    """
    生成包含两个Excel文件的ZIP压缩包

    Returns:
        ZIP文件的二进制内容
    """
    import zipfile

    # 生成Excel模板
    excel_content = generate_smart_templates(existing_knowledge_points)

    # 创建ZIP文件
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # 添加Excel文件
        zip_file.writestr("知识图谱批量导入模板.xlsx", excel_content)

        # 添加README文本
        readme_content = """
知识图谱批量导入模板使用说明
================================

📦 本压缩包包含：
  1. 知识图谱批量导入模板.xlsx - 包含两个sheet的Excel模板

📝 使用步骤：
  1. 解压本ZIP文件
  2. 打开Excel模板
  3. 阅读「使用说明」sheet
  4. 在「知识点主表」sheet填写知识点数据
  5. （可选）在「案例库表」sheet填写案例数据
  6. 保存文件
  7. 在系统中上传导入

✨ 核心特点：
  • 只需一个Excel文件，包含知识点和案例两张表
  • 不需要手动填写ID
  • 关系用自然语言表达（必须先学、建议同时学等）
  • 自动数据验证，防止输入错误
  • 智能错误提示

💡 提示：
  • 删除示例数据行（第3行）后再填写您的数据
  • 使用复制粘贴而不是手动输入知识点名称
  • 详细说明请查看Excel中的「使用说明」sheet

如有问题请联系技术支持。
"""
        zip_file.writestr("README.txt", readme_content.encode('utf-8'))

    zip_buffer.seek(0)
    return zip_buffer.getvalue()
