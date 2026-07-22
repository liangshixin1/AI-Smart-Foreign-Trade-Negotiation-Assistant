"""Neo4j integration layer for the knowledge graph features."""

from __future__ import annotations

import copy
import csv
import hashlib
import importlib.util
import io
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from neo4j import GraphDatabase, basic_auth
from neo4j.exceptions import IncompleteCommit, Neo4jError, ServiceUnavailable
from openpyxl import Workbook, load_workbook

import database
from services import rag_matcher
from services import rag_matcher


LOGGER = logging.getLogger(__name__)


class GraphUnavailableError(RuntimeError):
    """Raised when the Neo4j backend is not ready."""


class GraphEntityNotFoundError(RuntimeError):
    """Raised when a requested graph entity cannot be located."""


_DRIVER = None
_GRAPH_DISABLED = False
_GRAPH_DISABLED_REASON = ""
_DEFAULT_CATEGORIES_CACHE: Optional[Sequence[Dict[str, object]]] = None


INITIALIZATION_SETTING_KEY = "knowledge_graph.initialization"
LESSON_GRAPH_CACHE_PREFIX = "knowledge_graph.lesson_cache."


def _default_categories() -> Sequence[Dict[str, object]]:
    """Load the recommended category hierarchy from the migration preset."""

    global _DEFAULT_CATEGORIES_CACHE
    if _DEFAULT_CATEGORIES_CACHE is not None:
        return _DEFAULT_CATEGORIES_CACHE

    module_path = (
        Path(__file__).resolve().parent.parent / "migrations" / "001_enhance_knowledge_graph.py"
    )
    categories: Sequence[Dict[str, object]] = ()
    if module_path.exists():
        try:
            spec = importlib.util.spec_from_file_location(
                "migrations.enhance_knowledge_graph", str(module_path)
            )
            if spec and spec.loader:
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                categories = getattr(module, "DEFAULT_CATEGORIES", ())
        except Exception as exc:  # pragma: no cover - defensive logging
            LOGGER.warning("Failed to load default categories: %s", exc)
            categories = ()
    else:
        LOGGER.debug("Default category preset file not found at %s", module_path)

    _DEFAULT_CATEGORIES_CACHE = tuple(categories)
    return _DEFAULT_CATEGORIES_CACHE


def _neo4j_credentials() -> Tuple[str, str, str]:
    uri = os.getenv("NEO4J_URI")
    if not uri:
        raise GraphUnavailableError("NEO4J_URI environment variable is not configured")
    user = os.getenv("NEO4J_USER", "neo4j")
    password = os.getenv("NEO4J_PASSWORD", "")
    return uri, user, password


def _disable_graph(reason: str) -> None:
    """Disable knowledge graph operations after a fatal connectivity issue."""

    global _GRAPH_DISABLED, _GRAPH_DISABLED_REASON
    close_driver()
    _GRAPH_DISABLED = True
    _GRAPH_DISABLED_REASON = reason


def _get_driver():
    global _DRIVER, _GRAPH_DISABLED, _GRAPH_DISABLED_REASON
    if _GRAPH_DISABLED:
        raise GraphUnavailableError(_GRAPH_DISABLED_REASON or "Knowledge graph connectivity disabled")

    if _DRIVER is not None:
        return _DRIVER

    uri, user, password = _neo4j_credentials()
    try:
        driver = GraphDatabase.driver(uri, auth=basic_auth(user, password))
        driver.verify_connectivity()
    except (Neo4jError, ServiceUnavailable, OSError) as exc:  # pragma: no cover - network
        LOGGER.error("Failed to connect to Neo4j at %s: %s", uri, exc)
        _disable_graph(f"Failed to connect to Neo4j at {uri}: {exc}")
        raise GraphUnavailableError("Failed to connect to Neo4j") from exc

    _DRIVER = driver
    _GRAPH_DISABLED = False
    _GRAPH_DISABLED_REASON = ""
    return driver


def is_configured() -> bool:
    """Return True if Neo4j connection information is available."""

    return bool(os.getenv("NEO4J_URI"))


def graph_status() -> Dict[str, object]:
    """Return diagnostic information about the knowledge graph backend."""

    return {
        "configured": is_configured(),
        "available": is_configured() and not _GRAPH_DISABLED,
        "message": _GRAPH_DISABLED_REASON,
    }


def get_initialization_status() -> Dict[str, object]:
    """Return the persisted initialization state for the knowledge graph."""

    record = database.get_app_setting(INITIALIZATION_SETTING_KEY, {})
    initialized = bool(record.get("initialized")) if isinstance(record, dict) else False
    option = record.get("option") if isinstance(record, dict) else None
    completed_at = record.get("completedAt") if isinstance(record, dict) else None
    return {
        "initialized": initialized,
        "option": option,
        "completedAt": completed_at,
        "graph": graph_status(),
    }


def _set_initialization_status(option: str) -> Dict[str, object]:
    payload = {
        "initialized": True,
        "option": option,
        "completedAt": datetime.utcnow().isoformat() + "Z",
    }
    database.set_app_setting(INITIALIZATION_SETTING_KEY, payload)
    payload["graph"] = graph_status()
    return payload


def reset_initialization_status() -> None:
    """Clear the initialization status flag."""

    database.delete_app_setting(INITIALIZATION_SETTING_KEY)


def _lesson_cache_key(lesson_id: str) -> str:
    return f"{LESSON_GRAPH_CACHE_PREFIX}{lesson_id}"


def cache_lesson_graph_payload(lesson_id: str, payload: Dict[str, object]) -> None:
    """Persist precomputed lesson graph payload for student-side fast load."""

    database.set_app_setting(_lesson_cache_key(lesson_id), payload)


def get_cached_lesson_graph_payload(lesson_id: str) -> Optional[Dict[str, object]]:
    return database.get_app_setting(_lesson_cache_key(lesson_id))


def invalidate_lesson_graph_cache(lesson_id: str) -> None:
    database.delete_app_setting(_lesson_cache_key(lesson_id))


def get_initialization_defaults_preview() -> Dict[str, object]:
    """Return the recommended category tree and practice suggestions."""

    categories = copy.deepcopy(list(_default_categories()))
    knowledge_presets: List[Dict[str, object]] = []
    for practice_id, names in SECTION_KNOWLEDGE_PRESETS.items():
        section = database.get_section(practice_id)
        knowledge_presets.append(
            {
                "practiceId": practice_id,
                "practiceTitle": section.get("title") if section else "",
                "chapterId": section.get("chapter_id") if section else "",
                "count": len(names),
                "names": list(names),
            }
        )

    return {
        "categories": categories,
        "knowledgePresets": knowledge_presets,
    }


def close_driver() -> None:
    """Dispose of the cached Neo4j driver."""

    global _DRIVER, _GRAPH_DISABLED, _GRAPH_DISABLED_REASON
    if _DRIVER is not None:
        _DRIVER.close()
        _DRIVER = None
    _GRAPH_DISABLED = False
    _GRAPH_DISABLED_REASON = ""


class GraphService:
    """Wrapper class for Neo4j graph service operations."""

    @property
    def driver(self):
        """Get the Neo4j driver instance."""
        return _get_driver()


def _fallback_practice_detail(practice_id: str) -> Dict[str, object]:
    practice = database.get_section(practice_id)
    if not practice:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")
    return {
        "id": practice.get("id"),
        "title": practice.get("title"),
        "description": practice.get("description"),
        "orderIndex": practice.get("order_index"),
        "chapterId": practice.get("chapter_id"),
        "knowledgePoints": [],
    }


def _fallback_lesson_detail(lesson_id: str) -> Dict[str, object]:
    lesson = database.get_theory_lesson(lesson_id, include_unpublished=True)
    if not lesson:
        raise GraphEntityNotFoundError(f"Theory lesson {lesson_id} not found")
    return {
        "id": lesson.get("id"),
        "title": lesson.get("title"),
        "code": lesson.get("code"),
        "topicId": lesson.get("topicId"),
        "knowledgePoints": [],
    }


def _execute_read(query: str, parameters: Optional[Dict[str, object]] = None) -> List[Dict[str, object]]:
    driver = _get_driver()
    with driver.session() as session:
        result = session.run(query, parameters or {})
        return [record.data() for record in result]


def _execute_write(query: str, parameters: Optional[Dict[str, object]] = None) -> None:
    driver = _get_driver()
    with driver.session() as session:
        session.run(query, parameters or {})


def ensure_indexes() -> None:
    """Create the uniqueness constraints required by the knowledge graph."""

    driver = _get_driver()
    statements = [
        "CREATE CONSTRAINT chapter_id IF NOT EXISTS FOR (c:Chapter) REQUIRE c.id IS UNIQUE",
        "CREATE CONSTRAINT practice_id IF NOT EXISTS FOR (p:Practice) REQUIRE p.id IS UNIQUE",
        "CREATE CONSTRAINT theory_topic_id IF NOT EXISTS FOR (t:TheoryTopic) REQUIRE t.id IS UNIQUE",
        "CREATE CONSTRAINT lesson_id IF NOT EXISTS FOR (l:TheoryLesson) REQUIRE l.id IS UNIQUE",
        "CREATE CONSTRAINT knowledge_point_name IF NOT EXISTS FOR (k:KnowledgePoint) REQUIRE k.name IS UNIQUE",
        "CREATE CONSTRAINT process_step_id IF NOT EXISTS FOR (s:ProcessStep) REQUIRE s.id IS UNIQUE",
        "CREATE CONSTRAINT stage_id IF NOT EXISTS FOR (s:Stage) REQUIRE s.id IS UNIQUE",
        "CREATE CONSTRAINT culture_dimension_id IF NOT EXISTS FOR (c:CultureDimension) REQUIRE c.id IS UNIQUE",
    ]

    with driver.session() as session:
        for statement in statements:
            session.run(statement)


@dataclass(frozen=True)
class ProcessStep:
    """Metadata for the canonical negotiation process stages."""

    identifier: str
    name: str
    order_index: int


PROCESS_STEPS: Sequence[ProcessStep] = (
    ProcessStep("process-prologue", "课程导入", 0),
    ProcessStep("process-inquiry", "询盘", 1),
    ProcessStep("process-offer", "报盘", 2),
    ProcessStep("process-counter-offer", "还盘", 3),
    ProcessStep("process-acceptance", "接受与订货", 4),
    ProcessStep("process-logistics", "订舱与物流", 5),
    ProcessStep("process-payment", "付款与交货", 6),
    ProcessStep("process-inspection", "商检", 7),
    ProcessStep("process-risk", "保险与仲裁", 8),
    ProcessStep("process-complaint", "投诉处理", 9),
    ProcessStep("process-claim", "索赔与理赔", 10),
)


@dataclass(frozen=True)
class StandardStage:
    """Canonical customer-aligned foreign trade negotiation stage."""

    identifier: str
    name: str
    english_name: str
    order_index: int
    summary: str
    objectives: Sequence[str]
    concepts: Sequence[str]
    processes: Sequence[str]
    strategies: Sequence[str]
    culture_dimensions: Sequence[str]


@dataclass(frozen=True)
class CultureDimensionPreset:
    """Default cross-cultural teaching dimension."""

    identifier: str
    name: str
    theory: str
    summary: str
    teaching_tip: str


CULTURE_DIMENSIONS: Sequence[CultureDimensionPreset] = (
    CultureDimensionPreset("culture-power-distance", "权力距离（PDI）", "Hofstede", "谈判对象对层级、授权和正式背书的敏感程度。", "面对高权力距离客户时，突出正式授权、公司资质和高层承诺。"),
    CultureDimensionPreset("culture-individualism-collectivism", "个人主义/集体主义（IDV）", "Hofstede", "谈判决策更偏个人绩效还是组织关系与群体共识。", "根据对方文化调整个人收益、团队共赢和长期关系的表达比例。"),
    CultureDimensionPreset("culture-masculinity-femininity", "男性气质/女性气质（MAS）", "Hofstede", "谈判对象更偏竞争、成就和强势表达，还是更重视照顾、协商和关系平衡。", "根据对方偏好调整竞争性报价、合作式表达和让步语言。"),
    CultureDimensionPreset("culture-uncertainty-avoidance", "不确定性规避（UAI）", "Hofstede", "谈判对象对模糊条款、风险和不确定结果的容忍程度。", "对高不确定性规避客户提供更清晰的条款、证据和备选方案。"),
    CultureDimensionPreset("culture-long-short-term-orientation", "长期导向/短期导向（LTO）", "Hofstede", "谈判对象对短期收益与长期合作价值的取舍。", "长期导向客户适合强调合作路线图、复购机制和持续服务。"),
    CultureDimensionPreset("culture-indulgence-restraint", "放纵/约束（IVR）", "Hofstede", "谈判对象对灵活选择、即时满足、自我约束和规则限制的接受程度。", "根据对方文化调整促销承诺、灵活条款和规则边界的表达。"),
    CultureDimensionPreset("culture-high-context", "高情境文化", "Hall", "沟通更依赖关系、语境、暗示和礼貌缓冲。", "在高情境沟通中先铺垫关系，避免过度直接否定。"),
    CultureDimensionPreset("culture-low-context", "低情境文化", "Hall", "沟通更依赖明确文本、直接表达和可验证条款。", "在低情境沟通中明确条件、责任、时间和证据。"),
    CultureDimensionPreset("culture-door-in-the-face", "以退为进策略（Door-in-the-face）", "Negotiation Strategy", "先提出较高要求，再退回到真实目标以提升对方接受概率。", "用于让步设计时要控制初始要求的合理性，避免损害信任。"),
    CultureDimensionPreset("culture-foot-in-the-door", "得寸进尺法（Foot-in-the-door）", "Negotiation Strategy", "先取得小承诺，再逐步推进到更大的目标请求。", "用于分阶段推进付款、交期或服务承诺时，注意保持透明边界。"),
)


STANDARD_STAGES: Sequence[StandardStage] = (
    StandardStage("stage-inquiry", "询盘", "Inquiry", 1, "交易磋商启动阶段，围绕需求、规格、数量和初步条件进行信息获取。", ("识别询盘类型", "澄清客户真实需求", "形成专业询盘回复"), ("询盘定义", "一般询盘", "具体询盘"), ("询盘信息收集", "需求澄清流程", "询盘有效期判断"), ("范围询价策略", "需求澄清提问", "专业邮件礼仪"), ("culture-high-context", "culture-low-context")),
    StandardStage("stage-offer", "报盘", "Offer", 2, "卖方提出价格、数量、交期、付款等核心交易条件。", ("解释报盘法律性质", "设计报价结构", "控制报价锚点"), ("报盘定义", "实盘", "虚盘"), ("报价单结构", "有效期设置", "价格条款说明"), ("锚定策略", "价值陈述策略", "阶梯价格策略"), ("culture-uncertainty-avoidance", "culture-power-distance")),
    StandardStage("stage-counter-offer", "还盘", "Counter-offer", 3, "对原报盘进行修改或拒绝，并重新提出交易条件。", ("理解还盘法律后果", "设计让步节奏", "维护谈判底线"), ("还盘定义", "反要约", "底线"), ("还盘条件分析", "权利义务转移", "让步记录"), ("条件式让步", "反锚定", "打包交换"), ("culture-door-in-the-face", "culture-individualism-collectivism")),
    StandardStage("stage-acceptance-order", "接受与订货", "Acceptance & Order", 4, "确认交易条件并形成订单或合同文件。", ("判断有效接受", "核对订单条款", "控制确认风险"), ("接受", "订单确认", "形式发票"), ("接受函撰写", "订单审核", "合同确认"), ("风险提示", "条款复核", "关系维护"), ("culture-low-context", "culture-power-distance")),
    StandardStage("stage-packing-shipment", "包装与装运", "Packing & Shipment", 5, "围绕包装标准、唛头、装运安排和物流责任进行协调。", ("识别包装要求", "安排装运节点", "处理物流异常"), ("出口包装", "唛头", "装运通知"), ("包装确认", "订舱安排", "装运跟踪"), ("时效成本平衡", "异常预案", "多方协调"), ("culture-uncertainty-avoidance", "culture-long-short-term-orientation")),
    StandardStage("stage-payment-delivery", "付款与交货", "Payment & Delivery", 6, "围绕付款工具、交货节点、风险转移和资金安全展开谈判。", ("比较付款方式风险", "解释交货责任", "设计风险缓释方案"), ("信用证", "托收", "电汇"), ("审证", "改证", "交货确认"), ("付款条件博弈", "风险缓释", "分批交货"), ("culture-uncertainty-avoidance", "culture-long-short-term-orientation")),
    StandardStage("stage-inspection", "商检", "Inspection", 7, "围绕检验证书、质量标准、异议期限和责任划分进行协商。", ("识别检验标准", "设计检验条款", "处理质量异议"), ("商检", "检验证书", "异议期"), ("检验机构确认", "检验条款撰写", "异议处理"), ("证据链构建", "质量保证设计", "解决方案谈判"), ("culture-uncertainty-avoidance", "culture-low-context")),
    StandardStage("stage-insurance-arbitration", "保险与仲裁", "Insurance & Arbitration", 8, "围绕货运保险、争议解决、仲裁条款和法律适用进行安排。", ("理解保险责任", "撰写仲裁条款", "选择争议解决路径"), ("货运保险", "仲裁条款", "适用法律"), ("投保确认", "仲裁地选择", "法律适用判断"), ("争议预防", "条款谈判", "证据准备"), ("culture-uncertainty-avoidance", "culture-power-distance")),
    StandardStage("stage-complaint", "投诉处理", "Complaint", 9, "面对客户不满时进行事实核查、情绪管理和补救方案设计。", ("区分投诉类型", "管理客户情绪", "设计补救方案"), ("投诉", "客户情绪", "补救方案"), ("投诉受理", "内部核查", "方案反馈"), ("积极倾听", "补救谈判", "品牌声誉保护"), ("culture-high-context", "culture-individualism-collectivism")),
    StandardStage("stage-claim-settlement", "索赔与理赔", "Claim & Settlement", 10, "围绕损失证明、索赔时效、责任认定和赔偿方案进行谈判。", ("准备索赔证据", "计算损失金额", "完成理赔谈判"), ("索赔函", "理赔", "损失计算"), ("证据收集", "索赔函发送", "理赔审核"), ("事实锚定", "调解谈判", "赔偿方案设计"), ("culture-uncertainty-avoidance", "culture-low-context")),
)


BLOOM_LEVELS: Sequence[str] = ("remember", "understand", "apply", "analyze", "evaluate", "create")


ALLOWED_KNOWLEDGE_RELATION_TYPES = {
    "RELATED_TO",
    "SUGGESTS_CO_LEARNING",
    "CONTRASTS_WITH",
    "APPLIES_TO_SCENARIO",
    "SUGGESTS_STRATEGY",
    "HAS_EXCEPTION",
    "CULTURE_SENSITIVE_TO",
    "COMBINES_WITH",
    "CONFLICTS_WITH",
}


CHAPTER_PROCESS_MAPPING: Dict[str, str] = {
    "chapter-0": "process-prologue",
    "chapter-1": "process-inquiry",
    "chapter-2": "process-offer",
    "chapter-3": "process-counter-offer",
    "chapter-4": "process-acceptance",
    "chapter-5": "process-logistics",
    "chapter-6": "process-payment",
    "chapter-7": "process-inspection",
    "chapter-8": "process-risk",
    "chapter-9": "process-complaint",
    "chapter-10": "process-claim",
}


SECTION_KNOWLEDGE_PRESETS: Dict[str, Sequence[str]] = {
    "chapter-0-section-1": (
        "首轮报价锚点策略",
        "FOB 成本构成",
        "条件式让步技巧",
    ),
    "chapter-1-section-1": (
        "询盘结构要素",
        "产品规格描述",
        "专业邮件礼仪",
    ),
    "chapter-1-section-2": (
        "需求澄清提问",
        "跟进邮件结构",
        "跨文化沟通语气",
    ),
    "chapter-2-section-1": (
        "报盘组成项目",
        "价格梯度设计",
        "价值陈述技巧",
    ),
    "chapter-2-section-2": (
        "议价让步逻辑",
        "锚定与反锚定",
        "条件交换策略",
    ),
    "chapter-3-section-1": (
        "还盘框架设计",
        "底线管理",
        "谈判让步幅度",
    ),
    "chapter-4-section-1": (
        "报价单审阅要点",
        "价格条款辨识",
        "折扣条件分析",
    ),
    "chapter-4-section-2": (
        "形式发票结构",
        "支付条款匹配",
        "风险提示编写",
    ),
    "chapter-4-section-3": (
        "发盘流程",
        "条款协同",
        "谈判节奏控制",
    ),
    "chapter-4-section-4": (
        "国际支付工具",
        "信用证条款",
        "风险保障策略",
    ),
    "chapter-4-section-5": (
        "接受函撰写",
        "订单确认流程",
        "交付节点管理",
    ),
    "chapter-4-section-6": (
        "订单履行跟进",
        "客户关系维护",
        "跨部门协同",
    ),
    "chapter-5-section-1": (
        "出口包装标准",
        "唛头设计",
        "货物防护技巧",
    ),
    "chapter-5-section-2": (
        "运输方式比较",
        "运价谈判策略",
        "时效与成本平衡",
    ),
    "chapter-5-section-3": (
        "装运时间表",
        "港口协调",
        "突发事件预案",
    ),
    "chapter-5-section-4": (
        "全链路协调",
        "物流关键路径",
        "多方沟通机制",
    ),
    "chapter-6-section-1": (
        "信用证审证流程",
        "单证一致性",
        "改证谈判技巧",
    ),
    "chapter-6-section-2": (
        "托收风险识别",
        "银行交涉话术",
        "账款保障措施",
    ),
    "chapter-6-section-3": (
        "电汇操作流程",
        "收汇时间规划",
        "资金安全控制",
    ),
    "chapter-6-section-4": (
        "分批交货条款",
        "交货时间管理",
        "库存与产能匹配",
    ),
    "chapter-6-section-5": (
        "Incoterms 责任划分",
        "风险转移节点",
        "费用承担分析",
    ),
    "chapter-6-section-6": (
        "高风险客户评估",
        "付款条件博弈",
        "风险缓释方案",
    ),
    "chapter-7-section-1": (
        "检验证书类型",
        "法定与商检机构",
        "合规要求识别",
    ),
    "chapter-7-section-2": (
        "检验条款撰写",
        "合同风险控制",
        "检验责任划分",
    ),
    "chapter-7-section-3": (
        "检验谈判策略",
        "异议处理流程",
        "质量保证设计",
    ),
    "chapter-7-section-4": (
        "不合格争议应对",
        "证据链构建",
        "解决方案谈判",
    ),
    "chapter-8-section-1": (
        "货运保险类别",
        "保险金额计算",
        "风险敞口评估",
    ),
    "chapter-8-section-2": (
        "仲裁条款要素",
        "争议解决流程",
        "适用法律选择",
    ),
    "chapter-8-section-3": (
        "海运险理赔流程",
        "索赔资料准备",
        "损失证据整理",
    ),
    "chapter-8-section-4": (
        "仲裁申请准备",
        "律师沟通要点",
        "庭审应对策略",
    ),
    "chapter-9-section-1": (
        "投诉处理流程",
        "客户情绪管理",
        "补救方案设计",
    ),
    "chapter-9-section-2": (
        "投诉到争议升级",
        "内部协调机制",
        "品牌声誉保护",
    ),
    "chapter-10-section-1": (
        "索赔函结构",
        "损失计算方法",
        "法律条款引用",
    ),
    "chapter-10-section-2": (
        "理赔审核要点",
        "调解与谈判技巧",
        "赔偿决策机制",
    ),
}


LESSON_KNOWLEDGE_PRESETS: Dict[str, Sequence[str]] = {}


def bootstrap_graph() -> None:
    """Initialise constraints and ingest the static content hierarchy."""

    if not is_configured():
        LOGGER.warning("Neo4j connection is not configured; knowledge graph features disabled")
        return

    try:
        ensure_indexes()
        status = get_initialization_status()
        if status.get("initialized"):
            sync_static_content()
        else:
            LOGGER.info(
                "Knowledge graph initialization deferred; waiting for teacher confirmation"
            )
    except GraphUnavailableError as exc:  # pragma: no cover - depends on external service
        LOGGER.warning("Unable to bootstrap knowledge graph: %s", exc)
    except (Neo4jError, ServiceUnavailable, IncompleteCommit, OSError, TimeoutError) as exc:
        LOGGER.warning("Knowledge graph bootstrap failed: %s", exc)
        _disable_graph(f"Knowledge graph bootstrap failed: {exc}")


def sync_static_content(*, include_recommendations: bool = False) -> None:
    """Mirror the SQLite content hierarchy into Neo4j."""

    driver = _get_driver()

    chapters = database.list_level_hierarchy(include_prompts=True)
    theory_hierarchy = database.list_theory_hierarchy(
        include_content=True, published_only=False
    )

    try:
        with driver.session() as session:
            session.execute_write(_merge_process_steps)
            for chapter in chapters:
                session.execute_write(_merge_chapter, chapter)
                session.execute_write(_link_chapter_process, chapter["id"])
                for section in chapter.get("sections", []):
                    session.execute_write(_merge_practice, chapter, section)

            for theory_chapter in theory_hierarchy:
                for topic in theory_chapter.get("topics", []):
                    if not topic.get("id"):
                        LOGGER.warning(
                            "Skipping theory topic with missing id in chapter %s",
                            theory_chapter.get("chapterId"),
                        )
                        continue
                    session.execute_write(_merge_theory_topic, topic)
                    for lesson in topic.get("lessons", []):
                        if not lesson.get("id"):
                            LOGGER.warning(
                                "Skipping theory lesson with missing id for topic %s",
                                topic.get("id"),
                            )
                            continue
                        session.execute_write(_merge_theory_lesson, topic, lesson)
    except (Neo4jError, ServiceUnavailable, IncompleteCommit, OSError, TimeoutError) as exc:
        LOGGER.error("Failed to synchronise static content with Neo4j: %s", exc)
        _disable_graph(f"Failed to synchronise static content: {exc}")
        raise GraphUnavailableError("Failed to synchronise static content") from exc


def _merge_default_categories_tx(tx, categories: Sequence[Dict[str, object]]) -> int:
    created = 0
    stack: List[Tuple[Optional[str], Dict[str, object]]] = [
        (None, copy.deepcopy(category)) for category in categories
    ]
    while stack:
        parent_id, node = stack.pop()
        node_id = node.get("id")
        if not node_id:
            continue
        properties = {
            "id": node_id,
            "name": node.get("name"),
            "code": node.get("code"),
            "level": node.get("level", 1),
            "orderIndex": node.get("orderIndex", 0),
            "icon": node.get("icon"),
            "color": node.get("color"),
            "description": node.get("description"),
            "isActive": bool(node.get("isActive", True)),
            "isSystemRecommended": True,
        }
        tx.run(
            "MERGE (c:KnowledgeCategory {id: $id}) "
            "SET c.name = $name, c.code = $code, c.level = $level, "
            "    c.orderIndex = $orderIndex, c.icon = $icon, c.color = $color, "
            "    c.description = $description, c.isActive = $isActive, "
            "    c.isSystemRecommended = $isSystemRecommended, "
            "    c.updatedAt = datetime(), "
            "    c.createdAt = coalesce(c.createdAt, datetime())",
            properties,
        )
        if parent_id:
            tx.run(
                "MATCH (parent:KnowledgeCategory {id: $parent_id}), "
                "      (child:KnowledgeCategory {id: $child_id}) "
                "MERGE (parent)-[:PARENT_OF]->(child)",
                {"parent_id": parent_id, "child_id": node_id},
            )
        children = node.get("children") or []
        for child in children:
            stack.append((node_id, child))
        created += 1
    return created


def _merge_default_knowledge_points_tx(tx, names: Sequence[str]) -> int:
    created = 0
    for name in names:
        if not name:
            continue
        result = tx.run(
            """
            MERGE (k:KnowledgePoint {name: $name})
            ON CREATE SET k.createdAt = datetime(),
                          k.practiceCount = coalesce(k.practiceCount, 0),
                          k.lessonCount = coalesce(k.lessonCount, 0),
                          k.category = coalesce(k.category, 'uncategorized')
            SET k.isSystemRecommended = true,
                k.source = coalesce(k.source, 'preset'),
                k.updatedAt = datetime()
            """,
            {"name": name},
        )
        summary = result.consume()
        if summary.counters.nodes_created:
            created += 1
    return created


def _merge_culture_dimensions_tx(tx) -> int:
    created = 0
    for dim in CULTURE_DIMENSIONS:
        result = tx.run(
            """
            MERGE (c:CultureDimension {id: $id})
            ON CREATE SET c.createdAt = datetime()
            SET c.name = $name,
                c.theory = $theory,
                c.summary = $summary,
                c.teachingTip = $teaching_tip,
                c.isSystemRecommended = true,
                c.updatedAt = datetime()
            """,
            {
                "id": dim.identifier,
                "name": dim.name,
                "theory": dim.theory,
                "summary": dim.summary,
                "teaching_tip": dim.teaching_tip,
            },
        )
        if result.consume().counters.nodes_created:
            created += 1
    return created


def _merge_standard_stages_tx(tx) -> Dict[str, int]:
    counters = {"stages": 0, "topics": 0, "knowledgePoints": 0, "relations": 0}
    previous_id = None
    for stage in STANDARD_STAGES:
        result = tx.run(
            """
            MERGE (s:Stage {id: $id})
            ON CREATE SET s.createdAt = datetime()
            SET s.name = $name,
                s.englishName = $english_name,
                s.summary = $summary,
                s.description = $summary,
                s.orderIndex = $order_index,
                s.objectives = $objectives,
                s.isCustomerRoute = true,
                s.updatedAt = datetime()
            """,
            {
                "id": stage.identifier,
                "name": stage.name,
                "english_name": stage.english_name,
                "summary": stage.summary,
                "order_index": stage.order_index,
                "objectives": list(stage.objectives),
            },
        )
        if result.consume().counters.nodes_created:
            counters["stages"] += 1
        if previous_id:
            tx.run(
                """
                MATCH (a:Stage {id: $previous_id}), (b:Stage {id: $current_id})
                MERGE (a)-[:PRECEDES]->(b)
                """,
                {"previous_id": previous_id, "current_id": stage.identifier},
            )
            counters["relations"] += 1
        previous_id = stage.identifier

        tx.run(
            """
            MATCH (s:Stage {id: $stage_id}), (p:ProcessStep {id: $process_id})
            MERGE (p)-[:MAPS_TO_STAGE]->(s)
            """,
            {"stage_id": stage.identifier, "process_id": stage.identifier.replace("stage-", "process-")},
        )

        for topic_key, topic_name, names in (
            ("concept", "核心概念", stage.concepts),
            ("process", "流程逻辑", stage.processes),
            ("strategy", "策略技能", stage.strategies),
        ):
            topic_id = f"{stage.identifier}-{topic_key}"
            result = tx.run(
                """
                MERGE (t:Topic {id: $id})
                ON CREATE SET t.createdAt = datetime()
                SET t.name = $name,
                    t.stage = $stage_name,
                    t.stageId = $stage_id,
                    t.type = $topic_key,
                    t.orderIndex = $order_index,
                    t.updatedAt = datetime()
                WITH t
                MATCH (s:Stage {id: $stage_id})
                MERGE (s)-[:CONTAIN_TOPIC]->(t)
                """,
                {
                    "id": topic_id,
                    "name": topic_name,
                    "stage_name": stage.name,
                    "stage_id": stage.identifier,
                    "topic_key": topic_key,
                    "order_index": {"concept": 1, "process": 2, "strategy": 3}[topic_key],
                },
            )
            if result.consume().counters.nodes_created:
                counters["topics"] += 1
            for idx, kp_name in enumerate(names, start=1):
                result = tx.run(
                    """
                    MERGE (k:KnowledgePoint {name: $name})
                    ON CREATE SET k.createdAt = datetime(),
                                  k.description = $description,
                                  k.difficulty = 'beginner',
                                  k.importance = CASE WHEN $topic_key = 'concept' THEN 'high' ELSE 'medium' END,
                                  k.tags = []
                    SET k.category = $topic_name,
                        k.categoryPath = [$stage_name, $topic_name],
                        k.stage = $stage_name,
                        k.stageId = $stage_id,
                        k.type = coalesce(k.type, $topic_key),
                        k.bloomLevel = coalesce(k.bloomLevel, CASE WHEN $topic_key = 'concept' THEN 'understand' WHEN $topic_key = 'process' THEN 'apply' ELSE 'analyze' END),
                        k.teachingObjective = coalesce(k.teachingObjective, $objective),
                        k.orderIndex = coalesce(k.orderIndex, $order_index),
                        k.isSystemRecommended = true,
                        k.updatedAt = datetime()
                    WITH k
                    MATCH (t:Topic {id: $topic_id})
                    MERGE (t)-[:INCLUDE_POINT]->(k)
                    """,
                    {
                        "name": kp_name,
                        "description": f"{stage.name}环节的{topic_name}知识单元：{kp_name}",
                        "topic_name": topic_name,
                        "topic_key": topic_key,
                        "stage_name": stage.name,
                        "stage_id": stage.identifier,
                        "topic_id": topic_id,
                        "objective": f"能够在{stage.name}场景中理解并运用“{kp_name}”。",
                        "order_index": stage.order_index * 100 + idx,
                    },
                )
                if result.consume().counters.nodes_created:
                    counters["knowledgePoints"] += 1
        for dim_id in stage.culture_dimensions:
            tx.run(
                """
                MATCH (s:Stage {id: $stage_id}), (c:CultureDimension {id: $dim_id})
                MERGE (s)-[:HAS_CULTURAL_SENSITIVITY]->(c)
                WITH s, c
                MATCH (s)-[:CONTAIN_TOPIC]->(:Topic)-[:INCLUDE_POINT]->(k:KnowledgePoint)
                WHERE coalesce(k.type, '') IN ['strategy', 'skill'] OR k.category = '策略技能'
                MERGE (k)-[:INVOLVES_CULTURE]->(c)
                """,
                {"stage_id": stage.identifier, "dim_id": dim_id},
            )
    return counters


def apply_p0_standard_route() -> Dict[str, int]:
    """Apply customer-aligned P0 defaults: ten stages and culture dimensions."""

    driver = _get_driver()
    with driver.session() as session:
        culture_created = session.execute_write(_merge_culture_dimensions_tx)
        stage_summary = session.execute_write(_merge_standard_stages_tx)
    return {"cultureDimensions": culture_created, **stage_summary}


def apply_default_recommendations() -> Dict[str, int]:
    """Create the recommended categories and knowledge point shells."""

    categories = list(_default_categories())
    driver = _get_driver()
    knowledge_names = sorted(
        {name for names in SECTION_KNOWLEDGE_PRESETS.values() for name in names}
    )

    with driver.session() as session:
        created_categories = session.execute_write(_merge_default_categories_tx, categories)
        created_knowledge = 0
        if knowledge_names:
            created_knowledge = session.execute_write(
                _merge_default_knowledge_points_tx, knowledge_names
            )

    route_summary = apply_p0_standard_route()
    return {"categories": created_categories, "knowledgePoints": created_knowledge, **route_summary}


def reset_knowledge_categories_to_default() -> Dict[str, int]:
    """Remove existing categories and rebuild the default recommended tree."""

    driver = _get_driver()
    categories = list(_default_categories())

    with driver.session() as session:
        session.run("MATCH (c:KnowledgeCategory) DETACH DELETE c")
        created_categories = session.execute_write(_merge_default_categories_tx, categories)

    return {"categories": created_categories}


def initialize_graph(
    option: str,
    *,
    initiated_by: str = "system",
    force: bool = False,
) -> Dict[str, object]:
    """Execute the teacher-driven initialization workflow."""

    normalized = (option or "").strip().lower()
    if normalized not in {"default", "blank", "import"}:
        raise ValueError("Unsupported initialization option")

    current = get_initialization_status()
    if current.get("initialized") and not force:
        raise ValueError("Knowledge graph has already been initialized")

    if not is_configured():
        raise GraphUnavailableError("Knowledge graph backend is not configured")

    ensure_indexes()
    sync_static_content(include_recommendations=False)
    route_summary = apply_p0_standard_route()

    summary = {"categories": 0, "knowledgePoints": 0, **route_summary}
    if normalized == "default":
        summary = apply_default_recommendations()

    status = _set_initialization_status(normalized)
    status["initiatedBy"] = initiated_by
    status["summary"] = summary
    return status


def _merge_process_steps(tx) -> None:
    for step in PROCESS_STEPS:
        tx.run(
            "MERGE (s:ProcessStep {id: $id}) SET s.name = $name, s.orderIndex = $order",
            {"id": step.identifier, "name": step.name, "order": step.order_index},
        )

    for previous, current in zip(PROCESS_STEPS, PROCESS_STEPS[1:]):
        tx.run(
            "MATCH (a:ProcessStep {id: $a}), (b:ProcessStep {id: $b}) "
            "MERGE (a)-[:NEXT_STEP]->(b)",
            {"a": previous.identifier, "b": current.identifier},
        )


def _merge_chapter(tx, chapter: Dict[str, object]) -> None:
    tx.run(
        "MERGE (c:Chapter {id: $id}) "
        "SET c.title = $title, c.description = $description, "
        "    c.orderIndex = $orderIndex, c.isDefault = $isDefault",
        {
            "id": chapter.get("id"),
            "title": chapter.get("title"),
            "description": chapter.get("description"),
            "orderIndex": chapter.get("orderIndex", 0),
            "isDefault": bool(chapter.get("isDefault")),
        },
    )


def _link_chapter_process(tx, chapter_id: str) -> None:
    process_id = CHAPTER_PROCESS_MAPPING.get(chapter_id)
    if not process_id:
        return
    tx.run(
        "MATCH (c:Chapter {id: $chapter_id}), (p:ProcessStep {id: $process_id}) "
        "MERGE (c)-[:COVERS_PROCESS]->(p)",
        {"chapter_id": chapter_id, "process_id": process_id},
    )


def _merge_practice(tx, chapter: Dict[str, object], section: Dict[str, object]) -> None:
    properties = {
        "id": section.get("id"),
        "title": section.get("title"),
        "description": section.get("description"),
        "environmentPromptTemplate": section.get("environmentPromptTemplate"),
        "environmentUserMessage": section.get("environmentUserMessage"),
        "conversationPromptTemplate": section.get("conversationPromptTemplate"),
        "evaluationPromptTemplate": section.get("evaluationPromptTemplate"),
        "expectsBargaining": bool(section.get("expectsBargaining")),
        "orderIndex": section.get("orderIndex", 0),
    }
    tx.run(
        "MERGE (p:Practice {id: $id}) "
        "SET p.title = $title, p.description = $description, p.orderIndex = $orderIndex, "
        "    p.environmentPromptTemplate = $environmentPromptTemplate, "
        "    p.environmentUserMessage = $environmentUserMessage, "
        "    p.conversationPromptTemplate = $conversationPromptTemplate, "
        "    p.evaluationPromptTemplate = $evaluationPromptTemplate, "
        "    p.expectsBargaining = $expectsBargaining",
        properties,
    )
    tx.run(
        "MATCH (c:Chapter {id: $chapter_id}), (p:Practice {id: $practice_id}) "
        "MERGE (c)-[:HAS_PRACTICE]->(p)",
        {"chapter_id": chapter.get("id"), "practice_id": section.get("id")},
    )


def ensure_practice_node(practice_id: str) -> Dict[str, object]:
    """Ensure a Practice node exists for the given section identifier."""

    section = database.get_section(practice_id)
    if not section:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")
    chapter = database.get_chapter(section["chapter_id"])
    if not chapter:
        raise GraphEntityNotFoundError(
            f"Chapter {section['chapter_id']} for practice {practice_id} not found"
        )

    chapter_payload = {
        "id": chapter.get("id"),
        "title": chapter.get("title"),
        "description": chapter.get("description"),
        "orderIndex": chapter.get("orderIndex", 0),
        "isDefault": bool(chapter.get("isDefault")),
    }
    section_payload = {
        "id": section.get("id"),
        "chapterId": section.get("chapter_id"),
        "title": section.get("title"),
        "description": section.get("description"),
        "environmentPromptTemplate": section.get("environment_prompt_template"),
        "environmentUserMessage": section.get("environment_user_message"),
        "conversationPromptTemplate": section.get("conversation_prompt_template"),
        "evaluationPromptTemplate": section.get("evaluation_prompt_template"),
        "expectsBargaining": bool(section.get("expects_bargaining")),
        "orderIndex": section.get("order_index", 0),
    }

    driver = _get_driver()
    with driver.session() as session:
        session.execute_write(_merge_chapter, chapter_payload)
        session.execute_write(_link_chapter_process, chapter_payload["id"])
        session.execute_write(_merge_practice, chapter_payload, section_payload)

    return {"chapter": chapter_payload, "practice": section_payload}


def _merge_theory_topic(tx, topic: Dict[str, object]) -> None:
    tx.run(
        "MERGE (t:TheoryTopic {id: $id}) "
        "SET t.title = $title, t.code = $code, t.summary = $summary, t.orderIndex = $orderIndex",
        {
            "id": topic.get("id"),
            "title": topic.get("title"),
            "code": topic.get("code"),
            "summary": topic.get("summary"),
            "orderIndex": topic.get("orderIndex", 0),
        },
    )
    chapter_id = topic.get("chapterId")
    if chapter_id:
        tx.run(
            "MATCH (c:Chapter {id: $chapter_id}), (t:TheoryTopic {id: $topic_id}) "
            "MERGE (c)-[:HAS_TOPIC]->(t)",
            {"chapter_id": chapter_id, "topic_id": topic.get("id")},
        )


def _merge_theory_lesson(tx, topic: Dict[str, object], lesson: Dict[str, object]) -> None:
    tx.run(
        "MERGE (l:TheoryLesson {id: $id}) "
        "SET l.title = $title, l.code = $code, l.orderIndex = $orderIndex, "
        "    l.isPublished = $isPublished, l.contentHtml = $contentHtml",
        {
            "id": lesson.get("id"),
            "title": lesson.get("title"),
            "code": lesson.get("code"),
            "orderIndex": lesson.get("orderIndex", 0),
            "isPublished": bool(lesson.get("isPublished")),
            "contentHtml": lesson.get("contentHtml"),
        },
    )
    tx.run(
        "MATCH (t:TheoryTopic {id: $topic_id}), (l:TheoryLesson {id: $lesson_id}) "
        "MERGE (t)-[:HAS_LESSON]->(l)",
        {"topic_id": topic.get("id"), "lesson_id": lesson.get("id")},
    )


def _ensure_practice_knowledge(tx, practice_id: str, points: List[str]) -> None:
    existing = tx.run(
        "MATCH (:Practice {id: $id})-[:TESTS]->(k:KnowledgePoint) RETURN collect(k.name) AS names",
        {"id": practice_id},
    ).single()
    if existing and existing["names"]:
        return
    _set_practice_knowledge_tx(tx, practice_id, points)


def _ensure_lesson_knowledge(tx, lesson_id: str, points: List[object]) -> None:
    existing = tx.run(
        (
            "MATCH (:TheoryLesson {id: $id})-[rel]->(k:KnowledgePoint) "
            "WHERE type(rel) = 'EXPLAINS' RETURN collect(k.name) AS names"
        ),
        {"id": lesson_id},
    ).single()
    if existing and existing["names"]:
        return
    normalized = _normalize_knowledge_point_payloads(points)
    if not normalized:
        return
    _set_lesson_knowledge_tx(tx, lesson_id, normalized)


def set_practice_knowledge_points(practice_id: str, points: Sequence[object]) -> None:
    ensure_practice_node(practice_id)
    driver = _get_driver()
    normalized_payloads = _normalize_knowledge_point_payloads(points)
    names = [payload["name"] for payload in normalized_payloads if payload.get("name")]
    with driver.session() as session:
        session.execute_write(_set_practice_knowledge_tx, practice_id, names)


def _set_practice_knowledge_tx(tx, practice_id: str, points: Sequence[str]) -> None:
    record = tx.run(
        "MATCH (p:Practice {id: $id}) RETURN p", {"id": practice_id}
    ).single()
    if not record:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")

    tx.run(
        "MATCH (:Practice {id: $id})-[rel:TESTS]->(:KnowledgePoint) DELETE rel",
        {"id": practice_id},
    )
    for name in points:
        tx.run(
            "MATCH (p:Practice {id: $id}) "
            "MERGE (k:KnowledgePoint {name: $name}) "
            "MERGE (p)-[:TESTS]->(k)",
            {"id": practice_id, "name": name},
        )


def get_practice_knowledge_recommendations(practice_id: str) -> Dict[str, object]:
    """Return existing and recommended knowledge points for a practice."""

    details = ensure_practice_node(practice_id)
    recommended = list(SECTION_KNOWLEDGE_PRESETS.get(practice_id, ()))
    existing_records = _execute_read(
        "MATCH (:Practice {id: $id})-[:TESTS]->(k:KnowledgePoint) "
        "RETURN collect(k.name) AS names",
        {"id": practice_id},
    )
    existing = []
    if existing_records:
        existing = sorted({name for name in existing_records[0].get("names", []) if name})

    return {
        "practice": {
            "id": practice_id,
            "title": details["practice"].get("title"),
            "chapterId": details["practice"].get("chapterId"),
        },
        "existing": existing,
        "recommended": recommended,
    }


def set_lesson_knowledge_points(lesson_id: str, points: Sequence[object]) -> None:
    driver = _get_driver()
    normalized = _normalize_knowledge_point_payloads(points)
    with driver.session() as session:
        session.execute_write(_set_lesson_knowledge_tx, lesson_id, normalized)


def _set_lesson_knowledge_tx(tx, lesson_id: str, points: Sequence[Dict[str, object]]) -> None:
    record = tx.run(
        "MATCH (l:TheoryLesson {id: $id}) RETURN l", {"id": lesson_id}
    ).single()
    if not record:
        raise GraphEntityNotFoundError(f"Theory lesson {lesson_id} not found")

    tx.run(
        (
            "MATCH (:TheoryLesson {id: $id})-[rel]->(:KnowledgePoint) "
            "WHERE type(rel) = 'EXPLAINS' DELETE rel"
        ),
        {"id": lesson_id},
    )
    for payload in points:
        name = payload.get("name") if isinstance(payload, dict) else None
        if not name:
            continue
        summary_value = payload.get("summary", "")
        if summary_value is None:
            summary_value = ""
        elif not isinstance(summary_value, str):
            summary_value = str(summary_value)
        body_html_value = payload.get("bodyHtml", "")
        if body_html_value is None:
            body_html_value = ""
        image_url_value = payload.get("imageUrl", "")
        if image_url_value is None:
            image_url_value = ""
        elif not isinstance(image_url_value, str):
            image_url_value = str(image_url_value)
        image_alt_value = payload.get("imageAlt", "")
        if image_alt_value is None:
            image_alt_value = ""
        elif not isinstance(image_alt_value, str):
            image_alt_value = str(image_alt_value)
        content_value = payload.get("content", "") or body_html_value
        if content_value is None:
            content_value = ""
        elif not isinstance(content_value, str):
            content_value = str(content_value)
        anchor_value = payload.get("anchorId", "")
        if anchor_value is None:
            anchor_value = ""
        elif not isinstance(anchor_value, str):
            anchor_value = str(anchor_value)
        tags_value = payload.get("tags", [])
        if isinstance(tags_value, (list, tuple)):
            cleaned_tags = []
            for tag in tags_value:
                if tag is None:
                    continue
                tag_str = str(tag).strip()
                if tag_str and tag_str not in cleaned_tags:
                    cleaned_tags.append(tag_str)
            tags_value = cleaned_tags
        else:
            tags_value = []
        knowledge_id_value = payload.get("knowledgeId", "")
        if knowledge_id_value is None:
            knowledge_id_value = ""
        elif not isinstance(knowledge_id_value, str):
            knowledge_id_value = str(knowledge_id_value)
        tx.run(
            "MATCH (l:TheoryLesson {id: $id}) "
            "MERGE (k:KnowledgePoint {name: $name}) "
            "SET k.summary = CASE WHEN $summary = '' THEN k.summary ELSE $summary END "
            "SET k.imageUrl = CASE WHEN $imageUrl = '' THEN k.imageUrl ELSE $imageUrl END "
            "SET k.imageAlt = CASE WHEN $imageAlt = '' THEN k.imageAlt ELSE $imageAlt END "
            "SET k.content = CASE WHEN $content = '' THEN k.content ELSE $content END "
            "SET k.bodyHtml = CASE WHEN $bodyHtml = '' THEN k.bodyHtml ELSE $bodyHtml END "
            "SET k.sourceId = CASE WHEN $knowledgeId = '' THEN k.sourceId ELSE $knowledgeId END "
            "SET k.tags = CASE WHEN size($tags) = 0 THEN k.tags ELSE $tags END "
            "MERGE (l)-[rel:EXPLAINS]->(k) "
            "SET rel.anchorId = CASE WHEN $anchorId = '' THEN rel.anchorId ELSE $anchorId END "
            "SET rel.summary = CASE WHEN $summary = '' THEN rel.summary ELSE $summary END "
            "SET rel.bodyHtml = CASE WHEN $bodyHtml = '' THEN rel.bodyHtml ELSE $bodyHtml END "
            "SET rel.imageUrl = CASE WHEN $imageUrl = '' THEN rel.imageUrl ELSE $imageUrl END "
            "SET rel.imageAlt = CASE WHEN $imageAlt = '' THEN rel.imageAlt ELSE $imageAlt END "
            "SET rel.tags = CASE WHEN size($tags) = 0 THEN rel.tags ELSE $tags END",
            {
                "id": lesson_id,
                "name": name,
                "summary": summary_value,
                "bodyHtml": body_html_value,
                "content": content_value,
                "imageUrl": image_url_value,
                "imageAlt": image_alt_value,
                "anchorId": anchor_value,
                "tags": tags_value,
                "knowledgeId": knowledge_id_value,
            },
        )


def _normalize_knowledge_point_payloads(points: Sequence[object]) -> List[Dict[str, object]]:
    """Normalize arbitrary knowledge point payloads into consistent objects."""

    normalized: List[Dict[str, object]] = []
    by_name: Dict[str, Dict[str, object]] = {}

    def _clean_string(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _merge_payload(target: Dict[str, object], source: Dict[str, object]) -> None:
        for key, raw_value in source.items():
            if key == "name":
                continue
            if raw_value is None:
                continue
            if key == "tags":
                existing = target.get("tags")
                merged: List[str] = []
                if isinstance(existing, list):
                    merged.extend(str(tag) for tag in existing if str(tag))
                for tag in raw_value:
                    tag_value = _clean_string(tag)
                    if not tag_value:
                        continue
                    if tag_value not in merged:
                        merged.append(tag_value)
                if merged:
                    target["tags"] = merged
                continue
            cleaned_value = raw_value
            if isinstance(raw_value, str):
                cleaned_value = raw_value.strip()
            target[key] = cleaned_value

    for entry in points:
        if isinstance(entry, str):
            name = entry.strip()
            payload: Dict[str, object] = {"name": name}
        elif isinstance(entry, dict):
            candidate = (
                entry.get("name")
                or entry.get("title")
                or entry.get("label")
                or entry.get("id")
            )
            name = _clean_string(candidate)
            payload = {"name": name}
            summary = _clean_string(entry.get("summary") or entry.get("description"))
            if summary:
                payload["summary"] = summary
            body_html = entry.get("bodyHtml") or entry.get("html") or entry.get("body")
            if isinstance(body_html, str) and body_html.strip():
                payload["bodyHtml"] = body_html
            image_url = _clean_string(entry.get("imageUrl") or entry.get("image") or entry.get("coverUrl"))
            if image_url:
                payload["imageUrl"] = image_url
            image_alt = _clean_string(entry.get("imageAlt") or entry.get("alt"))
            if image_alt:
                payload["imageAlt"] = image_alt
            anchor_id = _clean_string(entry.get("anchorId") or entry.get("anchor"))
            if anchor_id:
                payload["anchorId"] = anchor_id
            source_id = _clean_string(entry.get("knowledgeId") or entry.get("sourceId"))
            if source_id:
                payload["knowledgeId"] = source_id
            tags_field = entry.get("tags")
            tags: List[str] = []
            if isinstance(tags_field, (list, tuple)):
                tags = [_clean_string(tag) for tag in tags_field if _clean_string(tag)]
            elif isinstance(tags_field, str):
                tags = [tag.strip() for tag in tags_field.split(",") if tag.strip()]
            if tags:
                payload["tags"] = tags
        else:
            continue

        name = payload.get("name", "")
        if not isinstance(name, str):
            name = str(name)
        cleaned_name = name.strip()
        if not cleaned_name:
            continue
        payload["name"] = cleaned_name

        existing = by_name.get(cleaned_name)
        if existing:
            _merge_payload(existing, payload)
            continue

        by_name[cleaned_name] = payload
        normalized.append(payload)

    return normalized


def _normalize_category_path_input(value: Optional[object]) -> List[str]:
    """Normalize category path inputs from APIs or imports into a list of segments."""

    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        result = [str(item).strip() for item in value if str(item).strip()]
        return [segment for segment in result if segment]
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return []
        separators = re.compile(r"[>/\\|]")
        return [segment.strip() for segment in separators.split(cleaned) if segment.strip()]
    return [str(value).strip()] if str(value).strip() else []


def _category_path_to_string(path: Sequence[str]) -> str:
    return "/".join(segment.strip() for segment in path if segment)


def get_practice_detail(practice_id: str) -> Dict[str, object]:
    try:
        records = _execute_read(
            """
            MATCH (p:Practice {id: $id})
            OPTIONAL MATCH (p)-[:TESTS]->(k:KnowledgePoint)
            OPTIONAL MATCH (c:Chapter)-[:HAS_PRACTICE]->(p)
            RETURN p AS practice, collect(DISTINCT k.name) AS knowledge, c.id AS chapterId
            """,
            {"id": practice_id},
        )
    except GraphUnavailableError:
        LOGGER.debug("Graph unavailable when fetching practice %s; returning fallback", practice_id)
        return _fallback_practice_detail(practice_id)

    if not records:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")

    record = records[0]
    practice = record["practice"]
    payload = {
        "id": practice.get("id"),
        "title": practice.get("title"),
        "description": practice.get("description"),
        "orderIndex": practice.get("orderIndex"),
        "chapterId": record.get("chapterId"),
        "knowledgePoints": sorted(filter(None, record.get("knowledge") or [])),
    }
    return payload


def get_lesson_detail(lesson_id: str) -> Dict[str, object]:
    try:
        records = _execute_read(
            """
            MATCH (l:TheoryLesson {id: $id})
            OPTIONAL MATCH (l)-[rel]->(k:KnowledgePoint)
            WHERE type(rel) = 'EXPLAINS'
            OPTIONAL MATCH (t:TheoryTopic)-[:HAS_LESSON]->(l)
            WITH l,
                 t,
                 k,
                 CASE WHEN rel IS NULL THEN {} ELSE properties(rel) END AS relProps,
                 CASE WHEN k IS NULL THEN {} ELSE properties(k) END AS kProps
            RETURN l AS lesson,
                   collect(DISTINCT CASE WHEN k IS NULL THEN NULL ELSE {
                     name: kProps['name'],
                     summary: coalesce(relProps['summary'], kProps['summary']),
                     bodyHtml: coalesce(relProps['bodyHtml'], kProps['bodyHtml']),
                     imageUrl: coalesce(relProps['imageUrl'], kProps['imageUrl']),
                     imageAlt: coalesce(relProps['imageAlt'], kProps['imageAlt']),
                     anchorId: relProps['anchorId'],
                     tags: relProps['tags'],
                     knowledgeId: kProps['sourceId']
                   } END) AS knowledge,
                   t.id AS topicId
            """,
            {"id": lesson_id},
        )
    except GraphUnavailableError:
        LOGGER.debug("Graph unavailable when fetching lesson %s; returning fallback", lesson_id)
        return _fallback_lesson_detail(lesson_id)

    if not records:
        raise GraphEntityNotFoundError(f"Theory lesson {lesson_id} not found")

    record = records[0]
    lesson = record["lesson"]
    raw_knowledge = [item for item in (record.get("knowledge") or []) if item]
    return {
        "id": lesson.get("id"),
        "title": lesson.get("title"),
        "code": lesson.get("code"),
        "topicId": record.get("topicId"),
        "knowledgePoints": _normalize_knowledge_point_payloads(raw_knowledge),
    }


def _list_knowledge_vocabulary_with_prereqs() -> List[Dict[str, object]]:
    records = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        OPTIONAL MATCH (k)-[:REQUIRES]->(p:KnowledgePoint)
        RETURN k.name AS name,
               k.summary AS summary,
               collect(DISTINCT p.name) AS prereqs
        """
    )
    vocab: List[Dict[str, object]] = []
    for rec in records:
        name = rec.get("name")
        if not name:
            continue
        vocab.append(
            {
                "name": str(name),
                "summary": rec.get("summary") or "",
                "prerequisites": [p for p in rec.get("prereqs") or [] if p],
            }
        )
    return vocab


def get_knowledge_prerequisite_map() -> Dict[str, List[str]]:
    """Return a mapping of knowledge point name -> prerequisites (REQUIRES)."""
    result: Dict[str, List[str]] = {}
    for item in _list_knowledge_vocabulary_with_prereqs():
        result[item["name"]] = item.get("prerequisites", [])
    return result


def detect_knowledge_points_in_text(html: str, max_results: int = 20) -> List[Dict[str, object]]:
    """Detect knowledge points via RAG匹配（嵌入 + 分段）。"""
    if not html:
        return []

    # 准备文本分段
    text = re.sub(r"<[^>]+>", " ", html or "")
    chunks = rag_matcher.chunk_text(text, max_len=420)
    if not chunks:
        return []

    # 知识库（含摘要/正文）
    knowledge_cards = list_knowledge_points()
    prereq_map = get_knowledge_prerequisite_map()

    best_scores: Dict[str, float] = {}
    best_meta: Dict[str, Dict[str, object]] = {}

    for chunk in chunks:
        best, score, _ = rag_matcher.match(chunk, knowledge_cards)
        if not best:
            continue
        name = (best.get("name") or "").strip()
        if not name:
            continue
        # 记录最高得分
        if score <= best_scores.get(name, 0.0):
            continue
        best_scores[name] = float(score)
        best_meta[name] = {
          "name": name,
          "summary": best.get("summary") or "",
          "bodyHtml": best.get("bodyHtml") or "",
          "prerequisites": prereq_map.get(name) or [],
          "source": "detected",
          "score": float(score),
        }

    ranked = sorted(best_meta.values(), key=lambda x: x.get("score", 0), reverse=True)
    return ranked[:max_results]


def compute_lesson_knowledge_and_graph(
    lesson_id: str,
    *,
    include_unpublished: bool = True,
    snapshot_limit: int = 800,
) -> Dict[str, object]:
    """Detect + persist lesson knowledge points, then cache subgraph/network."""

    lesson = database.get_theory_lesson(lesson_id, include_unpublished=include_unpublished)
    if not lesson:
        raise GraphEntityNotFoundError(f"Theory lesson {lesson_id} not found")

    html = lesson.get("contentHtml") or ""

    # 取当前图谱中的已关联知识点，防止覆盖教师手动配置
    existing_points: List[Dict[str, object]] = []
    try:
        detail = get_lesson_detail(lesson_id)
        existing_points = detail.get("knowledgePoints") or []
    except GraphUnavailableError:
        existing_points = []

    try:
        detected_points = detect_knowledge_points_in_text(html)
    except GraphUnavailableError:
        detected_points = []
    merged_points = _normalize_knowledge_point_payloads([*existing_points, *detected_points])
    detected_names = [p.get("name") for p in detected_points if isinstance(p, dict)]

    try:
        set_lesson_knowledge_points(lesson_id, merged_points)
    except GraphUnavailableError:
        LOGGER.warning("Graph unavailable when persisting lesson %s knowledge; cache only", lesson_id)

    # 构建学生端需要的子图/网络视图（无需二次检测）
    from services import lesson_graph_service  # lazy import to avoid circular dependency

    subgraph = lesson_graph_service.build_lesson_subgraph(
        lesson_id, detected_names=detected_names
    )

    highlight_names = [p.get("name") for p in merged_points if isinstance(p, dict) and p.get("name")]
    try:
        network_view = lesson_graph_service.build_lesson_network_view(
            lesson_id,
            highlight_names=highlight_names,
            limit=snapshot_limit,
        )
    except GraphUnavailableError:
        network_view = {"nodes": [], "edges": [], "highlights": []}

    payload = {
        "lessonId": lesson_id,
        "knowledgePoints": highlight_names,
        "detected": detected_points,
        "merged": merged_points,
        "subgraph": subgraph,
        "network": network_view,
        "updatedAt": datetime.utcnow().isoformat() + "Z",
    }

    try:
        cache_lesson_graph_payload(lesson_id, payload)
    except Exception as exc:  # pragma: no cover - cache must not break flow
        LOGGER.warning("Failed to cache lesson graph for %s: %s", lesson_id, exc)

    return payload


def get_practices_for_kp(name: str, limit: int = 5) -> List[Dict[str, object]]:
    records = _execute_read(
        """
        MATCH (p:Practice)-[:TESTS]->(k:KnowledgePoint {name: $name})
        OPTIONAL MATCH (c:Chapter)-[:HAS_PRACTICE]->(p)
        RETURN p.id AS id, p.title AS title, p.description AS description, c.id AS chapterId
        LIMIT $limit
        """,
        {"name": name, "limit": limit},
    )
    practices: List[Dict[str, object]] = []
    for rec in records:
        practices.append(
            {
                "id": rec.get("id"),
                "title": rec.get("title") or rec.get("id"),
                "description": rec.get("description") or "",
                "chapterId": rec.get("chapterId"),
            }
        )
    return practices


def list_knowledge_points() -> List[Dict[str, object]]:
    return _execute_read(
        """
        MATCH (k:KnowledgePoint)
        OPTIONAL MATCH (t:Topic)-[:INCLUDE_POINT]->(k)
        OPTIONAL MATCH (k)<-[:TESTS]-(p:Practice)
        OPTIONAL MATCH (k)<-[rel]-(l:TheoryLesson)
        WHERE rel IS NULL OR type(rel) = 'EXPLAINS'
        RETURN elementId(k) AS nodeId,
               k.name AS name,
               k.summary AS summary,
               k.bodyHtml AS bodyHtml,
               k.imageUrl AS imageUrl,
               k.imageAlt AS imageAlt,
               k.category AS category,
               k.categoryPath AS categoryPath,
               k.orderIndex AS orderIndex,
               k.sourceId AS knowledgeId,
               k.tags AS tags,
               k.bloomLevel AS bloom_level,
               k.cultureTags AS culture_tags,
               k.civicTags AS civic_tags,
               k.teachingObjective AS teaching_objective,
               k.assessmentHint AS assessment_hint,
               k.lex_role AS lex_role,
               t.name AS topic,
               CASE
                   WHEN 'Terminology' IN labels(k) THEN 'Terminology'
                   WHEN 'Skill' IN labels(k) THEN 'Skill'
                   ELSE coalesce(k.nodeType, 'KnowledgePoint')
               END AS nodeType,
               count(DISTINCT p) AS practiceCount,
               count(DISTINCT l) AS lessonCount
        ORDER BY name
        """,
    )


def get_related_practices_for_lesson(lesson_id: str) -> List[Dict[str, object]]:
    try:
        return _execute_read(
            """
            MATCH (l:TheoryLesson {id: $id})-[rel]->(k:KnowledgePoint)<-[:TESTS]-(p:Practice)
            WHERE type(rel) = 'EXPLAINS'
            OPTIONAL MATCH (c:Chapter)-[:HAS_PRACTICE]->(p)
            RETURN DISTINCT p.id AS id, p.title AS title, p.description AS description,
                   p.orderIndex AS orderIndex, c.id AS chapterId
            ORDER BY p.orderIndex, p.title
            """,
            {"id": lesson_id},
        )
    except GraphUnavailableError:
        LOGGER.debug(
            "Graph unavailable when listing practices for lesson %s; returning empty list",
            lesson_id,
        )
        # Ensure the lesson exists before returning an empty payload.
        _fallback_lesson_detail(lesson_id)
        return []


def get_related_lessons_for_practice(practice_id: str) -> List[Dict[str, object]]:
    try:
        return _execute_read(
            """
            MATCH (p:Practice {id: $id})-[:TESTS]->(k:KnowledgePoint)<-[rel]-(l:TheoryLesson)
            WHERE type(rel) = 'EXPLAINS'
            OPTIONAL MATCH (t:TheoryTopic)-[:HAS_LESSON]->(l)
            RETURN DISTINCT l.id AS id, l.title AS title, l.code AS code, l.orderIndex AS orderIndex,
                   l.isPublished AS isPublished, t.id AS topicId
            ORDER BY l.orderIndex, l.title
            """,
            {"id": practice_id},
        )
    except GraphUnavailableError:
        LOGGER.debug(
            "Graph unavailable when listing lessons for practice %s; returning empty list",
            practice_id,
        )
        _fallback_practice_detail(practice_id)
        return []


def fetch_graph_snapshot(limit: int = 800) -> Dict[str, object]:
    # _refresh_node_type_labels()  # 高频读接口避免触发全库写入导致锁/性能问题

    allowed_labels = [
        "Stage",
        "Topic",
        "Chapter",
        "Practice",
        "TheoryTopic",
        "TheoryLesson",
        "KnowledgePoint",
        "CultureDimension",
        "Skill",
        "Terminology",
        "ProcessStep",
    ]
    nodes = _execute_read(
        """
        MATCH (n)
        WHERE any(label IN labels(n) WHERE label IN $allowed)
        OPTIONAL MATCH (s1:Stage)-[:CONTAIN_TOPIC]->(:Topic)-[:HAS_CATEGORY]->(:KnowledgeCategory)-[:CONTAINS]->(kp:KnowledgePoint)
        WHERE kp = n
        OPTIONAL MATCH (s2:Stage)-[:CONTAIN_TOPIC]->(:Topic)-[:HAS_CATEGORY]->(kc:KnowledgeCategory)
        WHERE kc = n
        OPTIONAL MATCH (s3:Stage)-[:CONTAIN_TOPIC]->(t:Topic)
        WHERE t = n
        OPTIONAL MATCH (s4:Stage)
        WHERE s4 = n
        WITH
            labels(n) AS labels,
            n AS node,
            coalesce(s1.name, s2.name, s3.name, s4.name) AS stageName,
            CASE
                WHEN 'Stage' IN labels(n) THEN 0
                WHEN 'Topic' IN labels(n) THEN 1
                WHEN 'ProcessStep' IN labels(n) THEN 2
                WHEN 'Chapter' IN labels(n) THEN 3
                WHEN 'Practice' IN labels(n) THEN 4
                WHEN 'TheoryTopic' IN labels(n) THEN 5
                WHEN 'TheoryLesson' IN labels(n) THEN 6
                WHEN 'KnowledgePoint' IN labels(n) THEN 7
                ELSE 99
            END AS priority
        WITH DISTINCT labels, node, stageName, priority
        ORDER BY priority ASC
        LIMIT $limit
        RETURN labels, node, stageName
        """,
        {"allowed": allowed_labels, "limit": limit},
    )

    edges = _execute_read(
        """
        MATCH (a)-[r]->(b)
        WHERE any(label IN labels(a) WHERE label IN $allowed)
          AND any(label IN labels(b) WHERE label IN $allowed)
          AND type(r) <> 'HAS_TOPIC'
        RETURN labels(a) AS sourceLabels, a AS source,
               labels(b) AS targetLabels, b AS target,
               type(r) AS type, r AS relationship
        LIMIT $limit
        """,
        {"allowed": allowed_labels, "limit": limit * 3},
    )

    # 额外生成 Topic -> KnowledgePoint 的直连边（通过 Category 汇总）
    extra_edges = _execute_read(
        """
        MATCH (t:Topic)-[:HAS_CATEGORY]->(:KnowledgeCategory)-[:CONTAINS]->(k:KnowledgePoint)
        RETURN labels(t) AS sourceLabels, t AS source,
               labels(k) AS targetLabels, k AS target,
               'INCLUDE_POINT' AS type, {} AS relationship
        """
    )

    node_payload: Dict[str, Dict[str, object]] = {}
    for record in nodes:
        labels = record.get("labels") or []
        node = record.get("node") or {}
        primary = _select_primary_label(labels)
        if primary == "KnowledgeCategory":
            continue  # 不在总览中展示分类节点
        if not primary:
            continue
        identifier = _extract_node_identifier(primary, node)
        if not identifier:
            continue
        key = f"{primary}:{identifier}"

        # 计算层级和顺序
        level = _calculate_node_level(primary)
        stage_name = (
            record.get("stageName")
            or node.get("stage")
            or (node.get("name") if primary == "Stage" else None)
        )
        order = node.get("orderIndex", 0) or node.get("order", 0) or 0

        node_payload[key] = {
            "key": key,
            "label": primary,
            "nodeType": primary,
            "title": node.get("title") or node.get("name") or node.get("code") or identifier,
            "subtitle": _build_node_subtitle(primary, node),
            "level": level,
            "stage": stage_name,
            "stageName": stage_name,
            "order": order,
            "group": primary,
        }

    edge_payload: List[Dict[str, object]] = []
    for record in edges + extra_edges:
        source_primary = _select_primary_label(record.get("sourceLabels") or [])
        target_primary = _select_primary_label(record.get("targetLabels") or [])
        source_identifier = _extract_node_identifier(source_primary, record.get("source") or {})
        target_identifier = _extract_node_identifier(target_primary, record.get("target") or {})
        if not (source_primary and target_primary and source_identifier and target_identifier):
            continue
        source_key = f"{source_primary}:{source_identifier}"
        target_key = f"{target_primary}:{target_identifier}"
        if source_key not in node_payload or target_key not in node_payload:
            continue

        relationship = record.get("relationship")
        edge_type = record.get("type")

        # 转换 Neo4j Relationship 对象为字典
        rel_props = {}
        if relationship:
            try:
                # Neo4j Relationship 对象可以像字典一样访问
                if hasattr(relationship, 'items'):
                    rel_props = dict(relationship)
                elif hasattr(relationship, 'get'):
                    rel_props = {k: relationship.get(k) for k in getattr(relationship, 'keys', lambda: [])()}
            except (AttributeError, TypeError):
                pass

        edge_payload.append(
            {
                "source": source_key,
                "target": target_key,
                "type": edge_type,
                "label": _get_edge_label(edge_type, rel_props),
                "arrows": "to",
            }
        )

    return {"nodes": list(node_payload.values()), "edges": edge_payload}


def _refresh_node_type_labels() -> None:
    """根据节点属性 type/nodeType 自动补打 Skill/Terminology 标签, 保证前端着色正确。"""

    skill_aliases = [
        "skill",
        "skills",
        "技能",
        "技能型",
        "技能性",
        "业务流程",
        "流程",
        "流程型",
        "process",
        "practice",
    ]
    term_aliases = [
        "terminology",
        "term",
        "concept",
        "conceptual",
        "术语",
        "概念",
        "概念型",
        "概念性",
        "概念类",
    ]

    driver = _get_driver()
    with driver.session() as session:
        session.run(
            """
            MATCH (k:KnowledgePoint)
            WITH k, toLower(coalesce(k.nodeType, k.type, '')) AS t
            WHERE t <> ''
            FOREACH (_ IN CASE WHEN t IN $skill THEN [1] ELSE [] END |
                SET k:Skill SET k.nodeType = 'Skill'
            )
            FOREACH (_ IN CASE WHEN t IN $term THEN [1] ELSE [] END |
                SET k:Terminology SET k.nodeType = 'Terminology'
            )
            """,
            {"skill": skill_aliases, "term": term_aliases},
        )


def _select_primary_label(labels: Iterable[str]) -> Optional[str]:
    priority = [
        "Stage",
        "Topic",
        "KnowledgeCategory",
        "CultureDimension",
        "Chapter",
        "Practice",
        "TheoryTopic",
        "TheoryLesson",
        "Skill",
        "Terminology",
        "KnowledgePoint",
        "ProcessStep",
    ]
    for label in priority:
        if label in labels:
            return label
    return next(iter(labels), None) if labels else None


def _extract_node_identifier(label: Optional[str], node: Dict[str, object]) -> Optional[str]:
    if not label:
        return None
    if label == "KnowledgeCategory":
        base = node.get("name") or node.get("type")
        topic = node.get("topic") or node.get("topicName")
        stage = node.get("stage") or node.get("stageName")
        if topic:
            return f"{topic}:{base}"
        if stage and base:
            return f"{stage}:{base}"
        return base
    if label in {"Stage", "ProcessStep", "KnowledgePoint", "Skill", "Terminology", "Topic", "CultureDimension"}:
        if label == "Topic":
            stage = node.get("stage") or node.get("stageName") or ""
            name = node.get("name")
            return f"{stage}:{name}" if stage else name
        return node.get("name")
    return node.get("id") or node.get("code") or node.get("title") or node.get("name")


def _build_node_subtitle(label: str, node: Dict[str, object]) -> Optional[str]:
    if label == "Practice":
        return node.get("description")
    if label == "TheoryLesson":
        return node.get("code")
    if label == "Topic":
        return node.get("description")
    if label == "KnowledgePoint":
        return None
    if label == "ProcessStep":
        return f"顺序：{node.get('orderIndex')}"
    return node.get("description")


def _calculate_node_level(label: str) -> int:
    """
    计算节点的层级（用于层级布局）
    Level 0: Stage（阶段：询盘、报盘、还盘等）
    Level 1: Topic（二级主题）
    Level 2: KnowledgeCategory（类型分类）
    Level 3: KnowledgePoint（知识点）
    Level 4: ProcessStep（流程步骤）
    """
    level_map = {
        "Stage": 0,
        "Topic": 1,
        "KnowledgeCategory": 2,
        "Chapter": 1,
        "TheoryTopic": 2,
        "Practice": 2,
        "TheoryLesson": 2,
        "KnowledgePoint": 3,
        "Skill": 3,
        "Terminology": 3,
        "CultureDimension": 2,
        "ProcessStep": 4,
    }
    return level_map.get(label, 99)


def _get_edge_label(edge_type: Optional[str], relationship: Optional[Dict[str, object]]) -> Optional[str]:
    """
    获取关系边的标签（用于可视化）
    """
    # 知识点关系类型的中文映射
    if relationship:
        relation_type = relationship.get("relationType")
        if relation_type:
            type_labels = {
                "prerequisite": "前置",
                "similar": "相似",
                "contrast": "对比",
                "related": "相关",
            }
            label = type_labels.get(relation_type)
            if label:
                return label

    # 边类型的默认标签
    if edge_type:
        edge_labels = {
            "REQUIRES": "依赖",
            "RELATED_TO": "关联",
            "SUGGESTS_CO_LEARNING": "建议同时学",
            "RELATES_TO": "关联",
            "HAS_TOPIC": "包含",
            "CONTAIN_TOPIC": "包含",
            "INCLUDE_POINT": "收录",
            "TESTS": "考察",
            "EXPLAINS": "讲解",
            "HAS_PRACTICE": "练习",
            "NEXT": "下一步",
            "PRECEDES": "顺序",
            "HAS_CULTURAL_SENSITIVITY": "文化敏感",
            "INVOLVES_CULTURE": "涉及文化",
            "MAPS_TO_STAGE": "映射阶段",
            "CONTRASTS_WITH": "对比",
            "APPLIES_TO_SCENARIO": "情境适用",
            "SUGGESTS_STRATEGY": "策略建议",
            "HAS_EXCEPTION": "例外",
            "CULTURE_SENSITIVE_TO": "文化相关",
            "COMBINES_WITH": "组合",
            "CONFLICTS_WITH": "冲突",
        }
        return edge_labels.get(edge_type)

    return None


# ========== 知识点管理增强功能 ==========


def list_knowledge_points_enhanced(
    search: Optional[str] = None,
    category: Optional[str] = None,
    difficulty: Optional[str] = None,
) -> List[Dict[str, object]]:
    """获取知识点列表，支持过滤搜索。"""

    # 构建查询条件
    where_clauses = []
    params = {}

    if search:
        where_clauses.append("(k.name CONTAINS $search OR k.description CONTAINS $search)")
        params["search"] = search

    if category:
        where_clauses.append("k.category = $category")
        params["category"] = category

    if difficulty:
        where_clauses.append("k.difficulty = $difficulty")
        params["difficulty"] = difficulty

    where_clause = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    query = f"""
        MATCH (k:KnowledgePoint)
        {where_clause}
        OPTIONAL MATCH (k)<-[:TESTS]-(p:Practice)
        OPTIONAL MATCH (k)<-[rel]-(l:TheoryLesson)
        WHERE rel IS NULL OR type(rel) = 'EXPLAINS'
        OPTIONAL MATCH (k)<-[:REQUIRES]-(dependent:KnowledgePoint)
        OPTIONAL MATCH (k)-[:REQUIRES]->(prereq:KnowledgePoint)
        OPTIONAL MATCH (k)-[r:RELATED_TO|SUGGESTS_CO_LEARNING|CONTRASTS_WITH|APPLIES_TO_SCENARIO|SUGGESTS_STRATEGY|HAS_EXCEPTION|CULTURE_SENSITIVE_TO|COMBINES_WITH|CONFLICTS_WITH]-(related:KnowledgePoint)
        RETURN k.name AS name,
               k.description AS description,
               k.category AS category,
               k.categoryPath AS category_path,
               k.difficulty AS difficulty,
               k.importance AS importance,
               k.estimatedDuration AS estimated_duration,
               k.content AS content,
               k.orderIndex AS order_index,
               k.tags AS tags,
               k.bloomLevel AS bloom_level,
               k.cultureTags AS culture_tags,
               k.civicTags AS civic_tags,
               k.teachingObjective AS teaching_objective,
               k.assessmentHint AS assessment_hint,
               k.lex_role AS lex_role,
               CASE
                   WHEN 'Terminology' IN labels(k) THEN 'Terminology'
                   WHEN 'Skill' IN labels(k) THEN 'Skill'
                   ELSE coalesce(k.nodeType, 'KnowledgePoint')
               END AS nodeType,
               k.isSystemRecommended AS isSystemRecommended,
               k.isArchived AS isArchived,
               count(DISTINCT p) AS practiceCount,
               count(DISTINCT l) AS lessonCount,
                collect(DISTINCT prereq.name) AS prerequisites,
                collect(DISTINCT related.name) AS relations
        ORDER BY name
    """

    return _execute_read(query, params)


def get_knowledge_point(name: str) -> Dict[str, object]:
    """获取单个知识点的详细信息。"""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint {name: $name})
        OPTIONAL MATCH (t:Topic)-[:INCLUDE_POINT]->(k)
        OPTIONAL MATCH (k)<-[:TESTS]-(p:Practice)
        OPTIONAL MATCH (k)<-[rel]-(l:TheoryLesson)
        WHERE rel IS NULL OR type(rel) = 'EXPLAINS'
        OPTIONAL MATCH (k)-[:REQUIRES]->(prereq:KnowledgePoint)
        OPTIONAL MATCH (k)-[r:RELATED_TO|SUGGESTS_CO_LEARNING|CONTRASTS_WITH|APPLIES_TO_SCENARIO|SUGGESTS_STRATEGY|HAS_EXCEPTION|CULTURE_SENSITIVE_TO|COMBINES_WITH|CONFLICTS_WITH]-(related:KnowledgePoint)
        OPTIONAL MATCH (k)-[:HAS_CULTURAL_SENSITIVITY|INVOLVES_CULTURE]->(culture:CultureDimension)
        RETURN k.name AS name,
               k.description AS description,
               k.category AS category,
               k.categoryPath AS category_path,
               k.difficulty AS difficulty,
               k.importance AS importance,
               k.estimatedDuration AS estimated_duration,
               k.content AS content,
               k.orderIndex AS order_index,
               k.tags AS tags,
               k.bloomLevel AS bloom_level,
               k.cultureTags AS culture_tags,
               k.civicTags AS civic_tags,
               k.teachingObjective AS teaching_objective,
               k.assessmentHint AS assessment_hint,
               t.name AS topic,
               CASE
                   WHEN 'Terminology' IN labels(k) THEN 'Terminology'
                   WHEN 'Skill' IN labels(k) THEN 'Skill'
                   ELSE coalesce(k.nodeType, 'KnowledgePoint')
               END AS nodeType,
               collect(DISTINCT p.id) AS practices,
               collect(DISTINCT l.id) AS lessons,
               collect(DISTINCT prereq.name) AS prerequisites,
               collect(DISTINCT related.name) AS relations,
               collect(DISTINCT culture.name) AS culture_dimensions
        """,
        {"name": name},
    )

    if not records:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    record = records[0]
    return {
        "name": record.get("name"),
        "description": record.get("description"),
        "category": record.get("category"),
        "category_path": [segment for segment in (record.get("category_path") or []) if segment],
        "difficulty": record.get("difficulty"),
        "importance": record.get("importance"),
        "estimated_duration": record.get("estimated_duration"),
        "content": record.get("content"),
        "order_index": record.get("order_index"),
        "tags": [tag for tag in (record.get("tags") or []) if tag],
        "bloom_level": record.get("bloom_level"),
        "culture_tags": [tag for tag in (record.get("culture_tags") or []) if tag],
        "civic_tags": [tag for tag in (record.get("civic_tags") or []) if tag],
        "teaching_objective": record.get("teaching_objective"),
        "assessment_hint": record.get("assessment_hint"),
        "nodeType": record.get("nodeType") or "KnowledgePoint",
        "topic": record.get("topic"),
        "practices": [p for p in (record.get("practices") or []) if p],
        "lessons": [l for l in (record.get("lessons") or []) if l],
        "prerequisites": [p for p in (record.get("prerequisites") or []) if p],
        "relations": [r for r in (record.get("relations") or []) if r],
        "culture_dimensions": [c for c in (record.get("culture_dimensions") or []) if c],
    }


def _normalize_string_list(value: object) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [item.strip() for item in value.split(",") if item.strip()]
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _normalize_bloom_level(value: object, fallback: Optional[str] = None) -> Optional[str]:
    raw = str(value or fallback or "").strip()
    if not raw:
        return None
    aliases = {
        "识记": "remember",
        "记忆": "remember",
        "remember": "remember",
        "理解": "understand",
        "understand": "understand",
        "应用": "apply",
        "apply": "apply",
        "分析": "analyze",
        "analyse": "analyze",
        "analyze": "analyze",
        "评价": "evaluate",
        "评估": "evaluate",
        "evaluate": "evaluate",
        "创造": "create",
        "create": "create",
    }
    return aliases.get(raw.lower(), aliases.get(raw, raw if raw in BLOOM_LEVELS else fallback))


def create_knowledge_point(data: Dict[str, object]) -> Dict[str, object]:
    """创建新的知识点。"""

    name = data.get("name", "").strip()
    if not name:
        raise ValueError("Knowledge point name is required")

    # 检查是否已存在
    existing = _execute_read(
        "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
        {"name": name},
    )
    if existing:
        raise ValueError(f"Knowledge point '{name}' already exists")

    category_path = _normalize_category_path_input(
        data.get("category_path") or data.get("categoryPath") or data.get("category")
    )
    category_value = data.get("category") or (category_path[-1] if category_path else None)
    try:
        order_index = int(data.get("order_index") or data.get("orderIndex") or 0)
    except (ValueError, TypeError):
        order_index = 0

    # 创建节点
    _execute_write(
        """
        CREATE (k:KnowledgePoint {
            name: $name,
            description: $description,
            category: $category,
            categoryPath: $category_path,
            difficulty: $difficulty,
            importance: $importance,
            estimatedDuration: $estimated_duration,
            orderIndex: $order_index,
            content: $content,
            tags: $tags,
            bloomLevel: $bloom_level,
            cultureTags: $culture_tags,
            civicTags: $civic_tags,
            teachingObjective: $teaching_objective,
            assessmentHint: $assessment_hint
        })
        """,
        {
            "name": name,
            "description": data.get("description"),
            "category": category_value,
            "category_path": category_path,
            "difficulty": data.get("difficulty", "beginner"),
            "importance": data.get("importance", "medium"),
            "estimated_duration": data.get("estimated_duration"),
            "order_index": order_index,
            "content": data.get("content"),
            "tags": _normalize_string_list(data.get("tags", [])),
            "bloom_level": _normalize_bloom_level(data.get("bloom_level") or data.get("bloomLevel")),
            "culture_tags": _normalize_string_list(data.get("culture_tags") or data.get("cultureTags")),
            "civic_tags": _normalize_string_list(data.get("civic_tags") or data.get("civicTags")),
            "teaching_objective": data.get("teaching_objective") or data.get("teachingObjective"),
            "assessment_hint": data.get("assessment_hint") or data.get("assessmentHint"),
        },
    )

    return get_knowledge_point(name)


def update_knowledge_point(name: str, data: Dict[str, object]) -> Dict[str, object]:
    """更新知识点信息，支持部分字段更新。"""

    # 检查是否存在
    existing = _execute_read(
        "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
        {"name": name},
    )
    if not existing:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    node = existing[0]["k"]
    new_name = (data.get("name") or name).strip()

    # 如果改名，检查新名称是否冲突
    if new_name != name:
        conflict = _execute_read(
            "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
            {"name": new_name},
        )
        if conflict:
            raise ValueError(f"Knowledge point '{new_name}' already exists")

    category_field_provided = any(
        key in data for key in ("category_path", "categoryPath", "category")
    )
    if category_field_provided:
        category_path = _normalize_category_path_input(
            data.get("category_path") or data.get("categoryPath") or data.get("category")
        )
        category_value = data.get("category") or (category_path[-1] if category_path else None)
    else:
        category_path = [segment for segment in (node.get("categoryPath") or []) if segment]
        category_value = node.get("category")

    if "order_index" in data or "orderIndex" in data:
        try:
            order_index = int(data.get("order_index") or data.get("orderIndex") or 0)
        except (ValueError, TypeError):
            order_index = node.get("orderIndex") or 0
    else:
        order_index = node.get("orderIndex") or 0

    description = data.get("description") if "description" in data else node.get("description")
    difficulty = data.get("difficulty") if "difficulty" in data else node.get("difficulty")
    if difficulty is None:
        difficulty = "beginner"

    importance = data.get("importance") if "importance" in data else node.get("importance")
    if importance is None:
        importance = "medium"

    estimated_duration = (
        data.get("estimated_duration") if "estimated_duration" in data else node.get("estimatedDuration")
    )
    content = data.get("content") if "content" in data else node.get("content")

    if "tags" in data:
        tags_field = data.get("tags")
        if isinstance(tags_field, str):
            tags = [tag.strip() for tag in tags_field.split(",") if tag.strip()]
        elif isinstance(tags_field, (list, tuple)):
            tags = [str(tag).strip() for tag in tags_field if str(tag).strip()]
        else:
            tags = []
    else:
        tags = [tag for tag in (node.get("tags") or []) if tag]

    bloom_level = _normalize_bloom_level(
        data.get("bloom_level") if "bloom_level" in data else data.get("bloomLevel") if "bloomLevel" in data else None,
        node.get("bloomLevel"),
    )
    culture_tags = _normalize_string_list(
        data.get("culture_tags") if "culture_tags" in data else data.get("cultureTags") if "cultureTags" in data else node.get("cultureTags")
    )
    civic_tags = _normalize_string_list(
        data.get("civic_tags") if "civic_tags" in data else data.get("civicTags") if "civicTags" in data else node.get("civicTags")
    )
    teaching_objective = (
        data.get("teaching_objective") if "teaching_objective" in data else data.get("teachingObjective") if "teachingObjective" in data else node.get("teachingObjective")
    )
    assessment_hint = (
        data.get("assessment_hint") if "assessment_hint" in data else data.get("assessmentHint") if "assessmentHint" in data else node.get("assessmentHint")
    )

    # 更新节点
    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $old_name})
        SET k.name = $name,
            k.description = $description,
            k.category = $category,
            k.categoryPath = $category_path,
            k.difficulty = $difficulty,
            k.importance = $importance,
            k.estimatedDuration = $estimated_duration,
            k.orderIndex = $order_index,
            k.content = $content,
            k.tags = $tags,
            k.bloomLevel = $bloom_level,
            k.cultureTags = $culture_tags,
            k.civicTags = $civic_tags,
            k.teachingObjective = $teaching_objective,
            k.assessmentHint = $assessment_hint
        """,
        {
            "old_name": name,
            "name": new_name,
            "description": description,
            "category": category_value,
            "category_path": category_path,
            "difficulty": difficulty,
            "importance": importance,
            "estimated_duration": estimated_duration,
            "order_index": order_index,
            "content": content,
            "tags": tags,
            "bloom_level": bloom_level,
            "culture_tags": culture_tags,
            "civic_tags": civic_tags,
            "teaching_objective": teaching_objective,
            "assessment_hint": assessment_hint,
        },
    )

    return get_knowledge_point(new_name)


def update_knowledge_point_category(
    name: str,
    category_path_value: Optional[object],
    *,
    order_index: Optional[object] = None,
) -> Dict[str, object]:
    """快速更新知识点的分类与排序信息。"""

    # 确保知识点存在
    existing = _execute_read("MATCH (k:KnowledgePoint {name: $name}) RETURN k", {"name": name})
    if not existing:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    category_path = _normalize_category_path_input(category_path_value)
    category_value = category_path[-1] if category_path else None
    resolved_order: Optional[int]
    if order_index is None:
        resolved_order = None
    else:
        try:
            resolved_order = int(order_index)
        except (TypeError, ValueError):
            resolved_order = None

    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})
        SET k.category = $category,
            k.categoryPath = $category_path,
            k.orderIndex = COALESCE($order_index, k.orderIndex)
        """,
        {
            "name": name,
            "category": category_value,
            "category_path": category_path,
            "order_index": resolved_order,
        },
    )

    return get_knowledge_point(name)


def get_knowledge_management_overview() -> Dict[str, object]:
    """聚合知识点、分类树和知识卡索引，供前端统一加载。"""

    status = get_initialization_status()
    # 直接使用 Neo4j 为主数据源
    raw_points = list_knowledge_points_enhanced()
    card_by_name = {point.get("name"): point for point in raw_points if point.get("name")}

    # 如果已有数据，则认为图谱已完成初始化；否则返回默认预设
    if raw_points:
        if not status.get("initialized"):
            status["initialized"] = True
            status["option"] = status.get("option") or "import"
    else:
        return {
            "initialized": False,
            "initialization": status,
            "defaults": get_initialization_defaults_preview(),
        }

    overview_points: List[Dict[str, object]] = []
    category_paths: Dict[str, None] = {}
    difficulty_breakdown: Dict[str, int] = {}
    uncategorized_points: List[Dict[str, object]] = []
    metadata_suggestions: List[Dict[str, object]] = []
    uncategorized_count = 0
    unlinked_count = 0

    tree_root: Dict[str, object] = {
        "id": "__root__",
        "name": "全部",
        "path": [],
        "count": 0,
        "children": {},
        "knowledge": [],
    }

    def _ensure_child(parent: Dict[str, object], segment: str, path: Sequence[str]) -> Dict[str, object]:
        children = parent.setdefault("children", {})
        if segment not in children:
            key = _category_path_to_string(path)
            children[segment] = {
                "id": key or "未分类",
                "name": segment,
                "path": list(path),
                "count": 0,
                "children": {},
                "knowledge": [],
                "_order": len(children),
            }
        return children[segment]

    def _append_to_tree(path: Sequence[str], payload: Dict[str, object]) -> None:
        effective_path = list(path) if path else ["未分类"]
        current = tree_root
        current["count"] = current.get("count", 0) + 1
        for index, segment in enumerate(effective_path, start=1):
            sub_path = effective_path[:index]
            current = _ensure_child(current, segment, sub_path)
            current["count"] = current.get("count", 0) + 1
        current.setdefault("knowledge", []).append(payload)

    for point in raw_points:
        name = point.get("name")
        if not name:
            continue
        raw_category_path = point.get("category_path")
        category_path = _normalize_category_path_input(
            raw_category_path if raw_category_path is not None else point.get("category")
        )
        category_path_key = _category_path_to_string(category_path) or "未分类"
        category_paths[category_path_key] = None

        order_index = point.get("order_index")
        try:
            order_index_value = int(order_index) if order_index is not None else 0
        except (TypeError, ValueError):
            order_index_value = 0

        difficulty = point.get("difficulty") or "beginner"
        difficulty_breakdown[difficulty] = difficulty_breakdown.get(difficulty, 0) + 1

        tags = [tag for tag in (point.get("tags") or []) if tag]
        prerequisites = sorted({p for p in (point.get("prerequisites") or []) if p})
        relations = sorted({r for r in (point.get("relations") or []) if r})

        category_label = " / ".join(category_path) if category_path else "未分类"

        if not category_path:
            uncategorized_count += 1
            uncategorized_points.append(
                {
                    "name": name,
                    "practiceCount": point.get("practiceCount", 0),
                    "lessonCount": point.get("lessonCount", 0),
                    "isSystemRecommended": bool(point.get("isSystemRecommended")),
                }
            )

        overview = {
            "name": name,
            "description": point.get("description") or "",
            "category": category_path[-1] if category_path else None,
            "category_path": category_path,
            "category_path_key": category_path_key,
            "category_path_text": category_label,
            "difficulty": difficulty,
            "importance": point.get("importance"),
            "estimated_duration": point.get("estimated_duration"),
            "content": point.get("content") or "",
            "order_index": order_index_value,
            "tags": tags,
            "bloom_level": point.get("bloom_level"),
            "culture_tags": [tag for tag in (point.get("culture_tags") or []) if tag],
            "civic_tags": [tag for tag in (point.get("civic_tags") or []) if tag],
            "teaching_objective": point.get("teaching_objective"),
            "assessment_hint": point.get("assessment_hint"),
            "prerequisites": prerequisites,
            "relations": relations,
            "practiceCount": point.get("practiceCount", 0),
            "lessonCount": point.get("lessonCount", 0),
            "isSystemRecommended": bool(point.get("isSystemRecommended")),
            "isArchived": bool(point.get("isArchived")),
            "nodeType": point.get("nodeType") or "KnowledgePoint",
        }
        overview_points.append(overview)

        if not overview["practiceCount"] and not overview["lessonCount"]:
            unlinked_count += 1

        tree_payload = {
            "name": name,
            "difficulty": difficulty,
            "order_index": order_index_value,
            "tags": tags,
            "bloom_level": point.get("bloom_level"),
            "culture_tags": [tag for tag in (point.get("culture_tags") or []) if tag],
            "category_path_key": category_path_key,
        }
        _append_to_tree(category_path, tree_payload)

        card = card_by_name.get(name) or {}
        metadata_fields: Dict[str, object] = {}
        metadata_preview: Dict[str, object] = {}
        summary = card.get("summary")
        if not overview.get("description") and summary:
            metadata_fields["description"] = summary
            metadata_preview["description"] = summary
        card_tags = [tag for tag in (card.get("tags") or []) if tag]
        if not tags and card_tags:
            metadata_fields["tags"] = card_tags
            metadata_preview["tags"] = card_tags
        if metadata_fields:
            metadata_suggestions.append(
                {
                    "name": name,
                    "fields": metadata_fields,
                    "preview": metadata_preview,
                    "reason": "根据已存在的理论知识卡片信息自动补全",
                }
            )

    def _serialize_tree(node: Dict[str, object]) -> Dict[str, object]:
        children_map = node.get("children") or {}
        serialized_children = []
        for _, child in sorted(children_map.items(), key=lambda item: item[1].get("_order", 0)):
            child_payload = {
                "id": child.get("id"),
                "name": child.get("name"),
                "path": child.get("path", []),
                "count": child.get("count", 0),
                "knowledge": sorted(
                    child.get("knowledge", []),
                    key=lambda payload: (payload.get("order_index", 0), payload.get("name", "")),
                ),
                "children": [],
            }
            child_payload["children"] = _serialize_tree(child).get("children", [])
            serialized_children.append(child_payload)
        return {"children": serialized_children}

    tree_children = _serialize_tree(tree_root).get("children", [])

    category_options = sorted({path for path in category_paths.keys() if path}) or []
    if "未分类" not in category_options:
        category_options.append("未分类")

    stats = {
        "total": len(overview_points),
        "categories": len(category_options),
        "difficulty": difficulty_breakdown,
        "uncategorized": uncategorized_count,
        "unlinked": unlinked_count,
    }

    knowledge_cards: List[Dict[str, object]] = []
    for name, card in card_by_name.items():
        if not name:
            continue
        card_category_path = _normalize_category_path_input(
            card.get("categoryPath") if card.get("categoryPath") is not None else card.get("category")
        )
        summary = card.get("summary") or card.get("description") or card.get("content") or ""
        body_html = card.get("bodyHtml") or card.get("content") or card.get("description") or ""
        knowledge_cards.append(
            {
                "name": name,
                "summary": summary,
                "bodyHtml": body_html,
                "content": card.get("content") or "",
                "imageUrl": card.get("imageUrl") or "",
                "imageAlt": card.get("imageAlt") or "",
                "knowledgeId": card.get("knowledgeId") or "",
                "tags": [tag for tag in (card.get("tags") or []) if tag],
                "practiceCount": card.get("practiceCount", 0),
                "lessonCount": card.get("lessonCount", 0),
                "category_path": card_category_path,
            }
        )

    knowledge_cards.sort(key=lambda item: item.get("name", ""))

    overview_points.sort(key=lambda item: (item.get("order_index", 0), item.get("name", "")))

    smart_assist = {
        "uncategorized": sorted(
            uncategorized_points,
            key=lambda item: (item.get("practiceCount", 0) + item.get("lessonCount", 0), item.get("name", "")),
            reverse=True,
        ),
        "metadata_suggestions": metadata_suggestions,
    }

    category_paths_list = sorted(category_paths.keys())
    tree_children = list(tree_root.get("children", {}).values())

    return {
        "initialized": True,
        "initialization": status,
        "knowledge_points": overview_points,
        "tree": tree_children,
        "category_tree": tree_children,  # backward compatibility for UI
        "category_paths": category_paths_list,
        "uncategorized": uncategorized_points,
        "categoryOptions": category_options,
        "metadata": metadata_suggestions,
        "stats": stats,
        "knowledge_cards": knowledge_cards,
        "assist": smart_assist,
    }

def get_customer_route_alignment_report() -> Dict[str, object]:
    """Return a P0 customer-route alignment checklist for teachers."""

    stage_names = [stage.name for stage in STANDARD_STAGES]
    stage_ids = [stage.identifier for stage in STANDARD_STAGES]
    rows = _execute_read(
        """
        MATCH (s:Stage)
        WHERE s.name IN $stage_names OR s.id IN $stage_ids
        OPTIONAL MATCH (s)-[:CONTAIN_TOPIC]->(:Topic)-[:INCLUDE_POINT]->(k:KnowledgePoint)
        OPTIONAL MATCH (s)-[:HAS_CULTURAL_SENSITIVITY]->(c:CultureDimension)
        RETURN collect(DISTINCT s.name) AS stages,
               count(DISTINCT k) AS stageKnowledgeCount,
               collect(DISTINCT c.id) AS stageCultureIds
        """,
        {"stage_names": stage_names, "stage_ids": stage_ids},
    )
    row = rows[0] if rows else {}
    present_stages = set(row.get("stages") or [])
    present_culture_ids = set(row.get("stageCultureIds") or [])

    stats_rows = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        RETURN count(k) AS total,
               sum(CASE WHEN k.bloomLevel IS NOT NULL AND k.bloomLevel <> '' THEN 1 ELSE 0 END) AS withBloom,
               sum(CASE WHEN k.cultureTags IS NOT NULL AND size(k.cultureTags) > 0 THEN 1 ELSE 0 END) AS withCultureTags,
               sum(CASE WHEN k.civicTags IS NOT NULL AND size(k.civicTags) > 0 THEN 1 ELSE 0 END) AS withCivicTags,
               sum(CASE WHEN EXISTS { MATCH (k)<-[:TESTS]-(:Practice) } THEN 1 ELSE 0 END) AS withPractice,
               sum(CASE WHEN EXISTS { MATCH (k)<-[:EXPLAINS]-(:TheoryLesson) } THEN 1 ELSE 0 END) AS withLesson,
               sum(CASE WHEN EXISTS { MATCH (k)-[:REQUIRES]->(:KnowledgePoint) } THEN 1 ELSE 0 END) AS withPrereq
        """
    )
    stats = stats_rows[0] if stats_rows else {}
    culture_count_rows = _execute_read("MATCH (c:CultureDimension) RETURN count(c) AS total", {})
    culture_total = (culture_count_rows[0] or {}).get("total", 0) if culture_count_rows else 0
    total_points = int(stats.get("total") or 0)
    checks = [
        {"id": "ten_stages", "label": "十大谈判环节", "current": len(present_stages), "target": len(STANDARD_STAGES), "passed": len(present_stages) >= len(STANDARD_STAGES)},
        {"id": "stage_knowledge", "label": "环节知识点覆盖", "current": row.get("stageKnowledgeCount", 0), "target": 30, "passed": row.get("stageKnowledgeCount", 0) >= 30},
        {"id": "culture_dimensions", "label": "文化维度节点", "current": culture_total, "target": len(CULTURE_DIMENSIONS), "passed": culture_total >= len(CULTURE_DIMENSIONS)},
        {"id": "stage_culture_links", "label": "环节-跨文化映射", "current": len(present_culture_ids), "target": min(6, len(CULTURE_DIMENSIONS)), "passed": len(present_culture_ids) >= min(6, len(CULTURE_DIMENSIONS))},
        {"id": "bloom", "label": "布鲁姆认知层级", "current": stats.get("withBloom", 0), "target": total_points, "passed": bool(total_points) and stats.get("withBloom", 0) >= total_points},
        {"id": "civic", "label": "思政标签", "current": stats.get("withCivicTags", 0), "target": max(1, int(total_points * 0.3)), "passed": stats.get("withCivicTags", 0) >= max(1, int(total_points * 0.3))},
        {"id": "prereq", "label": "前置依赖", "current": stats.get("withPrereq", 0), "target": 1, "passed": stats.get("withPrereq", 0) >= 1},
        {"id": "practice_lesson_links", "label": "课时/练习绑定", "current": (stats.get("withPractice", 0) or 0) + (stats.get("withLesson", 0) or 0), "target": 1, "passed": ((stats.get("withPractice", 0) or 0) + (stats.get("withLesson", 0) or 0)) >= 1},
    ]
    score = round(sum(1 for check in checks if check["passed"]) / len(checks) * 100) if checks else 0
    return {"score": score, "checks": checks, "missingStages": [name for name in stage_names if name not in present_stages], "updatedAt": datetime.utcnow().isoformat() + "Z"}


def list_culture_dimensions() -> List[Dict[str, object]]:
    """List culture dimensions stored in Neo4j."""

    return _execute_read(
        """
        MATCH (c:CultureDimension)
        RETURN c.id AS id, c.name AS name, c.theory AS theory, c.summary AS summary, c.teachingTip AS teaching_tip
        ORDER BY c.name
        """
    )


def _stable_graph_id(prefix: str, value: object) -> str:
    text = str(value or "").strip() or prefix
    digest = hashlib.md5(text.encode("utf-8")).hexdigest()[:10]
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", text.lower()).strip("-")[:32]
    return f"{prefix}-{slug or digest}"


def _normalize_graph_json_rel_type(value: object) -> str:
    raw = str(value or "").strip()
    upper = raw.upper()
    mapping = {
        "req": "REQUIRES",
        "prerequisite": "REQUIRES",
        "requires": "REQUIRES",
        "scn": "SUGGESTS_CO_LEARNING",
        "similar": "SUGGESTS_CO_LEARNING",
        "suggests_co_learning": "SUGGESTS_CO_LEARNING",
        "con": "CONTRASTS_WITH",
        "contrast": "CONTRASTS_WITH",
        "contrasts_with": "CONTRASTS_WITH",
        "cul": "HAS_CULTURAL_SENSITIVITY",
        "culture": "HAS_CULTURAL_SENSITIVITY",
        "has_cultural_sensitivity": "HAS_CULTURAL_SENSITIVITY",
        "involves_culture": "HAS_CULTURAL_SENSITIVITY",
        "culture_sensitive_to": "HAS_CULTURAL_SENSITIVITY",
        "exc": "HAS_EXCEPTION",
        "has_exception": "HAS_EXCEPTION",
        "mig": "COMBINES_WITH",
        "combines_with": "COMBINES_WITH",
    }
    normalized = mapping.get(raw.lower(), upper or "RELATED_TO")
    allowed = {
        "REQUIRES",
        "SUGGESTS_CO_LEARNING",
        "CONTRASTS_WITH",
        "HAS_CULTURAL_SENSITIVITY",
        "INVOLVES_CULTURE",
        "HAS_EXCEPTION",
        "COMBINES_WITH",
        "RELATED_TO",
        "RELATES_TO",
        "APPLIES_TO_SCENARIO",
        "SUGGESTS_STRATEGY",
        "CONFLICTS_WITH",
        "CULTURE_SENSITIVE_TO",
    }
    return normalized if normalized in allowed else "RELATED_TO"


def _json_kp_type(node: Dict[str, object]) -> str:
    labels = node.get("labels") or []
    raw = str(node.get("nodeType") or node.get("type") or "").lower()
    if "Terminology" in labels or "terminology" in raw or "term" in raw or "术语" in raw:
        return "terminology"
    if "Skill" in labels or "skill" in raw or "技能" in raw:
        return "skill"
    return "knowledge"


def export_knowledge_graph_json() -> Dict[str, object]:
    """Export the maintainable teacher-facing graph model as JSON."""

    stage_rows = _execute_read(
        """
        MATCH (s:Stage)
        WITH properties(s) AS p
        RETURN p.id AS id, p.name AS name, p.englishName AS englishName,
               p.description AS description, p.orderIndex AS orderIndex,
               p.orderIndex AS order
        ORDER BY coalesce(p.orderIndex, 999), p.name
        """
    )
    topic_rows = _execute_read(
        """
        MATCH (s:Stage)-[:CONTAIN_TOPIC]->(t:Topic)
        WITH properties(s) AS ps, properties(t) AS pt
        RETURN pt.id AS id, pt.name AS name, pt.layer AS layer, pt.orderIndex AS orderIndex,
               ps.id AS stageId, ps.name AS stageName,
               coalesce(ps.orderIndex, 999) AS stageOrder
        ORDER BY stageOrder, coalesce(orderIndex, 999), name
        """
    )
    kp_rows = _execute_read(
        """
        MATCH (s:Stage)-[:CONTAIN_TOPIC]->(t:Topic)-[:HAS_CATEGORY]->(:KnowledgeCategory)-[:CONTAINS]->(k:KnowledgePoint)
        WITH DISTINCT properties(s) AS ps, properties(t) AS pt, properties(k) AS pk, labels(k) AS labels
        RETURN pk.id AS id, pk.name AS name, pk.description AS description,
               pk.summary AS summary, pk.content AS content, pk.category AS category,
               pk.type AS type, pk.nodeType AS nodeType, labels AS labels,
               pk.difficulty AS difficulty, pk.importance AS importance,
               pk.bloomLevel AS bloomLevel, pk.bloom_level AS bloom_level,
               pk.keywords AS keywords, pk.cultureTags AS cultureTags,
               pt.id AS topicId, pt.name AS topicName, pt.layer AS topicLayer,
               ps.id AS stageId, ps.name AS stageName,
               coalesce(ps.orderIndex, 999) AS stageOrder,
               coalesce(pt.orderIndex, 999) AS topicOrder
        ORDER BY stageOrder, topicOrder, topicName, name
        """
    )
    culture_rows = _execute_read(
        """
        MATCH (c:CultureDimension)
        WITH properties(c) AS p
        RETURN p.id AS id, p.name AS name, p.theory AS theory,
               p.summary AS summary, p.teachingTip AS teachingTip
        ORDER BY p.name
        """
    )
    semantic_relation_types = [
        "REQUIRES",
        "SUGGESTS_CO_LEARNING",
        "RELATED_TO",
        "RELATES_TO",
        "CONTRASTS_WITH",
        "APPLIES_TO_SCENARIO",
        "SUGGESTS_STRATEGY",
        "HAS_EXCEPTION",
        "COMBINES_WITH",
        "CONFLICTS_WITH",
        "CULTURE_SENSITIVE_TO",
        "HAS_CULTURAL_SENSITIVITY",
        "INVOLVES_CULTURE",
    ]
    rel_rows = _execute_read(
        """
        MATCH (a)-[r]->(b)
        WHERE type(r) IN $relationTypes
          AND ((a:KnowledgePoint AND b:KnowledgePoint) OR (a:KnowledgePoint AND b:CultureDimension))
        WITH labels(a) AS sourceLabels, properties(a) AS source,
             labels(b) AS targetLabels, properties(b) AS target,
             type(r) AS type, properties(r) AS rel
        RETURN sourceLabels, source.id AS sourceId, source.name AS sourceName,
               targetLabels, target.id AS targetId, target.name AS targetName,
               type, rel.relationType AS relationType
        ORDER BY sourceName, type, targetName
        """,
        {"relationTypes": semantic_relation_types},
    )

    stages = [
        {
            "id": row.get("id") or _stable_graph_id("stage", row.get("name")),
            "name": row.get("name"),
            "zh": row.get("name"),
            "en": row.get("englishName") or "",
            "order": row.get("orderIndex") or row.get("order") or index + 1,
            "description": row.get("description") or "",
        }
        for index, row in enumerate(stage_rows)
    ]
    stage_id_by_name = {stage["name"]: stage["id"] for stage in stages if stage.get("name")}

    topics = [
        {
            "id": row.get("id") or _stable_graph_id("topic", f"{row.get('stageName')}:{row.get('name')}"),
            "name": row.get("name"),
            "stage": row.get("stageId") or stage_id_by_name.get(row.get("stageName")),
            "layer": row.get("layer") or "concept",
            "order": row.get("orderIndex"),
        }
        for row in topic_rows
    ]
    topic_id_by_name_stage = {
        (topic.get("stage"), topic.get("name")): topic.get("id")
        for topic in topics
    }

    knowledge_points = []
    for row in kp_rows:
        topic_id = row.get("topicId") or topic_id_by_name_stage.get((row.get("stageId"), row.get("topicName")))
        knowledge_points.append({
            "id": row.get("id") or _stable_graph_id("kp", row.get("name")),
            "name": row.get("name"),
            "stage": row.get("stageId") or stage_id_by_name.get(row.get("stageName")),
            "topic": topic_id,
            "topicName": row.get("topicName"),
            "layer": row.get("topicLayer") or row.get("category") or "concept",
            "kpType": _json_kp_type(row),
            "nodeType": row.get("nodeType") or ("Terminology" if _json_kp_type(row) == "terminology" else "Skill" if _json_kp_type(row) == "skill" else "KnowledgePoint"),
            "difficulty": row.get("difficulty"),
            "importance": row.get("importance"),
            "bloom": row.get("bloomLevel") or row.get("bloom_level"),
            "summary": row.get("summary") or "",
            "description": row.get("description") or row.get("content") or "",
            "keywords": row.get("keywords") or [],
            "cultureTags": row.get("cultureTags") or [],
        })

    culture_dimensions = [
        {
            "id": row.get("id") or _stable_graph_id("culture", row.get("name")),
            "name": row.get("name"),
            "theory": row.get("theory") or "",
            "summary": row.get("summary") or "",
            "teachingTip": row.get("teachingTip") or "",
        }
        for row in culture_rows
    ]

    relations = [
        {
            "source": row.get("sourceId") or row.get("sourceName"),
            "sourceName": row.get("sourceName"),
            "target": row.get("targetId") or row.get("targetName"),
            "targetName": row.get("targetName"),
            "type": row.get("type"),
            "relationType": row.get("relationType") or row.get("type"),
        }
        for row in rel_rows
    ]

    return {
        "schema": "foreign-trade-knowledge-graph/v1",
        "exportedAt": datetime.utcnow().isoformat() + "Z",
        "stages": stages,
        "topics": topics,
        "knowledgePoints": knowledge_points,
        "cultureDimensions": culture_dimensions,
        "relations": relations,
    }


def import_knowledge_graph_json(payload: Dict[str, object], *, mode: str = "merge", created_by: str = "json-import") -> Dict[str, object]:
    """Import a graph JSON exported by export_knowledge_graph_json."""

    if not isinstance(payload, dict):
        raise ValueError("JSON 顶层必须是对象")

    stages = payload.get("stages") or []
    topics = payload.get("topics") or []
    knowledge_points = payload.get("knowledgePoints") or payload.get("kps") or []
    culture_dimensions = payload.get("cultureDimensions") or payload.get("culture") or []
    relations = payload.get("relations") or payload.get("rels") or []
    if not isinstance(stages, list) or not isinstance(topics, list) or not isinstance(knowledge_points, list):
        raise ValueError("JSON 必须包含数组字段 stages/topics/knowledgePoints")

    mode = (mode or "merge").lower()
    if mode not in {"merge", "replace"}:
        raise ValueError("mode 必须是 merge 或 replace")

    driver = _get_driver()
    counters = {"stages": 0, "topics": 0, "knowledgePoints": 0, "cultureDimensions": 0, "relations": 0}

    with driver.session() as session:
        if mode == "replace":
            session.run(
                """
                MATCH (n)
                WHERE any(label IN labels(n) WHERE label IN ['Stage', 'Topic', 'KnowledgeCategory', 'KnowledgePoint', 'CultureDimension'])
                DETACH DELETE n
                """
            )

        stage_ref: Dict[str, Dict[str, object]] = {}
        for index, stage in enumerate(stages):
            if not isinstance(stage, dict):
                continue
            name = (stage.get("name") or stage.get("zh") or "").strip()
            if not name:
                continue
            stage_id = stage.get("id") or _stable_graph_id("stage", name)
            order = stage.get("order") or stage.get("orderIndex") or index + 1
            session.run(
                """
                MERGE (s:Stage {id: $id})
                ON CREATE SET s.createdAt = datetime(), s.createdBy = $createdBy
                SET s.name = $name,
                    s.englishName = $englishName,
                    s.description = $description,
                    s.orderIndex = $order,
                    s.updatedAt = datetime(),
                    s.updatedBy = $createdBy
                """,
                {
                    "id": stage_id,
                    "name": name,
                    "englishName": stage.get("en") or stage.get("englishName") or "",
                    "description": stage.get("description") or stage.get("summary") or "",
                    "order": int(order) if str(order).isdigit() else index + 1,
                    "createdBy": created_by,
                },
            )
            stage_ref[str(stage.get("id") or stage_id)] = {"id": stage_id, "name": name}
            stage_ref[name] = {"id": stage_id, "name": name}
            counters["stages"] += 1

        topic_ref: Dict[str, Dict[str, object]] = {}
        for index, topic in enumerate(topics):
            if not isinstance(topic, dict):
                continue
            name = (topic.get("name") or "").strip()
            if not name:
                continue
            stage_key = topic.get("stage") or topic.get("stageId") or topic.get("stageName")
            stage = stage_ref.get(str(stage_key)) or stage_ref.get(str(topic.get("stageName") or ""))
            if not stage:
                continue
            topic_id = topic.get("id") or _stable_graph_id("topic", f"{stage['name']}:{name}")
            layer = topic.get("layer") or "concept"
            session.run(
                """
                MATCH (s:Stage {id: $stageId})
                MERGE (t:Topic {id: $id})
                ON CREATE SET t.createdAt = datetime(), t.createdBy = $createdBy
                SET t.name = $name,
                    t.layer = $layer,
                    t.orderIndex = $order,
                    t.updatedAt = datetime(),
                    t.updatedBy = $createdBy
                MERGE (s)-[:CONTAIN_TOPIC]->(t)
                """,
                {
                    "stageId": stage["id"],
                    "id": topic_id,
                    "name": name,
                    "layer": layer,
                    "order": topic.get("order") or topic.get("orderIndex") or index + 1,
                    "createdBy": created_by,
                },
            )
            topic_ref[str(topic.get("id") or topic_id)] = {"id": topic_id, "name": name, "stage": stage, "layer": layer}
            topic_ref[f"{stage['id']}:{name}"] = {"id": topic_id, "name": name, "stage": stage, "layer": layer}
            counters["topics"] += 1

        kp_ref: Dict[str, Dict[str, object]] = {}
        for kp in knowledge_points:
            if not isinstance(kp, dict):
                continue
            name = (kp.get("name") or "").strip()
            if not name:
                continue
            topic_key = kp.get("topic") or kp.get("topicId")
            stage_key = kp.get("stage") or kp.get("stageId") or kp.get("stageName")
            topic = topic_ref.get(str(topic_key))
            if not topic and kp.get("topicName"):
                stage = stage_ref.get(str(stage_key)) or stage_ref.get(str(kp.get("stageName") or ""))
                if stage:
                    topic = topic_ref.get(f"{stage['id']}:{kp.get('topicName')}")
            if not topic:
                continue
            kp_id = kp.get("id") or _stable_graph_id("kp", name)
            kp_type = (kp.get("kpType") or kp.get("type") or "").strip().lower()
            if kp_type in {"terminology", "term", "术语"}:
                node_type = "Terminology"
                category = "术语"
            elif kp_type in {"skill", "技能"}:
                node_type = "Skill"
                category = "技能"
            else:
                node_type = "KnowledgePoint"
                category = "知识"
            session.run(
                """
                MATCH (t:Topic {id: $topicId})
                MERGE (c:KnowledgeCategory {name: $category, topic: $topicName, stage: $stageName})
                MERGE (t)-[:HAS_CATEGORY]->(c)
                MERGE (k:KnowledgePoint {name: $name})
                ON CREATE SET k.createdAt = datetime(), k.createdBy = $createdBy
                SET k.id = $id,
                    k.description = $description,
                    k.summary = $summary,
                    k.type = $kpType,
                    k.nodeType = $nodeType,
                    k.category = $category,
                    k.difficulty = $difficulty,
                    k.importance = $importance,
                    k.bloomLevel = $bloom,
                    k.keywords = $keywords,
                    k.cultureTags = $cultureTags,
                    k.updatedAt = datetime(),
                    k.updatedBy = $createdBy
                MERGE (c)-[:CONTAINS]->(k)
                MERGE (t)-[:INCLUDE_POINT]->(k)
                """,
                {
                    "topicId": topic["id"],
                    "topicName": topic["name"],
                    "stageName": topic["stage"]["name"],
                    "category": category,
                    "name": name,
                    "id": kp_id,
                    "description": kp.get("description") or "",
                    "summary": kp.get("summary") or "",
                    "kpType": kp_type or "knowledge",
                    "nodeType": node_type,
                    "difficulty": kp.get("difficulty"),
                    "importance": kp.get("importance"),
                    "bloom": kp.get("bloom") or kp.get("bloomLevel"),
                    "keywords": kp.get("keywords") if isinstance(kp.get("keywords"), list) else [],
                    "cultureTags": kp.get("cultureTags") if isinstance(kp.get("cultureTags"), list) else [],
                    "createdBy": created_by,
                },
            )
            if node_type == "Terminology":
                session.run("MATCH (k:KnowledgePoint {name: $name}) SET k:Terminology", {"name": name})
            elif node_type == "Skill":
                session.run("MATCH (k:KnowledgePoint {name: $name}) SET k:Skill", {"name": name})
            kp_ref[str(kp.get("id") or kp_id)] = {"id": kp_id, "name": name}
            kp_ref[name] = {"id": kp_id, "name": name}
            counters["knowledgePoints"] += 1

        culture_ref: Dict[str, Dict[str, object]] = {}
        for culture in culture_dimensions:
            if not isinstance(culture, dict):
                continue
            name = (culture.get("name") or "").strip()
            if not name:
                continue
            culture_id = culture.get("id") or _stable_graph_id("culture", name)
            session.run(
                """
                MERGE (c:CultureDimension {id: $id})
                ON CREATE SET c.createdAt = datetime(), c.createdBy = $createdBy
                SET c.name = $name,
                    c.theory = $theory,
                    c.summary = $summary,
                    c.teachingTip = $teachingTip,
                    c.updatedAt = datetime(),
                    c.updatedBy = $createdBy
                """,
                {
                    "id": culture_id,
                    "name": name,
                    "theory": culture.get("theory") or "",
                    "summary": culture.get("summary") or "",
                    "teachingTip": culture.get("teachingTip") or culture.get("teaching_tip") or "",
                    "createdBy": created_by,
                },
            )
            culture_ref[str(culture.get("id") or culture_id)] = {"id": culture_id, "name": name}
            culture_ref[name] = {"id": culture_id, "name": name}
            counters["cultureDimensions"] += 1

        for relation in relations:
            if not isinstance(relation, dict):
                continue
            rel_type = _normalize_graph_json_rel_type(relation.get("type") or relation.get("r") or relation.get("relationType"))
            source = kp_ref.get(str(relation.get("source") or relation.get("s"))) or kp_ref.get(str(relation.get("sourceName") or ""))
            target_key = relation.get("target") or relation.get("t")
            target = kp_ref.get(str(target_key)) or kp_ref.get(str(relation.get("targetName") or ""))
            target_is_culture = False
            if not target:
                target = culture_ref.get(str(target_key)) or culture_ref.get(str(relation.get("targetName") or ""))
                target_is_culture = bool(target)
            if not source or not target:
                continue
            if rel_type in {"HAS_CULTURAL_SENSITIVITY", "INVOLVES_CULTURE"} or target_is_culture:
                session.run(
                    """
                    MATCH (a:KnowledgePoint {name: $sourceName})
                    MATCH (b:CultureDimension {id: $targetId})
                    MERGE (a)-[r:HAS_CULTURAL_SENSITIVITY]->(b)
                    SET r.relationType = 'HAS_CULTURAL_SENSITIVITY',
                        r.updatedAt = datetime(),
                        r.updatedBy = $createdBy
                    """,
                    {"sourceName": source["name"], "targetId": target["id"], "createdBy": created_by},
                )
            else:
                session.run(
                    f"""
                    MATCH (a:KnowledgePoint {{name: $sourceName}})
                    MATCH (b:KnowledgePoint {{name: $targetName}})
                    MERGE (a)-[r:{rel_type}]->(b)
                    SET r.relationType = $relType,
                        r.updatedAt = datetime(),
                        r.updatedBy = $createdBy
                    """,
                    {"sourceName": source["name"], "targetName": target["name"], "relType": rel_type, "createdBy": created_by},
                )
            counters["relations"] += 1

    return {"success": True, "mode": mode, "statistics": counters}

def delete_knowledge_point(name: str) -> None:
    """删除知识点及其所有关系。"""

    # 检查是否存在
    existing = _execute_read(
        "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
        {"name": name},
    )
    if not existing:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    # 删除节点及所有关系
    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})
        DETACH DELETE k
        """,
        {"name": name},
    )


def list_orphan_knowledge_points() -> List[Dict[str, object]]:
    """Return knowledge points that are not linked to practices or lessons."""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        WHERE NOT (k)<-[:TESTS]-(:Practice)
          AND NOT (k)<-[:EXPLAINS]-(:TheoryLesson)
        RETURN k.name AS name,
               k.description AS description,
               k.category AS category,
               k.updatedAt AS updatedAt,
               k.isSystemRecommended AS isSystemRecommended,
               k.source AS source,
               k.createdAt AS createdAt
        ORDER BY coalesce(k.updatedAt, k.createdAt) DESC, k.name
        """
    )
    return records


def cleanup_orphan_knowledge_points(
    names: Sequence[str], *, archive: bool = False
) -> Dict[str, object]:
    """Delete or archive orphaned knowledge points by name."""

    normalized = sorted({name for name in names if isinstance(name, str) and name.strip()})
    if not normalized:
        return {"affected": 0, "archived": archive}

    driver = _get_driver()
    with driver.session() as session:
        if archive:
            result = session.run(
                """
                MATCH (k:KnowledgePoint)
                WHERE k.name IN $names
                  AND NOT (k)<-[:TESTS]-(:Practice)
                  AND NOT (k)<-[:EXPLAINS]-(:TheoryLesson)
                SET k.isArchived = true,
                    k.archivedAt = datetime(),
                    k.updatedAt = datetime()
                RETURN count(k) AS count
                """,
                {"names": normalized},
            ).single()
        else:
            result = session.run(
                """
                MATCH (k:KnowledgePoint)
                WHERE k.name IN $names
                  AND NOT (k)<-[:TESTS]-(:Practice)
                  AND NOT (k)<-[:EXPLAINS]-(:TheoryLesson)
                DETACH DELETE k
                RETURN count(k) AS count
                """,
                {"names": normalized},
            ).single()
    affected = int(result["count"]) if result and result.get("count") is not None else 0
    return {"affected": affected, "archived": archive}


def add_knowledge_prerequisite(name: str, prerequisite_name: str) -> Dict[str, object]:
    """为知识点添加前置依赖关系。"""

    # 检查两个知识点是否存在
    for point_name in [name, prerequisite_name]:
        existing = _execute_read(
            "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
            {"name": point_name},
        )
        if not existing:
            raise GraphEntityNotFoundError(f"Knowledge point '{point_name}' not found")

    # 不能依赖自己
    if name == prerequisite_name:
        raise ValueError("A knowledge point cannot be a prerequisite of itself")

    # 创建REQUIRES关系
    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})
        MATCH (prereq:KnowledgePoint {name: $prerequisite})
        MERGE (k)-[:REQUIRES]->(prereq)
        """,
        {"name": name, "prerequisite": prerequisite_name},
    )

    return get_knowledge_point(name)


def remove_knowledge_prerequisite(name: str, prerequisite_name: str) -> Dict[str, object]:
    """移除知识点的前置依赖关系。"""

    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})-[r:REQUIRES]->(prereq:KnowledgePoint {name: $prerequisite})
        DELETE r
        """,
        {"name": name, "prerequisite": prerequisite_name},
    )

    return get_knowledge_point(name)


def add_knowledge_relation(
    name: str, related_name: str, relation_type: str = "RELATED_TO"
) -> Dict[str, object]:
    """为知识点添加关联关系。"""

    # 检查两个知识点是否存在
    for point_name in [name, related_name]:
        existing = _execute_read(
            "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
            {"name": point_name},
        )
        if not existing:
            raise GraphEntityNotFoundError(f"Knowledge point '{point_name}' not found")

    # 不能关联自己
    if name == related_name:
        raise ValueError("A knowledge point cannot be related to itself")

    normalized_type = (relation_type or "RELATED_TO").strip().upper()
    if normalized_type not in ALLOWED_KNOWLEDGE_RELATION_TYPES:
        normalized_type = "RELATED_TO"

    # 创建关联关系（双向）。Neo4j 不支持关系类型参数化，这里先通过白名单校验再插入。
    _execute_write(
        f"""
        MATCH (k1:KnowledgePoint {{name: $name}})
        MATCH (k2:KnowledgePoint {{name: $related}})
        MERGE (k1)-[r:{normalized_type}]-(k2)
        SET r.relationType = $relation_type, r.updatedAt = datetime()
        """,
        {"name": name, "related": related_name, "relation_type": normalized_type},
    )

    return get_knowledge_point(name)


def remove_knowledge_relation(name: str, related_name: str) -> Dict[str, object]:
    """移除知识点的关联关系。"""

    _execute_write(
        """
        MATCH (k1:KnowledgePoint {name: $name})-[r]-(k2:KnowledgePoint {name: $related})
        WHERE type(r) IN $allowed
        DELETE r
        """,
        {"name": name, "related": related_name, "allowed": list(ALLOWED_KNOWLEDGE_RELATION_TYPES)},
    )

    return get_knowledge_point(name)


def list_knowledge_categories() -> List[str]:
    """获取所有知识点分类列表。"""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        WHERE k.category IS NOT NULL AND k.category <> ''
        RETURN DISTINCT k.category AS category
        ORDER BY category
        """
    )

    return [record["category"] for record in records if record.get("category")]


def get_knowledge_categories_tree() -> List[Dict[str, object]]:
    """获取知识点分类树形结构（包含每个分类的知识点数量）。"""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        WHERE k.category IS NOT NULL AND k.category <> ''
        WITH k.category AS category, count(k) AS count
        RETURN category, count
        ORDER BY category
        """
    )

    return [
        {"name": record["category"], "count": record["count"]}
        for record in records
    ]


# ========== Excel/CSV 导入导出功能 ==========


def export_knowledge_points_to_excel() -> io.BytesIO:
    """将所有知识点导出为Excel文件。"""

    points = list_knowledge_points_enhanced()

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "知识点"

    # 写入表头
    headers = [
        "名称",
        "描述",
        "分类",
        "分类路径",
        "难度",
        "重要性",
        "预计学习时长(分钟)",
        "标签(逗号分隔)",
        "布鲁姆认知层级",
        "跨文化标签(逗号分隔)",
        "思政标签(逗号分隔)",
        "教学目标",
        "测评提示",
        "内容",
        "前置依赖(逗号分隔)",
        "关联知识点(逗号分隔)",
    ]
    ws.append(headers)

    # 写入数据
    for point in points:
        tags = ", ".join(point.get("tags") or [])
        prerequisites = ", ".join(point.get("prerequisites") or [])
        relations = ", ".join(point.get("relations") or [])

        row = [
            point.get("name", ""),
            point.get("description", ""),
            point.get("category", ""),
            " / ".join(point.get("category_path") or []) if point.get("category_path") else "",
            point.get("difficulty", ""),
            point.get("importance", ""),
            point.get("estimated_duration", ""),
            tags,
            point.get("bloom_level", ""),
            ", ".join(point.get("culture_tags") or []),
            ", ".join(point.get("civic_tags") or []),
            point.get("teaching_objective", ""),
            point.get("assessment_hint", ""),
            point.get("content", ""),
            prerequisites,
            relations,
        ]
        ws.append(row)

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_knowledge_points_to_csv() -> str:
    """将所有知识点导出为CSV字符串。"""

    points = list_knowledge_points_enhanced()

    output = io.StringIO()
    writer = csv.writer(output)

    # 写入表头
    headers = [
        "名称",
        "描述",
        "分类",
        "分类路径",
        "难度",
        "重要性",
        "预计学习时长(分钟)",
        "标签(逗号分隔)",
        "布鲁姆认知层级",
        "跨文化标签(逗号分隔)",
        "思政标签(逗号分隔)",
        "教学目标",
        "测评提示",
        "内容",
        "前置依赖(逗号分隔)",
        "关联知识点(逗号分隔)",
    ]
    writer.writerow(headers)

    # 写入数据
    for point in points:
        tags = ", ".join(point.get("tags") or [])
        prerequisites = ", ".join(point.get("prerequisites") or [])
        relations = ", ".join(point.get("relations") or [])

        row = [
            point.get("name", ""),
            point.get("description", ""),
            point.get("category", ""),
            " / ".join(point.get("category_path") or []) if point.get("category_path") else "",
            point.get("difficulty", ""),
            point.get("importance", ""),
            point.get("estimated_duration", ""),
            tags,
            point.get("bloom_level", ""),
            ", ".join(point.get("culture_tags") or []),
            ", ".join(point.get("civic_tags") or []),
            point.get("teaching_objective", ""),
            point.get("assessment_hint", ""),
            point.get("content", ""),
            prerequisites,
            relations,
        ]
        writer.writerow(row)

    return output.getvalue()


def import_knowledge_points_from_excel(file_content: bytes) -> Dict[str, int]:
    """
    从Excel文件导入知识点。
    返回导入统计：{"created": 数量, "updated": 数量, "failed": 数量, "errors": [错误列表]}
    """

    stats = {"created": 0, "updated": 0, "failed": 0, "errors": []}

    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        # 查找列索引
        col_indices = {}
        for idx, header in enumerate(headers):
            if header == "名称":
                col_indices["name"] = idx
            elif header == "描述":
                col_indices["description"] = idx
            elif header == "分类":
                col_indices["category"] = idx
            elif header == "分类路径":
                col_indices["category_path"] = idx
            elif header == "难度":
                col_indices["difficulty"] = idx
            elif header == "重要性":
                col_indices["importance"] = idx
            elif header == "预计学习时长(分钟)":
                col_indices["estimated_duration"] = idx
            elif header == "标签(逗号分隔)":
                col_indices["tags"] = idx
            elif header == "布鲁姆认知层级":
                col_indices["bloom_level"] = idx
            elif header == "跨文化标签(逗号分隔)":
                col_indices["culture_tags"] = idx
            elif header == "思政标签(逗号分隔)":
                col_indices["civic_tags"] = idx
            elif header == "教学目标":
                col_indices["teaching_objective"] = idx
            elif header == "测评提示":
                col_indices["assessment_hint"] = idx
            elif header == "内容":
                col_indices["content"] = idx
            elif header == "前置依赖(逗号分隔)":
                col_indices["prerequisites"] = idx
            elif header == "关联知识点(逗号分隔)":
                col_indices["relations"] = idx

        # 检查必需字段
        if "name" not in col_indices:
            stats["errors"].append("Excel文件缺少'名称'列")
            return stats

        # 处理每一行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = row[col_indices["name"]]
                if not name or not str(name).strip():
                    continue

                # 构建数据对象
                data = {
                    "name": str(name).strip(),
                    "description": str(row[col_indices.get("description")] or "").strip() or None,
                    "difficulty": str(row[col_indices.get("difficulty")] or "beginner").strip() or "beginner",
                    "importance": str(row[col_indices.get("importance")] or "medium").strip() or "medium",
                    "content": str(row[col_indices.get("content")] or "").strip() or None,
                }

                raw_category_value = row[col_indices.get("category")] if col_indices.get("category") is not None else None
                category_value = str(raw_category_value).strip() if raw_category_value else None
                raw_category_path = row[col_indices.get("category_path")] if col_indices.get("category_path") is not None else None
                category_path = _normalize_category_path_input(raw_category_path or category_value)
                if category_path:
                    data["category_path"] = category_path
                    data["category"] = category_value or category_path[-1]
                else:
                    data["category_path"] = []
                    data["category"] = category_value

                # 处理预计学习时长
                duration_value = row[col_indices.get("estimated_duration")]
                if duration_value:
                    try:
                        data["estimated_duration"] = int(duration_value)
                    except (ValueError, TypeError):
                        data["estimated_duration"] = None
                else:
                    data["estimated_duration"] = None

                def _cell(key):
                    idx = col_indices.get(key)
                    return row[idx] if idx is not None and idx < len(row) else None

                # 处理标签和P0教学属性
                tags_value = _cell("tags")
                data["tags"] = [tag.strip() for tag in str(tags_value).split(",") if tag.strip()] if tags_value else []
                data["bloom_level"] = str(_cell("bloom_level") or "").strip() or None
                data["culture_tags"] = _normalize_string_list(_cell("culture_tags"))
                data["civic_tags"] = _normalize_string_list(_cell("civic_tags"))
                data["teaching_objective"] = str(_cell("teaching_objective") or "").strip() or None
                data["assessment_hint"] = str(_cell("assessment_hint") or "").strip() or None

                # 检查是否已存在
                existing = _execute_read(
                    "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
                    {"name": data["name"]},
                )

                if existing:
                    # 更新
                    update_knowledge_point(data["name"], data)
                    stats["updated"] += 1
                else:
                    # 创建
                    create_knowledge_point(data)
                    stats["created"] += 1

                # 处理前置依赖
                prerequisites_value = _cell("prerequisites")
                if prerequisites_value:
                    prereqs = [p.strip() for p in str(prerequisites_value).split(",") if p.strip()]
                    for prereq in prereqs:
                        try:
                            add_knowledge_prerequisite(data["name"], prereq)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加前置依赖'{prereq}'失败: {str(e)}"
                            )

                # 处理关联关系
                relations_value = _cell("relations")
                if relations_value:
                    relations = [r.strip() for r in str(relations_value).split(",") if r.strip()]
                    for relation in relations:
                        try:
                            add_knowledge_relation(data["name"], relation)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加关联'{relation}'失败: {str(e)}"
                            )

            except Exception as e:
                stats["failed"] += 1
                stats["errors"].append(f"行{row_idx}: {str(e)}")

    except Exception as e:
        stats["errors"].append(f"读取Excel文件失败: {str(e)}")

    return stats


def import_knowledge_points_from_csv(file_content: str) -> Dict[str, int]:
    """
    从CSV文件导入知识点。
    返回导入统计：{"created": 数量, "updated": 数量, "failed": 数量, "errors": [错误列表]}
    """

    stats = {"created": 0, "updated": 0, "failed": 0, "errors": []}

    try:
        reader = csv.DictReader(io.StringIO(file_content))

        for row_idx, row in enumerate(reader, start=2):
            try:
                name = row.get("名称", "").strip()
                if not name:
                    continue

                # 构建数据对象
                data = {
                    "name": name,
                    "description": row.get("描述", "").strip() or None,
                    "difficulty": row.get("难度", "beginner").strip() or "beginner",
                    "importance": row.get("重要性", "medium").strip() or "medium",
                    "content": row.get("内容", "").strip() or None,
                }

                category_value = row.get("分类", "").strip() or None
                category_path = _normalize_category_path_input(row.get("分类路径", "").strip() or category_value)
                if category_path:
                    data["category_path"] = category_path
                    data["category"] = category_value or category_path[-1]
                else:
                    data["category_path"] = []
                    data["category"] = category_value

                # 处理预计学习时长
                duration_value = row.get("预计学习时长(分钟)", "").strip()
                if duration_value:
                    try:
                        data["estimated_duration"] = int(duration_value)
                    except (ValueError, TypeError):
                        data["estimated_duration"] = None
                else:
                    data["estimated_duration"] = None

                # 处理标签和P0教学属性
                tags_value = row.get("标签(逗号分隔)", "").strip()
                data["tags"] = [tag.strip() for tag in tags_value.split(",") if tag.strip()] if tags_value else []
                data["bloom_level"] = row.get("布鲁姆认知层级", "").strip() or None
                data["culture_tags"] = _normalize_string_list(row.get("跨文化标签(逗号分隔)", ""))
                data["civic_tags"] = _normalize_string_list(row.get("思政标签(逗号分隔)", ""))
                data["teaching_objective"] = row.get("教学目标", "").strip() or None
                data["assessment_hint"] = row.get("测评提示", "").strip() or None

                # 检查是否已存在
                existing = _execute_read(
                    "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
                    {"name": data["name"]},
                )

                if existing:
                    # 更新
                    update_knowledge_point(data["name"], data)
                    stats["updated"] += 1
                else:
                    # 创建
                    create_knowledge_point(data)
                    stats["created"] += 1

                # 处理前置依赖
                prerequisites_value = row.get("前置依赖(逗号分隔)", "").strip()
                if prerequisites_value:
                    prereqs = [p.strip() for p in prerequisites_value.split(",") if p.strip()]
                    for prereq in prereqs:
                        try:
                            add_knowledge_prerequisite(data["name"], prereq)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加前置依赖'{prereq}'失败: {str(e)}"
                            )

                # 处理关联关系
                relations_value = row.get("关联知识点(逗号分隔)", "").strip()
                if relations_value:
                    relations = [r.strip() for r in relations_value.split(",") if r.strip()]
                    for relation in relations:
                        try:
                            add_knowledge_relation(data["name"], relation)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加关联'{relation}'失败: {str(e)}"
                            )

            except Exception as e:
                stats["failed"] += 1
                stats["errors"].append(f"行{row_idx}: {str(e)}")

    except Exception as e:
        stats["errors"].append(f"读取CSV文件失败: {str(e)}")

    return stats


# ============================================
# 多节点类型架构支持 (Multi-Node Types Support)
# Migration 002 相关功能
# ============================================

def run_multi_node_types_migration(initiated_by: str = "system") -> Dict[str, object]:
    """
    运行多节点类型迁移 (Migration 002)

    创建 Stage, Skill, Terminology 等专用节点类型
    建立 PRECEDES 关系表达流程时序

    Args:
        initiated_by: 执行迁移的用户标识

    Returns:
        迁移统计信息
    """
    import importlib.util
    import sys
    from pathlib import Path

    # 动态加载迁移模块
    migration_path = Path(__file__).resolve().parent.parent / "migrations" / "002_multi_node_types.py"

    if not migration_path.exists():
        raise FileNotFoundError(f"Migration file not found: {migration_path}")

    spec = importlib.util.spec_from_file_location("migration_002", str(migration_path))
    if not spec or not spec.loader:
        raise RuntimeError("Failed to load migration module")

    migration_module = importlib.util.module_from_spec(spec)
    sys.modules["migration_002"] = migration_module
    spec.loader.exec_module(migration_module)

    # 执行迁移
    driver = _get_driver()
    stats = migration_module.upgrade(driver, initiated_by=initiated_by)

    LOGGER.info(f"Multi-node types migration completed: {stats}")
    return stats


def get_multi_node_types_migration_status() -> Dict[str, object]:
    """
    获取多节点类型迁移的状态

    Returns:
        {
            "applied": bool,
            "stages_count": int,
            "terminology_count": int,
            "precedes_count": int,
            ...
        }
    """
    import importlib.util
    import sys
    from pathlib import Path

    migration_path = Path(__file__).resolve().parent.parent / "migrations" / "002_multi_node_types.py"

    if not migration_path.exists():
        return {
            "applied": False,
            "error": "Migration file not found"
        }

    spec = importlib.util.spec_from_file_location("migration_002", str(migration_path))
    if not spec or not spec.loader:
        return {
            "applied": False,
            "error": "Failed to load migration module"
        }

    migration_module = importlib.util.module_from_spec(spec)
    sys.modules["migration_002"] = migration_module
    spec.loader.exec_module(migration_module)

    driver = _get_driver()
    return migration_module.get_migration_status(driver)


def get_process_flow() -> Dict[str, object]:
    """
    获取外贸谈判流程骨架

    返回所有 Stage 节点及其 PRECEDES 关系,用于前端渲染主流程轴

    Returns:
        {
            "stages": [
                {
                    "name": str,
                    "englishName": str,
                    "order": int,
                    "description": str,
                    "difficulty": str,
                    "icon": str,
                    "color": str,
                    ...
                }
            ],
            "flow": [
                {"from": str, "to": str, "description": str}
            ]
        }
    """
    # 获取所有 Stage 节点
    stages_query = """
    MATCH (s:Stage)
    OPTIONAL MATCH (s)-[:HAS_TOPIC]->(k)
    RETURN s {
        .*,
        topicsCount: count(DISTINCT k)
    } AS stage
    ORDER BY s.order
    """

    stages_records = _execute_read(stages_query, {})
    stages = [record["stage"] for record in stages_records] if stages_records else []

    # 获取 PRECEDES 关系
    flow_query = """
    MATCH (s1:Stage)-[r:PRECEDES]->(s2:Stage)
    RETURN s1.name AS from,
           s2.name AS to,
           r.description AS description,
           s1.order AS fromOrder,
           s2.order AS toOrder
    ORDER BY s1.order
    """

    flow_records = _execute_read(flow_query, {})
    flow = [
        {
            "from": record["from"],
            "to": record["to"],
            "description": record.get("description", ""),
            "fromOrder": record.get("fromOrder"),
            "toOrder": record.get("toOrder"),
        }
        for record in flow_records
    ] if flow_records else []

    return {
        "stages": stages,
        "flow": flow,
        "totalStages": len(stages),
        "totalFlows": len(flow),
    }


def list_stages(include_topics: bool = False) -> List[Dict[str, object]]:
    """
    获取所有 Stage (阶段) 节点列表

    Args:
        include_topics: 是否包含每个阶段下的知识点列表

    Returns:
        Stage 节点列表
    """
    if include_topics:
        query = """
        MATCH (s:Stage)
        OPTIONAL MATCH (s)-[:HAS_TOPIC]->(k)
        RETURN s {
            .*,
            topics: collect(DISTINCT k.name)
        } AS stage
        ORDER BY s.order
        """
    else:
        query = """
        MATCH (s:Stage)
        RETURN s AS stage
        ORDER BY s.order
        """

    records = _execute_read(query, {})
    return [record["stage"] for record in records] if records else []


def get_stage(name: str) -> Dict[str, object]:
    """
    获取单个 Stage 的详细信息

    Args:
        name: Stage 名称

    Returns:
        Stage 详细信息
    """
    query = """
    MATCH (s:Stage {name: $name})
    OPTIONAL MATCH (s)-[:HAS_TOPIC]->(k:KnowledgePoint)
    OPTIONAL MATCH (s)-[:PRECEDES]->(next:Stage)
    OPTIONAL MATCH (prev:Stage)-[:PRECEDES]->(s)
    WITH s, collect(DISTINCT k.name) AS topics, next, prev
    RETURN s {
        .*,
        topics: topics,
        nextStage: next.name,
        previousStage: prev.name
    } AS stage
    """

    records = _execute_read(query, {"name": name})
    if not records:
        raise GraphEntityNotFoundError(f"Stage '{name}' not found")

    return records[0]["stage"]


def list_terminology(category: Optional[str] = None) -> List[Dict[str, object]]:
    """
    获取术语列表

    Args:
        category: 可选的分类过滤 (如 "Incoterms", "Payment")

    Returns:
        Terminology 节点列表
    """
    if category:
        query = """
        MATCH (t:Terminology {category: $category})
        OPTIONAL MATCH (t)-[:RELATED_TO]-(related:Terminology)
        RETURN t {
            .*,
            relatedTerms: collect(DISTINCT related.name)
        } AS term
        ORDER BY t.name
        """
        params = {"category": category}
    else:
        query = """
        MATCH (t:Terminology)
        OPTIONAL MATCH (t)-[:RELATED_TO]-(related:Terminology)
        RETURN t {
            .*,
            relatedTerms: collect(DISTINCT related.name)
        } AS term
        ORDER BY t.category, t.name
        """
        params = {}

    records = _execute_read(query, params)
    return [record["term"] for record in records] if records else []


def get_terminology(name: str) -> Dict[str, object]:
    """
    获取单个术语的详细信息

    Args:
        name: 术语名称 (如 "FOB", "CIF")

    Returns:
        Terminology 详细信息
    """
    query = """
    MATCH (t:Terminology {name: $name})
    OPTIONAL MATCH (t)-[:RELATED_TO]-(related:Terminology)
    OPTIONAL MATCH (s:Stage)-[:HAS_TOPIC]->(k:KnowledgePoint)
           WHERE k.name CONTAINS t.name OR k.content CONTAINS t.name
    RETURN t {
        .*,
        relatedTerms: collect(DISTINCT related.name),
        usedInStages: collect(DISTINCT s.name)
    } AS term
    """

    records = _execute_read(query, {"name": name})
    if not records:
        raise GraphEntityNotFoundError(f"Terminology '{name}' not found")

    return records[0]["term"]


def link_knowledge_point_to_stage(knowledge_point_name: str, stage_name: str) -> Dict[str, object]:
    """
    将知识点关联到某个流程阶段

    创建 (Stage)-[:HAS_TOPIC]->(KnowledgePoint) 关系

    Args:
        knowledge_point_name: 知识点名称
        stage_name: 阶段名称

    Returns:
        更新后的 Stage 信息
    """
    query = """
    MATCH (s:Stage {name: $stage_name})
    MATCH (k:KnowledgePoint {name: $knowledge_point_name})
    MERGE (s)-[r:HAS_TOPIC]->(k)
    ON CREATE SET r.createdAt = datetime()
    RETURN s {
        .*,
        topics: [(s)-[:HAS_TOPIC]->(topic) | topic.name]
    } AS stage
    """

    records = _execute_read(
        query,
        {
            "stage_name": stage_name,
            "knowledge_point_name": knowledge_point_name,
        },
    )

    if not records:
        raise GraphEntityNotFoundError(
            f"Failed to link: Stage '{stage_name}' or KnowledgePoint '{knowledge_point_name}' not found"
        )

    LOGGER.info(f"Linked KnowledgePoint '{knowledge_point_name}' to Stage '{stage_name}'")
    return records[0]["stage"]


def unlink_knowledge_point_from_stage(
    knowledge_point_name: str, stage_name: str
) -> Dict[str, object]:
    """
    移除知识点与流程阶段的关联

    删除 (Stage)-[:HAS_TOPIC]->(KnowledgePoint) 关系

    Args:
        knowledge_point_name: 知识点名称
        stage_name: 阶段名称

    Returns:
        更新后的 Stage 信息
    """
    query = """
    MATCH (s:Stage {name: $stage_name})-[r:HAS_TOPIC]->(k:KnowledgePoint {name: $knowledge_point_name})
    DELETE r
    RETURN s {
        .*,
        topics: [(s)-[:HAS_TOPIC]->(topic) | topic.name]
    } AS stage
    """

    records = _execute_write(
        query,
        {
            "stage_name": stage_name,
            "knowledge_point_name": knowledge_point_name,
        },
    )

    if not records:
        raise GraphEntityNotFoundError(
            f"Relationship not found between Stage '{stage_name}' and KnowledgePoint '{knowledge_point_name}'"
        )

    LOGGER.info(f"Unlinked KnowledgePoint '{knowledge_point_name}' from Stage '{stage_name}'")
    return records[0]["stage"]


def get_enhanced_graph_visualization(
    node_types: Optional[List[str]] = None,
    max_nodes: int = 100,
) -> Dict[str, object]:
    """
    获取增强的图谱可视化数据,支持多节点类型

    Args:
        node_types: 要包含的节点类型列表,如 ["Stage", "KnowledgePoint", "Terminology"]
                   如果为 None,则包含所有类型
        max_nodes: 最大节点数量限制

    Returns:
        {
            "nodes": [
                {
                    "id": str,
                    "label": str,
                    "type": str,  # "Stage", "Topic", "KnowledgeCategory", "KnowledgePoint"
                    "properties": {...},
                    "group": str,  # 用于前端分组着色
                }
            ],
            "edges": [
                {
                    "from": str,
                    "to": str,
                    "label": str,  # "PRECEDES", "HAS_TOPIC", "REQUIRES", etc.
                    "type": str,
                }
            ],
            "statistics": {
                "nodesByType": {...},
                "edgesByType": {...},
            }
        }
    """
    # 默认包含所有节点类型
    if node_types is None:
        node_types = ["Stage", "Topic", "KnowledgeCategory", "KnowledgePoint"]

    # 查询节点
    nodes_query = """
    MATCH (s:Stage)-[:CONTAIN_TOPIC]->(t:Topic)
    OPTIONAL MATCH (t)-[:HAS_CATEGORY]->(c:KnowledgeCategory)
    OPTIONAL MATCH (c)-[:CONTAINS]->(k:KnowledgePoint)
    WITH collect(DISTINCT s) AS stages,
         collect(DISTINCT t) AS topics,
         [cat IN collect(DISTINCT c) WHERE cat IS NOT NULL] AS categories,
         [kp IN collect(DISTINCT k) WHERE kp IS NOT NULL][0..$max_nodes] AS points
    WITH stages + topics + categories + points AS nodes
    UNWIND nodes AS n
    WITH DISTINCT n
    WHERE any(label IN labels(n) WHERE label IN $labels)
    RETURN n,
           labels(n) AS labels,
           id(n) AS id
    """

    nodes_records = _execute_read(nodes_query, {"max_nodes": max_nodes, "labels": node_types})
    nodes = []
    node_ids = set()

    for record in nodes_records:
        node_data = record["n"]
        node_labels = record["labels"]
        internal_id = record["id"]

        # 确定主要标签 (优先级: Stage > Skill > Terminology > KnowledgePoint)
        primary_label = _select_primary_label_for_visualization(node_labels)

        # 提取节点标识符
        node_id = _extract_node_identifier_for_visualization(primary_label, node_data)

        nodes.append({
            "id": node_id,
            "label": node_data.get("name") or node_id,
            "type": primary_label,
            "properties": dict(node_data),
            "group": primary_label,  # 前端用于分组着色
            "internalId": internal_id,
        })

        node_ids.add(internal_id)

    # 查询这些节点之间的关系
    edges_query = """
    MATCH (a)-[r]->(b)
    WHERE id(a) IN $node_ids AND id(b) IN $node_ids
    RETURN id(a) AS fromId,
           id(b) AS toId,
           type(r) AS relType,
           properties(r) AS relProps,
           labels(a) AS fromLabels,
           labels(b) AS toLabels,
           a AS fromNode,
           b AS toNode
    """

    edges_records = _execute_read(edges_query, {"node_ids": list(node_ids)})
    edges = []

    for record in edges_records:
        from_labels = record["fromLabels"]
        to_labels = record["toLabels"]
        from_primary = _select_primary_label_for_visualization(from_labels)
        to_primary = _select_primary_label_for_visualization(to_labels)

        from_id = _extract_node_identifier_for_visualization(from_primary, record["fromNode"])
        to_id = _extract_node_identifier_for_visualization(to_primary, record["toNode"])

        edges.append({
            "from": from_id,
            "to": to_id,
            "label": record["relType"],
            "type": record["relType"],
            "properties": dict(record["relProps"]) if record["relProps"] else {},
        })

    # 统计信息
    nodes_by_type = {}
    for node in nodes:
        node_type = node["type"]
        nodes_by_type[node_type] = nodes_by_type.get(node_type, 0) + 1

    edges_by_type = {}
    for edge in edges:
        edge_type = edge["type"]
        edges_by_type[edge_type] = edges_by_type.get(edge_type, 0) + 1

    return {
        "nodes": nodes,
        "edges": edges,
        "statistics": {
            "totalNodes": len(nodes),
            "totalEdges": len(edges),
            "nodesByType": nodes_by_type,
            "edgesByType": edges_by_type,
        },
    }


def _select_primary_label_for_visualization(labels: List[str]) -> str:
    """
    从节点的多个标签中选择主要标签

    优先级: Stage > Topic > KnowledgeCategory > Skill > Terminology > KnowledgePoint > 其他
    """
    priority = ["Stage", "Topic", "KnowledgeCategory", "Skill", "Terminology", "KnowledgePoint"]

    for label in priority:
        if label in labels:
            return label

    return labels[0] if labels else "Unknown"


def _extract_node_identifier_for_visualization(primary_label: str, node_data: Dict) -> str:
    """
    从节点数据中提取标识符

    不同节点类型使用不同的标识字段
    """
    if primary_label == "KnowledgeCategory":
        base = node_data.get("name") or node_data.get("type") or "category"
        topic = node_data.get("topic") or node_data.get("topicName")
        stage = node_data.get("stage") or node_data.get("stageName")
        if topic:
            return f"{topic}:{base}"
        if stage:
            return f"{stage}:{base}"
        return str(base)

    if primary_label in ["Stage", "Skill", "Terminology", "KnowledgePoint"]:
        return node_data.get("name") or str(node_data.get("id", "unknown"))

    # 其他节点类型
    return str(node_data.get("id") or node_data.get("name", "unknown"))
