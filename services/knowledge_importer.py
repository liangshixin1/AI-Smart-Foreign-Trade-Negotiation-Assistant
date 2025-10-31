"""Knowledge point import utilities for Excel/CSV files.

Supports importing knowledge points from:
1. Excel (.xlsx, .xls) files
2. CSV files
3. Template generation for teachers
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Dict, List, Optional, BinaryIO

try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    LOGGER = logging.getLogger(__name__)
    LOGGER.warning("openpyxl not installed, Excel import will not be available")

from services import knowledge_service

LOGGER = logging.getLogger(__name__)


# ============================================
# Excel模板定义
# ============================================

EXCEL_TEMPLATE_HEADERS = [
    ("知识点名称", "name", True, "例如：FOB成本构成"),
    ("分类ID", "category", False, "例如：incoterms（留空则为uncategorized）"),
    ("类型", "type", False, "concept/skill/document/case/tool/theory/regulation"),
    ("难度", "difficulty", False, "beginner/intermediate/advanced"),
    ("重要性", "importance", False, "required/recommended/optional"),
    ("简介", "summary", False, "一句话描述，不超过100字"),
    ("详细描述", "description", False, "详细说明"),
    ("关键词", "keywords", False, "用逗号分隔，例如：FOB,成本,价格"),
    ("标签", "tags", False, "用逗号分隔"),
    ("预估学时(分钟)", "estimatedMinutes", False, "数字，例如：15"),
    ("图片URL", "imageUrl", False, "图片链接"),
    ("视频URL", "videoUrl", False, "视频链接"),
    ("文档URL", "documentUrl", False, "文档链接"),
    ("外部链接", "externalUrl", False, "参考资料链接"),
]


# ============================================
# 模板生成
# ============================================

def generate_excel_template() -> bytes:
    """生成Excel导入模板"""
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed. Please install it first: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "知识点导入模板"

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 15

    # 写入表头和说明
    for col_idx, (header, field, required, example) in enumerate(EXCEL_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = f"{header}{'*' if required else ''}"
        cell.font = openpyxl.styles.Font(bold=True, size=11)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 写入示例
        example_cell = ws.cell(row=2, column=col_idx)
        example_cell.value = example
        example_cell.font = openpyxl.styles.Font(italic=True, color="666666", size=9)

    # 添加说明sheet
    ws_guide = wb.create_sheet("使用说明")
    ws_guide.column_dimensions['A'].width = 80

    guide_content = [
        ("知识点批量导入说明", True),
        ("", False),
        ("1. 必填字段", True),
        ("   - 知识点名称：必须唯一，不能重复", False),
        ("", False),
        ("2. 分类ID参考（可在系统中查看完整列表）", True),
        ("   - incoterms: 贸易术语", False),
        ("   - payment-terms: 支付方式", False),
        ("   - trade-documents: 贸易文档", False),
        ("   - inquiry: 询盘阶段", False),
        ("   - offer: 报盘阶段", False),
        ("   - counter-offer: 还盘阶段", False),
        ("   - pricing-strategy: 价格策略", False),
        ("   - communication: 沟通技巧", False),
        ("   留空则默认为：uncategorized（未分类）", False),
        ("", False),
        ("3. 类型说明", True),
        ("   - concept: 概念型（定义、术语解释）", False),
        ("   - skill: 技能型（操作方法、技巧）", False),
        ("   - document: 文档型（表格、模板、格式）", False),
        ("   - case: 案例型（实际案例、情景）", False),
        ("   - tool: 工具型（计算器、检查清单）", False),
        ("   - theory: 理论型（理论框架、模型）", False),
        ("   - regulation: 法规型（法律、规则、标准）", False),
        ("", False),
        ("4. 难度等级", True),
        ("   - beginner: 初级（入门必学）", False),
        ("   - intermediate: 中级（进阶内容）", False),
        ("   - advanced: 高级（深度专题）", False),
        ("", False),
        ("5. 重要性级别", True),
        ("   - required: 必修（核心知识点）", False),
        ("   - recommended: 推荐（建议学习）", False),
        ("   - optional: 选修（扩展阅读）", False),
        ("", False),
        ("6. 注意事项", True),
        ("   - 关键词和标签用英文逗号分隔", False),
        ("   - 预估学时填写数字，单位为分钟", False),
        ("   - URL必须是完整的链接，包含http://或https://", False),
        ("   - 导入时，如果知识点名称已存在，将更新该知识点信息", False),
        ("", False),
        ("7. 示例数据", True),
        ("   请参考"知识点导入模板"sheet中的第2行示例", False),
        ("", False),
        ("8. 导入步骤", True),
        ("   1) 下载本模板", False),
        ("   2) 删除第2行的示例数据", False),
        ("   3) 填写您的知识点数据", False),
        ("   4) 保存文件", False),
        ("   5) 在系统中选择"导入Excel"并上传文件", False),
    ]

    for row_idx, (text, is_header) in enumerate(guide_content, start=1):
        cell = ws_guide.cell(row=row_idx, column=1)
        cell.value = text
        if is_header:
            cell.font = openpyxl.styles.Font(bold=True, size=12)
        else:
            cell.font = openpyxl.styles.Font(size=10)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # 保存到BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================
# Excel导入
# ============================================

def import_from_excel(file: BinaryIO, created_by: str = "excel-import") -> Dict[str, object]:
    """从Excel文件导入知识点"""
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed. Please install it first: pip install openpyxl")

    # 确保文件指针在开始位置
    if hasattr(file, 'seek'):
        file.seek(0)

    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    # 读取表头
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h.replace('*', '').strip() if h else '' for h in header_row]

    # 构建列名到索引的映射
    field_map = {}
    for idx, header in enumerate(headers):
        for template_header, field, _, _ in EXCEL_TEMPLATE_HEADERS:
            if header == template_header:
                field_map[field] = idx
                break

    # 检查必填字段
    if "name" not in field_map:
        raise ValueError("缺少必填列：知识点名称")

    # 读取数据行
    points_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):  # 跳过表头和示例
        # 跳过空行
        if not any(row):
            continue

        # 解析数据
        point = {}
        for field, col_idx in field_map.items():
            value = row[col_idx] if col_idx < len(row) else None
            if value is None or (isinstance(value, str) and not value.strip()):
                continue

            # 类型转换
            if field == "estimatedMinutes":
                try:
                    point[field] = int(value)
                except (ValueError, TypeError):
                    LOGGER.warning(f"Row {row_idx}: Invalid estimatedMinutes value: {value}")
                    continue
            elif field in ["keywords", "tags"]:
                # 分割逗号分隔的值
                if isinstance(value, str):
                    point[field] = [v.strip() for v in value.split(",") if v.strip()]
                else:
                    point[field] = []
            else:
                point[field] = str(value).strip() if value else ""

        # 必须有名称
        if not point.get("name"):
            continue

        points_data.append(point)

    wb.close()

    # 批量导入
    if not points_data:
        return {
            "success": False,
            "message": "没有找到有效的数据行",
            "created": 0,
            "updated": 0,
            "errors": 0,
        }

    result = knowledge_service.batch_import_knowledge_points(points_data, created_by=created_by)

    return {
        "success": result["errors"] == 0,
        "message": f"成功导入 {result['created'] + result['updated']} 个知识点",
        "created": result["created"],
        "updated": result["updated"],
        "errors": result["errors"],
        "errorMessages": result.get("errorMessages", []),
    }


# ============================================
# CSV导入（简化版）
# ============================================

def import_from_csv(file: BinaryIO, created_by: str = "csv-import") -> Dict[str, object]:
    """从CSV文件导入知识点"""
    # 确保文件指针在开始位置
    if hasattr(file, 'seek'):
        file.seek(0)

    # 尝试检测编码
    content = file.read()
    file.seek(0)

    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = content.decode('gbk')
        except UnicodeDecodeError:
            text = content.decode('latin-1')

    # 解析CSV
    reader = csv.DictReader(io.StringIO(text))

    # 构建字段映射
    field_map = {}
    for template_header, field, _, _ in EXCEL_TEMPLATE_HEADERS:
        if template_header in reader.fieldnames:
            field_map[template_header] = field

    points_data = []
    for row in reader:
        point = {}
        for csv_header, field in field_map.items():
            value = row.get(csv_header, "").strip()
            if not value:
                continue

            # 类型转换（同Excel）
            if field == "estimatedMinutes":
                try:
                    point[field] = int(value)
                except ValueError:
                    continue
            elif field in ["keywords", "tags"]:
                point[field] = [v.strip() for v in value.split(",") if v.strip()]
            else:
                point[field] = value

        if point.get("name"):
            points_data.append(point)

    # 批量导入
    if not points_data:
        return {
            "success": False,
            "message": "没有找到有效的数据行",
            "created": 0,
            "updated": 0,
            "errors": 0,
        }

    result = knowledge_service.batch_import_knowledge_points(points_data, created_by=created_by)

    return {
        "success": result["errors"] == 0,
        "message": f"成功导入 {result['created'] + result['updated']} 个知识点",
        "created": result["created"],
        "updated": result["updated"],
        "errors": result["errors"],
        "errorMessages": result.get("errorMessages", []),
    }


# ============================================
# 导出功能
# ============================================

def export_to_excel(
    category: Optional[str] = None,
    difficulty: Optional[str] = None,
    importance: Optional[str] = None,
) -> bytes:
    """导出知识点为Excel文件"""
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed")

    # 获取知识点列表
    points = knowledge_service.list_knowledge_points(
        category=category,
        difficulty=difficulty,
        importance=importance,
        limit=10000,  # 导出所有
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "知识点列表"

    # 写入表头
    for col_idx, (header, _, _, _) in enumerate(EXCEL_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # 写入数据
    for row_idx, point in enumerate(points, start=2):
        for col_idx, (_, field, _, _) in enumerate(EXCEL_TEMPLATE_HEADERS, start=1):
            value = point.get(field, "")

            # 格式化
            if field in ["keywords", "tags"] and isinstance(value, list):
                value = ", ".join(value)
            elif field == "estimatedMinutes":
                value = str(value) if value else ""

            ws.cell(row=row_idx, column=col_idx, value=value)

    # 保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
