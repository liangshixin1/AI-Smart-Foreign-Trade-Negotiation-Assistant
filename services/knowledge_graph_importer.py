"""
知识图谱导入服务 - 重构版

简洁的三表导入实现：
1. 谈判流程表 (Stage节点)
2. 知识点主表 (KnowledgePoint节点)
3. 案例库表 (Practice节点，可选)

设计原则：
- 单一职责：每个函数只做一件事
- 清晰的数据流：Excel -> 解析 -> 验证 -> 导入Neo4j
- 统一的错误处理
- 详细的日志记录
"""

from __future__ import annotations

import io
import logging
from typing import Dict, List, Optional, BinaryIO, Tuple
from dataclasses import dataclass, field
from datetime import datetime

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

LOGGER = logging.getLogger(__name__)


# ============================================
# 配置常量
# ============================================

# Excel 解析配置
EXCEL_HEADER_ROW = 1  # 表头行号
EXCEL_DATA_START_ROW = 2  # 数据起始行（跳过表头后的第一行，允许用户删除示例行）

# 默认节点属性
DEFAULT_STAGE_ICON = "🔵"
DEFAULT_STAGE_COLOR = "#3B82F6"
DEFAULT_STAGE_DIFFICULTY = "intermediate"
DEFAULT_STAGE_DURATION = 7

DEFAULT_POINT_TYPE = "concept"
DEFAULT_POINT_DIFFICULTY = "intermediate"
DEFAULT_POINT_IMPORTANCE = "recommended"

# 节点类型映射（用于将 Excel 类型列映射到节点标签）
POINT_TYPE_ALIASES = {
    "terminology": "Terminology",
    "term": "Terminology",
    "concept": "Terminology",
    "conceptual": "Terminology",
    "术语": "Terminology",
    "概念": "Terminology",
    "概念型": "Terminology",
    "概念性": "Terminology",
    "概念类": "Terminology",
    "知识点": "KnowledgePoint",
    "知识": "KnowledgePoint",
    "knowledge": "KnowledgePoint",
    "knowledgepoint": "KnowledgePoint",
    "skill": "Skill",
    "skills": "Skill",
    "技能": "Skill",
    "技能型": "Skill",
    "技能性": "Skill",
    "业务流程": "Skill",
    "流程": "Skill",
    "流程型": "Skill",
    "process": "Skill",
    "practice": "Skill",
}


# ============================================
# 数据结构定义
# ============================================

@dataclass
class ImportError:
    """导入错误"""
    level: str  # ERROR / WARNING
    sheet: str  # 谈判流程 / 知识点主表 / 案例库表
    row: int
    field: str = ""
    value: str = ""
    message: str = ""
    suggestion: str = ""


@dataclass
class ImportStats:
    """导入统计"""
    total: int = 0
    created: int = 0
    updated: int = 0
    failed: int = 0

    @property
    def success_rate(self) -> str:
        if self.total == 0:
            return "0%"
        return f"{(self.created + self.updated) / self.total * 100:.1f}%"


@dataclass
class ImportResult:
    """导入结果"""
    success: bool
    stages: ImportStats = field(default_factory=ImportStats)
    topics: ImportStats = field(default_factory=ImportStats)
    knowledge_points: ImportStats = field(default_factory=ImportStats)
    practices: ImportStats = field(default_factory=ImportStats)
    relations: ImportStats = field(default_factory=ImportStats)
    errors: List[ImportError] = field(default_factory=list)
    warnings: List[ImportError] = field(default_factory=list)
    duration_seconds: float = 0.0
    topics_by_stage: Dict[str, int] = field(default_factory=dict)

    def to_dict(self) -> Dict:
        """转换为字典格式（向后兼容：practices字段对外显示为examples）"""
        return {
            "success": self.success,
            "statistics": {
                "stages": {
                    "total": self.stages.total,
                    "created": self.stages.created,
                    "updated": self.stages.updated,
                    "failed": self.stages.failed,
                    "success_rate": self.stages.success_rate,
                },
                "topics": {
                    "total": self.topics.total,
                    "created": self.topics.created,
                    "updated": self.topics.updated,
                    "failed": self.topics.failed,
                    "success_rate": self.topics.success_rate,
                },
                "points": {
                    "total": self.knowledge_points.total,
                    "created": self.knowledge_points.created,
                    "updated": self.knowledge_points.updated,
                    "failed": self.knowledge_points.failed,
                    "success_rate": self.knowledge_points.success_rate,
                },
                "examples": {  # 向后兼容：前端期望examples字段
                    "total": self.practices.total,
                    "created": self.practices.created,
                    "failed": self.practices.failed,
                    "success_rate": self.practices.success_rate,
                },
                "relations": {
                    "total": self.relations.total,
                    "created": self.relations.created,
                    "failed": self.relations.failed,
                    "success_rate": self.relations.success_rate,
                },
            },
            "errors": [
                {
                    "level": e.level,
                    "sheet": e.sheet,
                    "row": e.row,
                    "field": e.field,
                    "value": e.value,
                    "message": e.message,
                    "suggestion": e.suggestion,
                }
                for e in self.errors
            ],
            "warnings": [
                {
                    "level": w.level,
                    "sheet": w.sheet,
                    "row": w.row,
                    "field": w.field,
                    "value": w.value,
                    "message": w.message,
                    "suggestion": w.suggestion,
                }
                for w in self.warnings
            ],
            "execution_time": f"{self.duration_seconds:.2f}s",
            "topicsByStage": self.topics_by_stage,
        }


# ============================================
# 主导入类
# ============================================

class KnowledgeGraphImporter:
    """知识图谱导入器"""

    def __init__(self, neo4j_driver):
        """
        初始化导入器

        Args:
            neo4j_driver: Neo4j数据库驱动
        """
        self.driver = neo4j_driver

    def _resolve_node_label(self, raw_type: Optional[str]) -> str:
        """根据导入类型返回节点标签,默认为 KnowledgePoint。"""

        if raw_type is None:
            return "KnowledgePoint"
        normalized = str(raw_type).strip().lower()
        return POINT_TYPE_ALIASES.get(normalized, "KnowledgePoint")

    def import_from_excel(
        self,
        excel_file: BinaryIO,
        created_by: str = "system",
    ) -> ImportResult:
        """
        从Excel文件导入知识图谱

        Args:
            excel_file: Excel文件流
            created_by: 创建者标识

        Returns:
            ImportResult: 导入结果
        """
        import time
        start_time = time.time()

        result = ImportResult(success=False)

        try:
            # 步骤1: 读取Excel文件内容
            LOGGER.info("步骤1: 读取Excel文件...")
            if hasattr(excel_file, 'read'):
                file_content = excel_file.read()
            else:
                file_content = excel_file

            # 步骤2: 解析三个表的数据
            LOGGER.info("步骤2: 解析Excel数据...")
            stages_data, stages_errors = self._parse_stages(file_content)
            result.errors.extend(stages_errors)

            points_data, points_errors = self._parse_knowledge_points(file_content)
            result.errors.extend(points_errors)

            practices_data, practices_errors = self._parse_practices(file_content)
            result.errors.extend(practices_errors)

            LOGGER.info(f"  解析结果: {len(stages_data)}个阶段, {len(points_data)}个知识点, {len(practices_data)}个案例")

            # 步骤3: 验证数据
            LOGGER.info("步骤3: 验证数据...")
            validation_errors, validation_warnings = self._validate_data(
                stages_data, points_data, practices_data
            )
            result.errors.extend(validation_errors)
            result.warnings.extend(validation_warnings)

            # 检查是否有致命错误
            if any(e.level == "ERROR" for e in result.errors):
                error_count = len([e for e in result.errors if e.level == "ERROR"])
                LOGGER.error(f"发现 {error_count} 个错误，停止导入")
                result.duration_seconds = time.time() - start_time
                return result

            # 步骤4: 导入到Neo4j（使用事务确保原子性）
            LOGGER.info("步骤4: 导入到Neo4j...")
            with self.driver.session() as session:
                # 使用 execute_write 确保事务正确管理（Neo4j driver 5.x+）
                def import_work(tx):
                    self._import_to_neo4j(
                        tx, stages_data, points_data, practices_data,
                        result, created_by
                    )

                # 兼容不同版本的 Neo4j driver
                if hasattr(session, 'execute_write'):
                    session.execute_write(import_work)
                elif hasattr(session, 'write_transaction'):
                    session.write_transaction(import_work)
                else:
                    # 降级到手动事务管理
                    tx = session.begin_transaction()
                    try:
                        import_work(tx)
                        tx.commit()
                    except Exception:
                        tx.rollback()
                        raise

                result.success = True
                LOGGER.info("事务提交成功")

            result.duration_seconds = time.time() - start_time

            LOGGER.info(
                f"导入完成: "
                f"阶段={result.stages.created}创建, "
                f"知识点={result.knowledge_points.created}创建/{result.knowledge_points.updated}更新, "
                f"案例={result.practices.created}创建, "
                f"关系={result.relations.created}创建"
            )

        except Exception as e:
            LOGGER.exception("导入失败")
            result.errors.append(ImportError(
                level="ERROR",
                sheet="系统",
                row=0,
                message=f"系统错误: {str(e)}",
            ))
            result.duration_seconds = time.time() - start_time

        return result

    # ========================================
    # 解析方法
    # ========================================

    def _parse_stages(self, file_content: bytes) -> Tuple[List[Dict], List[ImportError]]:
        """
        解析谈判流程表

        Returns:
            (阶段数据列表, 错误列表)
        """
        errors = []
        stages = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

            # 查找"谈判流程"Sheet
            if "谈判流程" not in wb.sheetnames:
                errors.append(ImportError(
                    level="ERROR",
                    sheet="谈判流程",
                    row=0,
                    message="未找到'谈判流程'Sheet",
                    suggestion="请确保Excel文件包含名为'谈判流程'的Sheet",
                ))
                wb.close()
                return [], errors

            ws = wb["谈判流程"]

            # 读取表头（第1行）
            headers = [str(cell).strip() if cell else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            # 构建列索引映射
            col_map = {}
            for idx, header in enumerate(headers):
                clean_header = header.replace('*', '').strip()
                if clean_header == "阶段名称":
                    col_map["name"] = idx
                elif clean_header == "英文名称":
                    col_map["englishName"] = idx
                elif clean_header == "阶段描述":
                    col_map["description"] = idx
                elif clean_header == "难度级别":
                    col_map["difficulty"] = idx
                elif clean_header == "预计时长(天)":
                    col_map["estimatedDuration"] = idx
                elif clean_header == "图标":
                    col_map["icon"] = idx
                elif clean_header == "颜色":
                    col_map["color"] = idx

            # 检查必填列
            if "name" not in col_map:
                errors.append(ImportError(
                    level="ERROR",
                    sheet="谈判流程",
                    row=1,
                    field="阶段名称",
                    message="缺少必填列'阶段名称'",
                ))
                wb.close()
                return [], errors

            # 读取数据（从第2行开始，跳过表头）
            order = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=EXCEL_DATA_START_ROW, values_only=True), start=EXCEL_DATA_START_ROW):
                # 跳过空行
                if not any(row):
                    continue

                # 提取字段
                name = str(row[col_map["name"]]).strip() if len(row) > col_map["name"] and row[col_map["name"]] else ""

                # 跳过没有名称的行
                if not name:
                    continue

                stage = {
                    "name": name,
                    "order": order,
                    "_row": row_idx,
                }
                order += 1

                # 可选字段
                if "englishName" in col_map and len(row) > col_map["englishName"] and row[col_map["englishName"]]:
                    stage["englishName"] = str(row[col_map["englishName"]]).strip()

                if "description" in col_map and len(row) > col_map["description"] and row[col_map["description"]]:
                    stage["description"] = str(row[col_map["description"]]).strip()

                if "difficulty" in col_map and len(row) > col_map["difficulty"] and row[col_map["difficulty"]]:
                    difficulty = str(row[col_map["difficulty"]]).strip()
                    # 中文转英文
                    difficulty_map = {"初级": "beginner", "中级": "intermediate", "高级": "advanced"}
                    stage["difficulty"] = difficulty_map.get(difficulty, difficulty)

                if "estimatedDuration" in col_map and len(row) > col_map["estimatedDuration"] and row[col_map["estimatedDuration"]]:
                    try:
                        stage["estimatedDuration"] = int(row[col_map["estimatedDuration"]])
                    except (ValueError, TypeError):
                        stage["estimatedDuration"] = 7

                if "icon" in col_map and len(row) > col_map["icon"] and row[col_map["icon"]]:
                    stage["icon"] = str(row[col_map["icon"]]).strip()

                if "color" in col_map and len(row) > col_map["color"] and row[col_map["color"]]:
                    stage["color"] = str(row[col_map["color"]]).strip()

                stages.append(stage)

            wb.close()

            if len(stages) == 0:
                errors.append(ImportError(
                    level="ERROR",
                    sheet="谈判流程",
                    row=0,
                    message="未找到任何阶段数据",
                    suggestion=f"请在'谈判流程'Sheet的第{EXCEL_DATA_START_ROW}行及以后添加阶段数据",
                ))

        except Exception as e:
            LOGGER.exception("解析谈判流程表失败")
            errors.append(ImportError(
                level="ERROR",
                sheet="谈判流程",
                row=0,
                message=f"解析失败: {str(e)}",
            ))

        return stages, errors

    def _parse_knowledge_points(self, file_content: bytes) -> Tuple[List[Dict], List[ImportError]]:
        """
        解析知识点主表

        Returns:
            (知识点数据列表, 错误列表)
        """
        errors = []
        points = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

            # 查找"知识点主表"Sheet（优先使用明确命名的sheet）
            sheet_name = None
            for possible_name in ["知识点主表", "知识点"]:
                if possible_name in wb.sheetnames:
                    sheet_name = possible_name
                    break

            # 如果没找到，尝试使用第一个不是"谈判流程"的sheet
            if not sheet_name:
                for name in wb.sheetnames:
                    if name not in ["谈判流程", "案例库表", "案例库"]:
                        sheet_name = name
                        break

            if not sheet_name:
                errors.append(ImportError(
                    level="WARNING",
                    sheet="知识点主表",
                    row=0,
                    message="未找到'知识点主表'Sheet",
                    suggestion="知识点导入将跳过",
                ))
                wb.close()
                return [], errors

            ws = wb[sheet_name]

            # 读取表头
            headers = [str(cell).strip() if cell else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            # 检查表头是否为空（可能是个空sheet）
            if not any(headers):
                errors.append(ImportError(
                    level="WARNING",
                    sheet="知识点主表",
                    row=0,
                    message=f"Sheet '{sheet_name}' 的表头为空",
                    suggestion="知识点导入将跳过",
                ))
                wb.close()
                return [], errors

            # 构建列索引映射
            col_map = {}
            for idx, header in enumerate(headers):
                clean_header = header.replace('*', '').strip()
                if clean_header == "知识点名称":
                    col_map["name"] = idx
                elif clean_header == "所属阶段":
                    col_map["stage"] = idx
                elif clean_header in {"二级主题", "主题", "业务主题"}:
                    col_map["topic"] = idx
                elif clean_header == "知识点类型":
                    col_map["type"] = idx
                elif clean_header == "难度":
                    col_map["difficulty"] = idx
                elif clean_header == "重要性":
                    col_map["importance"] = idx
                elif clean_header == "内容简介":
                    col_map["summary"] = idx
                elif clean_header == "详细描述":
                    col_map["description"] = idx
                elif clean_header == "章节":
                    col_map["chapter"] = idx

            # 检查必填列
            if "name" not in col_map:
                errors.append(ImportError(
                    level="ERROR",
                    sheet=f"知识点主表 ({sheet_name})",
                    row=1,
                    field="知识点名称",
                    message="缺少必填列'知识点名称'",
                    suggestion=f"请确保Sheet '{sheet_name}' 的第1行包含'知识点名称'或'*知识点名称'列",
                ))
                wb.close()
                return [], errors

            # 读取数据（从第2行开始，跳过表头）
            for row_idx, row in enumerate(ws.iter_rows(min_row=EXCEL_DATA_START_ROW, values_only=True), start=EXCEL_DATA_START_ROW):
                # 跳过空行
                if not any(row):
                    continue

                # 提取字段
                name = str(row[col_map["name"]]).strip() if len(row) > col_map["name"] and row[col_map["name"]] else ""

                # 跳过没有名称的行
                if not name:
                    continue

                point = {
                    "name": name,
                    "_row": row_idx,
                }

                # 可选字段
                if "stage" in col_map and len(row) > col_map["stage"] and row[col_map["stage"]]:
                    point["stage"] = str(row[col_map["stage"]]).strip()

                if "topic" in col_map and len(row) > col_map["topic"] and row[col_map["topic"]]:
                    point["topic"] = str(row[col_map["topic"]]).strip()

                if "type" in col_map and len(row) > col_map["type"] and row[col_map["type"]]:
                    point["type"] = str(row[col_map["type"]]).strip()

                if "difficulty" in col_map and len(row) > col_map["difficulty"] and row[col_map["difficulty"]]:
                    difficulty = str(row[col_map["difficulty"]]).strip()
                    difficulty_map = {"初级": "beginner", "中级": "intermediate", "高级": "advanced"}
                    point["difficulty"] = difficulty_map.get(difficulty, difficulty)

                if "importance" in col_map and len(row) > col_map["importance"] and row[col_map["importance"]]:
                    point["importance"] = str(row[col_map["importance"]]).strip()

                if "summary" in col_map and len(row) > col_map["summary"] and row[col_map["summary"]]:
                    point["summary"] = str(row[col_map["summary"]]).strip()

                if "description" in col_map and len(row) > col_map["description"] and row[col_map["description"]]:
                    point["description"] = str(row[col_map["description"]]).strip()

                if "chapter" in col_map and len(row) > col_map["chapter"] and row[col_map["chapter"]]:
                    point["chapter"] = str(row[col_map["chapter"]]).strip()

                points.append(point)

            wb.close()

        except Exception as e:
            LOGGER.exception("解析知识点主表失败")
            errors.append(ImportError(
                level="ERROR",
                sheet="知识点主表",
                row=0,
                message=f"解析失败: {str(e)}",
            ))

        return points, errors

    def _parse_practices(self, file_content: bytes) -> Tuple[List[Dict], List[ImportError]]:
        """
        解析案例库表（可选）

        Returns:
            (案例数据列表, 错误列表)
        """
        errors = []
        practices = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

            # 查找"案例库表"Sheet
            sheet_name = None
            for possible_name in ["案例库表", "案例库", "Sheet3"]:
                if possible_name in wb.sheetnames:
                    sheet_name = possible_name
                    break

            if not sheet_name:
                # 案例库是可选的，不报错
                wb.close()
                return [], errors

            ws = wb[sheet_name]

            # 读取表头
            headers = [str(cell).strip() if cell else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            # 构建列索引映射
            col_map = {}
            for idx, header in enumerate(headers):
                clean_header = header.replace('*', '').strip()
                if clean_header == "关联知识点":
                    col_map["knowledgePoint"] = idx
                elif clean_header == "案例标题":
                    col_map["title"] = idx
                elif clean_header == "案例场景":
                    col_map["scenario"] = idx
                elif clean_header == "案例内容":
                    col_map["content"] = idx

            # 读取数据（从第2行开始，跳过表头）
            for row_idx, row in enumerate(ws.iter_rows(min_row=EXCEL_DATA_START_ROW, values_only=True), start=EXCEL_DATA_START_ROW):
                # 跳过空行
                if not any(row):
                    continue

                practice = {"_row": row_idx}

                if "knowledgePoint" in col_map and len(row) > col_map["knowledgePoint"] and row[col_map["knowledgePoint"]]:
                    practice["knowledgePoint"] = str(row[col_map["knowledgePoint"]]).strip()

                if "title" in col_map and len(row) > col_map["title"] and row[col_map["title"]]:
                    practice["title"] = str(row[col_map["title"]]).strip()

                if "scenario" in col_map and len(row) > col_map["scenario"] and row[col_map["scenario"]]:
                    practice["scenario"] = str(row[col_map["scenario"]]).strip()

                if "content" in col_map and len(row) > col_map["content"] and row[col_map["content"]]:
                    practice["content"] = str(row[col_map["content"]]).strip()

                # 至少需要标题或内容
                if "title" in practice or "content" in practice:
                    practices.append(practice)

            wb.close()

        except Exception as e:
            LOGGER.exception("解析案例库表失败")
            errors.append(ImportError(
                level="WARNING",
                sheet="案例库表",
                row=0,
                message=f"解析失败: {str(e)}",
            ))

        return practices, errors

    # ========================================
    # 验证方法
    # ========================================

    def _validate_data(
        self,
        stages: List[Dict],
        points: List[Dict],
        practices: List[Dict],
    ) -> Tuple[List[ImportError], List[ImportError]]:
        """
        验证数据完整性

        Returns:
            (错误列表, 警告列表)
        """
        errors = []
        warnings = []

        # 建立阶段名称集合
        stage_names = {s["name"] for s in stages}

        # 建立知识点名称集合
        point_names = {p["name"] for p in points}

        # 验证知识点引用的阶段是否存在
        for point in points:
            if "stage" in point:
                if point["stage"] not in stage_names:
                    warnings.append(ImportError(
                        level="WARNING",
                        sheet="知识点主表",
                        row=point.get("_row", 0),
                        field="所属阶段",
                        value=point["stage"],
                        message=f"阶段'{point['stage']}'不存在",
                        suggestion=f"可用阶段: {', '.join(list(stage_names)[:5])}",
                    ))

        # 验证案例引用的知识点是否存在
        for practice in practices:
            if "knowledgePoint" in practice:
                if practice["knowledgePoint"] not in point_names:
                    warnings.append(ImportError(
                        level="WARNING",
                        sheet="案例库表",
                        row=practice.get("_row", 0),
                        field="关联知识点",
                        value=practice["knowledgePoint"],
                        message=f"知识点'{practice['knowledgePoint']}'不存在",
                        suggestion="该案例将被跳过",
                    ))

        return errors, warnings

    # ========================================
    # 导入方法
    # ========================================

    def _import_to_neo4j(
        self,
        tx,  # Neo4j Transaction对象
        stages: List[Dict],
        points: List[Dict],
        practices: List[Dict],
        result: ImportResult,
        created_by: str,
    ):
        """
        导入数据到Neo4j（在事务中执行）

        Args:
            tx: Neo4j事务对象
            stages: 阶段数据列表
            points: 知识点数据列表
            practices: 案例数据列表
            result: 导入结果对象
            created_by: 创建者标识
        """

        # 设置统计
        result.stages.total = len(stages)
        result.knowledge_points.total = len(points)
        result.practices.total = len(practices)

        # 第1步：创建Stage节点
        LOGGER.info(f"第1步: 创建 {len(stages)} 个Stage节点...")
        stage_names = set()
        for stage in stages:
            try:
                self._create_stage(tx, stage, created_by)
                result.stages.created += 1
                stage_names.add(stage["name"])
                LOGGER.debug(f"  创建Stage: {stage['name']}")
            except Exception as e:
                LOGGER.error(f"  创建Stage失败 {stage['name']}: {e}")
                result.stages.failed += 1

        # 第2步：创建Stage之间的PRECEDES关系
        LOGGER.info(f"第2步: 创建Stage顺序关系...")
        sorted_stages = sorted(stages, key=lambda s: s.get("order", 0))
        for i in range(len(sorted_stages) - 1):
            from_stage = sorted_stages[i]["name"]
            to_stage = sorted_stages[i + 1]["name"]

            if from_stage in stage_names and to_stage in stage_names:
                try:
                    self._create_precedes_relation(tx, from_stage, to_stage, created_by)
                    result.relations.created += 1
                    result.relations.total += 1
                    LOGGER.debug(f"  创建关系: {from_stage} -> {to_stage}")
                except Exception as e:
                    LOGGER.error(f"  创建关系失败 {from_stage}->{to_stage}: {e}")
                    result.relations.failed += 1
                    result.relations.total += 1

        # 第3步：创建KnowledgePoint节点
        LOGGER.info(f"第3步: 创建 {len(points)} 个KnowledgePoint节点...")
        point_names = set()
        point_stage_map = {}  # 记录知识点和阶段/主题的对应关系
        topics: Dict[str, Dict[str, str]] = {}  # topic_key -> {"stage": stage_name, "name": topic_name}

        for point in points:
            point_name = point["name"]
            stage_name = point.pop("stage", None)  # 移除stage字段，不存入节点
            topic_name = point.pop("topic", None)
            node_label = self._resolve_node_label(point.get("type"))
            point.pop("_row", None)  # 移除辅助字段

            try:
                if stage_name and stage_name not in stage_names:
                    # 确保不存在于本次stage列表的阶段也被创建, 保证Topic/Point可挂接
                    self._ensure_stage_stub(tx, stage_name, created_by)
                    stage_names.add(stage_name)

                # 检查是否已存在
                existing = self._get_knowledge_point(tx, point_name)
                if existing:
                    self._update_knowledge_point(tx, point_name, point, node_label, created_by)
                    result.knowledge_points.updated += 1
                    LOGGER.debug(f"  更新知识点: {point_name}")
                else:
                    self._create_knowledge_point(tx, point, node_label, created_by)
                    result.knowledge_points.created += 1
                    LOGGER.debug(f"  创建知识点: {point_name}")

                point_names.add(point_name)
                if stage_name:
                    point_stage_map[point_name] = {
                        "stage": stage_name,
                        "label": node_label,
                        "topic": topic_name,
                    }
                if topic_name and stage_name:
                    topic_key = f"{stage_name}::{topic_name}"
                    topics[topic_key] = {"stage": stage_name, "name": topic_name}

            except Exception as e:
                LOGGER.error(f"  创建/更新知识点失败 {point_name}: {e}")
                result.knowledge_points.failed += 1

        # 第3.5步：创建 Topic 节点并与 Stage 关联
        LOGGER.info("第3.5步: 创建 Topic 节点并关联 Stage ...")
        result.topics.total = len(topics)
        for topic_key, payload in topics.items():
            topic_name = payload.get("name")
            stage_name = payload.get("stage")
            if not stage_name or stage_name not in stage_names:
                continue
            try:
                self._create_topic(tx, topic_name, stage_name, created_by)
                result.topics.created += 1
                result.topics_by_stage[stage_name] = result.topics_by_stage.get(stage_name, 0) + 1
                result.relations.created += 1  # 计入 Stage-Topic 关系
                result.relations.total += 1
            except Exception as e:
                LOGGER.error(f"  创建 Topic 失败 {topic_name}: {e}")
                result.topics.failed += 1
                result.relations.failed += 1
                result.relations.total += 1

        # 第4步：创建Stage-KnowledgePoint的HAS_TOPIC关系
        LOGGER.info(f"第4步: 创建Topic-Point关系 (树形)")
        for point_name, stage_payload in point_stage_map.items():
            stage_name = stage_payload["stage"]
            topic_name = stage_payload.get("topic")
            if topic_name and stage_name in stage_names and point_name in point_names:
                try:
                    self._create_include_point_relation(
                        tx, topic_name, point_name, stage_name, created_by
                    )
                    result.relations.created += 1
                    result.relations.total += 1
                    LOGGER.debug(f"  创建关系: {topic_name} -> {point_name}")
                except Exception as e:
                    LOGGER.error(f"  创建关系失败 {topic_name}->{point_name}: {e}")
                    result.relations.failed += 1
                    result.relations.total += 1

        # 第5步：创建Practice节点和关系（如果有）
        if practices:
            LOGGER.info(f"第5步: 创建 {len(practices)} 个Practice节点...")
            for practice in practices:
                kp_name = practice.pop("knowledgePoint", None)
                practice.pop("_row", None)

                # 只有关联的知识点存在才创建
                if kp_name and kp_name in point_names:
                    try:
                        practice_id = self._create_practice(tx, practice, created_by)
                        result.practices.created += 1
                        LOGGER.debug(f"  创建案例: {practice.get('title', practice_id)}")

                        # 创建关联关系
                        self._create_has_practice_relation(tx, kp_name, practice_id, created_by)
                        result.relations.created += 1
                        result.relations.total += 1
                    except Exception as e:
                        LOGGER.error(f"  创建案例失败: {e}")
                        result.practices.failed += 1

    # ========================================
    # Neo4j操作方法
    # ========================================

    def _create_stage(self, tx, stage: Dict, created_by: str):
        """创建Stage节点（在事务中执行）"""
        query = """
        MERGE (s:Stage {name: $name})
        ON CREATE SET
            s.englishName = $englishName,
            s.description = $description,
            s.difficulty = $difficulty,
            s.estimatedDuration = $estimatedDuration,
            s.icon = $icon,
            s.color = $color,
            s.createdAt = datetime(),
            s.createdBy = $createdBy,
            s.updatedAt = datetime(),
            s.updatedBy = $createdBy
        SET
            s.englishName = coalesce(s.englishName, $englishName),
            s.description = $description,
            s.difficulty = $difficulty,
            s.estimatedDuration = $estimatedDuration,
            s.icon = $icon,
            s.color = $color,
            s.updatedAt = datetime(),
            s.updatedBy = $createdBy
        RETURN s.name AS name
        """
        params = {
            "name": stage.get("name"),
            "englishName": stage.get("englishName", ""),
            "description": stage.get("description", ""),
            "difficulty": stage.get("difficulty", DEFAULT_STAGE_DIFFICULTY),
            "estimatedDuration": stage.get("estimatedDuration", DEFAULT_STAGE_DURATION),
            "icon": stage.get("icon", DEFAULT_STAGE_ICON),
            "color": stage.get("color", DEFAULT_STAGE_COLOR),
            "createdBy": created_by,
        }

        tx.run(query, params)

    def _create_precedes_relation(self, tx, from_stage: str, to_stage: str, created_by: str):
        """创建PRECEDES关系（在事务中执行）"""
        query = """
        MATCH (s1:Stage {name: $from_stage})
        MATCH (s2:Stage {name: $to_stage})
        MERGE (s1)-[r:PRECEDES]->(s2)
        ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
        RETURN s1.name AS from, s2.name AS to
        """
        params = {
            "from_stage": from_stage,
            "to_stage": to_stage,
            "createdBy": created_by,
        }

        tx.run(query, params)

    def _ensure_stage_stub(self, tx, stage_name: str, created_by: str):
        """若Stage不存在则创建一个占位节点，避免引用失败"""
        tx.run(
            """
            MERGE (s:Stage {name: $name})
            ON CREATE SET s.createdAt = datetime(),
                          s.createdBy = $createdBy,
                          s.difficulty = $difficulty,
                          s.estimatedDuration = $estimatedDuration,
                          s.icon = $icon,
                          s.color = $color,
                          s.order = coalesce(s.order, 0),
                          s.updatedAt = datetime(),
                          s.updatedBy = $createdBy
            """,
            {
                "name": stage_name,
                "createdBy": created_by,
                "difficulty": DEFAULT_STAGE_DIFFICULTY,
                "estimatedDuration": DEFAULT_STAGE_DURATION,
                "icon": DEFAULT_STAGE_ICON,
                "color": DEFAULT_STAGE_COLOR,
            },
        )

    def _create_topic(self, tx, topic_name: str, stage_name: str, created_by: str):
        """创建 Topic 节点并与 Stage 建立 CONTAIN_TOPIC 关系"""
        query = """
        MERGE (s:Stage {name: $stage_name})
        MERGE (t:Topic {name: $name, stage: $stage_name})
        ON CREATE SET
            t.createdAt = datetime(),
            t.updatedAt = datetime(),
            t.createdBy = $createdBy,
            t.updatedBy = $createdBy
        SET
            t.updatedAt = datetime(),
            t.updatedBy = $createdBy
        MERGE (s)-[r:CONTAIN_TOPIC]->(t)
        ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
        RETURN t.name AS name
        """
        params = {
            "name": topic_name,
            "stage_name": stage_name,
            "createdBy": created_by,
        }
        tx.run(query, params)

    def _get_knowledge_point(self, tx, name: str) -> Optional[Dict]:
        """检查知识点是否存在（在事务中执行）"""
        query = """
        MATCH (k:KnowledgePoint {name: $name})
        RETURN k, labels(k) AS labels
        """
        result = tx.run(query, {"name": name})
        record = result.single()
        if not record:
            return None
        payload = dict(record["k"])
        payload["labels"] = record.get("labels") or []
        return payload

    def _create_knowledge_point(self, tx, point: Dict, label: str, created_by: str):
        """创建知识点节点（在事务中执行）"""
        labels_cypher = "KnowledgePoint" if label == "KnowledgePoint" else f"KnowledgePoint:{label}"
        query = f"""
        MERGE (k:{labels_cypher} {{
            name: $name
        }})
        ON CREATE SET
            k.type = $type,
            k.difficulty = $difficulty,
            k.importance = $importance,
            k.summary = $summary,
            k.description = $description,
            k.chapter = $chapter,
            k.nodeType = $nodeType,
            k.createdAt = datetime(),
            k.createdBy = $createdBy,
            k.updatedAt = datetime(),
            k.updatedBy = $createdBy
        SET
            k.type = $type,
            k.difficulty = $difficulty,
            k.importance = $importance,
            k.summary = $summary,
            k.description = $description,
            k.chapter = $chapter,
            k.nodeType = $nodeType,
            k.updatedAt = datetime(),
            k.updatedBy = $createdBy
        RETURN k.name AS name
        """
        point_type = str(point.get("type", DEFAULT_POINT_TYPE) or DEFAULT_POINT_TYPE).strip().lower()
        params = {
            "name": point.get("name"),
            "type": point_type,
            "difficulty": point.get("difficulty", DEFAULT_POINT_DIFFICULTY),
            "importance": point.get("importance", DEFAULT_POINT_IMPORTANCE),
            "summary": point.get("summary", ""),
            "description": point.get("description", ""),
            "chapter": point.get("chapter", ""),
            "nodeType": label,
            "createdBy": created_by,
        }

        tx.run(query, params)

    def _update_knowledge_point(self, tx, name: str, point: Dict, label: str, created_by: str):
        """更新知识点节点（在事务中执行）"""
        query = f"""
        MATCH (k:KnowledgePoint {{name: $name}})
        SET k:{label}
        SET k.type = $type,
            k.difficulty = $difficulty,
            k.importance = $importance,
            k.summary = $summary,
            k.description = $description,
            k.chapter = $chapter,
            k.nodeType = $nodeType,
            k.updatedAt = datetime(),
            k.updatedBy = $createdBy
        RETURN k.name AS name
        """
        point_type = str(point.get("type", DEFAULT_POINT_TYPE) or DEFAULT_POINT_TYPE).strip().lower()
        params = {
            "name": name,
            "type": point_type,
            "difficulty": point.get("difficulty", DEFAULT_POINT_DIFFICULTY),
            "importance": point.get("importance", DEFAULT_POINT_IMPORTANCE),
            "summary": point.get("summary", ""),
            "description": point.get("description", ""),
            "chapter": point.get("chapter", ""),
            "nodeType": label,
            "createdBy": created_by,
        }

        tx.run(query, params)

    def _create_has_topic_relation(self, tx, stage_name: str, point_name: str, target_label: str, created_by: str):
        """创建HAS_TOPIC关系（在事务中执行）"""
        query = """
        MATCH (s:Stage {name: $stage_name})
        MATCH (k {name: $point_name}) WHERE $target_label IN labels(k)
        MERGE (s)-[r:HAS_TOPIC]->(k)
        ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
        RETURN s.name AS stage, k.name AS point
        """
        params = {
            "stage_name": stage_name,
            "point_name": point_name,
            "target_label": target_label,
            "createdBy": created_by,
        }

        tx.run(query, params)

    def _create_include_point_relation(
        self, tx, topic_name: str, point_name: str, stage_name: str, created_by: str
    ):
        """创建 Topic->Point 的 INCLUDE_POINT 关系并确保 Stage 关联"""
        query = """
        MATCH (t:Topic {name: $topic_name, stage: $stage_name})
        MATCH (k {name: $point_name})
        MERGE (t)-[r:INCLUDE_POINT]->(k)
        ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
        WITH t
        MATCH (s:Stage {name: $stage_name})
        MERGE (s)-[:CONTAIN_TOPIC]->(t)
        """
        params = {
            "topic_name": topic_name,
            "point_name": point_name,
            "stage_name": stage_name,
            "createdBy": created_by,
        }
        tx.run(query, params)

    def _create_practice(self, tx, practice: Dict, created_by: str) -> str:
        """创建Practice节点（在事务中执行）"""
        import hashlib
        practice_id = hashlib.md5(
            f"{practice.get('title', '')}{practice.get('content', '')}".encode()
        ).hexdigest()[:12]

        query = """
        CREATE (p:Practice {
            id: $id,
            title: $title,
            scenario: $scenario,
            content: $content,
            createdAt: datetime(),
            createdBy: $createdBy
        })
        RETURN p.id AS id
        """
        params = {
            "id": practice_id,
            "title": practice.get("title", ""),
            "scenario": practice.get("scenario", ""),
            "content": practice.get("content", ""),
            "createdBy": created_by,
        }

        tx.run(query, params)

        return practice_id

    def _create_has_practice_relation(self, tx, point_name: str, practice_id: str, created_by: str):
        """创建HAS_PRACTICE关系（在事务中执行）"""
        query = """
        MATCH (k:KnowledgePoint {name: $point_name})
        MATCH (p:Practice {id: $practice_id})
        MERGE (k)-[r:HAS_PRACTICE]->(p)
        ON CREATE SET r.createdAt = datetime(), r.createdBy = $createdBy
        RETURN k.name AS point, p.id AS practice
        """
        params = {
            "point_name": point_name,
            "practice_id": practice_id,
            "createdBy": created_by,
        }

        tx.run(query, params)
