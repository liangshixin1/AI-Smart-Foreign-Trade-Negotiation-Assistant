"""Neo4j integration layer for the knowledge graph features."""

from __future__ import annotations

import copy
import csv
import importlib.util
import io
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from neo4j import GraphDatabase, basic_auth
from neo4j.exceptions import IncompleteCommit, Neo4jError, ServiceUnavailable
from openpyxl import Workbook, load_workbook

import database


LOGGER = logging.getLogger(__name__)


class GraphUnavailableError(RuntimeError):
    """Raised when the Neo4j backend is not ready."""


class GraphEntityNotFoundError(RuntimeError):
    """Raised when a requested graph entity cannot be located."""


_DRIVER = None
_GRAPH_DISABLED = False
_GRAPH_DISABLED_REASON = ""
_DEFAULT_CATEGORIES_CACHE: Optional[Sequence[Dict[str, object]]] = None


INITIALIZATION_SETTING_KEY = "knowledge_graph.initialization"


def _default_categories() -> Sequence[Dict[str, object]]:
    """Load the recommended category hierarchy from the migration preset."""

    global _DEFAULT_CATEGORIES_CACHE
    if _DEFAULT_CATEGORIES_CACHE is not None:
        return _DEFAULT_CATEGORIES_CACHE

    module_path = (
        Path(__file__).resolve().parent.parent / "migrations" / "001_enhance_knowledge_graph.py"
    )
    categories: Sequence[Dict[str, object]] = ()
    if module_path.exists():
        try:
            spec = importlib.util.spec_from_file_location(
                "migrations.enhance_knowledge_graph", str(module_path)
            )
            if spec and spec.loader:
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                categories = getattr(module, "DEFAULT_CATEGORIES", ())
        except Exception as exc:  # pragma: no cover - defensive logging
            LOGGER.warning("Failed to load default categories: %s", exc)
            categories = ()
    else:
        LOGGER.debug("Default category preset file not found at %s", module_path)

    _DEFAULT_CATEGORIES_CACHE = tuple(categories)
    return _DEFAULT_CATEGORIES_CACHE


def _neo4j_credentials() -> Tuple[str, str, str]:
    uri = os.getenv("NEO4J_URI")
    if not uri:
        raise GraphUnavailableError("NEO4J_URI environment variable is not configured")
    user = os.getenv("NEO4J_USER", "neo4j")
    password = os.getenv("NEO4J_PASSWORD", "")
    return uri, user, password


def _disable_graph(reason: str) -> None:
    """Disable knowledge graph operations after a fatal connectivity issue."""

    global _GRAPH_DISABLED, _GRAPH_DISABLED_REASON
    close_driver()
    _GRAPH_DISABLED = True
    _GRAPH_DISABLED_REASON = reason


def _get_driver():
    global _DRIVER, _GRAPH_DISABLED, _GRAPH_DISABLED_REASON
    if _GRAPH_DISABLED:
        raise GraphUnavailableError(_GRAPH_DISABLED_REASON or "Knowledge graph connectivity disabled")

    if _DRIVER is not None:
        return _DRIVER

    uri, user, password = _neo4j_credentials()
    try:
        driver = GraphDatabase.driver(uri, auth=basic_auth(user, password))
        driver.verify_connectivity()
    except (Neo4jError, ServiceUnavailable, OSError) as exc:  # pragma: no cover - network
        LOGGER.error("Failed to connect to Neo4j at %s: %s", uri, exc)
        _disable_graph(f"Failed to connect to Neo4j at {uri}: {exc}")
        raise GraphUnavailableError("Failed to connect to Neo4j") from exc

    _DRIVER = driver
    _GRAPH_DISABLED = False
    _GRAPH_DISABLED_REASON = ""
    return driver


def is_configured() -> bool:
    """Return True if Neo4j connection information is available."""

    return bool(os.getenv("NEO4J_URI"))


def graph_status() -> Dict[str, object]:
    """Return diagnostic information about the knowledge graph backend."""

    return {
        "configured": is_configured(),
        "available": is_configured() and not _GRAPH_DISABLED,
        "message": _GRAPH_DISABLED_REASON,
    }


def get_initialization_status() -> Dict[str, object]:
    """Return the persisted initialization state for the knowledge graph."""

    record = database.get_app_setting(INITIALIZATION_SETTING_KEY, {})
    initialized = bool(record.get("initialized")) if isinstance(record, dict) else False
    option = record.get("option") if isinstance(record, dict) else None
    completed_at = record.get("completedAt") if isinstance(record, dict) else None
    return {
        "initialized": initialized,
        "option": option,
        "completedAt": completed_at,
        "graph": graph_status(),
    }


def _set_initialization_status(option: str) -> Dict[str, object]:
    payload = {
        "initialized": True,
        "option": option,
        "completedAt": datetime.utcnow().isoformat() + "Z",
    }
    database.set_app_setting(INITIALIZATION_SETTING_KEY, payload)
    payload["graph"] = graph_status()
    return payload


def reset_initialization_status() -> None:
    """Clear the initialization status flag."""

    database.delete_app_setting(INITIALIZATION_SETTING_KEY)


def get_initialization_defaults_preview() -> Dict[str, object]:
    """Return the recommended category tree and practice suggestions."""

    categories = copy.deepcopy(list(_default_categories()))
    knowledge_presets: List[Dict[str, object]] = []
    for practice_id, names in SECTION_KNOWLEDGE_PRESETS.items():
        section = database.get_section(practice_id)
        knowledge_presets.append(
            {
                "practiceId": practice_id,
                "practiceTitle": section.get("title") if section else "",
                "chapterId": section.get("chapter_id") if section else "",
                "count": len(names),
                "names": list(names),
            }
        )

    return {
        "categories": categories,
        "knowledgePresets": knowledge_presets,
    }


def close_driver() -> None:
    """Dispose of the cached Neo4j driver."""

    global _DRIVER, _GRAPH_DISABLED, _GRAPH_DISABLED_REASON
    if _DRIVER is not None:
        _DRIVER.close()
        _DRIVER = None
    _GRAPH_DISABLED = False
    _GRAPH_DISABLED_REASON = ""


class GraphService:
    """Wrapper class for Neo4j graph service operations."""

    @property
    def driver(self):
        """Get the Neo4j driver instance."""
        return _get_driver()


def _fallback_practice_detail(practice_id: str) -> Dict[str, object]:
    practice = database.get_section(practice_id)
    if not practice:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")
    return {
        "id": practice.get("id"),
        "title": practice.get("title"),
        "description": practice.get("description"),
        "orderIndex": practice.get("order_index"),
        "chapterId": practice.get("chapter_id"),
        "knowledgePoints": [],
    }


def _fallback_lesson_detail(lesson_id: str) -> Dict[str, object]:
    lesson = database.get_theory_lesson(lesson_id, include_unpublished=True)
    if not lesson:
        raise GraphEntityNotFoundError(f"Theory lesson {lesson_id} not found")
    return {
        "id": lesson.get("id"),
        "title": lesson.get("title"),
        "code": lesson.get("code"),
        "topicId": lesson.get("topicId"),
        "knowledgePoints": [],
    }


def _execute_read(query: str, parameters: Optional[Dict[str, object]] = None) -> List[Dict[str, object]]:
    driver = _get_driver()
    with driver.session() as session:
        result = session.run(query, parameters or {})
        return [record.data() for record in result]


def _execute_write(query: str, parameters: Optional[Dict[str, object]] = None) -> None:
    driver = _get_driver()
    with driver.session() as session:
        session.run(query, parameters or {})


def ensure_indexes() -> None:
    """Create the uniqueness constraints required by the knowledge graph."""

    driver = _get_driver()
    statements = [
        "CREATE CONSTRAINT chapter_id IF NOT EXISTS FOR (c:Chapter) REQUIRE c.id IS UNIQUE",
        "CREATE CONSTRAINT practice_id IF NOT EXISTS FOR (p:Practice) REQUIRE p.id IS UNIQUE",
        "CREATE CONSTRAINT theory_topic_id IF NOT EXISTS FOR (t:TheoryTopic) REQUIRE t.id IS UNIQUE",
        "CREATE CONSTRAINT lesson_id IF NOT EXISTS FOR (l:TheoryLesson) REQUIRE l.id IS UNIQUE",
        "CREATE CONSTRAINT knowledge_point_name IF NOT EXISTS FOR (k:KnowledgePoint) REQUIRE k.name IS UNIQUE",
        "CREATE CONSTRAINT process_step_id IF NOT EXISTS FOR (s:ProcessStep) REQUIRE s.id IS UNIQUE",
    ]

    with driver.session() as session:
        for statement in statements:
            session.run(statement)


@dataclass(frozen=True)
class ProcessStep:
    """Metadata for the canonical negotiation process stages."""

    identifier: str
    name: str
    order_index: int


PROCESS_STEPS: Sequence[ProcessStep] = (
    ProcessStep("process-prologue", "课程导入", 0),
    ProcessStep("process-inquiry", "询盘", 1),
    ProcessStep("process-offer", "报盘", 2),
    ProcessStep("process-counter-offer", "还盘", 3),
    ProcessStep("process-acceptance", "接受与订货", 4),
    ProcessStep("process-logistics", "订舱与物流", 5),
    ProcessStep("process-payment", "付款与交货", 6),
    ProcessStep("process-inspection", "商检", 7),
    ProcessStep("process-risk", "保险与仲裁", 8),
    ProcessStep("process-complaint", "投诉处理", 9),
    ProcessStep("process-claim", "索赔与理赔", 10),
)


CHAPTER_PROCESS_MAPPING: Dict[str, str] = {
    "chapter-0": "process-prologue",
    "chapter-1": "process-inquiry",
    "chapter-2": "process-offer",
    "chapter-3": "process-counter-offer",
    "chapter-4": "process-acceptance",
    "chapter-5": "process-logistics",
    "chapter-6": "process-payment",
    "chapter-7": "process-inspection",
    "chapter-8": "process-risk",
    "chapter-9": "process-complaint",
    "chapter-10": "process-claim",
}


SECTION_KNOWLEDGE_PRESETS: Dict[str, Sequence[str]] = {
    "chapter-0-section-1": (
        "首轮报价锚点策略",
        "FOB 成本构成",
        "条件式让步技巧",
    ),
    "chapter-1-section-1": (
        "询盘结构要素",
        "产品规格描述",
        "专业邮件礼仪",
    ),
    "chapter-1-section-2": (
        "需求澄清提问",
        "跟进邮件结构",
        "跨文化沟通语气",
    ),
    "chapter-2-section-1": (
        "报盘组成项目",
        "价格梯度设计",
        "价值陈述技巧",
    ),
    "chapter-2-section-2": (
        "议价让步逻辑",
        "锚定与反锚定",
        "条件交换策略",
    ),
    "chapter-3-section-1": (
        "还盘框架设计",
        "底线管理",
        "谈判让步幅度",
    ),
    "chapter-4-section-1": (
        "报价单审阅要点",
        "价格条款辨识",
        "折扣条件分析",
    ),
    "chapter-4-section-2": (
        "形式发票结构",
        "支付条款匹配",
        "风险提示编写",
    ),
    "chapter-4-section-3": (
        "发盘流程",
        "条款协同",
        "谈判节奏控制",
    ),
    "chapter-4-section-4": (
        "国际支付工具",
        "信用证条款",
        "风险保障策略",
    ),
    "chapter-4-section-5": (
        "接受函撰写",
        "订单确认流程",
        "交付节点管理",
    ),
    "chapter-4-section-6": (
        "订单履行跟进",
        "客户关系维护",
        "跨部门协同",
    ),
    "chapter-5-section-1": (
        "出口包装标准",
        "唛头设计",
        "货物防护技巧",
    ),
    "chapter-5-section-2": (
        "运输方式比较",
        "运价谈判策略",
        "时效与成本平衡",
    ),
    "chapter-5-section-3": (
        "装运时间表",
        "港口协调",
        "突发事件预案",
    ),
    "chapter-5-section-4": (
        "全链路协调",
        "物流关键路径",
        "多方沟通机制",
    ),
    "chapter-6-section-1": (
        "信用证审证流程",
        "单证一致性",
        "改证谈判技巧",
    ),
    "chapter-6-section-2": (
        "托收风险识别",
        "银行交涉话术",
        "账款保障措施",
    ),
    "chapter-6-section-3": (
        "电汇操作流程",
        "收汇时间规划",
        "资金安全控制",
    ),
    "chapter-6-section-4": (
        "分批交货条款",
        "交货时间管理",
        "库存与产能匹配",
    ),
    "chapter-6-section-5": (
        "Incoterms 责任划分",
        "风险转移节点",
        "费用承担分析",
    ),
    "chapter-6-section-6": (
        "高风险客户评估",
        "付款条件博弈",
        "风险缓释方案",
    ),
    "chapter-7-section-1": (
        "检验证书类型",
        "法定与商检机构",
        "合规要求识别",
    ),
    "chapter-7-section-2": (
        "检验条款撰写",
        "合同风险控制",
        "检验责任划分",
    ),
    "chapter-7-section-3": (
        "检验谈判策略",
        "异议处理流程",
        "质量保证设计",
    ),
    "chapter-7-section-4": (
        "不合格争议应对",
        "证据链构建",
        "解决方案谈判",
    ),
    "chapter-8-section-1": (
        "货运保险类别",
        "保险金额计算",
        "风险敞口评估",
    ),
    "chapter-8-section-2": (
        "仲裁条款要素",
        "争议解决流程",
        "适用法律选择",
    ),
    "chapter-8-section-3": (
        "海运险理赔流程",
        "索赔资料准备",
        "损失证据整理",
    ),
    "chapter-8-section-4": (
        "仲裁申请准备",
        "律师沟通要点",
        "庭审应对策略",
    ),
    "chapter-9-section-1": (
        "投诉处理流程",
        "客户情绪管理",
        "补救方案设计",
    ),
    "chapter-9-section-2": (
        "投诉到争议升级",
        "内部协调机制",
        "品牌声誉保护",
    ),
    "chapter-10-section-1": (
        "索赔函结构",
        "损失计算方法",
        "法律条款引用",
    ),
    "chapter-10-section-2": (
        "理赔审核要点",
        "调解与谈判技巧",
        "赔偿决策机制",
    ),
}


LESSON_KNOWLEDGE_PRESETS: Dict[str, Sequence[str]] = {}


def bootstrap_graph() -> None:
    """Initialise constraints and ingest the static content hierarchy."""

    if not is_configured():
        LOGGER.warning("Neo4j connection is not configured; knowledge graph features disabled")
        return

    try:
        ensure_indexes()
        status = get_initialization_status()
        if status.get("initialized"):
            sync_static_content()
        else:
            LOGGER.info(
                "Knowledge graph initialization deferred; waiting for teacher confirmation"
            )
    except GraphUnavailableError as exc:  # pragma: no cover - depends on external service
        LOGGER.warning("Unable to bootstrap knowledge graph: %s", exc)
    except (Neo4jError, ServiceUnavailable, IncompleteCommit, OSError, TimeoutError) as exc:
        LOGGER.warning("Knowledge graph bootstrap failed: %s", exc)
        _disable_graph(f"Knowledge graph bootstrap failed: {exc}")


def sync_static_content(*, include_recommendations: bool = False) -> None:
    """Mirror the SQLite content hierarchy into Neo4j."""

    driver = _get_driver()

    chapters = database.list_level_hierarchy(include_prompts=True)
    theory_hierarchy = database.list_theory_hierarchy(
        include_content=True, published_only=False
    )

    try:
        with driver.session() as session:
            session.execute_write(_merge_process_steps)
            for chapter in chapters:
                session.execute_write(_merge_chapter, chapter)
                session.execute_write(_link_chapter_process, chapter["id"])
                for section in chapter.get("sections", []):
                    session.execute_write(_merge_practice, chapter, section)

            for theory_chapter in theory_hierarchy:
                for topic in theory_chapter.get("topics", []):
                    if not topic.get("id"):
                        LOGGER.warning(
                            "Skipping theory topic with missing id in chapter %s",
                            theory_chapter.get("chapterId"),
                        )
                        continue
                    session.execute_write(_merge_theory_topic, topic)
                    for lesson in topic.get("lessons", []):
                        if not lesson.get("id"):
                            LOGGER.warning(
                                "Skipping theory lesson with missing id for topic %s",
                                topic.get("id"),
                            )
                            continue
                        session.execute_write(_merge_theory_lesson, topic, lesson)
    except (Neo4jError, ServiceUnavailable, IncompleteCommit, OSError, TimeoutError) as exc:
        LOGGER.error("Failed to synchronise static content with Neo4j: %s", exc)
        _disable_graph(f"Failed to synchronise static content: {exc}")
        raise GraphUnavailableError("Failed to synchronise static content") from exc


def _merge_default_categories_tx(tx, categories: Sequence[Dict[str, object]]) -> int:
    created = 0
    stack: List[Tuple[Optional[str], Dict[str, object]]] = [
        (None, copy.deepcopy(category)) for category in categories
    ]
    while stack:
        parent_id, node = stack.pop()
        node_id = node.get("id")
        if not node_id:
            continue
        properties = {
            "id": node_id,
            "name": node.get("name"),
            "code": node.get("code"),
            "level": node.get("level", 1),
            "orderIndex": node.get("orderIndex", 0),
            "icon": node.get("icon"),
            "color": node.get("color"),
            "description": node.get("description"),
            "isActive": bool(node.get("isActive", True)),
            "isSystemRecommended": True,
        }
        tx.run(
            "MERGE (c:KnowledgeCategory {id: $id}) "
            "SET c.name = $name, c.code = $code, c.level = $level, "
            "    c.orderIndex = $orderIndex, c.icon = $icon, c.color = $color, "
            "    c.description = $description, c.isActive = $isActive, "
            "    c.isSystemRecommended = $isSystemRecommended, "
            "    c.updatedAt = datetime(), "
            "    c.createdAt = coalesce(c.createdAt, datetime())",
            properties,
        )
        if parent_id:
            tx.run(
                "MATCH (parent:KnowledgeCategory {id: $parent_id}), "
                "      (child:KnowledgeCategory {id: $child_id}) "
                "MERGE (parent)-[:PARENT_OF]->(child)",
                {"parent_id": parent_id, "child_id": node_id},
            )
        children = node.get("children") or []
        for child in children:
            stack.append((node_id, child))
        created += 1
    return created


def _merge_default_knowledge_points_tx(tx, names: Sequence[str]) -> int:
    created = 0
    for name in names:
        if not name:
            continue
        result = tx.run(
            """
            MERGE (k:KnowledgePoint {name: $name})
            ON CREATE SET k.createdAt = datetime(),
                          k.practiceCount = coalesce(k.practiceCount, 0),
                          k.lessonCount = coalesce(k.lessonCount, 0),
                          k.category = coalesce(k.category, 'uncategorized')
            SET k.isSystemRecommended = true,
                k.source = coalesce(k.source, 'preset'),
                k.updatedAt = datetime()
            """,
            {"name": name},
        )
        summary = result.consume()
        if summary.counters.nodes_created:
            created += 1
    return created


def apply_default_recommendations() -> Dict[str, int]:
    """Create the recommended categories and knowledge point shells."""

    categories = list(_default_categories())
    driver = _get_driver()
    knowledge_names = sorted(
        {name for names in SECTION_KNOWLEDGE_PRESETS.values() for name in names}
    )

    with driver.session() as session:
        created_categories = session.execute_write(_merge_default_categories_tx, categories)
        created_knowledge = 0
        if knowledge_names:
            created_knowledge = session.execute_write(
                _merge_default_knowledge_points_tx, knowledge_names
            )

    return {"categories": created_categories, "knowledgePoints": created_knowledge}


def reset_knowledge_categories_to_default() -> Dict[str, int]:
    """Remove existing categories and rebuild the default recommended tree."""

    driver = _get_driver()
    categories = list(_default_categories())

    with driver.session() as session:
        session.run("MATCH (c:KnowledgeCategory) DETACH DELETE c")
        created_categories = session.execute_write(_merge_default_categories_tx, categories)

    return {"categories": created_categories}


def initialize_graph(
    option: str,
    *,
    initiated_by: str = "system",
    force: bool = False,
) -> Dict[str, object]:
    """Execute the teacher-driven initialization workflow."""

    normalized = (option or "").strip().lower()
    if normalized not in {"default", "blank", "import"}:
        raise ValueError("Unsupported initialization option")

    current = get_initialization_status()
    if current.get("initialized") and not force:
        raise ValueError("Knowledge graph has already been initialized")

    if not is_configured():
        raise GraphUnavailableError("Knowledge graph backend is not configured")

    ensure_indexes()
    sync_static_content(include_recommendations=False)

    summary = {"categories": 0, "knowledgePoints": 0}
    if normalized == "default":
        summary = apply_default_recommendations()

    status = _set_initialization_status(normalized)
    status["initiatedBy"] = initiated_by
    status["summary"] = summary
    return status


def _merge_process_steps(tx) -> None:
    for step in PROCESS_STEPS:
        tx.run(
            "MERGE (s:ProcessStep {id: $id}) SET s.name = $name, s.orderIndex = $order",
            {"id": step.identifier, "name": step.name, "order": step.order_index},
        )

    for previous, current in zip(PROCESS_STEPS, PROCESS_STEPS[1:]):
        tx.run(
            "MATCH (a:ProcessStep {id: $a}), (b:ProcessStep {id: $b}) "
            "MERGE (a)-[:NEXT_STEP]->(b)",
            {"a": previous.identifier, "b": current.identifier},
        )


def _merge_chapter(tx, chapter: Dict[str, object]) -> None:
    tx.run(
        "MERGE (c:Chapter {id: $id}) "
        "SET c.title = $title, c.description = $description, "
        "    c.orderIndex = $orderIndex, c.isDefault = $isDefault",
        {
            "id": chapter.get("id"),
            "title": chapter.get("title"),
            "description": chapter.get("description"),
            "orderIndex": chapter.get("orderIndex", 0),
            "isDefault": bool(chapter.get("isDefault")),
        },
    )


def _link_chapter_process(tx, chapter_id: str) -> None:
    process_id = CHAPTER_PROCESS_MAPPING.get(chapter_id)
    if not process_id:
        return
    tx.run(
        "MATCH (c:Chapter {id: $chapter_id}), (p:ProcessStep {id: $process_id}) "
        "MERGE (c)-[:COVERS_PROCESS]->(p)",
        {"chapter_id": chapter_id, "process_id": process_id},
    )


def _merge_practice(tx, chapter: Dict[str, object], section: Dict[str, object]) -> None:
    properties = {
        "id": section.get("id"),
        "title": section.get("title"),
        "description": section.get("description"),
        "environmentPromptTemplate": section.get("environmentPromptTemplate"),
        "environmentUserMessage": section.get("environmentUserMessage"),
        "conversationPromptTemplate": section.get("conversationPromptTemplate"),
        "evaluationPromptTemplate": section.get("evaluationPromptTemplate"),
        "expectsBargaining": bool(section.get("expectsBargaining")),
        "orderIndex": section.get("orderIndex", 0),
    }
    tx.run(
        "MERGE (p:Practice {id: $id}) "
        "SET p.title = $title, p.description = $description, p.orderIndex = $orderIndex, "
        "    p.environmentPromptTemplate = $environmentPromptTemplate, "
        "    p.environmentUserMessage = $environmentUserMessage, "
        "    p.conversationPromptTemplate = $conversationPromptTemplate, "
        "    p.evaluationPromptTemplate = $evaluationPromptTemplate, "
        "    p.expectsBargaining = $expectsBargaining",
        properties,
    )
    tx.run(
        "MATCH (c:Chapter {id: $chapter_id}), (p:Practice {id: $practice_id}) "
        "MERGE (c)-[:HAS_PRACTICE]->(p)",
        {"chapter_id": chapter.get("id"), "practice_id": section.get("id")},
    )


def ensure_practice_node(practice_id: str) -> Dict[str, object]:
    """Ensure a Practice node exists for the given section identifier."""

    section = database.get_section(practice_id)
    if not section:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")
    chapter = database.get_chapter(section["chapter_id"])
    if not chapter:
        raise GraphEntityNotFoundError(
            f"Chapter {section['chapter_id']} for practice {practice_id} not found"
        )

    chapter_payload = {
        "id": chapter.get("id"),
        "title": chapter.get("title"),
        "description": chapter.get("description"),
        "orderIndex": chapter.get("orderIndex", 0),
        "isDefault": bool(chapter.get("isDefault")),
    }
    section_payload = {
        "id": section.get("id"),
        "chapterId": section.get("chapter_id"),
        "title": section.get("title"),
        "description": section.get("description"),
        "environmentPromptTemplate": section.get("environment_prompt_template"),
        "environmentUserMessage": section.get("environment_user_message"),
        "conversationPromptTemplate": section.get("conversation_prompt_template"),
        "evaluationPromptTemplate": section.get("evaluation_prompt_template"),
        "expectsBargaining": bool(section.get("expects_bargaining")),
        "orderIndex": section.get("order_index", 0),
    }

    driver = _get_driver()
    with driver.session() as session:
        session.execute_write(_merge_chapter, chapter_payload)
        session.execute_write(_link_chapter_process, chapter_payload["id"])
        session.execute_write(_merge_practice, chapter_payload, section_payload)

    return {"chapter": chapter_payload, "practice": section_payload}


def _merge_theory_topic(tx, topic: Dict[str, object]) -> None:
    tx.run(
        "MERGE (t:TheoryTopic {id: $id}) "
        "SET t.title = $title, t.code = $code, t.summary = $summary, t.orderIndex = $orderIndex",
        {
            "id": topic.get("id"),
            "title": topic.get("title"),
            "code": topic.get("code"),
            "summary": topic.get("summary"),
            "orderIndex": topic.get("orderIndex", 0),
        },
    )
    chapter_id = topic.get("chapterId")
    if chapter_id:
        tx.run(
            "MATCH (c:Chapter {id: $chapter_id}), (t:TheoryTopic {id: $topic_id}) "
            "MERGE (c)-[:HAS_TOPIC]->(t)",
            {"chapter_id": chapter_id, "topic_id": topic.get("id")},
        )


def _merge_theory_lesson(tx, topic: Dict[str, object], lesson: Dict[str, object]) -> None:
    tx.run(
        "MERGE (l:TheoryLesson {id: $id}) "
        "SET l.title = $title, l.code = $code, l.orderIndex = $orderIndex, "
        "    l.isPublished = $isPublished, l.contentHtml = $contentHtml",
        {
            "id": lesson.get("id"),
            "title": lesson.get("title"),
            "code": lesson.get("code"),
            "orderIndex": lesson.get("orderIndex", 0),
            "isPublished": bool(lesson.get("isPublished")),
            "contentHtml": lesson.get("contentHtml"),
        },
    )
    tx.run(
        "MATCH (t:TheoryTopic {id: $topic_id}), (l:TheoryLesson {id: $lesson_id}) "
        "MERGE (t)-[:HAS_LESSON]->(l)",
        {"topic_id": topic.get("id"), "lesson_id": lesson.get("id")},
    )


def _ensure_practice_knowledge(tx, practice_id: str, points: List[str]) -> None:
    existing = tx.run(
        "MATCH (:Practice {id: $id})-[:TESTS]->(k:KnowledgePoint) RETURN collect(k.name) AS names",
        {"id": practice_id},
    ).single()
    if existing and existing["names"]:
        return
    _set_practice_knowledge_tx(tx, practice_id, points)


def _ensure_lesson_knowledge(tx, lesson_id: str, points: List[object]) -> None:
    existing = tx.run(
        (
            "MATCH (:TheoryLesson {id: $id})-[rel]->(k:KnowledgePoint) "
            "WHERE type(rel) = 'EXPLAINS' RETURN collect(k.name) AS names"
        ),
        {"id": lesson_id},
    ).single()
    if existing and existing["names"]:
        return
    normalized = _normalize_knowledge_point_payloads(points)
    if not normalized:
        return
    _set_lesson_knowledge_tx(tx, lesson_id, normalized)


def set_practice_knowledge_points(practice_id: str, points: Sequence[object]) -> None:
    ensure_practice_node(practice_id)
    driver = _get_driver()
    normalized_payloads = _normalize_knowledge_point_payloads(points)
    names = [payload["name"] for payload in normalized_payloads if payload.get("name")]
    with driver.session() as session:
        session.execute_write(_set_practice_knowledge_tx, practice_id, names)


def _set_practice_knowledge_tx(tx, practice_id: str, points: Sequence[str]) -> None:
    record = tx.run(
        "MATCH (p:Practice {id: $id}) RETURN p", {"id": practice_id}
    ).single()
    if not record:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")

    tx.run(
        "MATCH (:Practice {id: $id})-[rel:TESTS]->(:KnowledgePoint) DELETE rel",
        {"id": practice_id},
    )
    for name in points:
        tx.run(
            "MATCH (p:Practice {id: $id}) "
            "MERGE (k:KnowledgePoint {name: $name}) "
            "MERGE (p)-[:TESTS]->(k)",
            {"id": practice_id, "name": name},
        )


def get_practice_knowledge_recommendations(practice_id: str) -> Dict[str, object]:
    """Return existing and recommended knowledge points for a practice."""

    details = ensure_practice_node(practice_id)
    recommended = list(SECTION_KNOWLEDGE_PRESETS.get(practice_id, ()))
    existing_records = _execute_read(
        "MATCH (:Practice {id: $id})-[:TESTS]->(k:KnowledgePoint) "
        "RETURN collect(k.name) AS names",
        {"id": practice_id},
    )
    existing = []
    if existing_records:
        existing = sorted({name for name in existing_records[0].get("names", []) if name})

    return {
        "practice": {
            "id": practice_id,
            "title": details["practice"].get("title"),
            "chapterId": details["practice"].get("chapterId"),
        },
        "existing": existing,
        "recommended": recommended,
    }


def set_lesson_knowledge_points(lesson_id: str, points: Sequence[object]) -> None:
    driver = _get_driver()
    normalized = _normalize_knowledge_point_payloads(points)
    with driver.session() as session:
        session.execute_write(_set_lesson_knowledge_tx, lesson_id, normalized)


def _set_lesson_knowledge_tx(tx, lesson_id: str, points: Sequence[Dict[str, object]]) -> None:
    record = tx.run(
        "MATCH (l:TheoryLesson {id: $id}) RETURN l", {"id": lesson_id}
    ).single()
    if not record:
        raise GraphEntityNotFoundError(f"Theory lesson {lesson_id} not found")

    tx.run(
        (
            "MATCH (:TheoryLesson {id: $id})-[rel]->(:KnowledgePoint) "
            "WHERE type(rel) = 'EXPLAINS' DELETE rel"
        ),
        {"id": lesson_id},
    )
    for payload in points:
        name = payload.get("name") if isinstance(payload, dict) else None
        if not name:
            continue
        summary_value = payload.get("summary", "")
        if summary_value is None:
            summary_value = ""
        elif not isinstance(summary_value, str):
            summary_value = str(summary_value)
        body_html_value = payload.get("bodyHtml", "")
        if body_html_value is None:
            body_html_value = ""
        image_url_value = payload.get("imageUrl", "")
        if image_url_value is None:
            image_url_value = ""
        elif not isinstance(image_url_value, str):
            image_url_value = str(image_url_value)
        image_alt_value = payload.get("imageAlt", "")
        if image_alt_value is None:
            image_alt_value = ""
        elif not isinstance(image_alt_value, str):
            image_alt_value = str(image_alt_value)
        anchor_value = payload.get("anchorId", "")
        if anchor_value is None:
            anchor_value = ""
        elif not isinstance(anchor_value, str):
            anchor_value = str(anchor_value)
        tags_value = payload.get("tags", [])
        if isinstance(tags_value, (list, tuple)):
            cleaned_tags = []
            for tag in tags_value:
                if tag is None:
                    continue
                tag_str = str(tag).strip()
                if tag_str and tag_str not in cleaned_tags:
                    cleaned_tags.append(tag_str)
            tags_value = cleaned_tags
        else:
            tags_value = []
        knowledge_id_value = payload.get("knowledgeId", "")
        if knowledge_id_value is None:
            knowledge_id_value = ""
        elif not isinstance(knowledge_id_value, str):
            knowledge_id_value = str(knowledge_id_value)
        tx.run(
            "MATCH (l:TheoryLesson {id: $id}) "
            "MERGE (k:KnowledgePoint {name: $name}) "
            "SET k.summary = CASE WHEN $summary = '' THEN k.summary ELSE $summary END "
            "SET k.imageUrl = CASE WHEN $imageUrl = '' THEN k.imageUrl ELSE $imageUrl END "
            "SET k.imageAlt = CASE WHEN $imageAlt = '' THEN k.imageAlt ELSE $imageAlt END "
            "SET k.bodyHtml = CASE WHEN $bodyHtml = '' THEN k.bodyHtml ELSE $bodyHtml END "
            "SET k.sourceId = CASE WHEN $knowledgeId = '' THEN k.sourceId ELSE $knowledgeId END "
            "SET k.tags = CASE WHEN size($tags) = 0 THEN k.tags ELSE $tags END "
            "MERGE (l)-[rel:EXPLAINS]->(k) "
            "SET rel.anchorId = CASE WHEN $anchorId = '' THEN rel.anchorId ELSE $anchorId END "
            "SET rel.summary = CASE WHEN $summary = '' THEN rel.summary ELSE $summary END "
            "SET rel.bodyHtml = CASE WHEN $bodyHtml = '' THEN rel.bodyHtml ELSE $bodyHtml END "
            "SET rel.imageUrl = CASE WHEN $imageUrl = '' THEN rel.imageUrl ELSE $imageUrl END "
            "SET rel.imageAlt = CASE WHEN $imageAlt = '' THEN rel.imageAlt ELSE $imageAlt END "
            "SET rel.tags = CASE WHEN size($tags) = 0 THEN rel.tags ELSE $tags END",
            {
                "id": lesson_id,
                "name": name,
                "summary": summary_value,
                "bodyHtml": body_html_value,
                "imageUrl": image_url_value,
                "imageAlt": image_alt_value,
                "anchorId": anchor_value,
                "tags": tags_value,
                "knowledgeId": knowledge_id_value,
            },
        )


def _normalize_knowledge_point_payloads(points: Sequence[object]) -> List[Dict[str, object]]:
    """Normalize arbitrary knowledge point payloads into consistent objects."""

    normalized: List[Dict[str, object]] = []
    by_name: Dict[str, Dict[str, object]] = {}

    def _clean_string(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _merge_payload(target: Dict[str, object], source: Dict[str, object]) -> None:
        for key, raw_value in source.items():
            if key == "name":
                continue
            if raw_value is None:
                continue
            if key == "tags":
                existing = target.get("tags")
                merged: List[str] = []
                if isinstance(existing, list):
                    merged.extend(str(tag) for tag in existing if str(tag))
                for tag in raw_value:
                    tag_value = _clean_string(tag)
                    if not tag_value:
                        continue
                    if tag_value not in merged:
                        merged.append(tag_value)
                if merged:
                    target["tags"] = merged
                continue
            cleaned_value = raw_value
            if isinstance(raw_value, str):
                cleaned_value = raw_value.strip()
            target[key] = cleaned_value

    for entry in points:
        if isinstance(entry, str):
            name = entry.strip()
            payload: Dict[str, object] = {"name": name}
        elif isinstance(entry, dict):
            candidate = (
                entry.get("name")
                or entry.get("title")
                or entry.get("label")
                or entry.get("id")
            )
            name = _clean_string(candidate)
            payload = {"name": name}
            summary = _clean_string(entry.get("summary") or entry.get("description"))
            if summary:
                payload["summary"] = summary
            body_html = entry.get("bodyHtml") or entry.get("html") or entry.get("body")
            if isinstance(body_html, str) and body_html.strip():
                payload["bodyHtml"] = body_html
            image_url = _clean_string(entry.get("imageUrl") or entry.get("image") or entry.get("coverUrl"))
            if image_url:
                payload["imageUrl"] = image_url
            image_alt = _clean_string(entry.get("imageAlt") or entry.get("alt"))
            if image_alt:
                payload["imageAlt"] = image_alt
            anchor_id = _clean_string(entry.get("anchorId") or entry.get("anchor"))
            if anchor_id:
                payload["anchorId"] = anchor_id
            source_id = _clean_string(entry.get("knowledgeId") or entry.get("sourceId"))
            if source_id:
                payload["knowledgeId"] = source_id
            tags_field = entry.get("tags")
            tags: List[str] = []
            if isinstance(tags_field, (list, tuple)):
                tags = [_clean_string(tag) for tag in tags_field if _clean_string(tag)]
            elif isinstance(tags_field, str):
                tags = [tag.strip() for tag in tags_field.split(",") if tag.strip()]
            if tags:
                payload["tags"] = tags
        else:
            continue

        name = payload.get("name", "")
        if not isinstance(name, str):
            name = str(name)
        cleaned_name = name.strip()
        if not cleaned_name:
            continue
        payload["name"] = cleaned_name

        existing = by_name.get(cleaned_name)
        if existing:
            _merge_payload(existing, payload)
            continue

        by_name[cleaned_name] = payload
        normalized.append(payload)

    return normalized


def _normalize_category_path_input(value: Optional[object]) -> List[str]:
    """Normalize category path inputs from APIs or imports into a list of segments."""

    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        result = [str(item).strip() for item in value if str(item).strip()]
        return [segment for segment in result if segment]
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return []
        separators = re.compile(r"[>/\\|]")
        return [segment.strip() for segment in separators.split(cleaned) if segment.strip()]
    return [str(value).strip()] if str(value).strip() else []


def _category_path_to_string(path: Sequence[str]) -> str:
    return "/".join(segment.strip() for segment in path if segment)


def get_practice_detail(practice_id: str) -> Dict[str, object]:
    try:
        records = _execute_read(
            """
            MATCH (p:Practice {id: $id})
            OPTIONAL MATCH (p)-[:TESTS]->(k:KnowledgePoint)
            OPTIONAL MATCH (c:Chapter)-[:HAS_PRACTICE]->(p)
            RETURN p AS practice, collect(DISTINCT k.name) AS knowledge, c.id AS chapterId
            """,
            {"id": practice_id},
        )
    except GraphUnavailableError:
        LOGGER.debug("Graph unavailable when fetching practice %s; returning fallback", practice_id)
        return _fallback_practice_detail(practice_id)

    if not records:
        raise GraphEntityNotFoundError(f"Practice {practice_id} not found")

    record = records[0]
    practice = record["practice"]
    payload = {
        "id": practice.get("id"),
        "title": practice.get("title"),
        "description": practice.get("description"),
        "orderIndex": practice.get("orderIndex"),
        "chapterId": record.get("chapterId"),
        "knowledgePoints": sorted(filter(None, record.get("knowledge") or [])),
    }
    return payload


def get_lesson_detail(lesson_id: str) -> Dict[str, object]:
    try:
        records = _execute_read(
            """
            MATCH (l:TheoryLesson {id: $id})
            OPTIONAL MATCH (l)-[rel]->(k:KnowledgePoint)
            WHERE type(rel) = 'EXPLAINS'
            OPTIONAL MATCH (t:TheoryTopic)-[:HAS_LESSON]->(l)
            WITH l,
                 t,
                 k,
                 CASE WHEN rel IS NULL THEN {} ELSE properties(rel) END AS relProps,
                 CASE WHEN k IS NULL THEN {} ELSE properties(k) END AS kProps
            RETURN l AS lesson,
                   collect(DISTINCT CASE WHEN k IS NULL THEN NULL ELSE {
                     name: kProps['name'],
                     summary: coalesce(relProps['summary'], kProps['summary']),
                     bodyHtml: coalesce(relProps['bodyHtml'], kProps['bodyHtml']),
                     imageUrl: coalesce(relProps['imageUrl'], kProps['imageUrl']),
                     imageAlt: coalesce(relProps['imageAlt'], kProps['imageAlt']),
                     anchorId: relProps['anchorId'],
                     tags: relProps['tags'],
                     knowledgeId: kProps['sourceId']
                   } END) AS knowledge,
                   t.id AS topicId
            """,
            {"id": lesson_id},
        )
    except GraphUnavailableError:
        LOGGER.debug("Graph unavailable when fetching lesson %s; returning fallback", lesson_id)
        return _fallback_lesson_detail(lesson_id)

    if not records:
        raise GraphEntityNotFoundError(f"Theory lesson {lesson_id} not found")

    record = records[0]
    lesson = record["lesson"]
    raw_knowledge = [item for item in (record.get("knowledge") or []) if item]
    return {
        "id": lesson.get("id"),
        "title": lesson.get("title"),
        "code": lesson.get("code"),
        "topicId": record.get("topicId"),
        "knowledgePoints": _normalize_knowledge_point_payloads(raw_knowledge),
    }


def list_knowledge_points() -> List[Dict[str, object]]:
    return _execute_read(
        """
        MATCH (k:KnowledgePoint)
        OPTIONAL MATCH (k)<-[:TESTS]-(p:Practice)
        OPTIONAL MATCH (k)<-[rel]-(l:TheoryLesson)
        WHERE rel IS NULL OR type(rel) = 'EXPLAINS'
        RETURN k.name AS name,
               k.summary AS summary,
               k.bodyHtml AS bodyHtml,
               k.imageUrl AS imageUrl,
               k.imageAlt AS imageAlt,
               k.category AS category,
               k.categoryPath AS categoryPath,
               k.orderIndex AS orderIndex,
               k.sourceId AS knowledgeId,
               k.tags AS tags,
               count(DISTINCT p) AS practiceCount,
               count(DISTINCT l) AS lessonCount
        ORDER BY name
        """,
    )


def get_related_practices_for_lesson(lesson_id: str) -> List[Dict[str, object]]:
    try:
        return _execute_read(
            """
            MATCH (l:TheoryLesson {id: $id})-[rel]->(k:KnowledgePoint)<-[:TESTS]-(p:Practice)
            WHERE type(rel) = 'EXPLAINS'
            OPTIONAL MATCH (c:Chapter)-[:HAS_PRACTICE]->(p)
            RETURN DISTINCT p.id AS id, p.title AS title, p.description AS description,
                   p.orderIndex AS orderIndex, c.id AS chapterId
            ORDER BY p.orderIndex, p.title
            """,
            {"id": lesson_id},
        )
    except GraphUnavailableError:
        LOGGER.debug(
            "Graph unavailable when listing practices for lesson %s; returning empty list",
            lesson_id,
        )
        # Ensure the lesson exists before returning an empty payload.
        _fallback_lesson_detail(lesson_id)
        return []


def get_related_lessons_for_practice(practice_id: str) -> List[Dict[str, object]]:
    try:
        return _execute_read(
            """
            MATCH (p:Practice {id: $id})-[:TESTS]->(k:KnowledgePoint)<-[rel]-(l:TheoryLesson)
            WHERE type(rel) = 'EXPLAINS'
            OPTIONAL MATCH (t:TheoryTopic)-[:HAS_LESSON]->(l)
            RETURN DISTINCT l.id AS id, l.title AS title, l.code AS code, l.orderIndex AS orderIndex,
                   l.isPublished AS isPublished, t.id AS topicId
            ORDER BY l.orderIndex, l.title
            """,
            {"id": practice_id},
        )
    except GraphUnavailableError:
        LOGGER.debug(
            "Graph unavailable when listing lessons for practice %s; returning empty list",
            practice_id,
        )
        _fallback_practice_detail(practice_id)
        return []


def fetch_graph_snapshot(limit: int = 800) -> Dict[str, object]:
    allowed_labels = [
        "Stage",
        "Chapter",
        "Practice",
        "TheoryTopic",
        "TheoryLesson",
        "KnowledgePoint",
        "ProcessStep",
    ]
    nodes = _execute_read(
        """
        MATCH (n)
        WHERE any(label IN labels(n) WHERE label IN $allowed)
        OPTIONAL MATCH (n)<-[:HAS_TOPIC]-(stage:Stage)
        WITH labels(n) AS labels, n AS node, stage.name AS stageName,
             CASE
                 WHEN 'Stage' IN labels(n) THEN 0
                 WHEN 'ProcessStep' IN labels(n) THEN 1
                 WHEN 'Chapter' IN labels(n) THEN 2
                 WHEN 'Practice' IN labels(n) THEN 3
                 WHEN 'TheoryTopic' IN labels(n) THEN 4
                 WHEN 'TheoryLesson' IN labels(n) THEN 5
                 WHEN 'KnowledgePoint' IN labels(n) THEN 6
                 ELSE 99
             END AS priority
        WITH DISTINCT labels, node, stageName, priority
        ORDER BY priority ASC
        LIMIT $limit
        RETURN labels, node, stageName
        """,
        {"allowed": allowed_labels, "limit": limit},
    )

    edges = _execute_read(
        """
        MATCH (a)-[r]->(b)
        WHERE any(label IN labels(a) WHERE label IN $allowed)
          AND any(label IN labels(b) WHERE label IN $allowed)
        RETURN labels(a) AS sourceLabels, a AS source,
               labels(b) AS targetLabels, b AS target,
               type(r) AS type, r AS relationship
        LIMIT $limit
        """,
        {"allowed": allowed_labels, "limit": limit * 3},
    )

    node_payload: Dict[str, Dict[str, object]] = {}
    for record in nodes:
        labels = record.get("labels") or []
        node = record.get("node") or {}
        primary = _select_primary_label(labels)
        if not primary:
            continue
        identifier = _extract_node_identifier(primary, node)
        if not identifier:
            continue
        key = f"{primary}:{identifier}"

        # 计算层级和顺序
        level = _calculate_node_level(primary)
        stage_name = record.get("stageName") or node.get("stage")
        order = node.get("orderIndex", 0) or node.get("order", 0) or 0

        node_payload[key] = {
            "key": key,
            "label": primary,
            "title": node.get("title") or node.get("name") or node.get("code") or identifier,
            "subtitle": _build_node_subtitle(primary, node),
            "level": level,
            "stage": stage_name,
            "order": order,
            "group": primary,
        }

    edge_payload: List[Dict[str, object]] = []
    for record in edges:
        source_primary = _select_primary_label(record.get("sourceLabels") or [])
        target_primary = _select_primary_label(record.get("targetLabels") or [])
        source_identifier = _extract_node_identifier(source_primary, record.get("source") or {})
        target_identifier = _extract_node_identifier(target_primary, record.get("target") or {})
        if not (source_primary and target_primary and source_identifier and target_identifier):
            continue
        source_key = f"{source_primary}:{source_identifier}"
        target_key = f"{target_primary}:{target_identifier}"
        if source_key not in node_payload or target_key not in node_payload:
            continue

        relationship = record.get("relationship")
        edge_type = record.get("type")

        # 转换 Neo4j Relationship 对象为字典
        rel_props = {}
        if relationship:
            try:
                # Neo4j Relationship 对象可以像字典一样访问
                if hasattr(relationship, 'items'):
                    rel_props = dict(relationship)
                elif hasattr(relationship, 'get'):
                    rel_props = {k: relationship.get(k) for k in getattr(relationship, 'keys', lambda: [])()}
            except (AttributeError, TypeError):
                pass

        edge_payload.append(
            {
                "source": source_key,
                "target": target_key,
                "type": edge_type,
                "label": _get_edge_label(edge_type, rel_props),
                "arrows": "to",
            }
        )

    return {"nodes": list(node_payload.values()), "edges": edge_payload}


def _select_primary_label(labels: Iterable[str]) -> Optional[str]:
    priority = [
        "Stage",
        "Chapter",
        "Practice",
        "TheoryTopic",
        "TheoryLesson",
        "KnowledgePoint",
        "ProcessStep",
    ]
    for label in priority:
        if label in labels:
            return label
    return next(iter(labels), None) if labels else None


def _extract_node_identifier(label: Optional[str], node: Dict[str, object]) -> Optional[str]:
    if not label:
        return None
    if label in {"Stage", "ProcessStep", "KnowledgePoint", "Skill", "Terminology"}:
        return node.get("name")
    return node.get("id") or node.get("code") or node.get("title") or node.get("name")


def _build_node_subtitle(label: str, node: Dict[str, object]) -> Optional[str]:
    if label == "Practice":
        return node.get("description")
    if label == "TheoryLesson":
        return node.get("code")
    if label == "KnowledgePoint":
        return None
    if label == "ProcessStep":
        return f"顺序：{node.get('orderIndex')}"
    return node.get("description")


def _calculate_node_level(label: str) -> int:
    """
    计算节点的层级（用于层级布局）
    Level 0: Stage（阶段：询盘、报盘、还盘等）
    Level 1: Chapter（章节）
    Level 2: TheoryTopic / Practice / TheoryLesson
    Level 3: KnowledgePoint（知识点）
    Level 4: ProcessStep（流程步骤）
    """
    level_map = {
        "Stage": 0,
        "Chapter": 1,
        "TheoryTopic": 2,
        "Practice": 2,
        "TheoryLesson": 2,
        "KnowledgePoint": 3,
        "ProcessStep": 4,
    }
    return level_map.get(label, 99)


def _get_edge_label(edge_type: Optional[str], relationship: Optional[Dict[str, object]]) -> Optional[str]:
    """
    获取关系边的标签（用于可视化）
    """
    # 知识点关系类型的中文映射
    if relationship:
        relation_type = relationship.get("relationType")
        if relation_type:
            type_labels = {
                "prerequisite": "前置",
                "similar": "相似",
                "contrast": "对比",
                "related": "相关",
            }
            label = type_labels.get(relation_type)
            if label:
                return label

    # 边类型的默认标签
    if edge_type:
        edge_labels = {
            "REQUIRES": "依赖",
            "RELATES_TO": "关联",
            "HAS_TOPIC": "包含",
            "TESTS": "考察",
            "EXPLAINS": "讲解",
            "HAS_PRACTICE": "练习",
            "NEXT": "下一步",
        }
        return edge_labels.get(edge_type)

    return None


# ========== 知识点管理增强功能 ==========


def list_knowledge_points_enhanced(
    search: Optional[str] = None,
    category: Optional[str] = None,
    difficulty: Optional[str] = None,
) -> List[Dict[str, object]]:
    """获取知识点列表，支持过滤搜索。"""

    # 构建查询条件
    where_clauses = []
    params = {}

    if search:
        where_clauses.append("(k.name CONTAINS $search OR k.description CONTAINS $search)")
        params["search"] = search

    if category:
        where_clauses.append("k.category = $category")
        params["category"] = category

    if difficulty:
        where_clauses.append("k.difficulty = $difficulty")
        params["difficulty"] = difficulty

    where_clause = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    query = f"""
        MATCH (k:KnowledgePoint)
        {where_clause}
        OPTIONAL MATCH (k)<-[:TESTS]-(p:Practice)
        OPTIONAL MATCH (k)<-[rel]-(l:TheoryLesson)
        WHERE rel IS NULL OR type(rel) = 'EXPLAINS'
        OPTIONAL MATCH (k)<-[:REQUIRES]-(dependent:KnowledgePoint)
        OPTIONAL MATCH (k)-[:REQUIRES]->(prereq:KnowledgePoint)
        OPTIONAL MATCH (k)-[r:RELATED_TO]-(related:KnowledgePoint)
        RETURN k.name AS name,
               k.description AS description,
               k.category AS category,
               k.categoryPath AS category_path,
               k.difficulty AS difficulty,
               k.importance AS importance,
               k.estimatedDuration AS estimated_duration,
               k.content AS content,
               k.orderIndex AS order_index,
               k.tags AS tags,
               k.isSystemRecommended AS isSystemRecommended,
               k.isArchived AS isArchived,
               count(DISTINCT p) AS practiceCount,
               count(DISTINCT l) AS lessonCount,
               collect(DISTINCT prereq.name) AS prerequisites,
               collect(DISTINCT related.name) AS relations
        ORDER BY name
    """

    return _execute_read(query, params)


def get_knowledge_point(name: str) -> Dict[str, object]:
    """获取单个知识点的详细信息。"""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint {name: $name})
        OPTIONAL MATCH (k)<-[:TESTS]-(p:Practice)
        OPTIONAL MATCH (k)<-[rel]-(l:TheoryLesson)
        WHERE rel IS NULL OR type(rel) = 'EXPLAINS'
        OPTIONAL MATCH (k)-[:REQUIRES]->(prereq:KnowledgePoint)
        OPTIONAL MATCH (k)-[:RELATED_TO]-(related:KnowledgePoint)
        RETURN k.name AS name,
               k.description AS description,
               k.category AS category,
               k.categoryPath AS category_path,
               k.difficulty AS difficulty,
               k.importance AS importance,
               k.estimatedDuration AS estimated_duration,
               k.content AS content,
               k.orderIndex AS order_index,
               k.tags AS tags,
               collect(DISTINCT p.id) AS practices,
               collect(DISTINCT l.id) AS lessons,
               collect(DISTINCT prereq.name) AS prerequisites,
               collect(DISTINCT related.name) AS relations
        """,
        {"name": name},
    )

    if not records:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    record = records[0]
    return {
        "name": record.get("name"),
        "description": record.get("description"),
        "category": record.get("category"),
        "category_path": [segment for segment in (record.get("category_path") or []) if segment],
        "difficulty": record.get("difficulty"),
        "importance": record.get("importance"),
        "estimated_duration": record.get("estimated_duration"),
        "content": record.get("content"),
        "order_index": record.get("order_index"),
        "tags": [tag for tag in (record.get("tags") or []) if tag],
        "practices": [p for p in (record.get("practices") or []) if p],
        "lessons": [l for l in (record.get("lessons") or []) if l],
        "prerequisites": [p for p in (record.get("prerequisites") or []) if p],
        "relations": [r for r in (record.get("relations") or []) if r],
    }


def create_knowledge_point(data: Dict[str, object]) -> Dict[str, object]:
    """创建新的知识点。"""

    name = data.get("name", "").strip()
    if not name:
        raise ValueError("Knowledge point name is required")

    # 检查是否已存在
    existing = _execute_read(
        "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
        {"name": name},
    )
    if existing:
        raise ValueError(f"Knowledge point '{name}' already exists")

    category_path = _normalize_category_path_input(
        data.get("category_path") or data.get("categoryPath") or data.get("category")
    )
    category_value = data.get("category") or (category_path[-1] if category_path else None)
    try:
        order_index = int(data.get("order_index") or data.get("orderIndex") or 0)
    except (ValueError, TypeError):
        order_index = 0

    # 创建节点
    _execute_write(
        """
        CREATE (k:KnowledgePoint {
            name: $name,
            description: $description,
            category: $category,
            categoryPath: $category_path,
            difficulty: $difficulty,
            importance: $importance,
            estimatedDuration: $estimated_duration,
            orderIndex: $order_index,
            content: $content,
            tags: $tags
        })
        """,
        {
            "name": name,
            "description": data.get("description"),
            "category": category_value,
            "category_path": category_path,
            "difficulty": data.get("difficulty", "beginner"),
            "importance": data.get("importance", "medium"),
            "estimated_duration": data.get("estimated_duration"),
            "order_index": order_index,
            "content": data.get("content"),
            "tags": data.get("tags", []),
        },
    )

    return get_knowledge_point(name)


def update_knowledge_point(name: str, data: Dict[str, object]) -> Dict[str, object]:
    """更新知识点信息，支持部分字段更新。"""

    # 检查是否存在
    existing = _execute_read(
        "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
        {"name": name},
    )
    if not existing:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    node = existing[0]["k"]
    new_name = (data.get("name") or name).strip()

    # 如果改名，检查新名称是否冲突
    if new_name != name:
        conflict = _execute_read(
            "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
            {"name": new_name},
        )
        if conflict:
            raise ValueError(f"Knowledge point '{new_name}' already exists")

    category_field_provided = any(
        key in data for key in ("category_path", "categoryPath", "category")
    )
    if category_field_provided:
        category_path = _normalize_category_path_input(
            data.get("category_path") or data.get("categoryPath") or data.get("category")
        )
        category_value = data.get("category") or (category_path[-1] if category_path else None)
    else:
        category_path = [segment for segment in (node.get("categoryPath") or []) if segment]
        category_value = node.get("category")

    if "order_index" in data or "orderIndex" in data:
        try:
            order_index = int(data.get("order_index") or data.get("orderIndex") or 0)
        except (ValueError, TypeError):
            order_index = node.get("orderIndex") or 0
    else:
        order_index = node.get("orderIndex") or 0

    description = data.get("description") if "description" in data else node.get("description")
    difficulty = data.get("difficulty") if "difficulty" in data else node.get("difficulty")
    if difficulty is None:
        difficulty = "beginner"

    importance = data.get("importance") if "importance" in data else node.get("importance")
    if importance is None:
        importance = "medium"

    estimated_duration = (
        data.get("estimated_duration") if "estimated_duration" in data else node.get("estimatedDuration")
    )
    content = data.get("content") if "content" in data else node.get("content")

    if "tags" in data:
        tags_field = data.get("tags")
        if isinstance(tags_field, str):
            tags = [tag.strip() for tag in tags_field.split(",") if tag.strip()]
        elif isinstance(tags_field, (list, tuple)):
            tags = [str(tag).strip() for tag in tags_field if str(tag).strip()]
        else:
            tags = []
    else:
        tags = [tag for tag in (node.get("tags") or []) if tag]

    # 更新节点
    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $old_name})
        SET k.name = $name,
            k.description = $description,
            k.category = $category,
            k.categoryPath = $category_path,
            k.difficulty = $difficulty,
            k.importance = $importance,
            k.estimatedDuration = $estimated_duration,
            k.orderIndex = $order_index,
            k.content = $content,
            k.tags = $tags
        """,
        {
            "old_name": name,
            "name": new_name,
            "description": description,
            "category": category_value,
            "category_path": category_path,
            "difficulty": difficulty,
            "importance": importance,
            "estimated_duration": estimated_duration,
            "order_index": order_index,
            "content": content,
            "tags": tags,
        },
    )

    return get_knowledge_point(new_name)


def update_knowledge_point_category(
    name: str,
    category_path_value: Optional[object],
    *,
    order_index: Optional[object] = None,
) -> Dict[str, object]:
    """快速更新知识点的分类与排序信息。"""

    # 确保知识点存在
    existing = _execute_read("MATCH (k:KnowledgePoint {name: $name}) RETURN k", {"name": name})
    if not existing:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    category_path = _normalize_category_path_input(category_path_value)
    category_value = category_path[-1] if category_path else None
    resolved_order: Optional[int]
    if order_index is None:
        resolved_order = None
    else:
        try:
            resolved_order = int(order_index)
        except (TypeError, ValueError):
            resolved_order = None

    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})
        SET k.category = $category,
            k.categoryPath = $category_path,
            k.orderIndex = COALESCE($order_index, k.orderIndex)
        """,
        {
            "name": name,
            "category": category_value,
            "category_path": category_path,
            "order_index": resolved_order,
        },
    )

    return get_knowledge_point(name)


def get_knowledge_management_overview() -> Dict[str, object]:
    """聚合知识点、分类树和知识卡索引，供前端统一加载。"""

    status = get_initialization_status()
    # 直接使用 Neo4j 为主数据源
    raw_points = list_knowledge_points_enhanced()
    card_by_name = {point.get("name"): point for point in raw_points if point.get("name")}

    # 如果已有数据，则认为图谱已完成初始化；否则返回默认预设
    if raw_points:
        if not status.get("initialized"):
            status["initialized"] = True
            status["option"] = status.get("option") or "import"
    else:
        return {
            "initialized": False,
            "initialization": status,
            "defaults": get_initialization_defaults_preview(),
        }

    overview_points: List[Dict[str, object]] = []
    category_paths: Dict[str, None] = {}
    difficulty_breakdown: Dict[str, int] = {}
    uncategorized_points: List[Dict[str, object]] = []
    metadata_suggestions: List[Dict[str, object]] = []
    uncategorized_count = 0
    unlinked_count = 0

    tree_root: Dict[str, object] = {
        "id": "__root__",
        "name": "全部",
        "path": [],
        "count": 0,
        "children": {},
        "knowledge": [],
    }

    def _ensure_child(parent: Dict[str, object], segment: str, path: Sequence[str]) -> Dict[str, object]:
        children = parent.setdefault("children", {})
        if segment not in children:
            key = _category_path_to_string(path)
            children[segment] = {
                "id": key or "未分类",
                "name": segment,
                "path": list(path),
                "count": 0,
                "children": {},
                "knowledge": [],
                "_order": len(children),
            }
        return children[segment]

    def _append_to_tree(path: Sequence[str], payload: Dict[str, object]) -> None:
        effective_path = list(path) if path else ["未分类"]
        current = tree_root
        current["count"] = current.get("count", 0) + 1
        for index, segment in enumerate(effective_path, start=1):
            sub_path = effective_path[:index]
            current = _ensure_child(current, segment, sub_path)
            current["count"] = current.get("count", 0) + 1
        current.setdefault("knowledge", []).append(payload)

    for point in raw_points:
        name = point.get("name")
        if not name:
            continue
        raw_category_path = point.get("category_path")
        category_path = _normalize_category_path_input(
            raw_category_path if raw_category_path is not None else point.get("category")
        )
        category_path_key = _category_path_to_string(category_path) or "未分类"
        category_paths[category_path_key] = None

        order_index = point.get("order_index")
        try:
            order_index_value = int(order_index) if order_index is not None else 0
        except (TypeError, ValueError):
            order_index_value = 0

        difficulty = point.get("difficulty") or "beginner"
        difficulty_breakdown[difficulty] = difficulty_breakdown.get(difficulty, 0) + 1

        tags = [tag for tag in (point.get("tags") or []) if tag]
        prerequisites = sorted({p for p in (point.get("prerequisites") or []) if p})
        relations = sorted({r for r in (point.get("relations") or []) if r})

        category_label = " / ".join(category_path) if category_path else "未分类"

        if not category_path:
            uncategorized_count += 1
            uncategorized_points.append(
                {
                    "name": name,
                    "practiceCount": point.get("practiceCount", 0),
                    "lessonCount": point.get("lessonCount", 0),
                    "isSystemRecommended": bool(point.get("isSystemRecommended")),
                }
            )

        overview = {
            "name": name,
            "description": point.get("description") or "",
            "category": category_path[-1] if category_path else None,
            "category_path": category_path,
            "category_path_key": category_path_key,
            "category_path_text": category_label,
            "difficulty": difficulty,
            "importance": point.get("importance"),
            "estimated_duration": point.get("estimated_duration"),
            "content": point.get("content") or "",
            "order_index": order_index_value,
            "tags": tags,
            "prerequisites": prerequisites,
            "relations": relations,
            "practiceCount": point.get("practiceCount", 0),
            "lessonCount": point.get("lessonCount", 0),
            "isSystemRecommended": bool(point.get("isSystemRecommended")),
            "isArchived": bool(point.get("isArchived")),
        }
        overview_points.append(overview)

        if not overview["practiceCount"] and not overview["lessonCount"]:
            unlinked_count += 1

        tree_payload = {
            "name": name,
            "difficulty": difficulty,
            "order_index": order_index_value,
            "tags": tags,
            "category_path_key": category_path_key,
        }
        _append_to_tree(category_path, tree_payload)

        card = card_by_name.get(name) or {}
        metadata_fields: Dict[str, object] = {}
        metadata_preview: Dict[str, object] = {}
        summary = card.get("summary")
        if not overview.get("description") and summary:
            metadata_fields["description"] = summary
            metadata_preview["description"] = summary
        card_tags = [tag for tag in (card.get("tags") or []) if tag]
        if not tags and card_tags:
            metadata_fields["tags"] = card_tags
            metadata_preview["tags"] = card_tags
        if metadata_fields:
            metadata_suggestions.append(
                {
                    "name": name,
                    "fields": metadata_fields,
                    "preview": metadata_preview,
                    "reason": "根据已存在的理论知识卡片信息自动补全",
                }
            )

    def _serialize_tree(node: Dict[str, object]) -> Dict[str, object]:
        children_map = node.get("children") or {}
        serialized_children = []
        for _, child in sorted(children_map.items(), key=lambda item: item[1].get("_order", 0)):
            child_payload = {
                "id": child.get("id"),
                "name": child.get("name"),
                "path": child.get("path", []),
                "count": child.get("count", 0),
                "knowledge": sorted(
                    child.get("knowledge", []),
                    key=lambda payload: (payload.get("order_index", 0), payload.get("name", "")),
                ),
                "children": [],
            }
            child_payload["children"] = _serialize_tree(child).get("children", [])
            serialized_children.append(child_payload)
        return {"children": serialized_children}

    tree_children = _serialize_tree(tree_root).get("children", [])

    category_options = sorted({path for path in category_paths.keys() if path}) or []
    if "未分类" not in category_options:
        category_options.append("未分类")

    stats = {
        "total": len(overview_points),
        "categories": len(category_options),
        "difficulty": difficulty_breakdown,
        "uncategorized": uncategorized_count,
        "unlinked": unlinked_count,
    }

    knowledge_cards: List[Dict[str, object]] = []
    for name, card in card_by_name.items():
        if not name:
            continue
        card_category_path = _normalize_category_path_input(
            card.get("categoryPath") if card.get("categoryPath") is not None else card.get("category")
        )
        knowledge_cards.append(
            {
                "name": name,
                "summary": card.get("summary") or "",
                "bodyHtml": card.get("bodyHtml") or "",
                "imageUrl": card.get("imageUrl") or "",
                "imageAlt": card.get("imageAlt") or "",
                "knowledgeId": card.get("knowledgeId") or "",
                "tags": [tag for tag in (card.get("tags") or []) if tag],
                "practiceCount": card.get("practiceCount", 0),
                "lessonCount": card.get("lessonCount", 0),
                "category_path": card_category_path,
            }
        )

    knowledge_cards.sort(key=lambda item: item.get("name", ""))

    overview_points.sort(key=lambda item: (item.get("order_index", 0), item.get("name", "")))

    smart_assist = {
        "uncategorized": sorted(
            uncategorized_points,
            key=lambda item: (item.get("practiceCount", 0) + item.get("lessonCount", 0), item.get("name", "")),
            reverse=True,
        ),
        "metadata_suggestions": metadata_suggestions,
    }

    category_paths_list = sorted(category_paths.keys())
    tree_children = list(tree_root.get("children", {}).values())

    return {
        "initialized": True,
        "initialization": status,
        "knowledge_points": overview_points,
        "tree": tree_children,
        "category_tree": tree_children,  # backward compatibility for UI
        "category_paths": category_paths_list,
        "uncategorized": uncategorized_points,
        "categoryOptions": category_options,
        "metadata": metadata_suggestions,
        "stats": stats,
        "knowledge_cards": knowledge_cards,
        "assist": smart_assist,
    }
def delete_knowledge_point(name: str) -> None:
    """删除知识点及其所有关系。"""

    # 检查是否存在
    existing = _execute_read(
        "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
        {"name": name},
    )
    if not existing:
        raise GraphEntityNotFoundError(f"Knowledge point '{name}' not found")

    # 删除节点及所有关系
    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})
        DETACH DELETE k
        """,
        {"name": name},
    )


def list_orphan_knowledge_points() -> List[Dict[str, object]]:
    """Return knowledge points that are not linked to practices or lessons."""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        WHERE NOT (k)<-[:TESTS]-(:Practice)
          AND NOT (k)<-[:EXPLAINS]-(:TheoryLesson)
        RETURN k.name AS name,
               k.description AS description,
               k.category AS category,
               k.updatedAt AS updatedAt,
               k.isSystemRecommended AS isSystemRecommended,
               k.source AS source,
               k.createdAt AS createdAt
        ORDER BY coalesce(k.updatedAt, k.createdAt) DESC, k.name
        """
    )
    return records


def cleanup_orphan_knowledge_points(
    names: Sequence[str], *, archive: bool = False
) -> Dict[str, object]:
    """Delete or archive orphaned knowledge points by name."""

    normalized = sorted({name for name in names if isinstance(name, str) and name.strip()})
    if not normalized:
        return {"affected": 0, "archived": archive}

    driver = _get_driver()
    with driver.session() as session:
        if archive:
            result = session.run(
                """
                MATCH (k:KnowledgePoint)
                WHERE k.name IN $names
                  AND NOT (k)<-[:TESTS]-(:Practice)
                  AND NOT (k)<-[:EXPLAINS]-(:TheoryLesson)
                SET k.isArchived = true,
                    k.archivedAt = datetime(),
                    k.updatedAt = datetime()
                RETURN count(k) AS count
                """,
                {"names": normalized},
            ).single()
        else:
            result = session.run(
                """
                MATCH (k:KnowledgePoint)
                WHERE k.name IN $names
                  AND NOT (k)<-[:TESTS]-(:Practice)
                  AND NOT (k)<-[:EXPLAINS]-(:TheoryLesson)
                DETACH DELETE k
                RETURN count(k) AS count
                """,
                {"names": normalized},
            ).single()
    affected = int(result["count"]) if result and result.get("count") is not None else 0
    return {"affected": affected, "archived": archive}


def add_knowledge_prerequisite(name: str, prerequisite_name: str) -> Dict[str, object]:
    """为知识点添加前置依赖关系。"""

    # 检查两个知识点是否存在
    for point_name in [name, prerequisite_name]:
        existing = _execute_read(
            "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
            {"name": point_name},
        )
        if not existing:
            raise GraphEntityNotFoundError(f"Knowledge point '{point_name}' not found")

    # 不能依赖自己
    if name == prerequisite_name:
        raise ValueError("A knowledge point cannot be a prerequisite of itself")

    # 创建REQUIRES关系
    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})
        MATCH (prereq:KnowledgePoint {name: $prerequisite})
        MERGE (k)-[:REQUIRES]->(prereq)
        """,
        {"name": name, "prerequisite": prerequisite_name},
    )

    return get_knowledge_point(name)


def remove_knowledge_prerequisite(name: str, prerequisite_name: str) -> Dict[str, object]:
    """移除知识点的前置依赖关系。"""

    _execute_write(
        """
        MATCH (k:KnowledgePoint {name: $name})-[r:REQUIRES]->(prereq:KnowledgePoint {name: $prerequisite})
        DELETE r
        """,
        {"name": name, "prerequisite": prerequisite_name},
    )

    return get_knowledge_point(name)


def add_knowledge_relation(
    name: str, related_name: str, relation_type: str = "RELATED_TO"
) -> Dict[str, object]:
    """为知识点添加关联关系。"""

    # 检查两个知识点是否存在
    for point_name in [name, related_name]:
        existing = _execute_read(
            "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
            {"name": point_name},
        )
        if not existing:
            raise GraphEntityNotFoundError(f"Knowledge point '{point_name}' not found")

    # 不能关联自己
    if name == related_name:
        raise ValueError("A knowledge point cannot be related to itself")

    # 创建关联关系（双向）
    _execute_write(
        """
        MATCH (k1:KnowledgePoint {name: $name})
        MATCH (k2:KnowledgePoint {name: $related})
        MERGE (k1)-[:RELATED_TO]-(k2)
        """,
        {"name": name, "related": related_name},
    )

    return get_knowledge_point(name)


def remove_knowledge_relation(name: str, related_name: str) -> Dict[str, object]:
    """移除知识点的关联关系。"""

    _execute_write(
        """
        MATCH (k1:KnowledgePoint {name: $name})-[r:RELATED_TO]-(k2:KnowledgePoint {name: $related})
        DELETE r
        """,
        {"name": name, "related": related_name},
    )

    return get_knowledge_point(name)


def list_knowledge_categories() -> List[str]:
    """获取所有知识点分类列表。"""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        WHERE k.category IS NOT NULL AND k.category <> ''
        RETURN DISTINCT k.category AS category
        ORDER BY category
        """
    )

    return [record["category"] for record in records if record.get("category")]


def get_knowledge_categories_tree() -> List[Dict[str, object]]:
    """获取知识点分类树形结构（包含每个分类的知识点数量）。"""

    records = _execute_read(
        """
        MATCH (k:KnowledgePoint)
        WHERE k.category IS NOT NULL AND k.category <> ''
        WITH k.category AS category, count(k) AS count
        RETURN category, count
        ORDER BY category
        """
    )

    return [
        {"name": record["category"], "count": record["count"]}
        for record in records
    ]


# ========== Excel/CSV 导入导出功能 ==========


def export_knowledge_points_to_excel() -> io.BytesIO:
    """将所有知识点导出为Excel文件。"""

    points = list_knowledge_points_enhanced()

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "知识点"

    # 写入表头
    headers = [
        "名称",
        "描述",
        "分类",
        "分类路径",
        "难度",
        "重要性",
        "预计学习时长(分钟)",
        "标签(逗号分隔)",
        "内容",
        "前置依赖(逗号分隔)",
        "关联知识点(逗号分隔)",
    ]
    ws.append(headers)

    # 写入数据
    for point in points:
        tags = ", ".join(point.get("tags") or [])
        prerequisites = ", ".join(point.get("prerequisites") or [])
        relations = ", ".join(point.get("relations") or [])

        row = [
            point.get("name", ""),
            point.get("description", ""),
            point.get("category", ""),
            " / ".join(point.get("category_path") or []) if point.get("category_path") else "",
            point.get("difficulty", ""),
            point.get("importance", ""),
            point.get("estimated_duration", ""),
            tags,
            point.get("content", ""),
            prerequisites,
            relations,
        ]
        ws.append(row)

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_knowledge_points_to_csv() -> str:
    """将所有知识点导出为CSV字符串。"""

    points = list_knowledge_points_enhanced()

    output = io.StringIO()
    writer = csv.writer(output)

    # 写入表头
    headers = [
        "名称",
        "描述",
        "分类",
        "分类路径",
        "难度",
        "重要性",
        "预计学习时长(分钟)",
        "标签(逗号分隔)",
        "内容",
        "前置依赖(逗号分隔)",
        "关联知识点(逗号分隔)",
    ]
    writer.writerow(headers)

    # 写入数据
    for point in points:
        tags = ", ".join(point.get("tags") or [])
        prerequisites = ", ".join(point.get("prerequisites") or [])
        relations = ", ".join(point.get("relations") or [])

        row = [
            point.get("name", ""),
            point.get("description", ""),
            point.get("category", ""),
            " / ".join(point.get("category_path") or []) if point.get("category_path") else "",
            point.get("difficulty", ""),
            point.get("importance", ""),
            point.get("estimated_duration", ""),
            tags,
            point.get("content", ""),
            prerequisites,
            relations,
        ]
        writer.writerow(row)

    return output.getvalue()


def import_knowledge_points_from_excel(file_content: bytes) -> Dict[str, int]:
    """
    从Excel文件导入知识点。
    返回导入统计：{"created": 数量, "updated": 数量, "failed": 数量, "errors": [错误列表]}
    """

    stats = {"created": 0, "updated": 0, "failed": 0, "errors": []}

    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        # 查找列索引
        col_indices = {}
        for idx, header in enumerate(headers):
            if header == "名称":
                col_indices["name"] = idx
            elif header == "描述":
                col_indices["description"] = idx
            elif header == "分类":
                col_indices["category"] = idx
            elif header == "分类路径":
                col_indices["category_path"] = idx
            elif header == "难度":
                col_indices["difficulty"] = idx
            elif header == "重要性":
                col_indices["importance"] = idx
            elif header == "预计学习时长(分钟)":
                col_indices["estimated_duration"] = idx
            elif header == "标签(逗号分隔)":
                col_indices["tags"] = idx
            elif header == "内容":
                col_indices["content"] = idx
            elif header == "前置依赖(逗号分隔)":
                col_indices["prerequisites"] = idx
            elif header == "关联知识点(逗号分隔)":
                col_indices["relations"] = idx

        # 检查必需字段
        if "name" not in col_indices:
            stats["errors"].append("Excel文件缺少'名称'列")
            return stats

        # 处理每一行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = row[col_indices["name"]]
                if not name or not str(name).strip():
                    continue

                # 构建数据对象
                data = {
                    "name": str(name).strip(),
                    "description": str(row[col_indices.get("description")] or "").strip() or None,
                    "difficulty": str(row[col_indices.get("difficulty")] or "beginner").strip() or "beginner",
                    "importance": str(row[col_indices.get("importance")] or "medium").strip() or "medium",
                    "content": str(row[col_indices.get("content")] or "").strip() or None,
                }

                raw_category_value = row[col_indices.get("category")] if col_indices.get("category") is not None else None
                category_value = str(raw_category_value).strip() if raw_category_value else None
                raw_category_path = row[col_indices.get("category_path")] if col_indices.get("category_path") is not None else None
                category_path = _normalize_category_path_input(raw_category_path or category_value)
                if category_path:
                    data["category_path"] = category_path
                    data["category"] = category_value or category_path[-1]
                else:
                    data["category_path"] = []
                    data["category"] = category_value

                # 处理预计学习时长
                duration_value = row[col_indices.get("estimated_duration")]
                if duration_value:
                    try:
                        data["estimated_duration"] = int(duration_value)
                    except (ValueError, TypeError):
                        data["estimated_duration"] = None
                else:
                    data["estimated_duration"] = None

                # 处理标签
                tags_value = row[col_indices.get("tags")]
                if tags_value:
                    tags = [tag.strip() for tag in str(tags_value).split(",") if tag.strip()]
                    data["tags"] = tags
                else:
                    data["tags"] = []

                # 检查是否已存在
                existing = _execute_read(
                    "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
                    {"name": data["name"]},
                )

                if existing:
                    # 更新
                    update_knowledge_point(data["name"], data)
                    stats["updated"] += 1
                else:
                    # 创建
                    create_knowledge_point(data)
                    stats["created"] += 1

                # 处理前置依赖
                prerequisites_value = row[col_indices.get("prerequisites")]
                if prerequisites_value:
                    prereqs = [p.strip() for p in str(prerequisites_value).split(",") if p.strip()]
                    for prereq in prereqs:
                        try:
                            add_knowledge_prerequisite(data["name"], prereq)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加前置依赖'{prereq}'失败: {str(e)}"
                            )

                # 处理关联关系
                relations_value = row[col_indices.get("relations")]
                if relations_value:
                    relations = [r.strip() for r in str(relations_value).split(",") if r.strip()]
                    for relation in relations:
                        try:
                            add_knowledge_relation(data["name"], relation)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加关联'{relation}'失败: {str(e)}"
                            )

            except Exception as e:
                stats["failed"] += 1
                stats["errors"].append(f"行{row_idx}: {str(e)}")

    except Exception as e:
        stats["errors"].append(f"读取Excel文件失败: {str(e)}")

    return stats


def import_knowledge_points_from_csv(file_content: str) -> Dict[str, int]:
    """
    从CSV文件导入知识点。
    返回导入统计：{"created": 数量, "updated": 数量, "failed": 数量, "errors": [错误列表]}
    """

    stats = {"created": 0, "updated": 0, "failed": 0, "errors": []}

    try:
        reader = csv.DictReader(io.StringIO(file_content))

        for row_idx, row in enumerate(reader, start=2):
            try:
                name = row.get("名称", "").strip()
                if not name:
                    continue

                # 构建数据对象
                data = {
                    "name": name,
                    "description": row.get("描述", "").strip() or None,
                    "difficulty": row.get("难度", "beginner").strip() or "beginner",
                    "importance": row.get("重要性", "medium").strip() or "medium",
                    "content": row.get("内容", "").strip() or None,
                }

                category_value = row.get("分类", "").strip() or None
                category_path = _normalize_category_path_input(row.get("分类路径", "").strip() or category_value)
                if category_path:
                    data["category_path"] = category_path
                    data["category"] = category_value or category_path[-1]
                else:
                    data["category_path"] = []
                    data["category"] = category_value

                # 处理预计学习时长
                duration_value = row.get("预计学习时长(分钟)", "").strip()
                if duration_value:
                    try:
                        data["estimated_duration"] = int(duration_value)
                    except (ValueError, TypeError):
                        data["estimated_duration"] = None
                else:
                    data["estimated_duration"] = None

                # 处理标签
                tags_value = row.get("标签(逗号分隔)", "").strip()
                if tags_value:
                    tags = [tag.strip() for tag in tags_value.split(",") if tag.strip()]
                    data["tags"] = tags
                else:
                    data["tags"] = []

                # 检查是否已存在
                existing = _execute_read(
                    "MATCH (k:KnowledgePoint {name: $name}) RETURN k",
                    {"name": data["name"]},
                )

                if existing:
                    # 更新
                    update_knowledge_point(data["name"], data)
                    stats["updated"] += 1
                else:
                    # 创建
                    create_knowledge_point(data)
                    stats["created"] += 1

                # 处理前置依赖
                prerequisites_value = row.get("前置依赖(逗号分隔)", "").strip()
                if prerequisites_value:
                    prereqs = [p.strip() for p in prerequisites_value.split(",") if p.strip()]
                    for prereq in prereqs:
                        try:
                            add_knowledge_prerequisite(data["name"], prereq)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加前置依赖'{prereq}'失败: {str(e)}"
                            )

                # 处理关联关系
                relations_value = row.get("关联知识点(逗号分隔)", "").strip()
                if relations_value:
                    relations = [r.strip() for r in relations_value.split(",") if r.strip()]
                    for relation in relations:
                        try:
                            add_knowledge_relation(data["name"], relation)
                        except Exception as e:
                            stats["errors"].append(
                                f"行{row_idx}: 添加关联'{relation}'失败: {str(e)}"
                            )

            except Exception as e:
                stats["failed"] += 1
                stats["errors"].append(f"行{row_idx}: {str(e)}")

    except Exception as e:
        stats["errors"].append(f"读取CSV文件失败: {str(e)}")

    return stats


# ============================================
# 多节点类型架构支持 (Multi-Node Types Support)
# Migration 002 相关功能
# ============================================

def run_multi_node_types_migration(initiated_by: str = "system") -> Dict[str, object]:
    """
    运行多节点类型迁移 (Migration 002)

    创建 Stage, Skill, Terminology 等专用节点类型
    建立 PRECEDES 关系表达流程时序

    Args:
        initiated_by: 执行迁移的用户标识

    Returns:
        迁移统计信息
    """
    import importlib.util
    import sys
    from pathlib import Path

    # 动态加载迁移模块
    migration_path = Path(__file__).resolve().parent.parent / "migrations" / "002_multi_node_types.py"

    if not migration_path.exists():
        raise FileNotFoundError(f"Migration file not found: {migration_path}")

    spec = importlib.util.spec_from_file_location("migration_002", str(migration_path))
    if not spec or not spec.loader:
        raise RuntimeError("Failed to load migration module")

    migration_module = importlib.util.module_from_spec(spec)
    sys.modules["migration_002"] = migration_module
    spec.loader.exec_module(migration_module)

    # 执行迁移
    driver = _get_driver()
    stats = migration_module.upgrade(driver, initiated_by=initiated_by)

    LOGGER.info(f"Multi-node types migration completed: {stats}")
    return stats


def get_multi_node_types_migration_status() -> Dict[str, object]:
    """
    获取多节点类型迁移的状态

    Returns:
        {
            "applied": bool,
            "stages_count": int,
            "terminology_count": int,
            "precedes_count": int,
            ...
        }
    """
    import importlib.util
    import sys
    from pathlib import Path

    migration_path = Path(__file__).resolve().parent.parent / "migrations" / "002_multi_node_types.py"

    if not migration_path.exists():
        return {
            "applied": False,
            "error": "Migration file not found"
        }

    spec = importlib.util.spec_from_file_location("migration_002", str(migration_path))
    if not spec or not spec.loader:
        return {
            "applied": False,
            "error": "Failed to load migration module"
        }

    migration_module = importlib.util.module_from_spec(spec)
    sys.modules["migration_002"] = migration_module
    spec.loader.exec_module(migration_module)

    driver = _get_driver()
    return migration_module.get_migration_status(driver)


def get_process_flow() -> Dict[str, object]:
    """
    获取外贸谈判流程骨架

    返回所有 Stage 节点及其 PRECEDES 关系,用于前端渲染主流程轴

    Returns:
        {
            "stages": [
                {
                    "name": str,
                    "englishName": str,
                    "order": int,
                    "description": str,
                    "difficulty": str,
                    "icon": str,
                    "color": str,
                    ...
                }
            ],
            "flow": [
                {"from": str, "to": str, "description": str}
            ]
        }
    """
    # 获取所有 Stage 节点
    stages_query = """
    MATCH (s:Stage)
    OPTIONAL MATCH (s)-[:HAS_TOPIC]->(k)
    RETURN s {
        .*,
        topicsCount: count(DISTINCT k)
    } AS stage
    ORDER BY s.order
    """

    stages_records = _execute_read(stages_query, {})
    stages = [record["stage"] for record in stages_records] if stages_records else []

    # 获取 PRECEDES 关系
    flow_query = """
    MATCH (s1:Stage)-[r:PRECEDES]->(s2:Stage)
    RETURN s1.name AS from,
           s2.name AS to,
           r.description AS description,
           s1.order AS fromOrder,
           s2.order AS toOrder
    ORDER BY s1.order
    """

    flow_records = _execute_read(flow_query, {})
    flow = [
        {
            "from": record["from"],
            "to": record["to"],
            "description": record.get("description", ""),
            "fromOrder": record.get("fromOrder"),
            "toOrder": record.get("toOrder"),
        }
        for record in flow_records
    ] if flow_records else []

    return {
        "stages": stages,
        "flow": flow,
        "totalStages": len(stages),
        "totalFlows": len(flow),
    }


def list_stages(include_topics: bool = False) -> List[Dict[str, object]]:
    """
    获取所有 Stage (阶段) 节点列表

    Args:
        include_topics: 是否包含每个阶段下的知识点列表

    Returns:
        Stage 节点列表
    """
    if include_topics:
        query = """
        MATCH (s:Stage)
        OPTIONAL MATCH (s)-[:HAS_TOPIC]->(k)
        RETURN s {
            .*,
            topics: collect(DISTINCT k.name)
        } AS stage
        ORDER BY s.order
        """
    else:
        query = """
        MATCH (s:Stage)
        RETURN s AS stage
        ORDER BY s.order
        """

    records = _execute_read(query, {})
    return [record["stage"] for record in records] if records else []


def get_stage(name: str) -> Dict[str, object]:
    """
    获取单个 Stage 的详细信息

    Args:
        name: Stage 名称

    Returns:
        Stage 详细信息
    """
    query = """
    MATCH (s:Stage {name: $name})
    OPTIONAL MATCH (s)-[:HAS_TOPIC]->(k:KnowledgePoint)
    OPTIONAL MATCH (s)-[:PRECEDES]->(next:Stage)
    OPTIONAL MATCH (prev:Stage)-[:PRECEDES]->(s)
    WITH s, collect(DISTINCT k.name) AS topics, next, prev
    RETURN s {
        .*,
        topics: topics,
        nextStage: next.name,
        previousStage: prev.name
    } AS stage
    """

    records = _execute_read(query, {"name": name})
    if not records:
        raise GraphEntityNotFoundError(f"Stage '{name}' not found")

    return records[0]["stage"]


def list_terminology(category: Optional[str] = None) -> List[Dict[str, object]]:
    """
    获取术语列表

    Args:
        category: 可选的分类过滤 (如 "Incoterms", "Payment")

    Returns:
        Terminology 节点列表
    """
    if category:
        query = """
        MATCH (t:Terminology {category: $category})
        OPTIONAL MATCH (t)-[:RELATED_TO]-(related:Terminology)
        RETURN t {
            .*,
            relatedTerms: collect(DISTINCT related.name)
        } AS term
        ORDER BY t.name
        """
        params = {"category": category}
    else:
        query = """
        MATCH (t:Terminology)
        OPTIONAL MATCH (t)-[:RELATED_TO]-(related:Terminology)
        RETURN t {
            .*,
            relatedTerms: collect(DISTINCT related.name)
        } AS term
        ORDER BY t.category, t.name
        """
        params = {}

    records = _execute_read(query, params)
    return [record["term"] for record in records] if records else []


def get_terminology(name: str) -> Dict[str, object]:
    """
    获取单个术语的详细信息

    Args:
        name: 术语名称 (如 "FOB", "CIF")

    Returns:
        Terminology 详细信息
    """
    query = """
    MATCH (t:Terminology {name: $name})
    OPTIONAL MATCH (t)-[:RELATED_TO]-(related:Terminology)
    OPTIONAL MATCH (s:Stage)-[:HAS_TOPIC]->(k:KnowledgePoint)
           WHERE k.name CONTAINS t.name OR k.content CONTAINS t.name
    RETURN t {
        .*,
        relatedTerms: collect(DISTINCT related.name),
        usedInStages: collect(DISTINCT s.name)
    } AS term
    """

    records = _execute_read(query, {"name": name})
    if not records:
        raise GraphEntityNotFoundError(f"Terminology '{name}' not found")

    return records[0]["term"]


def link_knowledge_point_to_stage(knowledge_point_name: str, stage_name: str) -> Dict[str, object]:
    """
    将知识点关联到某个流程阶段

    创建 (Stage)-[:HAS_TOPIC]->(KnowledgePoint) 关系

    Args:
        knowledge_point_name: 知识点名称
        stage_name: 阶段名称

    Returns:
        更新后的 Stage 信息
    """
    query = """
    MATCH (s:Stage {name: $stage_name})
    MATCH (k:KnowledgePoint {name: $knowledge_point_name})
    MERGE (s)-[r:HAS_TOPIC]->(k)
    ON CREATE SET r.createdAt = datetime()
    RETURN s {
        .*,
        topics: [(s)-[:HAS_TOPIC]->(topic) | topic.name]
    } AS stage
    """

    records = _execute_read(
        query,
        {
            "stage_name": stage_name,
            "knowledge_point_name": knowledge_point_name,
        },
    )

    if not records:
        raise GraphEntityNotFoundError(
            f"Failed to link: Stage '{stage_name}' or KnowledgePoint '{knowledge_point_name}' not found"
        )

    LOGGER.info(f"Linked KnowledgePoint '{knowledge_point_name}' to Stage '{stage_name}'")
    return records[0]["stage"]


def unlink_knowledge_point_from_stage(
    knowledge_point_name: str, stage_name: str
) -> Dict[str, object]:
    """
    移除知识点与流程阶段的关联

    删除 (Stage)-[:HAS_TOPIC]->(KnowledgePoint) 关系

    Args:
        knowledge_point_name: 知识点名称
        stage_name: 阶段名称

    Returns:
        更新后的 Stage 信息
    """
    query = """
    MATCH (s:Stage {name: $stage_name})-[r:HAS_TOPIC]->(k:KnowledgePoint {name: $knowledge_point_name})
    DELETE r
    RETURN s {
        .*,
        topics: [(s)-[:HAS_TOPIC]->(topic) | topic.name]
    } AS stage
    """

    records = _execute_write(
        query,
        {
            "stage_name": stage_name,
            "knowledge_point_name": knowledge_point_name,
        },
    )

    if not records:
        raise GraphEntityNotFoundError(
            f"Relationship not found between Stage '{stage_name}' and KnowledgePoint '{knowledge_point_name}'"
        )

    LOGGER.info(f"Unlinked KnowledgePoint '{knowledge_point_name}' from Stage '{stage_name}'")
    return records[0]["stage"]


def get_enhanced_graph_visualization(
    node_types: Optional[List[str]] = None,
    max_nodes: int = 100,
) -> Dict[str, object]:
    """
    获取增强的图谱可视化数据,支持多节点类型

    Args:
        node_types: 要包含的节点类型列表,如 ["Stage", "KnowledgePoint", "Terminology"]
                   如果为 None,则包含所有类型
        max_nodes: 最大节点数量限制

    Returns:
        {
            "nodes": [
                {
                    "id": str,
                    "label": str,
                    "type": str,  # "Stage", "Skill", "Terminology", "KnowledgePoint"
                    "properties": {...},
                    "group": str,  # 用于前端分组着色
                }
            ],
            "edges": [
                {
                    "from": str,
                    "to": str,
                    "label": str,  # "PRECEDES", "HAS_TOPIC", "REQUIRES", etc.
                    "type": str,
                }
            ],
            "statistics": {
                "nodesByType": {...},
                "edgesByType": {...},
            }
        }
    """
    # 默认包含所有节点类型
    if node_types is None:
        node_types = ["Stage", "Skill", "Terminology", "KnowledgePoint"]

    # 构建节点类型过滤条件
    labels_condition = " OR ".join([f'"{label}" IN labels(n)' for label in node_types])

    # 查询节点
    nodes_query = f"""
    MATCH (n)
    WHERE {labels_condition}
    RETURN n,
           labels(n) AS labels,
           id(n) AS id
    LIMIT $max_nodes
    """

    nodes_records = _execute_read(nodes_query, {"max_nodes": max_nodes})
    nodes = []
    node_ids = set()

    for record in nodes_records:
        node_data = record["n"]
        node_labels = record["labels"]
        internal_id = record["id"]

        # 确定主要标签 (优先级: Stage > Skill > Terminology > KnowledgePoint)
        primary_label = _select_primary_label_for_visualization(node_labels)

        # 提取节点标识符
        node_id = _extract_node_identifier_for_visualization(primary_label, node_data)

        nodes.append({
            "id": node_id,
            "label": node_data.get("name") or node_id,
            "type": primary_label,
            "properties": dict(node_data),
            "group": primary_label,  # 前端用于分组着色
            "internalId": internal_id,
        })

        node_ids.add(internal_id)

    # 查询这些节点之间的关系
    edges_query = """
    MATCH (a)-[r]->(b)
    WHERE id(a) IN $node_ids AND id(b) IN $node_ids
    RETURN id(a) AS fromId,
           id(b) AS toId,
           type(r) AS relType,
           properties(r) AS relProps,
           labels(a) AS fromLabels,
           labels(b) AS toLabels,
           a AS fromNode,
           b AS toNode
    """

    edges_records = _execute_read(edges_query, {"node_ids": list(node_ids)})
    edges = []

    for record in edges_records:
        from_labels = record["fromLabels"]
        to_labels = record["toLabels"]
        from_primary = _select_primary_label_for_visualization(from_labels)
        to_primary = _select_primary_label_for_visualization(to_labels)

        from_id = _extract_node_identifier_for_visualization(from_primary, record["fromNode"])
        to_id = _extract_node_identifier_for_visualization(to_primary, record["toNode"])

        edges.append({
            "from": from_id,
            "to": to_id,
            "label": record["relType"],
            "type": record["relType"],
            "properties": dict(record["relProps"]) if record["relProps"] else {},
        })

    # 统计信息
    nodes_by_type = {}
    for node in nodes:
        node_type = node["type"]
        nodes_by_type[node_type] = nodes_by_type.get(node_type, 0) + 1

    edges_by_type = {}
    for edge in edges:
        edge_type = edge["type"]
        edges_by_type[edge_type] = edges_by_type.get(edge_type, 0) + 1

    return {
        "nodes": nodes,
        "edges": edges,
        "statistics": {
            "totalNodes": len(nodes),
            "totalEdges": len(edges),
            "nodesByType": nodes_by_type,
            "edgesByType": edges_by_type,
        },
    }


def _select_primary_label_for_visualization(labels: List[str]) -> str:
    """
    从节点的多个标签中选择主要标签

    优先级: Stage > Skill > Terminology > KnowledgePoint > 其他
    """
    priority = ["Stage", "Skill", "Terminology", "KnowledgePoint"]

    for label in priority:
        if label in labels:
            return label

    return labels[0] if labels else "Unknown"


def _extract_node_identifier_for_visualization(primary_label: str, node_data: Dict) -> str:
    """
    从节点数据中提取标识符

    不同节点类型使用不同的标识字段
    """
    if primary_label in ["Stage", "Skill", "Terminology", "KnowledgePoint"]:
        return node_data.get("name") or str(node_data.get("id", "unknown"))

    # 其他节点类型
    return str(node_data.get("id") or node_data.get("name", "unknown"))
